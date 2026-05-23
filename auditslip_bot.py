#!/usr/bin/env python3
"""Auditslip Telegram bot.

A single production-style Telegram bot that can use two OCR providers
(Gemini + OpenAI) through one provider router.
"""
from __future__ import annotations

import base64
import hashlib
import html
import json
import mimetypes
import os
import re
import sqlite3
import sys
import tempfile
import threading
import time
import traceback
from collections import defaultdict
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

APP_NAME = os.environ.get("BOT_DISPLAY_NAME", "Auditslip")
APP_DIR = Path(os.environ.get("AUDITSLIP_HOME", Path(__file__).resolve().parent))
DATA_DIR = Path(os.environ.get("AUDITSLIP_DATA_DIR", APP_DIR / "data"))
EXPORT_DIR = Path(os.environ.get("AUDITSLIP_EXPORT_DIR", APP_DIR / "exports"))
DB_PATH = Path(os.environ.get("AUDITSLIP_DB", DATA_DIR / "auditslip.db"))
BOT_TOKEN = os.environ.get("BOT_TOKEN") or os.environ.get("TELEGRAM_BOT_TOKEN", "")
TELEGRAM_API = os.environ.get("TELEGRAM_API", "https://api.telegram.org")

OCR_PROVIDERS = os.environ.get("OCR_PROVIDERS", "gemini,openai")
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY") or os.environ.get("GOOGLE_API_KEY", "")
GEMINI_MODEL = os.environ.get("GEMINI_MODEL", "gemini-2.5-flash")
GEMINI_FALLBACK_MODELS = [
    item.strip()
    for item in os.environ.get("GEMINI_FALLBACK_MODELS", "gemini-2.5-flash-lite,gemini-2.0-flash-lite").split(",")
    if item.strip()
]
OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY", "")
OPENAI_MODEL = os.environ.get("OPENAI_MODEL", "gpt-4o-mini")
OCR_RETRY_ATTEMPTS = max(1, int(os.environ.get("OCR_RETRY_ATTEMPTS", "3")))
OCR_RETRY_BASE_DELAY = float(os.environ.get("OCR_RETRY_BASE_DELAY", "2"))
POLL_TIMEOUT = int(os.environ.get("AUDITSLIP_POLL_TIMEOUT", "30"))
MAX_SLIPS_PER_POLL = max(1, int(os.environ.get("AUDITSLIP_MAX_SLIPS_PER_POLL", "100")))
OCR_WORKERS = max(1, int(os.environ.get("AUDITSLIP_OCR_WORKERS", "4")))
OCR_JOB_MAX_ATTEMPTS = max(1, int(os.environ.get("AUDITSLIP_OCR_JOB_MAX_ATTEMPTS", "3")))
OCR_JOB_STALE_MS = max(60_000, int(os.environ.get("AUDITSLIP_OCR_JOB_STALE_MS", "600000")))
OCR_WORKER_IDLE_SLEEP = float(os.environ.get("AUDITSLIP_OCR_WORKER_IDLE_SLEEP", "0.5"))
REPLY_ON_QUEUE = os.environ.get("AUDITSLIP_REPLY_ON_QUEUE", "0").strip().lower() in {"1", "true", "yes", "y"}
REPLY_ON_RESULT = os.environ.get("AUDITSLIP_REPLY_ON_RESULT", "1").strip().lower() in {"1", "true", "yes", "y"}
UNCLEAR_MIN_CONFIDENCE = float(os.environ.get("AUDITSLIP_UNCLEAR_MIN_CONFIDENCE", "0.65"))
ADMIN_IDS = {item.strip() for item in os.environ.get("AUDITSLIP_ADMIN_IDS", "").split(",") if item.strip()}
BKK = timezone(timedelta(hours=7))

COMMANDS = [
    {"command": "start", "description": "เริ่มใช้งาน Auditslip"},
    {"command": "help", "description": "ดูคำสั่งทั้งหมด"},
    {"command": "summary", "description": "สรุปยอด (/summary today|open|all|DD/MM/YY)"},
    {"command": "today", "description": "สรุปยอดวันนี้"},
    {"command": "daily", "description": "สรุปยอดแยกตามวัน"},
    {"command": "names", "description": "สรุปยอดแยกตามชื่อผู้โอน"},
    {"command": "userall", "description": "alias: สรุปยอดแยกตามชื่อผู้โอน"},
    {"command": "excel", "description": "ส่งออก Excel (/excel today|open|all|DD/MM/YY)"},
    {"command": "close", "description": "เคลียร์ยอด/ปิดรอบปัจจุบันแบบเก็บประวัติ"},
    {"command": "clear", "description": "ล้างข้อมูลห้องนี้แบบถาวร ต้อง /clear confirm"},
    {"command": "queue", "description": "ดูคิว fail/อ่านไม่ชัด"},
    {"command": "failed", "description": "ดูรายการ fail/unclear ที่รอ reprocess"},
    {"command": "reprocess", "description": "ประมวลผลรายการ fail/unclear ใหม่"},
    {"command": "recent", "description": "ดูรายการล่าสุด"},
    {"command": "stats", "description": "ดูสถิติบอทในห้องนี้"},
    {"command": "dupes", "description": "ดูรายการซ้ำ"},
    {"command": "providers", "description": "ดูสถานะ OCR providers"},
    {"command": "usage", "description": "สรุปการใช้งาน OCR API"},
]

SCHEMA = """
PRAGMA journal_mode=WAL;
CREATE TABLE IF NOT EXISTS slips (
  id TEXT PRIMARY KEY,
  update_id INTEGER,
  bot_key TEXT NOT NULL DEFAULT 'default',
  company_name TEXT,
  chat_id TEXT NOT NULL,
  chat_title TEXT,
  user_id TEXT,
  username TEXT,
  sender_name TEXT,
  message_id INTEGER,
  file_id TEXT,
  caption TEXT,
  status TEXT NOT NULL DEFAULT 'success',
  error TEXT,
  slip_date_display TEXT,
  slip_date_iso TEXT,
  slip_time TEXT,
  issuer_bank TEXT,
  seq TEXT,
  location TEXT,
  transaction_type TEXT,
  transferor_name TEXT,
  recipient_name TEXT,
  from_bank TEXT,
  from_account TEXT,
  to_bank TEXT,
  to_account TEXT,
  account_name TEXT,
  amount REAL DEFAULT 0,
  fee REAL DEFAULT 0,
  reference_no TEXT,
  aid TEXT,
  label TEXT,
  raw_text TEXT,
  confidence REAL DEFAULT 0,
  ocr_provider TEXT,
  ocr_model TEXT,
  is_duplicate INTEGER DEFAULT 0,
  duplicate_of TEXT,
  settlement_id TEXT,
  created_at INTEGER NOT NULL,
  created_at_iso TEXT NOT NULL,
  raw_json TEXT
);
CREATE INDEX IF NOT EXISTS idx_slips_chat_date ON slips(chat_id, slip_date_iso, slip_date_display);
CREATE INDEX IF NOT EXISTS idx_slips_chat_status ON slips(chat_id, status);
CREATE INDEX IF NOT EXISTS idx_slips_chat_transferor ON slips(chat_id, transferor_name);
CREATE INDEX IF NOT EXISTS idx_slips_settlement ON slips(chat_id, settlement_id);
CREATE INDEX IF NOT EXISTS idx_slips_update ON slips(update_id);
CREATE TABLE IF NOT EXISTS processed_updates (
  bot_key TEXT NOT NULL DEFAULT 'default',
  update_id INTEGER NOT NULL,
  processed_at INTEGER NOT NULL,
  PRIMARY KEY(bot_key, update_id)
);
CREATE TABLE IF NOT EXISTS bot_state (
  key TEXT PRIMARY KEY,
  value TEXT,
  updated_at INTEGER NOT NULL
);
CREATE TABLE IF NOT EXISTS settlements (
  settlement_id TEXT PRIMARY KEY,
  bot_key TEXT NOT NULL DEFAULT 'default',
  company_name TEXT,
  chat_id TEXT NOT NULL,
  closed_by TEXT,
  note TEXT,
  closed_at INTEGER NOT NULL,
  closed_at_iso TEXT NOT NULL,
  total_amount REAL DEFAULT 0,
  total_fee REAL DEFAULT 0,
  closed_count INTEGER DEFAULT 0,
  excel_path TEXT,
  raw_json TEXT
);
CREATE TABLE IF NOT EXISTS ocr_jobs (
  job_id TEXT PRIMARY KEY,
  slip_id TEXT NOT NULL,
  update_id INTEGER,
  bot_key TEXT NOT NULL DEFAULT 'default',
  company_name TEXT,
  chat_id TEXT NOT NULL,
  chat_title TEXT,
  user_id TEXT,
  username TEXT,
  sender_name TEXT,
  message_id INTEGER,
  file_id TEXT NOT NULL,
  caption TEXT,
  mime TEXT,
  status TEXT NOT NULL DEFAULT 'queued',
  attempts INTEGER NOT NULL DEFAULT 0,
  max_attempts INTEGER NOT NULL DEFAULT 3,
  locked_by TEXT,
  locked_at INTEGER,
  next_run_at INTEGER NOT NULL DEFAULT 0,
  created_at INTEGER NOT NULL,
  updated_at INTEGER NOT NULL,
  error TEXT,
  raw_json TEXT
);
CREATE INDEX IF NOT EXISTS idx_ocr_jobs_status_next ON ocr_jobs(status, next_run_at, created_at);
CREATE INDEX IF NOT EXISTS idx_ocr_jobs_chat_status ON ocr_jobs(chat_id, status);
CREATE TABLE IF NOT EXISTS company_accounts (
  bot_key TEXT NOT NULL DEFAULT 'default',
  chat_id TEXT NOT NULL,
  account_key TEXT NOT NULL,
  company_name TEXT NOT NULL,
  bank TEXT,
  account_no TEXT,
  account_name TEXT,
  daily_limit REAL DEFAULT 0,
  active INTEGER DEFAULT 1,
  updated_at INTEGER NOT NULL,
  PRIMARY KEY(bot_key, chat_id, account_key)
);
"""

SLIP_FIELDS = [
    "slip_date_display",
    "slip_date_iso",
    "slip_time",
    "issuer_bank",
    "seq",
    "location",
    "transaction_type",
    "transferor_name",
    "recipient_name",
    "from_bank",
    "from_account",
    "to_bank",
    "to_account",
    "account_name",
    "amount",
    "fee",
    "reference_no",
    "aid",
    "label",
    "raw_text",
    "confidence",
    "ocr_provider",
    "ocr_model",
]


def log(*args: Any) -> None:
    print(datetime.now(BKK).isoformat(), *args, flush=True)


def now_ms() -> int:
    return int(time.time() * 1000)


def bkk_now() -> datetime:
    return datetime.now(BKK)


def ensure_dirs() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)


def h(value: Any) -> str:
    return html.escape("" if value is None else str(value))


def fmt_money(value: Any) -> str:
    try:
        return f"{float(value):,.2f}"
    except Exception:
        return "0.00"


def parse_number(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").strip()
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(match.group(0)) if match else 0.0


def clean_text(value: Any) -> str:
    return " ".join(str(value or "").strip().split())


BANK_ALIASES = {
    "KRUNGTHAI": {"krungthai", "krungthaibank", "ktb", "ktbnext", "กรุงไทย", "ธกรุงไทย", "ธนาคารกรุงไทย"},
    "KBANK": {"kbank", "kasikorn", "kasikornbank", "kasikornthai", "กสิกร", "กสิกรไทย", "ธกสิกรไทย", "ธนาคารกสิกรไทย"},
    "SCB": {"scb", "scbeasy", "siamcommercialbank", "siamcommercial", "ไทยพาณิชย์", "ธไทยพาณิชย์", "ธนาคารไทยพาณิชย์"},
    "BANGKOK BANK": {"bangkokbank", "bangkok", "bbl", "bualuang", "กรุงเทพ", "ธกรุงเทพ", "ธนาคารกรุงเทพ"},
    "GSB": {"gsb", "governmentsavingsbank", "governmentsavings", "ออมสิน", "ธออมสิน", "ธนาคารออมสิน"},
    "TTB": {"ttb", "tmb", "tmbthanachart", "ttbtmbthanachart", "ทีเอ็มบีธนชาต", "ทหารไทยธนชาต"},
    "BAAC": {"baac", "ธกส", "ธกสธนาคารเพื่อการเกษตรและสหกรณ์การเกษตร", "เพื่อการเกษตร"},
    "KRUNGSRI": {"krungsri", "bay", "bankofayudhya", "กรุงศรี", "อยุธยา", "กรุงศรีอยุธยา", "ธกรุงศรี", "ธนาคารกรุงศรี"},
    "KKP": {"kkp", "kiatnakinphatra", "เกียรตินาคินภัทร", "เกียรตินาคิน"},
    "UOB": {"uob", "uobtmrw", "ยูโอบี", "ธยูโอบี", "ธนาคารยูโอบี"},
    "CIMB": {"cimb", "cimbthai", "ซีไอเอ็มบี", "ซีไอเอ็มบีไทย", "ธนาคารซีไอเอ็มบีไทย"},
    "LH BANK": {"lhbank", "landandhouse", "landandhouses", "แลนด์แอนด์เฮ้าส์", "ธนาคารแลนด์แอนด์เฮ้าส์"},
    "GHB": {"ghb", "governmenthousingbank", "ธอส", "อาคารสงเคราะห์", "ธนาคารอาคารสงเคราะห์"},
    "THAI CREDIT": {"thaicredit", "ไทยเครดิต", "ธนาคารไทยเครดิต"},
    "TISCO": {"tisco", "ทิสโก้", "ทิสโก", "ธนาคารทิสโก้"},
    "ICBC": {"icbc", "icbcthai", "ไอซีบีซี", "ธนาคารไอซีบีซี"},
    "STANDARD CHARTERED": {"standardchartered", "standardcharteredbank", "สแตนดาร์ดชาร์เตอร์ด"},
}
MISSING_BANK_VALUES = {"", "unknown", "unknownbank", "n/a", "na", "none", "null", "-", "ไม่ทราบ", "xxx", "xxxx", "xxxbank", "masked"}


def bank_key(value: Any) -> str:
    return re.sub(r"[\s\*\u200b\u200c\u200d.\-_/|(),]+", "", clean_text(value).lower())


def bank_is_missing(value: Any) -> bool:
    text = clean_text(value)
    key = bank_key(text)
    return (not key) or key in MISSING_BANK_VALUES or bool(re.fullmatch(r"x+(?:bank)?", key, flags=re.I)) or "ไม่ทราบ" in text


def canonical_bank(value: Any) -> str:
    bank = clean_text(value)
    key = bank_key(bank)
    if bank_is_missing(bank):
        return ""
    for canonical, aliases in BANK_ALIASES.items():
        normalized_aliases = {bank_key(alias) for alias in aliases}
        if key in normalized_aliases or any(alias and ((len(alias) >= 3 and key.startswith(alias)) or (len(alias) >= 4 and alias in key)) for alias in normalized_aliases):
            return canonical
    return bank.upper() if re.fullmatch(r"[A-Za-z0-9 ._-]+", bank) else bank


def stable_id(*parts: Any, prefix: str = "SLIP") -> str:
    raw = "|".join(str(p or "") for p in parts)
    return f"{prefix}-{hashlib.sha256(raw.encode()).hexdigest()[:16]}"


def safe_bot_key(value: Any) -> str:
    key = re.sub(r"[^A-Za-z0-9_.-]+", "-", clean_text(value)).strip("-._")
    return key or "default"


def telegram_bot_configs() -> List[Dict[str, str]]:
    """Parse multi-bot config without exposing tokens to the dashboard.

    AUDITSLIP_TELEGRAM_BOTS supports either JSON list objects or CSV entries:
    bot1:BOT_TOKEN_1:บริษัท 1,bot2:BOT_TOKEN_2:บริษัท 2
    The second field is treated as an env-var name when it exists in os.environ.
    """
    raw = os.environ.get("AUDITSLIP_TELEGRAM_BOTS", "").strip()
    configs: List[Dict[str, str]] = []
    if raw:
        if raw[0] in "[{":
            parsed = json.loads(raw)
            items = parsed.values() if isinstance(parsed, dict) else parsed
            for idx, item in enumerate(items, start=1):
                if not isinstance(item, dict):
                    continue
                bot_key = safe_bot_key(item.get("bot_key") or item.get("key") or f"bot{idx}")
                token_env = clean_text(item.get("token_env") or "")
                token = clean_text(os.environ.get(token_env, "") if token_env else item.get("token") or "")
                company_name = clean_text(item.get("company_name") or item.get("company") or bot_key)
                if token:
                    configs.append({"bot_key": bot_key, "token_env": token_env, "token": token, "company_name": company_name})
        else:
            for idx, part in enumerate(raw.split(","), start=1):
                part = part.strip()
                if not part:
                    continue
                pieces = part.split(":", 2)
                bot_key = safe_bot_key(pieces[0] if len(pieces) >= 1 else f"bot{idx}")
                token_ref = clean_text(pieces[1] if len(pieces) >= 2 else "")
                company_name = clean_text(pieces[2] if len(pieces) >= 3 else bot_key)
                token = clean_text(os.environ.get(token_ref, "")) if token_ref else ""
                if not token and token_ref and token_ref.startswith("bot"):
                    token = token_ref
                if token:
                    configs.append({"bot_key": bot_key, "token_env": token_ref if token_ref in os.environ else "", "token": token, "company_name": company_name})
    if not configs and BOT_TOKEN:
        configs.append({"bot_key": "default", "token_env": "BOT_TOKEN" if os.environ.get("BOT_TOKEN") else "TELEGRAM_BOT_TOKEN", "token": BOT_TOKEN, "company_name": APP_NAME})
    return configs


def ocr_provider_candidates(
    provider_string: Optional[str] = None,
    gemini_key: Optional[str] = None,
    openai_key: Optional[str] = None,
) -> List[str]:
    raw = provider_string if provider_string is not None else OCR_PROVIDERS
    gkey = GEMINI_API_KEY if gemini_key is None else gemini_key
    okey = OPENAI_API_KEY if openai_key is None else openai_key
    candidates: List[str] = []
    for item in raw.split(","):
        provider = item.strip().lower()
        if not provider or provider in candidates:
            continue
        if provider == "gemini" and gkey:
            candidates.append(provider)
        elif provider == "openai" and okey:
            candidates.append(provider)
    return candidates


def provider_status() -> List[Dict[str, Any]]:
    configured = [p.strip().lower() for p in OCR_PROVIDERS.split(",") if p.strip()]
    active = set(ocr_provider_candidates())
    return [
        {
            "provider": p,
            "active": p in active,
            "model": GEMINI_MODEL if p == "gemini" else OPENAI_MODEL if p == "openai" else "",
            "has_key": bool(GEMINI_API_KEY if p == "gemini" else OPENAI_API_KEY if p == "openai" else ""),
        }
        for p in configured
    ]


def gemini_model_candidates() -> List[str]:
    models: List[str] = []
    for model in [GEMINI_MODEL, *GEMINI_FALLBACK_MODELS]:
        model = model.strip()
        if model and model not in models:
            models.append(model)
    return models


def is_transient_status(status_code: int, body: str = "") -> bool:
    text = body.lower()
    return status_code in {408, 409, 425, 429, 500, 502, 503, 504} or any(
        token in text for token in ["temporar", "timeout", "try again", "rate", "quota", "overload", "high demand"]
    )


def ocr_prompt() -> str:
    return """
You are Auditslip OCR. Extract a Thai bank transfer / ATM transaction slip into JSON only.
Return one JSON object with these keys:
{
  "slip_date_display": "DD/MM/YY or visible date",
  "slip_date_iso": "YYYY-MM-DD if inferable, else empty",
  "slip_time": "HH:MM or HH:MM:SS",
  "issuer_bank": "bank/issuer printed on receipt",
  "seq": "SEQ or sequence number",
  "location": "ATM/location/branch if visible",
  "transaction_type": "transfer/payment/withdrawal/deposit/etc",
  "transferor_name": "name of payer / sender / ผู้โอน if visible",
  "recipient_name": "name of receiver / recipient / ผู้รับ if visible",
  "from_bank": "source bank if visible",
  "from_account": "source account/card number masked as visible",
  "to_bank": "destination bank if visible",
  "to_account": "destination account masked as visible",
  "account_name": "merchant/account name if visible",
  "amount": number,
  "fee": number,
  "reference_no": "transaction/reference number if visible",
  "aid": "AID/terminal reference if visible",
  "label": "short human label",
  "raw_text": "important OCR text lines joined with newline",
  "confidence": number between 0 and 1
}
Rules:
- JSON only. No markdown.
- Preserve Thai names exactly.
- Preserve slip_date_display exactly as printed/visible on the slip; use slip_date_iso only as a normalized helper when confidently inferable.
- Do not invent names or numbers. If missing, use empty string or 0.
- For Thai ATM dates like 26/05/09, infer Buddhist/AD carefully from surrounding context; if unsure preserve display and leave iso empty.
""".strip()


def parse_json_from_text(text: str) -> Dict[str, Any]:
    text = text.strip()
    if not text:
        raise ValueError("empty OCR response")
    try:
        return json.loads(text)
    except Exception:
        pass
    match = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", text, flags=re.S)
    if match:
        return json.loads(match.group(1))
    start = text.find("{")
    end = text.rfind("}")
    if start >= 0 and end > start:
        return json.loads(text[start : end + 1])
    raise ValueError("OCR response did not contain JSON")


THAI_MONTHS = {
    "มค": 1,
    "มกราคม": 1,
    "กพ": 2,
    "กุมภาพันธ์": 2,
    "มีค": 3,
    "มีนาคม": 3,
    "เมย": 4,
    "เมษายน": 4,
    "พค": 5,
    "พฤษภาคม": 5,
    "มิย": 6,
    "มิถุนายน": 6,
    "กค": 7,
    "กรกฎาคม": 7,
    "สค": 8,
    "สิงหาคม": 8,
    "กย": 9,
    "กันยายน": 9,
    "ตค": 10,
    "ตุลาคม": 10,
    "พย": 11,
    "พฤศจิกายน": 11,
    "ธค": 12,
    "ธันวาคม": 12,
}


def _normalize_year(year: int) -> int:
    if year < 100:
        year += 2500 if year >= 40 else 2000
    if year > 2400:
        year -= 543
    return year


def _format_date_parts(day: int, month: int, year: int, raw: str) -> Tuple[str, str]:
    try:
        year = _normalize_year(year)
        if not (1 <= month <= 12 and 1 <= day <= 31):
            return raw, ""
        display = f"{day:02d}/{month:02d}/{str(year)[-2:]}"
        return display, f"{year:04d}-{month:02d}-{day:02d}"
    except Exception:
        return raw, ""


def normalize_date_parts(value: Any) -> Tuple[str, str]:
    raw = clean_text(value)
    if not raw:
        return "", ""
    normalized = raw.replace("-", "/").replace(".", "/")
    match = re.search(r"(\d{1,4})/(\d{1,2})/(\d{1,4})", normalized)
    if match:
        a, b, c = [int(x) for x in match.groups()]
        if a > 1900:
            year, month, day = a, b, c
        else:
            day, month, year = a, b, c
        return _format_date_parts(day, month, year, raw)

    thai_match = re.search(r"(\d{1,2})\s*([^\d]+?)\s*(\d{2,4})", normalized)
    if thai_match:
        day = int(thai_match.group(1))
        month_key = re.sub(r"[^ก-๙A-Za-z]", "", thai_match.group(2)).lower()
        month = THAI_MONTHS.get(month_key)
        if month:
            year = int(thai_match.group(3))
            return _format_date_parts(day, month, year, raw)
    return raw, ""


def scope_to_date(scope: str) -> Tuple[str, str]:
    s = (scope or "open").strip().lower()
    if s in {"", "open", "current"}:
        return "open", "รอบที่ยังไม่ปิด"
    if s in {"today", "วันนี้"}:
        return bkk_now().strftime("%Y-%m-%d"), "วันนี้"
    if s in {"all", "ทั้งหมด"}:
        return "all", "ทั้งหมด"
    display, iso = normalize_date_parts(s)
    if iso:
        return iso, display
    return s, s


def normalize_record(data: Dict[str, Any]) -> Dict[str, Any]:
    aliases = {
        "transferor_name": ["transferor_name", "payer_name", "sender_name", "from_name", "ผู้โอน"],
        "recipient_name": ["recipient_name", "receiver_name", "to_name", "payee_name", "ผู้รับ"],
        "reference_no": ["reference_no", "reference", "ref", "transaction_id", "transaction_ref"],
        "slip_date_display": ["slip_date_display", "date", "transaction_date"],
        "slip_time": ["slip_time", "time", "transaction_time"],
    }
    out: Dict[str, Any] = {}
    for field in SLIP_FIELDS:
        keys = aliases.get(field, [field])
        value = ""
        for key in keys:
            if key in data and data.get(key) not in (None, ""):
                value = data.get(key)
                break
        out[field] = value
    raw_display = clean_text(out.get("slip_date_display"))
    raw_iso = clean_text(data.get("slip_date_iso"))
    display_from_display, iso_from_display = normalize_date_parts(raw_display)
    display_from_iso, iso_from_iso = normalize_date_parts(raw_iso)
    # If the model returns both a visible slip date and a normalized ISO date,
    # prefer the visible date when they disagree. Thai slips often show short
    # years like 22/05/26; a model can hallucinate 2022-05-22 as ISO while the
    # visible date clearly means 2026-05-22 in this dataset.
    iso = iso_from_display or iso_from_iso
    out["slip_date_display"] = raw_display or display_from_iso or display_from_display
    out["slip_date_iso"] = iso
    out["slip_time"] = clean_text(out.get("slip_time"))
    for field in [
        "issuer_bank",
        "seq",
        "location",
        "transaction_type",
        "transferor_name",
        "recipient_name",
        "from_bank",
        "from_account",
        "to_bank",
        "to_account",
        "account_name",
        "reference_no",
        "aid",
        "label",
        "raw_text",
        "ocr_provider",
        "ocr_model",
    ]:
        out[field] = clean_text(out.get(field))
    for field in ["issuer_bank", "from_bank", "to_bank"]:
        out[field] = canonical_bank(out.get(field))
    out["amount"] = parse_number(out.get("amount"))
    out["fee"] = parse_number(out.get("fee"))
    conf = parse_number(out.get("confidence"))
    out["confidence"] = max(0.0, min(1.0, conf if conf else 0.0))
    return out


def unclear_reason(data: Dict[str, Any]) -> str:
    if parse_number(data.get("amount")) <= 0:
        return "missing amount"
    if not (data.get("slip_date_iso") or data.get("slip_date_display")):
        return "missing date"
    if parse_number(data.get("confidence")) and parse_number(data.get("confidence")) < UNCLEAR_MIN_CONFIDENCE:
        return f"low confidence {parse_number(data.get('confidence')):.2f}"
    return ""


def mime_for_path(path: Path) -> str:
    mime, _ = mimetypes.guess_type(str(path))
    return mime or "image/jpeg"


def is_image_message(msg: Optional[Dict[str, Any]]) -> bool:
    if not msg:
        return False
    if msg.get("photo"):
        return True
    doc = msg.get("document") or {}
    return str(doc.get("mime_type", "")).startswith("image/")


def _gemini_response_text(obj: Any) -> str:
    """Extract text from a Gemini generateContent response defensively.
    The API occasionally nests a list where a dict is expected (or vice versa)."""
    if isinstance(obj, list):
        obj = obj[0] if obj else {}
    if not isinstance(obj, dict):
        return ""
    candidates = obj.get("candidates") or []
    if not isinstance(candidates, list) or not candidates:
        return ""
    candidate = candidates[0]
    if not isinstance(candidate, dict):
        return ""
    content = candidate.get("content") or {}
    if isinstance(content, list):
        content = content[0] if content else {}
    if not isinstance(content, dict):
        return ""
    parts = content.get("parts") or []
    if not isinstance(parts, list):
        return ""
    texts = [str(p["text"]) for p in parts if isinstance(p, dict) and p.get("text")]
    return "".join(texts)


def gemini_extract(image_path: Path, mime: Optional[str] = None) -> Tuple[Dict[str, Any], Dict[str, str]]:
    if not GEMINI_API_KEY:
        raise RuntimeError("GEMINI_API_KEY missing")
    image_bytes = image_path.read_bytes()
    payload_base = {
        "contents": [
            {
                "role": "user",
                "parts": [
                    {"text": ocr_prompt()},
                    {"inline_data": {"mime_type": mime or mime_for_path(image_path), "data": base64.b64encode(image_bytes).decode()}},
                ],
            }
        ],
        "generationConfig": {"temperature": 0, "response_mime_type": "application/json"},
    }
    last_error = ""
    for model in gemini_model_candidates():
        url = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={GEMINI_API_KEY}"
        for attempt in range(1, OCR_RETRY_ATTEMPTS + 1):
            resp = requests.post(url, json=payload_base, timeout=75)
            if resp.status_code < 400:
                obj = resp.json()
                text = _gemini_response_text(obj)
                return parse_json_from_text(text), {"provider": "gemini", "model": model}
            last_error = f"gemini {model} HTTP {resp.status_code}: {resp.text[:300]}"
            if not is_transient_status(resp.status_code, resp.text):
                break
            if attempt < OCR_RETRY_ATTEMPTS:
                time.sleep(OCR_RETRY_BASE_DELAY * attempt)
    raise RuntimeError(last_error or "Gemini OCR failed")


def openai_extract(image_path: Path, mime: Optional[str] = None) -> Tuple[Dict[str, Any], Dict[str, str]]:
    if not OPENAI_API_KEY:
        raise RuntimeError("OPENAI_API_KEY missing")
    image_b64 = base64.b64encode(image_path.read_bytes()).decode()
    payload = {
        "model": OPENAI_MODEL,
        "temperature": 0,
        "response_format": {"type": "json_object"},
        "messages": [
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": ocr_prompt()},
                    {"type": "image_url", "image_url": {"url": f"data:{mime or mime_for_path(image_path)};base64,{image_b64}"}},
                ],
            }
        ],
    }
    last_error = ""
    for attempt in range(1, OCR_RETRY_ATTEMPTS + 1):
        resp = requests.post(
            "https://api.openai.com/v1/chat/completions",
            headers={"Authorization": f"Bearer {OPENAI_API_KEY}", "Content-Type": "application/json"},
            json=payload,
            timeout=90,
        )
        if resp.status_code < 400:
            obj = resp.json()
            text = obj.get("choices", [{}])[0].get("message", {}).get("content", "")
            return parse_json_from_text(text), {"provider": "openai", "model": OPENAI_MODEL}
        last_error = f"openai {OPENAI_MODEL} HTTP {resp.status_code}: {resp.text[:300]}"
        if not is_transient_status(resp.status_code, resp.text):
            break
        if attempt < OCR_RETRY_ATTEMPTS:
            time.sleep(OCR_RETRY_BASE_DELAY * attempt)
    raise RuntimeError(last_error or "OpenAI OCR failed")


def extract_with_provider(provider: str, image_path: Path, mime: Optional[str] = None) -> Tuple[Dict[str, Any], Dict[str, str]]:
    provider = provider.lower().strip()
    if provider == "gemini":
        return gemini_extract(image_path, mime)
    if provider == "openai":
        return openai_extract(image_path, mime)
    raise ValueError(f"unknown OCR provider: {provider}")


def ocr_extract(image_path: Path, mime: Optional[str] = None) -> Tuple[str, Dict[str, Any]]:
    errors: List[str] = []
    candidates = ocr_provider_candidates()
    if not candidates:
        raise RuntimeError("No OCR provider is active. Configure GEMINI_API_KEY and/or OPENAI_API_KEY.")
    for provider in candidates:
        try:
            data, meta = extract_with_provider(provider, image_path, mime)
            normalized = normalize_record(data)
            normalized["ocr_provider"] = meta.get("provider", provider)
            normalized["ocr_model"] = meta.get("model", "")
            normalized["raw_json"] = json.dumps(data, ensure_ascii=False)
            return provider, normalized
        except Exception as exc:
            errors.append(f"{provider}: {exc}")
            log("OCR provider failed", provider, str(exc)[:300])
    raise RuntimeError("; ".join(errors))


class AuditslipBot:
    def __init__(
        self,
        token: str = "",
        db_path: Path = DB_PATH,
        dry_run: bool = False,
        bot_key: str = "default",
        company_name: str = "",
        reply_on_result: Optional[bool] = None,
    ) -> None:
        self.token = token or BOT_TOKEN
        self.db_path = Path(db_path)
        self.dry_run = dry_run
        self.bot_key = safe_bot_key(bot_key)
        self.company_name = clean_text(company_name) or APP_NAME
        self.reply_on_result = REPLY_ON_RESULT if reply_on_result is None else bool(reply_on_result)
        self.api_base = f"{TELEGRAM_API}/bot{self.token}" if self.token else ""
        ensure_dirs()
        self.db_path.parent.mkdir(parents=True, exist_ok=True)

    def connect(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self.db_path, timeout=30.0)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA busy_timeout=30000;")
        conn.execute("PRAGMA journal_mode=WAL;")
        conn.execute("PRAGMA synchronous=NORMAL;")
        return conn

    def init_db(self) -> None:
        with self.connect() as conn:
            conn.executescript(SCHEMA)
            self.migrate_db(conn)
            conn.commit()

    @staticmethod
    def table_columns(conn: sqlite3.Connection, table: str) -> set[str]:
        return {str(row[1]) for row in conn.execute(f"PRAGMA table_info({table})").fetchall()}

    def migrate_db(self, conn: sqlite3.Connection) -> None:
        for table in ["slips", "ocr_jobs", "settlements"]:
            cols = self.table_columns(conn, table)
            if "bot_key" not in cols:
                conn.execute(f"ALTER TABLE {table} ADD COLUMN bot_key TEXT NOT NULL DEFAULT 'default'")
            if "company_name" not in cols:
                conn.execute(f"ALTER TABLE {table} ADD COLUMN company_name TEXT")
        cols = self.table_columns(conn, "processed_updates")
        if "bot_key" not in cols:
            legacy = f"processed_updates_legacy_{int(time.time())}"
            conn.execute(f"ALTER TABLE processed_updates RENAME TO {legacy}")
            conn.execute("""
                CREATE TABLE processed_updates (
                  bot_key TEXT NOT NULL DEFAULT 'default',
                  update_id INTEGER NOT NULL,
                  processed_at INTEGER NOT NULL,
                  PRIMARY KEY(bot_key, update_id)
                )
            """)
            conn.execute(f"INSERT OR IGNORE INTO processed_updates(bot_key, update_id, processed_at) SELECT 'default', update_id, processed_at FROM {legacy}")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_slips_bot_chat_date ON slips(bot_key, chat_id, slip_date_iso, slip_date_display)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_ocr_jobs_bot_status_next ON ocr_jobs(bot_key, status, next_run_at, created_at)")

    def telegram_call(self, method: str, data: Optional[Dict[str, Any]] = None, files: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        if self.dry_run:
            return {"ok": True, "result": {}}
        if not self.token:
            raise RuntimeError("BOT_TOKEN missing")
        try:
            resp = requests.post(f"{self.api_base}/{method}", data=data, files=files, timeout=90)
            try:
                obj = resp.json()
            except ValueError:
                obj = {"ok": False, "description": resp.text}
        except (requests.ConnectionError, requests.Timeout, requests.RequestException) as exc:
            log(f"telegram network error: {exc}")
            raise RuntimeError(f"Telegram {method} network error: {exc}") from exc
        if resp.status_code >= 400 or not obj.get("ok", False):
            raise RuntimeError(f"Telegram {method} failed: {resp.status_code} {obj}")
        return obj

    def telegram_get(self, method: str, params: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        if self.dry_run:
            return {"ok": True, "result": {}}
        try:
            resp = requests.get(f"{self.api_base}/{method}", params=params, timeout=90)
            resp.raise_for_status()
            obj = resp.json()
        except (requests.ConnectionError, requests.Timeout, requests.RequestException, ValueError) as exc:
            log(f"telegram network error: {exc}")
            raise RuntimeError(f"Telegram {method} network error: {exc}") from exc
        if not obj.get("ok", False):
            raise RuntimeError(f"Telegram {method} failed: {resp.status_code} {obj}")
        return obj

    def clear_webhook_for_polling(self) -> None:
        # Polling-mode bot: stale webhooks make getUpdates fail.
        self.telegram_call("deleteWebhook", {"drop_pending_updates": "false"})

    def set_commands(self) -> None:
        self.telegram_call("setMyCommands", {"commands": json.dumps(COMMANDS, ensure_ascii=False)})

    def reply(self, chat_id: Any, text: str, reply_to_message_id: Optional[int] = None) -> None:
        payload: Dict[str, Any] = {"chat_id": chat_id, "text": text, "parse_mode": "HTML", "disable_web_page_preview": True}
        if reply_to_message_id:
            payload["reply_to_message_id"] = reply_to_message_id
        self.telegram_call("sendMessage", payload)

    def send_document(self, chat_id: Any, path: Path, caption: str = "") -> None:
        with open(path, "rb") as fh:
            self.telegram_call("sendDocument", {"chat_id": chat_id, "caption": caption, "parse_mode": "HTML"}, {"document": fh})

    def state_key(self, key: str) -> str:
        return key if self.bot_key == "default" else f"{key}:{self.bot_key}"

    def get_offset(self) -> int:
        with self.connect() as conn:
            row = conn.execute("SELECT value FROM bot_state WHERE key=?", (self.state_key("offset"),)).fetchone()
            return int(row["value"]) if row and row["value"] else 0

    def persist_offset(self, offset: int) -> None:
        with self.connect() as conn:
            conn.execute(
                "INSERT INTO bot_state(key,value,updated_at) VALUES(?,?,?) ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at",
                (self.state_key("offset"), str(offset), now_ms()),
            )
            conn.commit()

    def mark_processed(self, update_id: int) -> None:
        with self.connect() as conn:
            conn.execute("INSERT OR IGNORE INTO processed_updates(bot_key, update_id, processed_at) VALUES (?,?,?)", (self.bot_key, update_id, now_ms()))
            conn.commit()

    def already_processed(self, update_id: int) -> bool:
        with self.connect() as conn:
            row = conn.execute("SELECT 1 FROM processed_updates WHERE bot_key=? AND update_id=?", (self.bot_key, update_id)).fetchone()
            return bool(row)

    def save_slip(self, row: Dict[str, Any]) -> str:
        normalized = normalize_record(row)
        merged = {**row, **normalized}
        ts = int(merged.get("created_at") or now_ms())
        if not merged.get("bot_key"):
            merged["bot_key"] = self.bot_key
        if not merged.get("company_name"):
            merged["company_name"] = self.company_name
        merged.setdefault("id", stable_id(merged.get("bot_key"), merged.get("chat_id"), merged.get("message_id"), merged.get("file_id")))
        merged.setdefault("created_at", ts)
        merged.setdefault("created_at_iso", datetime.fromtimestamp(ts / 1000, BKK).isoformat())
        merged.setdefault("status", "success")
        merged.setdefault("raw_json", json.dumps(row, ensure_ascii=False))
        if merged.get("status") == "success" and not merged.get("is_duplicate"):
            duplicate_of = self.find_duplicate(merged)
            if duplicate_of:
                merged["is_duplicate"] = 1
                merged["duplicate_of"] = duplicate_of
        columns = [
            "id",
            "update_id",
            "bot_key",
            "company_name",
            "chat_id",
            "chat_title",
            "user_id",
            "username",
            "sender_name",
            "message_id",
            "file_id",
            "caption",
            "status",
            "error",
            *SLIP_FIELDS,
            "is_duplicate",
            "duplicate_of",
            "settlement_id",
            "created_at",
            "created_at_iso",
            "raw_json",
        ]
        values = [merged.get(col) for col in columns]
        placeholders = ",".join("?" for _ in columns)
        updates = ",".join(f"{col}=excluded.{col}" for col in columns if col != "id")
        with self.connect() as conn:
            conn.execute(
                f"INSERT INTO slips({','.join(columns)}) VALUES ({placeholders}) ON CONFLICT(id) DO UPDATE SET {updates}",
                values,
            )
            conn.commit()
        return str(merged["id"])

    def enqueue_ocr_job(self, base_row: Dict[str, Any], mime: str = "image/jpeg") -> str:
        """Persist a queued slip + OCR job without running OCR in the Telegram polling loop."""
        ts = now_ms()
        base_row = {**base_row, "bot_key": base_row.get("bot_key") or self.bot_key, "company_name": base_row.get("company_name") or self.company_name}
        slip_id = str(base_row.get("id") or stable_id(base_row.get("bot_key"), base_row.get("chat_id"), base_row.get("message_id"), base_row.get("file_id")))
        job_id = stable_id("JOB", base_row.get("bot_key"), slip_id, base_row.get("file_id"), prefix="JOB")
        queued_row = {
            **base_row,
            "id": slip_id,
            "status": "queued",
            "error": "waiting for OCR worker",
            "created_at": ts,
            "created_at_iso": datetime.fromtimestamp(ts / 1000, BKK).isoformat(),
        }
        self.save_slip(queued_row)
        with self.connect() as conn:
            conn.execute(
                """
                INSERT INTO ocr_jobs(
                  job_id, slip_id, update_id, bot_key, company_name, chat_id, chat_title, user_id, username, sender_name,
                  message_id, file_id, caption, mime, status, attempts, max_attempts,
                  next_run_at, created_at, updated_at, raw_json
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                ON CONFLICT(job_id) DO UPDATE SET
                  status=CASE WHEN ocr_jobs.status IN ('done','processing') THEN ocr_jobs.status ELSE 'queued' END,
                  updated_at=excluded.updated_at,
                  error=NULL
                """,
                (
                    job_id,
                    slip_id,
                    base_row.get("update_id"),
                    base_row.get("bot_key"),
                    base_row.get("company_name"),
                    str(base_row.get("chat_id")),
                    base_row.get("chat_title"),
                    base_row.get("user_id"),
                    base_row.get("username"),
                    base_row.get("sender_name"),
                    base_row.get("message_id"),
                    base_row.get("file_id"),
                    base_row.get("caption"),
                    mime,
                    "queued",
                    0,
                    OCR_JOB_MAX_ATTEMPTS,
                    0,
                    ts,
                    ts,
                    json.dumps(base_row, ensure_ascii=False),
                ),
            )
            conn.commit()
        return job_id

    def requeue_stale_ocr_jobs(self) -> int:
        cutoff = now_ms() - OCR_JOB_STALE_MS
        with self.connect() as conn:
            count = conn.execute(
                """
                UPDATE ocr_jobs
                SET status='queued', locked_by=NULL, locked_at=NULL, updated_at=?
                WHERE status='processing' AND COALESCE(locked_at,0) < ? AND COALESCE(bot_key,'default')=?
                """,
                (now_ms(), cutoff, self.bot_key),
            ).rowcount
            conn.commit()
            return int(count or 0)

    def resume_incomplete_ocr_jobs(self) -> int:
        """Make in-flight OCR jobs claimable after process/VPS restart."""
        ts = now_ms()
        with self.connect() as conn:
            count = conn.execute(
                """
                UPDATE ocr_jobs
                SET status='queued', locked_by=NULL, locked_at=NULL, next_run_at=0, updated_at=?, error=NULL
                WHERE status='processing' AND COALESCE(bot_key,'default')=?
                """,
                (ts, self.bot_key),
            ).rowcount
            conn.commit()
            return int(count or 0)

    def claim_ocr_job(self, worker_id: str) -> Optional[sqlite3.Row]:
        self.requeue_stale_ocr_jobs()
        ts = now_ms()
        with self.connect() as conn:
            conn.execute("BEGIN IMMEDIATE")
            row = conn.execute(
                """
                SELECT * FROM ocr_jobs
                WHERE status='queued' AND next_run_at <= ? AND COALESCE(bot_key,'default')=?
                ORDER BY created_at ASC
                LIMIT 1
                """,
                (ts, self.bot_key),
            ).fetchone()
            if not row:
                conn.commit()
                return None
            conn.execute(
                """
                UPDATE ocr_jobs
                SET status='processing', locked_by=?, locked_at=?, attempts=attempts+1, updated_at=?
                WHERE job_id=? AND status='queued'
                """,
                (worker_id, ts, ts, row["job_id"]),
            )
            claimed = conn.execute("SELECT * FROM ocr_jobs WHERE job_id=?", (row["job_id"],)).fetchone()
            conn.commit()
            return claimed

    def mark_ocr_job_done(self, job_id: str) -> None:
        with self.connect() as conn:
            conn.execute(
                "UPDATE ocr_jobs SET status='done', locked_by=NULL, locked_at=NULL, updated_at=?, error=NULL WHERE job_id=?",
                (now_ms(), job_id),
            )
            conn.commit()

    def fail_ocr_job(self, job: sqlite3.Row, error: str) -> None:
        ts = now_ms()
        attempts = int(job["attempts"] or 0)
        max_attempts = int(job["max_attempts"] or OCR_JOB_MAX_ATTEMPTS)
        final = attempts >= max_attempts
        status = "failed" if final else "queued"
        next_run_at = ts + min(300_000, 30_000 * max(1, attempts))
        with self.connect() as conn:
            conn.execute(
                """
                UPDATE ocr_jobs
                SET status=?, locked_by=NULL, locked_at=NULL, next_run_at=?, updated_at=?, error=?
                WHERE job_id=?
                """,
                (status, next_run_at if not final else ts, ts, error[:500], job["job_id"]),
            )
            conn.commit()
        if final:
            self.save_slip(
                {
                    "id": job["slip_id"],
                    "update_id": job["update_id"],
                    "bot_key": job["bot_key"],
                    "company_name": job["company_name"],
                    "chat_id": job["chat_id"],
                    "chat_title": job["chat_title"],
                    "user_id": job["user_id"],
                    "username": job["username"],
                    "sender_name": job["sender_name"],
                    "message_id": job["message_id"],
                    "file_id": job["file_id"],
                    "caption": job["caption"],
                    "status": "error",
                    "error": error[:500],
                }
            )

    def process_ocr_job(self, job: sqlite3.Row, worker_id: str = "worker") -> None:
        try:
            image_path, mime = self.download_file(job["file_id"])
            provider, data = ocr_extract(image_path, mime or job["mime"] or "image/jpeg")
            reason = unclear_reason(data)
            status = "unclear" if reason else "success"
            row = {
                "id": job["slip_id"],
                "update_id": job["update_id"],
                "bot_key": job["bot_key"],
                "company_name": job["company_name"],
                "chat_id": job["chat_id"],
                "chat_title": job["chat_title"],
                "user_id": job["user_id"],
                "username": job["username"],
                "sender_name": job["sender_name"],
                "message_id": job["message_id"],
                "file_id": job["file_id"],
                "caption": job["caption"],
                **data,
                "status": status,
                "error": reason,
                "ocr_provider": provider,
            }
            self.save_slip(row)
            self.mark_ocr_job_done(job["job_id"])
            if self.reply_on_result:
                if status == "success":
                    open_summary = self.summary_by_transferor(job["chat_id"], "open")
                    self.reply(job["chat_id"], self.success_reply_text(row, open_summary), job["message_id"])
                else:
                    self.reply(job["chat_id"], f"อ่านได้ไม่ชัด ⚠️ เก็บเข้า /queue แล้ว\nเหตุผล: {h(reason)}\nใช้ /reprocess {h(job['slip_id'])} ได้ค่ะ", job["message_id"])
        except Exception as exc:
            log("OCR job failed", worker_id, job["job_id"], str(exc)[:300])
            self.fail_ocr_job(job, str(exc))

    def worker_loop(self, worker_id: str) -> None:
        while True:
            try:
                job = self.claim_ocr_job(worker_id)
                if not job:
                    time.sleep(OCR_WORKER_IDLE_SLEEP)
                    continue
                self.process_ocr_job(job, worker_id=worker_id)
            except Exception:
                log("worker loop error", worker_id, traceback.format_exc())
                time.sleep(2)

    def start_workers(self) -> None:
        for i in range(OCR_WORKERS):
            worker_id = f"{self.bot_key}-ocr-worker-{i + 1}"
            t = threading.Thread(target=self.worker_loop, args=(worker_id,), daemon=True, name=worker_id)
            t.start()
        log(APP_NAME, "OCR workers started", OCR_WORKERS, "max_slips_per_poll", MAX_SLIPS_PER_POLL)

    def find_duplicate(self, row: Dict[str, Any]) -> str:
        """Return original slip id only when the full transaction signature matches.

        Duplicate matching is intentionally strict: same bot, same chat, same payer
        and recipient names, same slip date, same amount, same time, and same transaction ref.
        Partial matches are review evidence, not automatic duplicates.
        """
        amount = parse_number(row.get("amount"))
        date_key = clean_text(row.get("slip_date_iso") or row.get("slip_date_display"))
        slip_time = clean_text(row.get("slip_time"))
        name = clean_text(row.get("transferor_name"))
        recipient = clean_text(row.get("recipient_name"))
        reference_no = clean_text(row.get("reference_no"))
        if not amount or not date_key or not slip_time or not name or not recipient or not reference_no:
            return ""
        with self.connect() as conn:
            rows = conn.execute(
                """
                SELECT id
                FROM slips
                WHERE chat_id=? AND COALESCE(bot_key,'default')=? AND status='success' AND COALESCE(is_duplicate,0)=0 AND id<>?
                  AND amount=?
                  AND COALESCE(slip_date_iso, slip_date_display)=?
                  AND slip_time=?
                  AND TRIM(COALESCE(transferor_name,''))=?
                  AND TRIM(COALESCE(recipient_name,''))=?
                  AND TRIM(COALESCE(reference_no,''))=?
                ORDER BY created_at ASC
                LIMIT 1
                """,
                (row.get("chat_id"), row.get("bot_key") or self.bot_key, row.get("id"), amount, date_key, slip_time, name, recipient, reference_no),
            ).fetchone()
        return str(rows["id"]) if rows else ""

    def scope_clause(self, scope: str, success_only: bool = True) -> Tuple[str, List[Any], str]:
        normalized, label = scope_to_date(scope)
        clause = "chat_id=? AND COALESCE(bot_key,'default')=?"
        params: List[Any] = [self.bot_key]
        if success_only:
            clause += " AND status='success' AND COALESCE(is_duplicate,0)=0"
        if normalized == "open":
            clause += " AND settlement_id IS NULL"
        elif normalized == "all":
            pass
        elif re.match(r"^\d{4}-\d{2}-\d{2}$", normalized):
            clause += " AND slip_date_iso=?"
            params.append(normalized)
        else:
            clause += " AND (slip_date_display=? OR slip_date_iso=?)"
            params.extend([normalized, normalized])
        return clause, params, label

    def summary_by_transferor(self, chat_id: Any, scope: str = "open") -> Dict[str, Any]:
        clause, params, label = self.scope_clause(scope, success_only=True)
        query = f"""
            SELECT
              COALESCE(NULLIF(transferor_name,''), NULLIF(sender_name,''), '(ไม่ทราบชื่อผู้โอน)') AS name,
              COUNT(*) AS count,
              SUM(amount) AS amount,
              SUM(fee) AS fee
            FROM slips
            WHERE {clause}
            GROUP BY name
            ORDER BY amount DESC, count DESC, name ASC
        """
        with self.connect() as conn:
            rows = conn.execute(query, [str(chat_id), *params]).fetchall()
        by_name: Dict[str, Dict[str, Any]] = {}
        total_amount = 0.0
        total_fee = 0.0
        total_count = 0
        for r in rows:
            amount = float(r["amount"] or 0)
            fee = float(r["fee"] or 0)
            count = int(r["count"] or 0)
            by_name[str(r["name"])] = {"count": count, "amount": amount, "fee": fee}
            total_amount += amount
            total_fee += fee
            total_count += count
        return {"scope": scope, "label": label, "total_amount": total_amount, "total_fee": total_fee, "total_count": total_count, "by_name": by_name}

    def daily_summary(self, chat_id: Any, scope: str = "all") -> List[sqlite3.Row]:
        clause, params, _ = self.scope_clause(scope, success_only=True)
        with self.connect() as conn:
            return conn.execute(
                f"""
                SELECT COALESCE(NULLIF(slip_date_display,''), NULLIF(slip_date_iso,''), '') AS day, COUNT(*) AS count, SUM(amount) AS amount, SUM(fee) AS fee
                FROM slips
                WHERE {clause}
                GROUP BY day
                ORDER BY day DESC
                """,
                [str(chat_id), *params],
            ).fetchall()

    def issue_rows(self, chat_id: Any, limit: int = 20) -> List[sqlite3.Row]:
        with self.connect() as conn:
            return conn.execute(
                """
                SELECT * FROM slips
                WHERE chat_id=? AND COALESCE(bot_key,'default')=? AND status IN ('error','unclear')
                ORDER BY created_at DESC
                LIMIT ?
                """,
                (str(chat_id), self.bot_key, limit),
            ).fetchall()

    def recent_rows(self, chat_id: Any, limit: int = 5) -> List[sqlite3.Row]:
        with self.connect() as conn:
            return conn.execute(
                "SELECT * FROM slips WHERE chat_id=? AND COALESCE(bot_key,'default')=? ORDER BY created_at DESC LIMIT ?",
                (str(chat_id), self.bot_key, limit),
            ).fetchall()

    def duplicate_rows(self, chat_id: Any, limit: int = 10) -> List[sqlite3.Row]:
        with self.connect() as conn:
            return conn.execute(
                "SELECT * FROM slips WHERE chat_id=? AND COALESCE(bot_key,'default')=? AND is_duplicate=1 ORDER BY created_at DESC LIMIT ?",
                (str(chat_id), self.bot_key, limit),
            ).fetchall()

    def settlement_rows(self, chat_id: Any) -> List[sqlite3.Row]:
        with self.connect() as conn:
            return conn.execute(
                "SELECT * FROM settlements WHERE chat_id=? AND COALESCE(bot_key,'default')=? ORDER BY closed_at DESC LIMIT 100",
                (str(chat_id), self.bot_key),
            ).fetchall()

    def export_excel(self, chat_id: Any, scope: str = "open", settlement_id: str = "") -> str:
        EXPORT_DIR.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Slips"
        headers = [
            "chat_title",
            "username",
            "sender_name",
            "message_id",
            "caption",
            "error",
            "slip_date_display",
            "slip_date_iso",
            "slip_time",
            "issuer_bank",
            "seq",
            "location",
            "transaction_type",
            "transferor_name",
            "recipient_name",
            "from_bank",
            "from_account",
            "to_bank",
            "to_account",
            "account_name",
            "amount",
            "fee",
            "reference_no",
            "aid",
            "label",
            "raw_text",
            "confidence",
            "is_duplicate",
            "duplicate_of",
            "settlement_id",
            "created_at_iso",
        ]
        bank_fields = {"issuer_bank", "from_bank", "to_bank"}

        def cell_value(row: sqlite3.Row, header: str) -> Any:
            value = row[header]
            return canonical_bank(value) if header in bank_fields else value

        self.write_sheet(ws, headers, [])
        clause, params, _ = self.scope_clause(scope, success_only=False)
        if settlement_id:
            clause = "chat_id=? AND COALESCE(bot_key,'default')=? AND settlement_id=?"
            params = [self.bot_key, settlement_id]
        with self.connect() as conn:
            rows = conn.execute(
                f"SELECT * FROM slips WHERE {clause} ORDER BY slip_date_iso, slip_time, created_at",
                [str(chat_id), *params],
            ).fetchall()
            duplicate_ids = [str(r["duplicate_of"]) for r in rows if int(r["is_duplicate"] or 0) and r["duplicate_of"]]
            original_by_id: Dict[str, sqlite3.Row] = {}
            if duplicate_ids:
                placeholders = ",".join("?" for _ in duplicate_ids)
                original_rows = conn.execute(f"SELECT * FROM slips WHERE id IN ({placeholders})", duplicate_ids).fetchall()
                original_by_id = {str(r["id"]): r for r in original_rows}
        for row in rows:
            if int(row["is_duplicate"] or 0):
                continue
            ws.append([cell_value(row, header) for header in headers])

        dup_ws = wb.create_sheet("DuplicateSlips")
        dup_headers = [
            "duplicate_message_id",
            "matched_message_id",
            "duplicate_of",
            "slip_date_display",
            "slip_time",
            "transferor_name",
            "from_bank",
            "from_account",
            "to_bank",
            "to_account",
            "amount",
            "reference_no",
            "matched_reference_no",
            "sender_name",
            "matched_sender_name",
            "created_at_iso",
        ]
        self.write_sheet(dup_ws, dup_headers, [])
        for row in rows:
            if not int(row["is_duplicate"] or 0):
                continue
            original = original_by_id.get(str(row["duplicate_of"] or ""))
            dup_ws.append(
                [
                    row["message_id"],
                    original["message_id"] if original else "",
                    row["duplicate_of"],
                    row["slip_date_display"],
                    row["slip_time"],
                    row["transferor_name"],
                    canonical_bank(row["from_bank"]),
                    row["from_account"],
                    canonical_bank(row["to_bank"]),
                    row["to_account"],
                    row["amount"],
                    row["reference_no"],
                    original["reference_no"] if original else "",
                    row["sender_name"],
                    original["sender_name"] if original else "",
                    row["created_at_iso"],
                ]
            )

        ws2 = wb.create_sheet("SummaryByTransferor")
        summary = self.summary_by_transferor(chat_id, scope=scope if not settlement_id else "all")
        self.write_sheet(ws2, ["transferor_name", "count", "amount", "fee"], [])
        if settlement_id:
            with self.connect() as conn:
                group_rows = conn.execute(
                    """
                    SELECT COALESCE(NULLIF(transferor_name,''), NULLIF(sender_name,''), '(ไม่ทราบชื่อผู้โอน)') AS name,
                           COUNT(*) AS count, SUM(amount) AS amount, SUM(fee) AS fee
                    FROM slips
                    WHERE chat_id=? AND settlement_id=? AND status='success' AND COALESCE(is_duplicate,0)=0
                    GROUP BY name ORDER BY amount DESC
                    """,
                    (str(chat_id), settlement_id),
                ).fetchall()
            for r in group_rows:
                ws2.append([r["name"], r["count"], r["amount"], r["fee"]])
        else:
            for name, item in summary["by_name"].items():
                ws2.append([name, item["count"], item["amount"], item["fee"]])

        ws3 = wb.create_sheet("DailySummary")
        self.write_sheet(ws3, ["date", "count", "amount", "fee"], [])
        for r in self.daily_summary(chat_id, scope="all" if settlement_id else scope):
            ws3.append([r["day"], r["count"], r["amount"], r["fee"]])

        ws4 = wb.create_sheet("Issues")
        issue_headers = ["error", "date", "time", "amount", "message_id", "created_at", "raw_text"]
        self.write_sheet(ws4, issue_headers, [])
        with self.connect() as conn:
            issues = conn.execute(
                "SELECT * FROM slips WHERE chat_id=? AND COALESCE(bot_key,'default')=? AND status IN ('error','unclear') ORDER BY created_at DESC",
                (str(chat_id), self.bot_key),
            ).fetchall()
        for r in issues:
            ws4.append([r["error"], r["slip_date_display"], r["slip_time"], r["amount"], r["message_id"], r["created_at_iso"], r["raw_text"]])

        ws5 = wb.create_sheet("Settlements")
        self.write_sheet(ws5, ["settlement_id", "closed_at", "closed_by", "note", "count", "amount", "fee", "excel_path"], [])
        for r in self.settlement_rows(chat_id):
            ws5.append([r["settlement_id"], r["closed_at_iso"], r["closed_by"], r["note"], r["closed_count"], r["total_amount"], r["total_fee"], r["excel_path"]])

        for sheet in wb.worksheets:
            self.autofit(sheet)
        label = re.sub(r"[^A-Za-z0-9_.-]+", "-", str(scope or "open"))[:40]
        sid = settlement_id or label
        path = EXPORT_DIR / f"auditslip-{chat_id}-{sid}-{int(time.time())}.xlsx"
        wb.save(path)
        return str(path)

    @staticmethod
    def write_sheet(ws: Any, headers: Sequence[str], rows: Iterable[Sequence[Any]]) -> None:
        ws.append(list(headers))
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        for row in rows:
            ws.append(list(row))

    @staticmethod
    def autofit(ws: Any) -> None:
        for col in ws.columns:
            max_len = 0
            letter = get_column_letter(col[0].column)
            for cell in col:
                max_len = max(max_len, len(str(cell.value or "")))
            ws.column_dimensions[letter].width = min(max_len + 2, 48)

    def close_period(self, chat_id: Any, closed_by: str = "", note: str = "") -> Dict[str, Any]:
        ts = now_ms()
        settlement_id = stable_id(chat_id, ts, note, prefix="SET")
        with self.connect() as conn:
            rows = conn.execute(
                """
                SELECT COUNT(*) AS count, SUM(amount) AS amount, SUM(fee) AS fee
                FROM slips
                WHERE chat_id=? AND COALESCE(bot_key,'default')=? AND status='success' AND COALESCE(is_duplicate,0)=0 AND settlement_id IS NULL
                """,
                (str(chat_id), self.bot_key),
            ).fetchone()
            count = int(rows["count"] or 0)
            amount = float(rows["amount"] or 0)
            fee = float(rows["fee"] or 0)
            if count:
                conn.execute(
                    """
                    UPDATE slips SET settlement_id=?
                    WHERE chat_id=? AND COALESCE(bot_key,'default')=? AND status='success' AND COALESCE(is_duplicate,0)=0 AND settlement_id IS NULL
                    """,
                    (settlement_id, str(chat_id), self.bot_key),
                )
            closed_at_iso = datetime.fromtimestamp(ts / 1000, BKK).isoformat()
            conn.execute(
                """
                INSERT INTO settlements(settlement_id, bot_key, company_name, chat_id, closed_by, note, closed_at, closed_at_iso, total_amount, total_fee, closed_count, raw_json)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (settlement_id, self.bot_key, self.company_name, str(chat_id), closed_by, note, ts, closed_at_iso, amount, fee, count, json.dumps({"note": note}, ensure_ascii=False)),
            )
            conn.commit()
        excel_path = self.export_excel(chat_id, scope="all", settlement_id=settlement_id) if count else ""
        if excel_path:
            with self.connect() as conn:
                conn.execute("UPDATE settlements SET excel_path=? WHERE settlement_id=?", (excel_path, settlement_id))
                conn.commit()
        return {"settlement_id": settlement_id, "closed_count": count, "total_amount": amount, "total_fee": fee, "excel_path": excel_path}

    def clear_chat(self, chat_id: Any) -> Dict[str, int]:
        with self.connect() as conn:
            slips = conn.execute("DELETE FROM slips WHERE chat_id=? AND COALESCE(bot_key,'default')=?", (str(chat_id), self.bot_key)).rowcount
            jobs = conn.execute("DELETE FROM ocr_jobs WHERE chat_id=? AND COALESCE(bot_key,'default')=?", (str(chat_id), self.bot_key)).rowcount
            settlements = conn.execute("DELETE FROM settlements WHERE chat_id=? AND COALESCE(bot_key,'default')=?", (str(chat_id), self.bot_key)).rowcount
            conn.commit()
        return {"slips": slips, "jobs": jobs, "settlements": settlements}

    def is_admin(self, user_id: Any) -> bool:
        if not ADMIN_IDS:
            return True
        return str(user_id) in ADMIN_IDS

    def summary_text(self, chat_id: Any, scope: str = "open") -> str:
        summary = self.summary_by_transferor(chat_id, scope)
        lines = [
            f"<b>{h(APP_NAME)} สรุปยอด: {h(summary['label'])}</b>",
            f"รายการ: <b>{summary['total_count']}</b>",
            f"ยอดรวม: <b>{fmt_money(summary['total_amount'])}</b>",
        ]
        if summary["by_name"]:
            lines.append("")
            lines.append("<b>แยกตามชื่อผู้โอน</b>")
            for name, item in list(summary["by_name"].items())[:30]:
                lines.append(f"• {h(name)} — {item['count']} รายการ — <b>{fmt_money(item['amount'])}</b>")
        return "\n".join(lines)

    def daily_text(self, chat_id: Any, scope: str = "all") -> str:
        rows = self.daily_summary(chat_id, scope)
        if not rows:
            return "ยังไม่มีรายการสำเร็จค่ะ"
        lines = [f"<b>{h(APP_NAME)} สรุปรายวัน</b>"]
        for r in rows[:60]:
            lines.append(f"• {h(r['day'])}: {int(r['count'] or 0)} รายการ — <b>{fmt_money(r['amount'])}</b>")
        return "\n".join(lines)

    def providers_text(self) -> str:
        lines = [f"<b>{h(APP_NAME)} OCR providers</b>"]
        for item in provider_status():
            mark = "✅ active" if item["active"] else "⚠️ inactive"
            key = "key ok" if item["has_key"] else "missing key"
            lines.append(f"• {h(item['provider'])}: {mark} — {h(item['model'])} — {key}")
        lines.append(f"คิว OCR ต่อรอบ: <b>{MAX_SLIPS_PER_POLL}</b> สลิป")
        return "\n".join(lines)

    def success_reply_text(self, row: Dict[str, Any], open_summary: Optional[Dict[str, Any]] = None) -> str:
        summary = open_summary or {"total_amount": 0, "total_count": 0}
        name = row.get("transferor_name") or row.get("sender_name") or "(ไม่ทราบชื่อผู้โอน)"
        date_text = " ".join(x for x in [clean_text(row.get("slip_date_display") or row.get("slip_date_iso")), clean_text(row.get("slip_time"))] if x)
        lines = [
            "บันทึกแล้ว ✅",
            f"ผู้โอน: <b>{h(name)}</b>",
            f"ยอดสลิปนี้: <b>{fmt_money(row.get('amount'))}</b>",
            f"ยอดรวมที่จับได้: <b>{fmt_money(summary.get('total_amount'))}</b> ({int(summary.get('total_count') or 0)} สลิป)",
        ]
        if date_text:
            lines.append(f"วันที่: {h(date_text)}")
        ref = row.get("reference_no") or row.get("aid") or row.get("seq")
        if ref:
            lines.append(f"อ้างอิง: <code>{h(ref)}</code>")
        return "\n".join(lines)

    def usage_text(self, chat_id: Any, scope: str = "today") -> str:
        clause, params, label = self.scope_clause(scope, success_only=False)
        with self.connect() as conn:
            rows = conn.execute(
                f"""
                SELECT COALESCE(NULLIF(ocr_provider,''), '(unknown)') AS provider,
                       COALESCE(NULLIF(ocr_model,''), '') AS model,
                       status,
                       COUNT(*) AS count
                FROM slips
                WHERE {clause}
                GROUP BY provider, model, status
                ORDER BY provider, model, status
                """,
                [str(chat_id), *params],
            ).fetchall()
        totals: Dict[str, int] = defaultdict(int)
        status_totals: Dict[str, int] = defaultdict(int)
        lines = [f"<b>{h(APP_NAME)} API usage: {h(label)}</b>", "นับจากรายการที่บอทบันทึกไว้ ไม่ใช่ billing dashboard ของ provider"]
        if not rows:
            lines.append("ยังไม่มีรายการ OCR ในช่วงนี้ค่ะ")
        for r in rows:
            provider = str(r["provider"] or "(unknown)")
            count = int(r["count"] or 0)
            totals[provider] += count
            status_totals[str(r["status"])] += count
            model = f" / {r['model']}" if r["model"] else ""
            lines.append(f"• {h(provider)}{h(model)} — {h(r['status'])}: <b>{count}</b>")
        if rows:
            lines.append("")
            lines.append("<b>รวมตาม provider</b>")
            for provider, count in sorted(totals.items()):
                lines.append(f"• {h(provider)}: <b>{count}</b> calls/สลิป")
            lines.append("<b>รวมตามสถานะ</b>")
            for status, count in sorted(status_totals.items()):
                lines.append(f"• {h(status)}: <b>{count}</b>")
        lines.append("")
        lines.append(f"คิว OCR ต่อรอบ: <b>{MAX_SLIPS_PER_POLL}</b> สลิป")
        return "\n".join(lines)

    def help_text(self) -> str:
        return f"""<b>{h(APP_NAME)}</b> — บอทตรวจสลิป + audit ยอด

ส่งรูปสลิปเข้าห้องนี้ได้เลย บอทจะ OCR และเก็บเข้าระบบ

คำสั่งหลัก:
/summary [open|today|all|DD/MM/YY] — สรุปยอด
/names [open|today|all|DD/MM/YY] — สรุปแยกตามชื่อผู้โอน
/excel [open|today|all|DD/MM/YY] — ส่งออก Excel
/close [note] — เคลียร์ยอด/ปิดรอบปัจจุบันแบบเก็บประวัติ
/clear — ดูวิธีล้างข้อมูลถาวร
/queue — ดูรายการ fail/อ่านไม่ชัด
/reprocess [id] — OCR ใหม่จากรายการที่ fail/unclear
/providers — ดูสถานะ Gemini/OpenAI ในบอทเดียว
/usage [today|open|all] — ดูการใช้งาน OCR API ที่บันทึกไว้

ระบบจะประมวลผล OCR เป็นคิว รอบละไม่เกิน {MAX_SLIPS_PER_POLL} สลิป เพื่อกันยอดมั่วและกัน provider หน่วง
""".strip()

    @staticmethod
    def chat_title(msg: Dict[str, Any]) -> str:
        chat = msg.get("chat", {})
        return chat.get("title") or " ".join(x for x in [chat.get("first_name"), chat.get("last_name")] if x) or str(chat.get("id", ""))

    @staticmethod
    def sender_name(msg: Dict[str, Any]) -> str:
        user = msg.get("from", {})
        return clean_text(" ".join(x for x in [user.get("first_name"), user.get("last_name")] if x)) or user.get("username") or str(user.get("id", ""))

    def handle_command(self, msg: Dict[str, Any], command: str, args: List[str]) -> None:
        chat_id = msg.get("chat", {}).get("id")
        user_id = msg.get("from", {}).get("id")
        reply_to = msg.get("message_id")
        cmd = command.split("@", 1)[0].lower()
        scope = args[0] if args else "open"
        if cmd in {"start", "help"}:
            self.reply(chat_id, self.help_text(), reply_to)
        elif cmd in {"summary"}:
            self.reply(chat_id, self.summary_text(chat_id, scope), reply_to)
        elif cmd == "today":
            self.reply(chat_id, self.summary_text(chat_id, "today"), reply_to)
        elif cmd == "daily":
            self.reply(chat_id, self.daily_text(chat_id, scope if args else "all"), reply_to)
        elif cmd in {"names", "userall"}:
            self.reply(chat_id, self.summary_text(chat_id, scope), reply_to)
        elif cmd == "excel":
            path = Path(self.export_excel(chat_id, scope))
            self.send_document(chat_id, path, f"{APP_NAME} Excel — {scope}")
        elif cmd == "providers":
            self.reply(chat_id, self.providers_text(), reply_to)
        elif cmd == "usage":
            self.reply(chat_id, self.usage_text(chat_id, scope), reply_to)
        elif cmd in {"queue", "failed"}:
            self.reply(chat_id, self.queue_text(chat_id), reply_to)
        elif cmd == "recent":
            self.reply(chat_id, self.recent_text(chat_id), reply_to)
        elif cmd == "stats":
            self.reply(chat_id, self.stats_text(chat_id), reply_to)
        elif cmd == "dupes":
            self.reply(chat_id, self.dupes_text(chat_id), reply_to)
        elif cmd == "close":
            if not self.is_admin(user_id):
                self.reply(chat_id, "คำสั่งนี้ใช้ได้เฉพาะ admin ที่ตั้งไว้ค่ะ", reply_to)
                return
            note = " ".join(args)
            settlement = self.close_period(chat_id, closed_by=str(user_id or ""), note=note)
            text = (
                f"<b>ปิดรอบ/เคลียร์ยอดแล้ว</b>\n"
                f"รหัสรอบ: <code>{h(settlement['settlement_id'])}</code>\n"
                f"รายการ: <b>{settlement['closed_count']}</b>\n"
                f"ยอดรวม: <b>{fmt_money(settlement['total_amount'])}</b>"
            )
            self.reply(chat_id, text, reply_to)
            if settlement.get("excel_path"):
                self.send_document(chat_id, Path(settlement["excel_path"]), f"{APP_NAME} close batch {settlement['settlement_id']}")
        elif cmd == "clear":
            if not self.is_admin(user_id):
                self.reply(chat_id, "คำสั่งนี้ใช้ได้เฉพาะ admin ที่ตั้งไว้ค่ะ", reply_to)
                return
            if args and args[0].lower() == "confirm":
                result = self.clear_chat(chat_id)
                self.reply(chat_id, f"ล้างข้อมูลห้องนี้แล้ว: slips {result['slips']} รายการ, settlements {result['settlements']} รอบ", reply_to)
            else:
                self.reply(chat_id, "ถ้าต้องการล้างข้อมูลถาวรของห้องนี้ ใช้ <code>/clear confirm</code>\nถ้าต้องการแค่เคลียร์ยอดรอบปัจจุบัน ใช้ <code>/close</code> แทนค่ะ", reply_to)
        elif cmd == "reprocess":
            self.reply(chat_id, self.reprocess_latest(chat_id, args[0] if args else ""), reply_to)
        else:
            self.reply(chat_id, "ไม่รู้จักคำสั่งนี้ค่ะ ใช้ /help ได้เลย", reply_to)

    def queue_text(self, chat_id: Any) -> str:
        lines = ["<b>คิว OCR / รายการที่ต้องตรวจ</b>"]
        with self.connect() as conn:
            job_counts = conn.execute(
                "SELECT status, COUNT(*) AS count FROM ocr_jobs WHERE chat_id=? AND COALESCE(bot_key,'default')=? GROUP BY status ORDER BY status",
                (str(chat_id), self.bot_key),
            ).fetchall()
            jobs = conn.execute(
                """
                SELECT * FROM ocr_jobs
                WHERE chat_id=? AND COALESCE(bot_key,'default')=? AND status IN ('queued','processing','failed')
                ORDER BY created_at ASC
                LIMIT 20
                """,
                (str(chat_id), self.bot_key),
            ).fetchall()
        if job_counts:
            lines.append("<b>สถานะคิว</b>")
            for r in job_counts:
                lines.append(f"• {h(r['status'])}: <b>{int(r['count'] or 0)}</b>")
        if jobs:
            lines.append("")
            lines.append("<b>งานคิวล่าสุด</b>")
            for r in jobs:
                lines.append(f"• <code>{h(r['slip_id'])}</code> — {h(r['status'])} — attempts {int(r['attempts'] or 0)}/{int(r['max_attempts'] or 0)}")
        rows = self.issue_rows(chat_id, 20)
        if rows:
            lines.append("")
            lines.append("<b>อ่านไม่ชัด/ผิดพลาด</b>")
            for r in rows:
                lines.append(f"• <code>{h(r['id'])}</code> — {h(r['status'])}: {h(r['error'])} — msg {h(r['message_id'])}")
        if len(lines) == 1:
            return "ไม่มีคิวค้างหรือรายการ fail/อ่านไม่ชัดค่ะ"
        return "\n".join(lines)

    def recent_text(self, chat_id: Any) -> str:
        rows = self.recent_rows(chat_id, 10)
        if not rows:
            return "ยังไม่มีรายการค่ะ"
        lines = ["<b>รายการล่าสุด</b>"]
        for r in rows:
            name = r["transferor_name"] or r["sender_name"] or "(ไม่ทราบชื่อ)"
            lines.append(f"• {h(r['status'])} {h(r['slip_date_display'] or r['slip_date_iso'])} {h(r['slip_time'])} — {h(name)} — <b>{fmt_money(r['amount'])}</b>")
        return "\n".join(lines)

    def stats_text(self, chat_id: Any) -> str:
        with self.connect() as conn:
            rows = conn.execute("SELECT status, COUNT(*) AS count FROM slips WHERE chat_id=? AND COALESCE(bot_key,'default')=? GROUP BY status", (str(chat_id), self.bot_key)).fetchall()
        lines = [f"<b>{h(APP_NAME)} stats</b>"]
        for r in rows:
            lines.append(f"• {h(r['status'])}: {int(r['count'])}")
        return "\n".join(lines) if len(lines) > 1 else "ยังไม่มีข้อมูลค่ะ"

    def dupes_text(self, chat_id: Any) -> str:
        rows = self.duplicate_rows(chat_id, 20)
        if not rows:
            return "ยังไม่พบรายการซ้ำค่ะ"
        lines = ["<b>รายการซ้ำ</b>"]
        for r in rows:
            lines.append(f"• {h(r['slip_date_display'])} {h(r['slip_time'])} — {h(r['transferor_name'])} — {fmt_money(r['amount'])} duplicate_of <code>{h(r['duplicate_of'])}</code>")
        return "\n".join(lines)

    def reprocess_latest(self, chat_id: Any, target_id: str = "") -> str:
        with self.connect() as conn:
            if target_id:
                row = conn.execute("SELECT * FROM slips WHERE chat_id=? AND COALESCE(bot_key,'default')=? AND id=?", (str(chat_id), self.bot_key, target_id)).fetchone()
            else:
                row = conn.execute("SELECT * FROM slips WHERE chat_id=? AND COALESCE(bot_key,'default')=? AND status IN ('error','unclear') ORDER BY created_at DESC LIMIT 1", (str(chat_id), self.bot_key)).fetchone()
        if not row:
            return "ไม่พบรายการที่ต้อง reprocess ค่ะ"
        if not row["file_id"]:
            return "รายการนี้ไม่มี file_id ให้ดึงรูปกลับมา OCR ใหม่ค่ะ"
        try:
            path, mime = self.download_file(row["file_id"])
            provider, data = ocr_extract(path, mime)
            reason = unclear_reason(data)
            status = "unclear" if reason else "success"
            saved = dict(row)
            saved.update(data)
            saved["status"] = status
            saved["error"] = reason
            saved["ocr_provider"] = provider
            saved["raw_json"] = data.get("raw_json") or json.dumps(data, ensure_ascii=False)
            self.save_slip(saved)
            return f"reprocess แล้ว: {status} ผ่าน {provider}" + (f" — {reason}" if reason else "")
        except Exception as exc:
            saved = dict(row)
            saved["status"] = "error"
            saved["error"] = str(exc)[:500]
            self.save_slip(saved)
            return f"reprocess ยังไม่ผ่าน: {h(exc)}"

    def process_image_message(self, update_id: int, msg: Dict[str, Any]) -> None:
        chat_id = msg.get("chat", {}).get("id")
        user = msg.get("from", {})
        message_id = msg.get("message_id")
        file_id = ""
        mime = "image/jpeg"
        if msg.get("photo"):
            file_id = msg["photo"][-1]["file_id"]
        elif msg.get("document") and str(msg["document"].get("mime_type", "")).startswith("image/"):
            file_id = msg["document"]["file_id"]
            mime = msg["document"].get("mime_type") or mime
        if not file_id:
            return
        slip_id = stable_id(self.bot_key, chat_id, message_id, file_id)
        base_row = {
            "id": slip_id,
            "update_id": update_id,
            "bot_key": self.bot_key,
            "company_name": self.company_name,
            "chat_id": str(chat_id),
            "chat_title": self.chat_title(msg),
            "user_id": str(user.get("id", "")),
            "username": user.get("username", ""),
            "sender_name": self.sender_name(msg),
            "message_id": message_id,
            "file_id": file_id,
            "caption": msg.get("caption", ""),
        }
        try:
            job_id = self.enqueue_ocr_job(base_row, mime)
            if REPLY_ON_QUEUE:
                self.reply(chat_id, f"รับเข้าคิว OCR แล้ว ✅\nรหัสงาน: <code>{h(job_id)}</code>", message_id)
        except Exception as exc:
            err = str(exc)[:500]
            self.save_slip({**base_row, "status": "error", "error": err})
            self.reply(chat_id, f"รับสลิปไม่สำเร็จ เก็บเข้า /queue ไม่ได้ค่ะ\n<code>{h(err)}</code>", message_id)

    def download_file(self, file_id: str) -> Tuple[Path, str]:
        obj = self.telegram_get("getFile", {"file_id": file_id})
        file_path = obj.get("result", {}).get("file_path")
        if not file_path:
            raise RuntimeError("Telegram getFile did not return file_path")
        url = f"{TELEGRAM_API}/file/bot{self.token}/{file_path}"
        resp = requests.get(url, timeout=90)
        resp.raise_for_status()
        suffix = Path(file_path).suffix or ".jpg"
        fd, tmp = tempfile.mkstemp(prefix="auditslip-", suffix=suffix)
        os.close(fd)
        path = Path(tmp)
        path.write_bytes(resp.content)
        return path, mime_for_path(path)

    def process_update(self, update: Dict[str, Any]) -> None:
        update_id = int(update.get("update_id", 0))
        if self.already_processed(update_id):
            return
        msg = update.get("message") or update.get("edited_message")
        if not msg:
            self.mark_processed(update_id)
            return
        text = msg.get("text") or ""
        if text.startswith("/"):
            parts = text.strip().split()
            self.handle_command(msg, parts[0][1:], parts[1:])
        else:
            self.process_image_message(update_id, msg)
        self.mark_processed(update_id)

    def poll_once(self, offset: int) -> int:
        obj = self.telegram_get("getUpdates", {"offset": offset, "timeout": POLL_TIMEOUT, "allowed_updates": json.dumps(["message", "edited_message"])})
        updates = obj.get("result", [])
        next_offset = offset
        slips_this_round = 0
        for update in updates:
            msg = update.get("message") or update.get("edited_message") or {}
            if is_image_message(msg) and slips_this_round >= MAX_SLIPS_PER_POLL:
                break
            update_id = int(update.get("update_id", 0))
            try:
                self.process_update(update)
                if is_image_message(msg):
                    slips_this_round += 1
                next_offset = max(next_offset, update_id + 1)
                self.persist_offset(next_offset)
            except Exception:
                log("process_update failed", update_id, traceback.format_exc())
                raise
        return next_offset

    def run(self) -> None:
        self.init_db()
        self.clear_webhook_for_polling()
        self.set_commands()
        resumed = self.resume_incomplete_ocr_jobs()
        if resumed:
            log(APP_NAME, "resumed incomplete OCR jobs", resumed)
        self.start_workers()
        offset = self.get_offset()
        log(APP_NAME, "started", "db", self.db_path, "providers", provider_status())
        while True:
            try:
                offset = self.poll_once(offset)
            except KeyboardInterrupt:
                raise
            except Exception:
                log("poll loop error", traceback.format_exc())
                time.sleep(5)


def main() -> None:
    configs = telegram_bot_configs()
    if not configs:
        raise SystemExit("BOT_TOKEN or AUDITSLIP_TELEGRAM_BOTS is required")
    if len(configs) == 1:
        cfg = configs[0]
        AuditslipBot(token=cfg["token"], bot_key=cfg["bot_key"], company_name=cfg["company_name"]).run()
        return
    threads: List[threading.Thread] = []
    for cfg in configs:
        bot = AuditslipBot(token=cfg["token"], bot_key=cfg["bot_key"], company_name=cfg["company_name"])
        t = threading.Thread(target=bot.run, daemon=True, name=f"auditslip-{cfg['bot_key']}")
        t.start()
        threads.append(t)
    try:
        while True:
            time.sleep(3600)
    except KeyboardInterrupt:
        return


if __name__ == "__main__":
    main()
