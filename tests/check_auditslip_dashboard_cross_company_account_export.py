#!/usr/bin/env python3
"""Guard: cross-company account slip search can be exported as one Excel workbook."""
from __future__ import annotations

import importlib.util
import os
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
os.environ["BOT_TOKEN"] = "TEST_TOKEN"
os.environ["AUDITSLIP_TELEGRAM_BOTS"] = "botA:BOT_TOKEN:บริษัท A,botB:BOT_TOKEN:บริษัท B,botC:BOT_TOKEN:บริษัท C"
os.environ["AUDITSLIP_DB"] = str(Path(tempfile.mkdtemp(prefix="auditslip-cross-account-export-")) / "auditslip.db")
os.environ["AUDITSLIP_HOME"] = str(ROOT)
os.environ["AUDITSLIP_EXPORT_DIR"] = str(Path(tempfile.mkdtemp(prefix="auditslip-cross-account-export-out-")))
os.environ["AUDITSLIP_DASHBOARD_TOKEN"] = "test-token"

bot_spec = importlib.util.spec_from_file_location("auditslip_bot", ROOT / "auditslip_bot.py")
assert bot_spec and bot_spec.loader
bot_mod = importlib.util.module_from_spec(bot_spec)
sys.modules["auditslip_bot"] = bot_mod
bot_spec.loader.exec_module(bot_mod)

db_path = Path(os.environ["AUDITSLIP_DB"])
bots = {
    "botA": bot_mod.AuditslipBot(token="TEST_TOKEN", db_path=db_path, dry_run=True, bot_key="botA", company_name="บริษัท A"),
    "botB": bot_mod.AuditslipBot(token="TEST_TOKEN", db_path=db_path, dry_run=True, bot_key="botB", company_name="บริษัท B"),
    "botC": bot_mod.AuditslipBot(token="TEST_TOKEN", db_path=db_path, dry_run=True, bot_key="botC", company_name="บริษัท C"),
}
for bot in bots.values():
    bot.init_db()


def save(bot_key: str, slip_id: str, company: str, chat_id: str, title: str, amount: float, to_account: str, ref: str, duplicate: int = 0, date_iso: str = "2026-05-22", from_account: str = "") -> None:
    bots[bot_key].save_slip({
        "id": slip_id,
        "bot_key": bot_key,
        "company_name": company,
        "chat_id": chat_id,
        "chat_title": title,
        "message_id": int(ref[-2:]) if ref[-2:].isdigit() else 1,
        "file_id": f"FILE_{slip_id}",
        "sender_name": "Uploader",
        "status": "success",
        "slip_date_display": "22/05/26" if date_iso == "2026-05-22" else "21/05/26",
        "slip_date_iso": date_iso,
        "slip_time": "10:00",
        "transferor_name": f"ลูกค้า {company}",
        "recipient_name": company,
        "from_bank": "SCB",
        "from_account": from_account or f"FROM-{bot_key}",
        "to_bank": "KBANK",
        "to_account": to_account,
        "account_name": company,
        "amount": amount,
        "fee": 0.0,
        "reference_no": ref,
        "is_duplicate": duplicate,
    })


save("botA", "A_SHARED", "บริษัท A", "A_WD", "บริษัท A ถอน", 100.0, "DEST-A", "RA01", from_account="SHARED-ACC-789")
save("botB", "B_SHARED", "บริษัท B", "B_WD", "บริษัท B ถอน", 250.0, "DEST-B", "RB02", from_account="SHARED-ACC-789")
save("botC", "C_OTHER", "บริษัท C", "C_WD", "บริษัท C ถอน", 999.0, "DEST-C", "RC03", from_account="OTHER-ACC")
save("botB", "B_DUP", "บริษัท B", "B_WD", "บริษัท B ถอน", 888.0, "DEST-B", "RB04", duplicate=1, from_account="SHARED-ACC-789")
save("botA", "A_OLD", "บริษัท A", "A_WD", "บริษัท A ถอน", 777.0, "DEST-A", "RA05", date_iso="2026-05-21", from_account="SHARED-ACC-789")
# คู่ที่ "ตรงกัน": ยอด+วันเดียวกัน ปรากฏทั้ง A และ B -> ต้องยุบเป็นบรรทัดเดียวเทียบกัน
save("botA", "A_MATCH", "บริษัท A", "A_WD", "บริษัท A ถอน", 500.0, "DEST-A", "RA10", from_account="SHARED-ACC-789")
save("botB", "B_MATCH", "บริษัท B", "B_WD", "บริษัท B ถอน", 500.0, "DEST-B", "RB10", from_account="SHARED-ACC-789")

spec = importlib.util.spec_from_file_location("auditslip_dashboard", ROOT / "auditslip_dashboard.py")
assert spec and spec.loader
Dash = importlib.util.module_from_spec(spec)
sys.modules["auditslip_dashboard"] = Dash
spec.loader.exec_module(Dash)

xlsx = Dash.export_cross_company_account_slips_excel(
    db_path,
    flow_type="withdraw",
    scope="2026-05-22",
    search="SHAREDACC789",
)
assert xlsx.suffix == ".xlsx" and xlsx.exists(), xlsx
wb = load_workbook(xlsx, data_only=True)
assert {"SummaryByCompany", "CrossCompanyAccountSlips"}.issubset(set(wb.sheetnames)), wb.sheetnames

summary_rows = list(wb["SummaryByCompany"].iter_rows(values_only=True))
summary_headers = list(summary_rows[0])
summary = [dict(zip(summary_headers, row)) for row in summary_rows[1:]]
assert {str(row["company_name"] or "") for row in summary} == {"บริษัท A", "บริษัท B"}, summary
# A: RA01(100)+RA10(500)=2, B: RB02(250)+RB10(500)=2 -> รวม 4 slip, 1350 บาท
assert sum(int(float(str(row["count"] or 0))) for row in summary) == 4, summary
assert sum(float(str(row["amount"] or 0)) for row in summary) == 1350.0, summary

# Layout ใหม่: บรรทัดเดียวเทียบยอดข้ามบริษัท (ไม่ใช่ 1 slip 1 แถวแบบเดิม)
slip_rows = list(wb["CrossCompanyAccountSlips"].iter_rows(values_only=True))
headers = list(slip_rows[0])
records = [dict(zip(headers, row)) for row in slip_rows[1:]]

# หัวคอลัมน์: คอลัมน์คงที่ + block ต่อบริษัท (จำนวน/เวลา/อ้างอิง/ผู้โอน) + คอลัมน์สถานะ
assert headers[:2] == ["slip_date_display", "amount"], headers
assert headers[-4:] == ["companies_present", "match_status", "count_diff", "amount_diff"], headers
for company in ("บริษัท A", "บริษัท B"):
    assert f"{company} - จำนวน" in headers, headers
    assert f"{company} - เวลา" in headers, headers
    assert f"{company} - อ้างอิง" in headers, headers
    assert f"{company} - ผู้โอน" in headers, headers
# บริษัท C ไม่มียอดข้ามบริษัทกับ account นี้ -> ไม่มี block
assert "บริษัท C - จำนวน" not in headers, headers
# ไม่มีคอลัมน์ระดับ slip เดี่ยวแบบเดิม
assert "reference_no" not in headers and "company_name" not in headers, headers

# 1 บรรทัด = 1 ยอด/วัน. มี 3 ยอด: 500(ตรงกัน A+B), 250(เฉพาะ B), 100(เฉพาะ A)
by_amount = {float(r["amount"]): r for r in records}
assert set(by_amount) == {500.0, 250.0, 100.0}, records

matched = by_amount[500.0]
assert matched["match_status"] == "ตรงกัน", matched
assert matched["บริษัท A - จำนวน"] == 1 and matched["บริษัท B - จำนวน"] == 1, matched
assert matched["บริษัท A - อ้างอิง"] == "RA10" and matched["บริษัท B - อ้างอิง"] == "RB10", matched
assert matched["บริษัท A - เวลา"] == "10:00" and matched["บริษัท B - เวลา"] == "10:00", matched
assert matched["count_diff"] == 0 and matched["amount_diff"] == 0, matched
assert "บริษัท A" in str(matched["companies_present"]) and "บริษัท B" in str(matched["companies_present"]), matched

only_b = by_amount[250.0]
assert only_b["match_status"] == "เฉพาะฝั่งเดียว", only_b
assert only_b["บริษัท B - อ้างอิง"] == "RB02" and (only_b["บริษัท A - อ้างอิง"] in ("", None)), only_b

only_a = by_amount[100.0]
assert only_a["match_status"] == "เฉพาะฝั่งเดียว", only_a
assert only_a["บริษัท A - อ้างอิง"] == "RA01", only_a

# duplicate / นอกวัน / นอก account ต้องไม่โผล่
all_refs = " ".join(str(r["บริษัท A - อ้างอิง"]) + " " + str(r["บริษัท B - อ้างอิง"]) for r in records)
assert "RB04" not in all_refs and "RA05" not in all_refs and "RC03" not in all_refs, records

html = Dash.render_dashboard_html("test-token")
for marker in [
    "buildCrossCompanyAccountExcelUrl",
    "cross_account_search",
    "ส่งออก Excel ข้ามบริษัท",
    "target=\"exportDownloadFrame\"",
]:
    assert marker in html, marker

print("ok: cross-company account slip search exports one Excel workbook")
