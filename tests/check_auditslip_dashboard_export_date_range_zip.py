#!/usr/bin/env python3
"""Guard: dashboard Excel export supports selected date ranges and all-company zip separation."""
from __future__ import annotations

import importlib.util
import os
import sys
import tempfile
import zipfile
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
os.environ["BOT_TOKEN"] = "TOKEN1"
os.environ["BOT_TOKEN_2"] = "TOKEN2"
os.environ["AUDITSLIP_TELEGRAM_BOTS"] = "bot1:BOT_TOKEN:บริษัท 1,bot2:BOT_TOKEN_2:บริษัท 2"
os.environ["AUDITSLIP_DB"] = str(Path(tempfile.mkdtemp(prefix="auditslip-export-range-")) / "auditslip.db")
os.environ["AUDITSLIP_HOME"] = str(ROOT)
os.environ["AUDITSLIP_EXPORT_DIR"] = str(Path(tempfile.mkdtemp(prefix="auditslip-export-range-out-")))
os.environ["AUDITSLIP_DASHBOARD_TOKEN"] = "test-token"
os.environ["GEMINI_API_KEY"] = "gemini-test-key"
os.environ["OPENAI_API_KEY"] = "openai-test-key"

bot_spec = importlib.util.spec_from_file_location("auditslip_bot", ROOT / "auditslip_bot.py")
assert bot_spec and bot_spec.loader
bot_mod = importlib.util.module_from_spec(bot_spec)
sys.modules["auditslip_bot"] = bot_mod
bot_spec.loader.exec_module(bot_mod)

db_path = Path(os.environ["AUDITSLIP_DB"])
bot1 = bot_mod.AuditslipBot(token="TOKEN1", db_path=db_path, dry_run=True, bot_key="bot1", company_name="บริษัท 1")
bot2 = bot_mod.AuditslipBot(token="TOKEN2", db_path=db_path, dry_run=True, bot_key="bot2", company_name="บริษัท 2")
bot1.init_db(); bot2.init_db()
base = {
    "sender_name": "Uploader",
    "status": "success",
    "slip_time": "10:00",
    "transferor_name": "ลูกค้า",
    "recipient_name": "ร้านค้า",
    "from_bank": "KBANK",
    "to_bank": "SCB",
    "amount": 100.0,
    "reference_no": "REF",
}
bot1.save_slip({**base, "id": "B1_IN", "bot_key": "bot1", "company_name": "บริษัท 1", "chat_id": "CHAT1", "chat_title": "บริษัท 1 ถอน", "message_id": 11, "file_id": "F1", "slip_date_display": "22/05/26", "slip_date_iso": "2026-05-22", "amount": 111.0, "reference_no": "R1"})
bot1.save_slip({**base, "id": "B1_OUT", "bot_key": "bot1", "company_name": "บริษัท 1", "chat_id": "CHAT1", "chat_title": "บริษัท 1 ถอน", "message_id": 12, "file_id": "F2", "slip_date_display": "25/05/26", "slip_date_iso": "2026-05-25", "amount": 999.0, "reference_no": "R2"})
bot2.save_slip({**base, "id": "B2_IN", "bot_key": "bot2", "company_name": "บริษัท 2", "chat_id": "CHAT2", "chat_title": "บริษัท 2 เติมมือ", "message_id": 21, "file_id": "F3", "slip_date_display": "23/05/26", "slip_date_iso": "2026-05-23", "amount": 222.0, "reference_no": "R3"})

spec = importlib.util.spec_from_file_location("auditslip_dashboard", ROOT / "auditslip_dashboard.py")
assert spec and spec.loader
Dash = importlib.util.module_from_spec(spec)
sys.modules["auditslip_dashboard"] = Dash
spec.loader.exec_module(Dash)

xlsx = Dash.export_dashboard_excel(db_path, bot_key="bot1", chat_id="", flow_type="withdraw", scope="all", start_date="2026-05-22", end_date="2026-05-23")
wb = load_workbook(xlsx, data_only=True)
ws = wb["Slips"]
rows = list(ws.iter_rows(values_only=True))
headers = list(rows[0])
records = [dict(zip(headers, row)) for row in rows[1:]]
refs = [row["reference_no"] for row in records]
assert refs == ["R1"], rows
assert "R2" not in refs, rows
assert "id" not in headers and "bot_key" not in headers and "file_id" not in headers, headers

zip_path = Dash.export_dashboard_zip_by_company(db_path, bot_key="__all__", flow_type="all", scope="all", start_date="2026-05-22", end_date="2026-05-23")
with zipfile.ZipFile(zip_path) as zf:
    names = sorted(zf.namelist())
assert len(names) == 2 and any("bot1" in n for n in names) and any("bot2" in n for n in names), names

html = Dash.render_dashboard_html("test-token")
for marker in ["exportStartDate", "exportEndDate", "start_date", "end_date", "export_dashboard_zip_by_company"]:
    assert marker in html, marker

print("ok: dashboard export supports date range and all-company zip")
