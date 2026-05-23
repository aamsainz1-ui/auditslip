#!/usr/bin/env python3
"""Guard: dashboard export has operator-ready sheets and date ranges include visible slip dates without ISO."""
from __future__ import annotations

import importlib.util
import os
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
os.environ["BOT_TOKEN"] = "TOKEN1"
os.environ["AUDITSLIP_TELEGRAM_BOTS"] = "bot1:BOT_TOKEN:บริษัท 1"
os.environ["AUDITSLIP_DB"] = str(Path(tempfile.mkdtemp(prefix="auditslip-export-format-")) / "auditslip.db")
os.environ["AUDITSLIP_HOME"] = str(ROOT)
os.environ["AUDITSLIP_EXPORT_DIR"] = str(Path(tempfile.mkdtemp(prefix="auditslip-export-format-out-")))
os.environ["AUDITSLIP_DASHBOARD_TOKEN"] = "test-token"
os.environ["GEMINI_API_KEY"] = "gemini-test-key"
os.environ["OPENAI_API_KEY"] = "openai-test-key"

bot_spec = importlib.util.spec_from_file_location("auditslip_bot", ROOT / "auditslip_bot.py")
assert bot_spec and bot_spec.loader
bot_mod = importlib.util.module_from_spec(bot_spec)
sys.modules["auditslip_bot"] = bot_mod
bot_spec.loader.exec_module(bot_mod)

db_path = Path(os.environ["AUDITSLIP_DB"])
bot = bot_mod.AuditslipBot(token="TOKEN1", db_path=db_path, dry_run=True, bot_key="bot1", company_name="บริษัท 1")
bot.init_db()
base = {
    "bot_key": "bot1",
    "company_name": "บริษัท 1",
    "chat_id": "CHAT1",
    "chat_title": "บริษัท 1 ถอน",
    "sender_name": "Uploader",
    "status": "success",
    "slip_time": "10:00",
    "transferor_name": "ลูกค้า",
    "recipient_name": "ร้านค้า",
    "from_bank": "KBANK",
    "to_bank": "SCB",
    "amount": 100.0,
}
bot.save_slip({**base, "id": "IN_RANGE_VISIBLE_DATE", "message_id": 11, "file_id": "F1", "slip_date_display": "22/05/26", "slip_date_iso": "", "amount": 111.0, "reference_no": "R-IN"})
bot.save_slip({**base, "id": "OUT_RANGE_VISIBLE_DATE", "message_id": 12, "file_id": "F2", "slip_date_display": "25/05/26", "slip_date_iso": "", "amount": 999.0, "reference_no": "R-OUT"})

spec = importlib.util.spec_from_file_location("auditslip_dashboard", ROOT / "auditslip_dashboard.py")
assert spec and spec.loader
Dash = importlib.util.module_from_spec(spec)
sys.modules["auditslip_dashboard"] = Dash
spec.loader.exec_module(Dash)

xlsx = Dash.export_dashboard_excel(db_path, bot_key="bot1", chat_id="", flow_type="withdraw", scope="all", start_date="2026-05-22", end_date="2026-05-22")
wb = load_workbook(xlsx, data_only=True)
for sheet in ["SummaryByCompany", "SummaryByTransferor", "DailySummary", "Slips", "DuplicateSlips"]:
    assert sheet in wb.sheetnames, wb.sheetnames

headers = [cell.value for cell in wb["Slips"][1]]
for forbidden_header in ["id", "chat_id", "file_id", "status", "bot_key"]:
    assert forbidden_header not in headers, f"dashboard export still exposes/internal old column: {forbidden_header}"
for required_header in ["company_name", "chat_title", "message_id", "slip_date_display", "slip_date_iso", "transferor_name", "from_bank", "to_bank", "amount", "reference_no"]:
    assert required_header in headers, f"missing useful operator column: {required_header}"
slip_rows = [dict(zip(headers, [cell.value for cell in row])) for row in wb["Slips"].iter_rows(min_row=2)]
assert [row["reference_no"] for row in slip_rows] == ["R-IN"], slip_rows
assert slip_rows[0]["amount"] == 111.0, slip_rows

summary_headers = [cell.value for cell in wb["SummaryByCompany"][1]]
summary_rows = [dict(zip(summary_headers, [cell.value for cell in row])) for row in wb["SummaryByCompany"].iter_rows(min_row=2)]
assert summary_rows == [{"company_name": "บริษัท 1", "count": 1, "amount": 111.0, "fee": 0}], summary_rows

daily_headers = [cell.value for cell in wb["DailySummary"][1]]
daily_rows = [dict(zip(daily_headers, [cell.value for cell in row])) for row in wb["DailySummary"].iter_rows(min_row=2)]
assert daily_rows == [{"date": "22/05/26", "count": 1, "amount": 111.0, "fee": 0}], daily_rows

print("ok: dashboard export format is operator-ready and visible dates work in ranges")
