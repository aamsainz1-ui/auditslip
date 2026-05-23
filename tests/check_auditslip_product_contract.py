#!/usr/bin/env python3
"""Product contract checks for Auditslip.

This is intentionally framework-free so it can run on a minimal VPS:
    python3 tests/check_auditslip_product_contract.py
"""
from __future__ import annotations

import importlib.util
import os
import sqlite3
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
MODULE = ROOT / "auditslip_bot.py"

assert MODULE.exists(), "auditslip_bot.py must exist"

os.environ["BOT_TOKEN"] = "TEST_TOKEN"
os.environ["AUDITSLIP_DB"] = str(Path(tempfile.mkdtemp(prefix="auditslip-test-")) / "auditslip.db")
os.environ["AUDITSLIP_HOME"] = str(ROOT)
os.environ["AUDITSLIP_EXPORT_DIR"] = str(Path(tempfile.mkdtemp(prefix="auditslip-export-")))
os.environ["GEMINI_API_KEY"] = "gemini-test-key"
os.environ["OPENAI_API_KEY"] = "openai-test-key"
os.environ["OCR_PROVIDERS"] = "gemini,openai"

spec = importlib.util.spec_from_file_location("auditslip_bot", MODULE)
assert spec is not None
app = importlib.util.module_from_spec(spec)
sys.modules["auditslip_bot"] = app
assert spec.loader is not None
spec.loader.exec_module(app)

assert app.APP_NAME == "Auditslip"
providers = app.ocr_provider_candidates(
    provider_string="gemini,openai",
    gemini_key="gemini-test-key",
    openai_key="openai-test-key",
)
assert providers == ["gemini", "openai"], providers
providers_without_openai = app.ocr_provider_candidates(
    provider_string="gemini,openai",
    gemini_key="gemini-test-key",
    openai_key="",
)
assert providers_without_openai == ["gemini"], providers_without_openai

commands = {item["command"] for item in app.COMMANDS}
for required in [
    "summary",
    "today",
    "daily",
    "names",
    "userall",
    "excel",
    "close",
    "clear",
    "queue",
    "failed",
    "reprocess",
    "providers",
    "usage",
]:
    assert required in commands, f"missing command /{required}"

bot = app.AuditslipBot(token="TEST_TOKEN", db_path=Path(os.environ["AUDITSLIP_DB"]), dry_run=True)
bot.init_db()

base = {
    "chat_id": "CHAT1",
    "chat_title": "Audit Room",
    "user_id": "U1",
    "username": "alice",
    "sender_name": "Alice Telegram",
    "message_id": 1,
    "file_id": "FILE1",
    "caption": "caption sample",
    "status": "success",
    "slip_date_display": "22/05/26",
    "slip_date_iso": "2026-05-22",
    "slip_time": "10:00",
    "issuer_bank": "SCB",
    "seq": "123456",
    "location": "ATM TEST",
    "transaction_type": "TRANSFER",
    "transferor_name": "คุณเอ",
    "recipient_name": "ร้านค้า",
    "from_bank": "SCB",
    "from_account": "123-xxx-4567",
    "to_bank": "KBank",
    "to_account": "987-xxx-6543",
    "account_name": "ร้านค้า จำกัด",
    "amount": 1000.0,
    "fee": 0.0,
    "reference_no": "REF1",
    "aid": "AID123",
    "label": "SCB ATM",
    "raw_text": "full raw OCR text",
    "confidence": 0.95,
    "ocr_provider": "gemini",
    "ocr_model": "gemini-2.5-flash",
}

bot.save_slip(dict(base, id="S1"))
bot.save_slip(dict(base, id="S2", message_id=2, file_id="FILE2", amount=500.0, transferor_name="คุณบี", reference_no="REF2", ocr_provider="openai", ocr_model="gpt-4o-mini"))
bot.save_slip(dict(base, id="S3", message_id=3, file_id="FILE3", amount=250.0, transferor_name="คุณเอ", reference_no="REF3"))
bot.save_slip(dict(base, id="E1", message_id=4, file_id="FILE4", status="unclear", amount=0.0, transferor_name="", error="missing amount", confidence=0.2))
bot.save_slip(dict(base, id="DUP1", message_id=5, file_id="FILE5", sender_name="Bob Telegram"))
duplicates = bot.duplicate_rows("CHAT1", 10)
assert len(duplicates) == 1, duplicates
assert duplicates[0]["duplicate_of"] == "S1", duplicates[0]

summary = bot.summary_by_transferor("CHAT1", scope="all")
assert summary["total_amount"] == 1750.0, summary
assert summary["by_name"]["คุณเอ"]["amount"] == 1250.0, summary
assert summary["by_name"]["คุณเอ"]["count"] == 2, summary
assert summary["by_name"]["คุณบี"]["amount"] == 500.0, summary

success_reply = bot.success_reply_text(dict(base), {"total_amount": 1750.0, "total_count": 3})
assert "บันทึกแล้ว" in success_reply
assert "ยอดสลิปนี้" in success_reply
assert "ยอดรวมที่จับได้" in success_reply
assert "1,000.00" in success_reply
assert "1,750.00" in success_reply
assert "OCR:" not in success_reply

usage = bot.usage_text("CHAT1", "all")
assert "API usage" in usage
assert "gemini" in usage
assert "openai" in usage
assert "คิว OCR ต่อรอบ" in usage

xlsx = bot.export_excel("CHAT1", scope="all")
assert Path(xlsx).exists(), xlsx
wb = load_workbook(xlsx)
for sheet in ["Slips", "DuplicateSlips", "SummaryByTransferor", "DailySummary", "Issues", "Settlements"]:
    assert sheet in wb.sheetnames, wb.sheetnames
slips_headers = [cell.value for cell in wb["Slips"][1]]
for forbidden_header in ["id", "update_id", "chat_id", "user_id", "file_id", "status"]:
    assert forbidden_header not in slips_headers, f"internal Slips Excel column should be hidden: {forbidden_header}"
for required_header in [
    "chat_title",
    "username",
    "sender_name",
    "message_id",
    "caption",
    "error",
    "slip_date_display",
    "slip_date_iso",
    "slip_time",
    "issuer_bank",
    "seq",
    "location",
    "transaction_type",
    "transferor_name",
    "recipient_name",
    "from_bank",
    "from_account",
    "to_bank",
    "to_account",
    "account_name",
    "amount",
    "fee",
    "reference_no",
    "aid",
    "label",
    "raw_text",
    "confidence",
    "is_duplicate",
    "duplicate_of",
    "settlement_id",
    "created_at_iso",
]:
    assert required_header in slips_headers, f"missing Slips Excel column {required_header}"
row_values = dict(zip(slips_headers, [cell.value for cell in wb["Slips"][2]]))
slip_rows = [dict(zip(slips_headers, [cell.value for cell in row])) for row in wb["Slips"].iter_rows(min_row=2)]
assert all(row.get("is_duplicate") != 1 for row in slip_rows), slip_rows
assert all(row.get("message_id") != 5 for row in slip_rows), slip_rows
assert row_values["issuer_bank"] == "SCB"
assert row_values["seq"] == "123456"
assert row_values["location"] == "ATM TEST"
assert row_values["from_account"] == "123-xxx-4567"
assert row_values["to_account"] == "987-xxx-6543"
assert row_values["raw_text"] == "full raw OCR text"
assert "ocr_provider" not in slips_headers
assert "ocr_model" not in slips_headers

dup_headers = [cell.value for cell in wb["DuplicateSlips"][1]]
dup_rows = [dict(zip(dup_headers, [cell.value for cell in row])) for row in wb["DuplicateSlips"].iter_rows(min_row=2)]
assert len(dup_rows) == 1, dup_rows
assert dup_rows[0]["duplicate_message_id"] == 5, dup_rows
assert dup_rows[0]["matched_message_id"] == 1, dup_rows
assert dup_rows[0]["duplicate_of"] == "S1", dup_rows
assert dup_rows[0]["matched_reference_no"] == "REF1", dup_rows

settlement = bot.close_period("CHAT1", closed_by="U1", note="test close")
assert settlement["closed_count"] == 3, settlement
assert settlement["total_amount"] == 1750.0, settlement
with sqlite3.connect(os.environ["AUDITSLIP_DB"]) as conn:
    rows = conn.execute("SELECT DISTINCT settlement_id FROM slips WHERE chat_id='CHAT1' AND status='success' AND COALESCE(is_duplicate,0)=0").fetchall()
assert len(rows) == 1 and rows[0][0] == settlement["settlement_id"], rows

print("ok: Auditslip product contract")
