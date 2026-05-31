#!/usr/bin/env python3
"""Auditslip lightweight VPS dashboard.

Serves a public read-only dashboard for accounting totals, OCR queue health,
recent slips, and Excel exports. Admin mutations require token or owner login.
"""
from __future__ import annotations

import warnings

warnings.filterwarnings("ignore", "'cgi' is deprecated", DeprecationWarning)
import cgi
import csv
import datetime as dt
import hashlib
import json
import logging
import mimetypes
import os
import re
import sqlite3
import subprocess
import sys
import tempfile
import time
import uuid
import zipfile
from copy import copy

logger = logging.getLogger("auditslip.dashboard")
if not logger.handlers:
    _dh = logging.StreamHandler()
    _dh.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(name)s %(message)s"))
    logger.addHandler(_dh)
    logger.setLevel(logging.INFO)


def safe_error(exc: Exception) -> str:
    """Non-sensitive error label for HTTP responses. Full traceback goes to logger only."""
    return type(exc).__name__


import requests
from openpyxl import Workbook
from collections import defaultdict
from difflib import SequenceMatcher
from functools import lru_cache
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any, Dict, List, Tuple
from urllib.parse import parse_qs, urlencode, urlparse

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import auditslip_bank_ledger as bank_ledger_component
import auditslip_audit_employee as audit_employee_component

from auditslip_bot import (
    APP_NAME,
    DATA_DIR,
    DB_PATH,
    EXPORT_DIR,
    TELEGRAM_API,
    AuditslipBot,
    bkk_iso_from_ms,
    fmt_money,
    h,
    normalize_date_parts,
    normalize_record,
    ocr_extract,
    openai_extract,
    parse_number,
    provider_status,
    scope_to_date,
    telegram_bot_configs,
    unclear_reason,
)

HOST = os.environ.get("AUDITSLIP_DASHBOARD_HOST", "0.0.0.0")
PORT = int(os.environ.get("AUDITSLIP_DASHBOARD_PORT", "8095"))
DASHBOARD_TOKEN = os.environ.get("AUDITSLIP_DASHBOARD_TOKEN", "")
DASHBOARD_OWNER_USER = os.environ.get("AUDITSLIP_DASHBOARD_OWNER_USER", "owner").strip()
DASHBOARD_OWNER_PASSWORD = os.environ.get("AUDITSLIP_DASHBOARD_OWNER_PASSWORD", "").strip()
COOKIE_NAME = "auditslip_dashboard_token"
TWALLET_DASHBOARD_URL = os.environ.get("AUDITSLIP_TWALLET_DASHBOARD_URL", "http://76.13.190.65:3051").rstrip("/")
TWALLET_TIMEOUT_SECONDS = float(os.environ.get("AUDITSLIP_TWALLET_TIMEOUT", "2.5"))
TWALLET_CACHE_TTL_SECONDS = float(os.environ.get("AUDITSLIP_TWALLET_CACHE_TTL", "20"))
_TWALLET_SUMMARY_CACHE: Dict[str, Any] = {"ts": 0.0, "base_url": "", "data": None}


def connect(db_path: Path) -> sqlite3.Connection:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn


REQUIRED_HEALTH_TABLES = ("slips", "ocr_jobs", "processed_updates", "bot_state", "company_accounts")
OPTIONAL_HEALTH_TABLES = ("pending_actions", "dashboard_mutation_log")


def health_env_bool(name: str, default: bool = True) -> bool:
    raw = os.environ.get(name)
    if raw is None:
        return default
    return clean_display(raw).lower() not in {"0", "false", "no", "n", "off", "skip"}


def connect_readonly(db_path: Path) -> sqlite3.Connection:
    """Open SQLite in read-only mode so health checks cannot create or mutate DB files."""
    conn = sqlite3.connect(f"file:{Path(db_path)}?mode=ro", uri=True, timeout=5.0)
    conn.row_factory = sqlite3.Row
    return conn


def _status_counts(conn: sqlite3.Connection, table: str, column: str) -> Dict[str, int]:
    rows = conn.execute(
        f"SELECT COALESCE(NULLIF({column},''),'unknown') AS status, COUNT(*) AS count FROM {table} GROUP BY COALESCE(NULLIF({column},''),'unknown') ORDER BY status"
    ).fetchall()
    return {str(r["status"]): int(r["count"] or 0) for r in rows}


def _file_age_seconds(path: Path, now: float | None = None) -> int | None:
    if not path.exists():
        return None
    return max(0, int((time.time() if now is None else now) - path.stat().st_mtime))


def _systemctl_state(unit: str) -> Dict[str, Any]:
    unit = clean_display(unit)
    if not unit:
        return {"ok": False, "state": "missing", "unit": ""}
    try:
        proc = subprocess.run(["systemctl", "is-active", unit], capture_output=True, text=True, timeout=1.5)
        state = clean_display(proc.stdout or proc.stderr) or "unknown"
        return {"ok": proc.returncode == 0 and state == "active", "state": state, "unit": unit}
    except Exception as exc:
        return {"ok": False, "state": "unknown", "unit": unit, "error_type": type(exc).__name__}


def dashboard_quick_health(db_path: Path = DB_PATH) -> Dict[str, Any]:
    """Cheap health check for watchdog reachability probes.

    The full operational health intentionally includes dashboard/queue/provider
    aggregates. Under production load those aggregate scans can exceed the
    watchdog's short HTTP timeout, causing a false "dashboard down" alert even
    while the dashboard is alive. The watchdog already performs queue/service
    checks separately, so this quick mode only verifies the dashboard process can
    open the DB read-only and see the required tables.
    """
    generated_at = dt.datetime.now(dt.timezone.utc).isoformat()
    checks: Dict[str, Any] = {}
    criticals: List[Dict[str, Any]] = []
    path = Path(db_path)
    db_check: Dict[str, Any] = {"ok": False, "exists": path.exists(), "readable": False, "path": "configured"}
    checks["db"] = db_check
    if not path.exists():
        criticals.append({"code": "db_missing", "message": "database file is missing"})
    else:
        try:
            with connect_readonly(path) as conn:
                conn.execute("SELECT 1").fetchone()
                db_check["readable"] = True
                db_check["ok"] = True
                required = {name: sqlite_table_exists(conn, name) for name in REQUIRED_HEALTH_TABLES}
                schema_ok = all(required.values())
                checks["schema"] = {"ok": schema_ok, "required_tables": required}
                if not schema_ok:
                    missing = [name for name, exists in required.items() if not exists]
                    criticals.append({"code": "schema_missing", "message": "required tables are missing", "tables": missing})
        except Exception as exc:
            db_check["readable"] = False
            db_check["ok"] = False
            checks.setdefault("schema", {"ok": False, "required_tables": {}})
            criticals.append({"code": "db_unreadable", "message": "database could not be read", "error_type": type(exc).__name__})
    status = "critical" if criticals else "ok"
    return {
        "ok": not criticals,
        "app": APP_NAME,
        "status": status,
        "quick": True,
        "generated_at": generated_at,
        "checks": checks,
        "warning_count": 0,
        "critical_count": len(criticals),
        "warnings": [],
        "criticals": criticals[:10],
    }


def dashboard_operational_health(db_path: Path = DB_PATH) -> Dict[str, Any]:
    """Return public-safe operational health without writing to DB or exposing secrets."""
    generated_at = dt.datetime.now(dt.timezone.utc).isoformat()
    now_ms_value = int(time.time() * 1000)
    stale_minutes = int(os.environ.get("AUDITSLIP_HEALTH_STALE_MINUTES") or os.environ.get("AUDITSLIP_WATCHDOG_STALE_MINUTES") or "15")
    stale_cutoff = now_ms_value - stale_minutes * 60_000
    checks: Dict[str, Any] = {}
    warnings_list: List[Dict[str, Any]] = []
    criticals: List[Dict[str, Any]] = []
    path = Path(db_path)

    db_check: Dict[str, Any] = {"ok": False, "exists": path.exists(), "readable": False, "path": "configured"}
    if path.exists():
        try:
            db_check["file_size_bytes"] = path.stat().st_size
        except OSError:
            pass
    checks["db"] = db_check

    if not path.exists():
        criticals.append({"code": "db_missing", "message": "database file is missing"})
    else:
        try:
            with connect_readonly(path) as conn:
                conn.execute("SELECT 1").fetchone()
                db_check["readable"] = True
                db_check["ok"] = True

                required = {name: sqlite_table_exists(conn, name) for name in REQUIRED_HEALTH_TABLES}
                optional = {name: sqlite_table_exists(conn, name) for name in OPTIONAL_HEALTH_TABLES}
                schema_ok = all(required.values())
                checks["schema"] = {"ok": schema_ok, "required_tables": required, "optional_tables": optional}
                if not schema_ok:
                    missing = [name for name, exists in required.items() if not exists]
                    criticals.append({"code": "schema_missing", "message": "required tables are missing", "tables": missing})

                if required.get("slips"):
                    totals = conn.execute(
                        """
                        SELECT COUNT(*) AS total_count,
                               COALESCE(SUM(CASE WHEN COALESCE(status,'success')='success' AND COALESCE(is_duplicate,0)=0 THEN 1 ELSE 0 END),0) AS success_nonduplicate_count,
                               COALESCE(SUM(CASE WHEN COALESCE(status,'success')='success' AND COALESCE(is_duplicate,0)=0 THEN COALESCE(amount,0) ELSE 0 END),0) AS success_nonduplicate_amount,
                               MAX(created_at) AS newest_created_at
                        FROM slips
                        """
                    ).fetchone()
                    today = os.environ.get("AUDITSLIP_HEALTH_TODAY") or os.environ.get("AUDITSLIP_WATCHDOG_TODAY") or dt.datetime.now(dt.timezone(dt.timedelta(hours=7))).date().isoformat()
                    today_row = conn.execute(
                        """
                        SELECT COUNT(*) AS count, COALESCE(SUM(amount),0) AS amount
                        FROM slips
                        WHERE COALESCE(status,'success')='success' AND COALESCE(is_duplicate,0)=0 AND COALESCE(slip_date_iso,'')=?
                        """,
                        (today,),
                    ).fetchone()
                    checks["slips"] = {
                        "ok": True,
                        "total_count": int(totals["total_count"] or 0),
                        "success_nonduplicate_count": int(totals["success_nonduplicate_count"] or 0),
                        "success_nonduplicate_amount": float(totals["success_nonduplicate_amount"] or 0),
                        "newest_created_age_seconds": max(0, int((now_ms_value - int(totals["newest_created_at"] or now_ms_value)) / 1000)) if totals else None,
                        "today": {"date": today, "success_nonduplicate_count": int(today_row["count"] or 0), "success_nonduplicate_amount": float(today_row["amount"] or 0)},
                    }
                else:
                    checks["slips"] = {"ok": False, "error": "missing_table"}

                if required.get("ocr_jobs"):
                    counts = _status_counts(conn, "ocr_jobs", "status")
                    active_count = int(counts.get("queued", 0) + counts.get("processing", 0))
                    stale_rows = conn.execute(
                        """
                        SELECT status, MIN(COALESCE(NULLIF(updated_at,0), created_at)) AS oldest_ts, COUNT(*) AS count
                        FROM ocr_jobs
                        WHERE status IN ('queued','processing') AND COALESCE(NULLIF(updated_at,0), created_at) < ?
                        GROUP BY status
                        ORDER BY status
                        """,
                        (stale_cutoff,),
                    ).fetchall()
                    stale_active_count = sum(int(r["count"] or 0) for r in stale_rows)
                    oldest_active = conn.execute(
                        "SELECT MIN(COALESCE(NULLIF(updated_at,0), created_at)) AS oldest_ts FROM ocr_jobs WHERE status IN ('queued','processing')"
                    ).fetchone()
                    oldest_ts = int(oldest_active["oldest_ts"] or 0) if oldest_active else 0
                    queue_check = {
                        "ok": stale_active_count == 0,
                        "counts": counts,
                        "active_count": active_count,
                        "stale_active_count": stale_active_count,
                        "stale_after_minutes": stale_minutes,
                        "oldest_active_age_seconds": max(0, int((now_ms_value - oldest_ts) / 1000)) if oldest_ts else None,
                        "stale_by_status": {str(r["status"]): int(r["count"] or 0) for r in stale_rows},
                    }
                    checks["ocr_queue"] = queue_check
                    if stale_active_count:
                        warnings_list.append({"code": "ocr_queue_stale", "message": "queued/processing OCR jobs are stale", "count": stale_active_count})
                    failed_threshold = int(os.environ.get("AUDITSLIP_HEALTH_FAILED_THRESHOLD") or os.environ.get("AUDITSLIP_WATCHDOG_FAILED_THRESHOLD") or "1")
                    failed_count = int(counts.get("failed", 0))
                    if failed_threshold and failed_count >= failed_threshold:
                        warnings_list.append({"code": "ocr_jobs_failed", "message": "OCR jobs are failed", "count": failed_count})
                else:
                    checks["ocr_queue"] = {"ok": False, "error": "missing_table", "counts": {}}

                if optional.get("pending_actions"):
                    pending_counts = _status_counts(conn, "pending_actions", "status")
                    checks["pending_actions"] = {
                        "ok": True,
                        "counts": pending_counts,
                        "pending_count": int(pending_counts.get("pending", 0)),
                    }
                else:
                    checks["pending_actions"] = {"ok": True, "present": False, "pending_count": 0, "counts": {}}
        except Exception as exc:
            db_check["readable"] = False
            db_check["ok"] = False
            checks.setdefault("schema", {"ok": False, "required_tables": {}, "optional_tables": {}})
            checks.setdefault("slips", {"ok": False, "error": "db_unreadable"})
            checks.setdefault("ocr_queue", {"ok": False, "error": "db_unreadable", "counts": {}})
            checks.setdefault("pending_actions", {"ok": False, "error": "db_unreadable"})
            criticals.append({"code": "db_unreadable", "message": "database could not be read", "error_type": type(exc).__name__})

    try:
        providers = provider_status()
        active_count = sum(1 for item in providers if item.get("active"))
        circuit_open_count = sum(1 for item in providers if item.get("circuit_open"))
        checks["ocr_providers"] = {
            "ok": bool(active_count),
            "active_count": active_count,
            "configured_count": len(providers),
            "circuit_open_count": circuit_open_count,
            "providers": providers,
        }
        if providers and not active_count:
            warnings_list.append({"code": "ocr_providers_unavailable", "message": "no OCR provider is currently available"})
        elif circuit_open_count:
            warnings_list.append({"code": "ocr_provider_circuit_open", "message": "one or more OCR providers are in circuit-breaker cooldown", "count": circuit_open_count})
    except Exception as exc:
        checks["ocr_providers"] = {"ok": False, "error_type": type(exc).__name__, "providers": []}
        warnings_list.append({"code": "ocr_provider_health_unavailable", "message": "OCR provider health could not be read"})

    watchdog_state = Path(os.environ.get("AUDITSLIP_WATCHDOG_STATE") or DATA_DIR / "watchdog-state.json")
    state_file = {"exists": watchdog_state.exists(), "age_seconds": _file_age_seconds(watchdog_state)}
    watchdog_check: Dict[str, Any] = {"ok": True, "state_file": state_file}
    if not state_file["exists"]:
        watchdog_check["ok"] = False
        warnings_list.append({"code": "watchdog_state_missing", "message": "watchdog state file is missing"})
    if health_env_bool("AUDITSLIP_HEALTH_SYSTEMCTL", True):
        watchdog_check["services"] = {
            "dashboard": _systemctl_state(os.environ.get("AUDITSLIP_DASHBOARD_SERVICE", "auditslip-dashboard.service")),
            "bot": _systemctl_state(os.environ.get("AUDITSLIP_BOT_SERVICE", "auditslip-bot.service")),
            "watchdog_timer": _systemctl_state(os.environ.get("AUDITSLIP_WATCHDOG_TIMER", "auditslip-bot-watchdog.timer")),
        }
        if not all(item.get("ok") for item in watchdog_check["services"].values()):
            watchdog_check["ok"] = False
            warnings_list.append({"code": "service_state_warning", "message": "one or more service units are not active"})
    else:
        watchdog_check["services"] = {"skipped": True}
    checks["watchdog"] = watchdog_check

    status = "critical" if criticals else ("degraded" if warnings_list else "ok")
    return {
        "ok": not criticals,
        "app": APP_NAME,
        "status": status,
        "generated_at": generated_at,
        "checks": checks,
        "warning_count": len(warnings_list),
        "critical_count": len(criticals),
        "warnings": warnings_list[:10],
        "criticals": criticals[:10],
    }


def sqlite_table_exists(conn: sqlite3.Connection, table_name: str) -> bool:
    row = conn.execute("SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (table_name,)).fetchone()
    return row is not None


def ensure_dashboard_performance_indexes(conn: sqlite3.Connection) -> None:
    """Create idempotent indexes for high-frequency dashboard summary/search paths."""
    if not sqlite_table_exists(conn, "slips"):
        return
    conn.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_slips_bot_chat_date_amount
        ON slips(bot_key, chat_id, slip_date_iso, amount)
        WHERE COALESCE(status,'success')='success' AND COALESCE(is_duplicate,0)=0
        """
    )
    conn.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_slips_status
        ON slips(status)
        WHERE status IN ('unclear','error','success')
        """
    )
    conn.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_slips_duplicate_created
        ON slips(created_at DESC, bot_key, chat_id, duplicate_of)
        WHERE status='success' AND COALESCE(is_duplicate,0)=1
        """
    )
    if sqlite_table_exists(conn, "ocr_jobs"):
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_ocr_jobs_slip_bot_created
            ON ocr_jobs(slip_id, bot_key, created_at)
            """
        )


def rows_to_dicts(rows: List[sqlite3.Row]) -> List[Dict[str, Any]]:
    return [dict(r) for r in rows]


def aggregate_dicts(rows: Any) -> List[Dict[str, Any]]:
    if rows and isinstance(rows[0], sqlite3.Row):
        return rows_to_dicts(rows)
    return list(rows or [])


def clean_display(value: Any) -> str:
    return " ".join(str(value or "").strip().split())


def mask_phone(value: Any) -> str:
    digits = re.sub(r"\D+", "", clean_display(value))
    if len(digits) >= 5:
        return f"{digits[:3]}***{digits[-2:]}"
    return clean_display(value)


def satang_to_baht(value: Any) -> float:
    return round(parse_number(value) / 100.0, 2)


def twallet_json(base_url: str, path: str) -> Dict[str, Any]:
    response = requests.get(f"{base_url.rstrip('/')}{path}", timeout=TWALLET_TIMEOUT_SECONDS)
    response.raise_for_status()
    data = response.json()
    return data if isinstance(data, dict) else {}


def fetch_twallet_summary(force: bool = False, base_url: str = "") -> Dict[str, Any]:
    """Read a small, safe True Wallet dashboard summary for embedding in Auditslip."""
    url = clean_display(base_url or TWALLET_DASHBOARD_URL).rstrip("/")
    if not url:
        return {"enabled": False, "ok": False, "error": "not configured"}
    now = time.time()
    cached = _TWALLET_SUMMARY_CACHE.get("data")
    if (
        not force
        and cached
        and _TWALLET_SUMMARY_CACHE.get("base_url") == url
        and now - float(_TWALLET_SUMMARY_CACHE.get("ts") or 0) < TWALLET_CACHE_TTL_SECONDS
    ):
        return dict(cached)
    try:
        health = twallet_json(url, "/health")
        balance = twallet_json(url, "/api/tm/balance")
        daily = twallet_json(url, "/api/stats/daily?days=1")
        last = twallet_json(url, "/api/tm/my-last-receive")
        balance_data = balance.get("data") if isinstance(balance.get("data"), dict) else {}
        daily_items = daily.get("items") if isinstance(daily.get("items"), list) else []
        today = daily_items[-1] if daily_items else {}
        last_data = last.get("data") if isinstance(last.get("data"), dict) else {}
        summary = {
            "enabled": True,
            "ok": bool(health.get("ok")) and balance.get("status") == "ok" and bool(daily.get("ok")),
            "tokens_loaded": bool(health.get("tokensLoaded")),
            "tx_count": int(health.get("txCount") or 0),
            "balance_amount": satang_to_baht(balance_data.get("balance")),
            "mobile_masked": mask_phone(balance_data.get("mobile_no")),
            "balance_updated_at": clean_display(balance_data.get("updated_at")),
            "today_date": clean_display(today.get("date")),
            "today_count": int(today.get("count") or 0),
            "today_total": float(today.get("total") or 0),
            "last_receive": {
                "amount": satang_to_baht(last_data.get("amount")),
                "sender_mobile": mask_phone(last_data.get("sender_mobile")),
                "receiver_mobile": mask_phone(last_data.get("receiver_mobile")),
                "received_time": clean_display(last_data.get("received_time")),
                "event_type": clean_display(last_data.get("event_type") or "P2P"),
            } if last_data else {},
            "error": "",
        }
    except Exception as exc:
        summary = {"enabled": True, "ok": False, "error": type(exc).__name__, "today_count": 0, "today_total": 0.0, "balance_amount": 0.0, "last_receive": {}}
    _TWALLET_SUMMARY_CACHE.update({"ts": now, "base_url": url, "data": dict(summary)})
    return summary


def clean_company_name(value: Any, bot_key: Any = "") -> str:
    """Operator-facing company name: remove noisy suffixes such as Audit."""
    text = clean_display(value)
    text = re.sub(r"(?i)\baudit\b", "", text)
    text = clean_display(text).strip(" -_/|")
    return text or clean_display(bot_key)


def company_number(value: Any, bot_key: Any = "") -> int:
    text = clean_display(f"{value} {bot_key}")
    for pattern in [r"บริษัท\s*0*(\d+)", r"company\s*0*(\d+)", r"bot\s*0*(\d+)"]:
        match = re.search(pattern, text, flags=re.I)
        if match:
            return int(match.group(1))
    match = re.search(r"(?<!\d)0*([1-9]\d?)(?!\d)", text)
    return int(match.group(1)) if match else 999999


def company_sort_key(value: Any, bot_key: Any = "") -> Tuple[int, int, str, str]:
    name = clean_company_name(value, bot_key)
    number = company_number(name, bot_key)
    return (0 if number != 999999 else 1, number, name.lower(), clean_display(bot_key).lower())


def dict_company_sort_key(row: Dict[str, Any]) -> Tuple[int, int, str, str]:
    return company_sort_key(row.get("company_name") or row.get("bot_key"), row.get("bot_key"))


def clean_company_fields(row: Dict[str, Any]) -> Dict[str, Any]:
    row["bot_key"] = clean_display(row.get("bot_key")) or "default"
    row["company_name"] = clean_company_name(row.get("company_name"), row.get("bot_key"))
    return row


BANK_ALIASES = {
    "KRUNGTHAI": {"krungthai", "krungthaibank", "ktb", "ktbnext", "กรุงไทย", "ธกรุงไทย", "ธนาคารกรุงไทย"},
    "KBANK": {"kbank", "kasikorn", "kasikornbank", "kasikornthai", "กสิกร", "กสิกรไทย", "ธกสิกรไทย", "ธนาคารกสิกรไทย"},
    "SCB": {"scb", "scbeasy", "siamcommercialbank", "siamcommercial", "ไทยพาณิชย์", "ธไทยพาณิชย์", "ธนาคารไทยพาณิชย์"},
    "BANGKOK BANK": {"bangkokbank", "bangkok", "bbl", "bualuang", "กรุงเทพ", "ธกรุงเทพ", "ธนาคารกรุงเทพ"},
    "GSB": {"gsb", "governmentsavingsbank", "governmentsavings", "ออมสิน", "ธออมสิน", "ธนาคารออมสิน"},
    "TTB": {"ttb", "tmb", "tmbthanachart", "ttbtmbthanachart", "ทีเอ็มบีธนชาต", "ทหารไทยธนชาต"},
    "BAAC": {"baac", "ธกส", "ธกสธนาคารเพื่อการเกษตรและสหกรณ์การเกษตร", "เพื่อการเกษตร"},
    "KRUNGSRI": {"krungsri", "bay", "bankofayudhya", "กรุงศรี", "อยุธยา", "กรุงศรีอยุธยา", "ธกรุงศรี", "ธนาคารกรุงศรี"},
    "KKP": {"kkp", "kiatnakinphatra", "เกียรตินาคินภัทร", "เกียรตินาคิน"},
    "UOB": {"uob", "uobtmrw", "ยูโอบี", "ธยูโอบี", "ธนาคารยูโอบี"},
    "CIMB": {"cimb", "cimbthai", "ซีไอเอ็มบี", "ซีไอเอ็มบีไทย", "ธนาคารซีไอเอ็มบีไทย"},
    "LH BANK": {"lhbank", "landandhouse", "landandhouses", "แลนด์แอนด์เฮ้าส์", "ธนาคารแลนด์แอนด์เฮ้าส์"},
    "GHB": {"ghb", "governmenthousingbank", "ธอส", "อาคารสงเคราะห์", "ธนาคารอาคารสงเคราะห์"},
    "THAI CREDIT": {"thaicredit", "ไทยเครดิต", "ธนาคารไทยเครดิต"},
    "TISCO": {"tisco", "ทิสโก้", "ทิสโก", "ธนาคารทิสโก้"},
    "ICBC": {"icbc", "icbcthai", "ไอซีบีซี", "ธนาคารไอซีบีซี"},
    "STANDARD CHARTERED": {"standardchartered", "standardcharteredbank", "สแตนดาร์ดชาร์เตอร์ด"},
}

BANK_LIMITS = {
    "SCB": 200000.0,
    "KRUNGTHAI": 50000.0,
}


MISSING_BANK_VALUES = {"", "unknown", "unknownbank", "n/a", "na", "none", "null", "-", "ไม่ทราบ", "xxx", "xxxx", "xxxbank", "masked"}


@lru_cache(maxsize=8192)
def _bank_key_cached(text: str) -> str:
    return re.sub(r"[\s\*\u200b\u200c\u200d.\-_/|(),]+", "", text.lower())


def bank_key(value: Any) -> str:
    return _bank_key_cached(clean_display(value))


BANK_ALIAS_KEYS = {
    canonical: {_bank_key_cached(clean_display(alias)) for alias in aliases}
    for canonical, aliases in BANK_ALIASES.items()
}


def bank_needs_review(value: Any) -> bool:
    text = clean_display(value)
    key = bank_key(text)
    return (not key) or key in MISSING_BANK_VALUES or bool(re.fullmatch(r"x+(?:bank)?", key, flags=re.I)) or "ไม่ทราบ" in text


@lru_cache(maxsize=4096)
def _display_bank_cached(bank: str) -> str:
    key = bank_key(bank)
    if bank_needs_review(bank):
        return ""
    for canonical, normalized_aliases in BANK_ALIAS_KEYS.items():
        if key in normalized_aliases or any(alias and ((len(alias) >= 3 and key.startswith(alias)) or (len(alias) >= 4 and alias in key)) for alias in normalized_aliases):
            return canonical
    return bank.upper() if re.fullmatch(r"[A-Za-z0-9 ._-]+", bank) else bank


def display_bank(value: Any) -> str:
    return _display_bank_cached(clean_display(value))


display_bank.cache_clear = _display_bank_cached.cache_clear  # type: ignore[attr-defined]


def bank_limit_amount(value: Any) -> float:
    return float(BANK_LIMITS.get(display_bank(value), 0.0))


def is_known_bank(value: Any) -> bool:
    return display_bank(value) in BANK_ALIASES


def limit_check_enabled_for_flow(flow_type: str) -> bool:
    return normalize_flow_type(flow_type) != "deposit"


def limit_key_for(bank: Any, account: Any, name_key: str) -> str:
    bank_part = bank_key(display_bank(bank)) or "unknown-bank"
    account_part = bank_key(account) or name_key or "unknown-account"
    return f"{bank_part}|{account_part}"


def ensure_account_limit_table(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS account_limits (
          chat_id TEXT NOT NULL,
          limit_key TEXT NOT NULL,
          display_name TEXT,
          bank TEXT,
          account TEXT,
          limit_amount REAL DEFAULT 0,
          updated_at INTEGER NOT NULL,
          PRIMARY KEY(chat_id, limit_key)
        )
        """
    )


def account_limit_scope_key(chat_id: str = "", bot_key: str = "") -> str:
    """Return the storage scope for account limits.

    Historical limits were saved per Telegram chat_id. Operators now often select a
    whole company/bot (for example บริษัท 6) and set limits from the company-level
    account rows, where there is no single chat_id selected.  Store those as
    bot-scoped rows (`bot:<bot_key>`) while keeping chat-scoped rows as an override.
    """
    chat = clean_display(chat_id)
    if chat:
        return chat
    bot = clean_display(bot_key)
    if bot and bot not in {"__all__", "all"}:
        return f"bot:{bot}"
    return ""


def account_limit_scoped_key(scope: Any, limit_key: Any) -> str:
    return f"{clean_display(scope)}\x1f{clean_display(limit_key)}"


def load_account_limits(conn: sqlite3.Connection, chat_id: str, bot_key: str = "") -> Dict[str, Dict[str, Any]]:
    ensure_account_limit_table(conn)
    scopes: List[str] = []
    bot_scope = account_limit_scope_key("", bot_key)
    chat_scope = account_limit_scope_key(chat_id, "")
    # Bot-level limits are defaults; chat-level limits override the same limit_key.
    for scope in [bot_scope, chat_scope]:
        if scope and scope not in scopes:
            scopes.append(scope)
    if not scopes:
        return {}
    out: Dict[str, Dict[str, Any]] = {}
    for scope in scopes:
        rows = conn.execute("SELECT * FROM account_limits WHERE chat_id=?", (scope,)).fetchall()
        for r in rows:
            row = dict(r)
            out[str(r["limit_key"])] = row
            out[account_limit_scoped_key(scope, r["limit_key"])] = row
    return out


def load_bot_account_limits(conn: sqlite3.Connection) -> Dict[str, Dict[str, Any]]:
    """Load bot-scoped limits for the all-company overview.

    The overview groups rows from multiple bots/companies.  A plain dict keyed only
    by limit_key can collide across companies, so store only scoped keys here.
    """
    ensure_account_limit_table(conn)
    out: Dict[str, Dict[str, Any]] = {}
    rows = conn.execute("SELECT * FROM account_limits WHERE chat_id LIKE 'bot:%'").fetchall()
    for r in rows:
        out[account_limit_scoped_key(r["chat_id"], r["limit_key"])] = dict(r)
    return out


def account_limit_for(account_limits: Dict[str, Dict[str, Any]], limit_key: Any, bot_key: Any = "") -> Dict[str, Any]:
    key = clean_display(limit_key)
    if not key:
        return {}
    direct = account_limits.get(key)
    if direct:
        return direct
    bot_scope = account_limit_scope_key("", clean_display(bot_key))
    if bot_scope:
        return account_limits.get(account_limit_scoped_key(bot_scope, key), {})
    return {}


def account_limit_payload_error(payload: Dict[str, Any]) -> str:
    if not clean_display(payload.get("chat_id")):
        return "chat_id required"
    if not clean_display(payload.get("limit_key")):
        return "limit_key required"
    if not clean_display(payload.get("account")):
        return "account required"
    return ""


def save_account_limit(db_path: Path, chat_id: str, limit_key: str, display_name: str, bank: str, account: str, limit_amount: float) -> Dict[str, Any]:
    chat_scope = clean_display(chat_id)
    key = clean_display(limit_key)
    account_clean = clean_display(account)
    err = account_limit_payload_error({"chat_id": chat_scope, "limit_key": key, "account": account_clean})
    if err:
        return {"ok": False, "error": err, "status_code": 400}
    with connect(db_path) as conn:
        ensure_account_limit_table(conn)
        conn.execute(
            """
            INSERT INTO account_limits(chat_id, limit_key, display_name, bank, account, limit_amount, updated_at)
            VALUES (?,?,?,?,?,?,?)
            ON CONFLICT(chat_id, limit_key) DO UPDATE SET
              display_name=excluded.display_name,
              bank=excluded.bank,
              account=excluded.account,
              limit_amount=excluded.limit_amount,
              updated_at=excluded.updated_at
            """,
            (chat_scope, key, clean_display(display_name), display_bank(bank), account_clean, float(limit_amount or 0), int(time.time())),
        )
        conn.commit()
    return {"ok": True, "chat_id": chat_scope, "limit_key": key, "limit_amount": float(limit_amount or 0)}


def account_key_for(company_name: Any, bank: Any, account_no: Any, account_name: Any) -> str:
    raw = "|".join([clean_display(company_name), display_bank(bank), bank_key(account_no), clean_display(account_name)])
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:16]


def ensure_company_account_table(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS company_accounts (
          bot_key TEXT NOT NULL DEFAULT 'default',
          chat_id TEXT NOT NULL,
          account_key TEXT NOT NULL,
          company_name TEXT NOT NULL,
          bank TEXT,
          account_no TEXT,
          account_name TEXT,
          daily_limit REAL DEFAULT 0,
          active INTEGER DEFAULT 1,
          updated_at INTEGER NOT NULL,
          opening_balance REAL DEFAULT 0,
          opening_balance_date TEXT,
          PRIMARY KEY(bot_key, chat_id, account_key)
        )
        """
    )
    # migration: add columns if older schema
    try:
        cols = {row[1] for row in conn.execute("PRAGMA table_info(company_accounts)").fetchall()}
        if "opening_balance" not in cols:
            conn.execute("ALTER TABLE company_accounts ADD COLUMN opening_balance REAL DEFAULT 0")
        if "opening_balance_date" not in cols:
            conn.execute("ALTER TABLE company_accounts ADD COLUMN opening_balance_date TEXT")
    except sqlite3.OperationalError:
        pass


def ensure_slip_reviews_table(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS slip_reviews (
          slip_id TEXT PRIMARY KEY,
          reviewed_by TEXT,
          reviewed_at INTEGER NOT NULL,
          note TEXT
        )
        """
    )


def save_company_account(
    db_path: Path,
    bot_key: str,
    chat_id: str,
    company_name: str,
    bank: str,
    account_no: str,
    account_name: str,
    daily_limit: float = 0.0,
) -> Dict[str, Any]:
    bot_key = clean_display(bot_key) or "default"
    chat_id = str(chat_id or "")
    company_name = clean_company_name(company_name, bot_key) or bot_key
    bank = display_bank(bank)
    account_no = clean_display(account_no)
    account_name = clean_display(account_name)
    account_key = account_key_for(company_name, bank, account_no, account_name)
    with connect(db_path) as conn:
        ensure_company_account_table(conn)
        conn.execute(
            """
            INSERT INTO company_accounts(bot_key, chat_id, account_key, company_name, bank, account_no, account_name, daily_limit, active, updated_at)
            VALUES (?,?,?,?,?,?,?,?,1,?)
            ON CONFLICT(bot_key, chat_id, account_key) DO UPDATE SET
              company_name=excluded.company_name,
              bank=excluded.bank,
              account_no=excluded.account_no,
              account_name=excluded.account_name,
              daily_limit=excluded.daily_limit,
              active=1,
              updated_at=excluded.updated_at
            """,
            (bot_key, chat_id, account_key, company_name, bank, account_no, account_name, float(daily_limit or 0), int(time.time())),
        )
        conn.commit()
    return {
        "ok": True,
        "bot_key": bot_key,
        "chat_id": chat_id,
        "account_key": account_key,
        "company_name": company_name,
        "bank": bank,
        "account_no": account_no,
        "account_name": account_name,
        "daily_limit": float(daily_limit or 0),
    }


def load_company_accounts(conn: sqlite3.Connection, bot_key: str, chat_id: str) -> List[Dict[str, Any]]:
    ensure_company_account_table(conn)
    rows = conn.execute(
        """
        SELECT * FROM company_accounts
        WHERE bot_key=? AND chat_id=? AND COALESCE(active,1)=1
        ORDER BY company_name ASC, bank ASC, account_no ASC
        """,
        (clean_display(bot_key) or "default", str(chat_id or "")),
    ).fetchall()
    out = [clean_company_fields(dict(r)) for r in rows]
    return sorted(out, key=lambda r: (*dict_company_sort_key(r), clean_display(r.get("bank")), clean_display(r.get("account_no"))))


def public_telegram_bots() -> List[Dict[str, Any]]:
    rows = [
        {
            "bot_key": clean_display(cfg.get("bot_key", "default")) or "default",
            "company_name": clean_company_name(cfg.get("company_name", ""), cfg.get("bot_key", "default")),
            "token_env": cfg.get("token_env", ""),
            "has_token": bool(cfg.get("token")),
        }
        for cfg in telegram_bot_configs()
    ]
    return sorted(rows, key=dict_company_sort_key)


def token_for_bot(bot_key: str) -> str:
    bot_key = clean_display(bot_key) or "default"
    for cfg in telegram_bot_configs():
        if cfg.get("bot_key") == bot_key:
            return cfg.get("token", "")
    return ""


def fetch_slip_image(db_path: Path, slip_id: str) -> Tuple[bytes, str]:
    with connect(db_path) as conn:
        row = conn.execute("SELECT id, bot_key, file_id FROM slips WHERE id=?", (slip_id,)).fetchone()
    if not row or not row["file_id"]:
        raise FileNotFoundError("slip image not found")
    bot_key = clean_display(row["bot_key"]) or "default"
    file_id = clean_display(row["file_id"])
    token = token_for_bot(bot_key)
    if not token:
        raise FileNotFoundError("bot token not configured for this slip")
    cache_dir = DATA_DIR / "slip-images"
    cache_dir.mkdir(parents=True, exist_ok=True)
    cache_key = hashlib.sha256(f"{bot_key}|{file_id}".encode("utf-8")).hexdigest()
    cached = sorted(cache_dir.glob(cache_key + ".*"))
    if cached:
        body = cached[0].read_bytes()
        mime = mimetypes.guess_type(str(cached[0]))[0] or "image/jpeg"
        return body, mime
    meta = requests.get(f"{TELEGRAM_API}/bot{token}/getFile", params={"file_id": file_id}, timeout=30)
    meta.raise_for_status()
    obj = meta.json()
    file_path = obj.get("result", {}).get("file_path")
    if not file_path:
        raise FileNotFoundError("Telegram getFile did not return file_path")
    image = requests.get(f"{TELEGRAM_API}/file/bot{token}/{file_path}", timeout=90)
    image.raise_for_status()
    suffix = Path(file_path).suffix or ".jpg"
    out = cache_dir / f"{cache_key}{suffix}"
    out.write_bytes(image.content)
    mime = mimetypes.guess_type(str(out))[0] or "image/jpeg"
    return image.content, mime


def display_transferor_name(value: Any) -> str:
    text = clean_display(value)
    text = re.sub(r"^(นาย|นางสาว|นาง|น\.ส\.|MR\.?|MRS\.?|MS\.?|MISS)\s+", "", text, flags=re.I)
    text = re.sub(r"\s*\*\s*", "*", text)
    text = re.sub(r"\s+([.,])", r"\1", text)
    return clean_display(text).strip(" -_/|")


def transferor_key(value: Any) -> str:
    text = display_transferor_name(value).lower()
    text = re.sub(r"[\s\*\u200b\u200c\u200d.\-_/|()]+", "", text)
    return text


def names_similar(a: str, b: str) -> bool:
    if not a or not b:
        return False
    if a == b:
        return True
    if min(len(a), len(b)) < 4:
        return False
    ratio = SequenceMatcher(None, a, b).ratio()
    return ratio >= (0.78 if min(len(a), len(b)) <= 8 else 0.84)


def banks_compatible(a: Any, b: Any) -> bool:
    """Merge the same account when OCR missed one side's bank, but not across two known different banks."""
    left = display_bank(a)
    right = display_bank(b)
    return (not left) or (not right) or bank_key(left) == bank_key(right)


def masked_account_pattern(value: Any) -> str:
    """Account number pattern with digits plus x wildcards, stripping formatting.

    Slip OCRs sometimes expose the same bank account with different masks, e.g.
    `8XX-2-XXX31-5` vs `XXX-X-XX931-5`.  Keep only real digits and common mask
    markers so compatible masked formats can be grouped without inventing a full
    account number.
    """
    text = clean_display(value).lower()
    if not text:
        return ""
    text = re.sub(r"[\*•●○]+", "x", text)
    chars: List[str] = []
    for ch in text:
        if ch.isdigit():
            chars.append(ch)
        elif ch in {"x", "×"}:
            chars.append("x")
    return "".join(chars)


def masked_account_known_count(pattern: str) -> int:
    return sum(1 for ch in pattern if ch.isdigit())


def masked_account_search_token(pattern: str) -> str:
    """Best conservative SQL prefilter token from a masked account pattern."""
    runs = re.findall(r"\d{3,}", pattern or "")
    if not runs:
        return ""
    # Prefer the longest/right-most known digit run; account masks usually keep a suffix.
    return max(reversed(runs), key=len)


def masked_account_patterns_compatible(left: str, right: str) -> bool:
    """True when two same-length account masks do not conflict on known digits."""
    left = clean_display(left)
    right = clean_display(right)
    if not left or not right or len(left) != len(right):
        return False
    known_overlap = 0
    for a, b in zip(left, right):
        if a != "x" and b != "x":
            if a != b:
                return False
            known_overlap += 1
    if known_overlap < 3:
        return False
    return max(masked_account_known_count(left), masked_account_known_count(right)) >= 4


def masked_account_merge_pattern(current: str, incoming: str) -> str:
    """Merge compatible patterns into a stricter union; preserve current on conflict."""
    if not current:
        return incoming
    if not incoming or not masked_account_patterns_compatible(current, incoming):
        return current
    merged: List[str] = []
    for a, b in zip(current, incoming):
        merged.append(a if a != "x" else b)
    return "".join(merged)


def account_display_score(value: Any) -> Tuple[int, int, int]:
    pattern = masked_account_pattern(value)
    return (masked_account_known_count(pattern), len(pattern), len(clean_display(value)))


def better_account_display(candidate: Any, current: Any) -> bool:
    return account_display_score(candidate) > account_display_score(current)


def transferor_totals(conn: sqlite3.Connection, where_clause: str, params: List[Any], account_limits: Dict[str, Dict[str, Any]] | None = None, limit: int = 50) -> List[Dict[str, Any]]:
    account_limits = account_limits or {}
    rows = conn.execute(
        f"""
        SELECT COALESCE(NULLIF(transferor_name,''), NULLIF(sender_name,''), '(ไม่ทราบชื่อผู้โอน)') AS raw_name,
               COALESCE(NULLIF(company_name,''), NULLIF(bot_key,''), '') AS company_name,
               COALESCE(NULLIF(bot_key,''),'default') AS bot_key,
               COALESCE(NULLIF(from_bank,''), NULLIF(issuer_bank,''), '') AS bank,
               COALESCE(NULLIF(from_account,''), '') AS account,
               slip_date_display,
               slip_date_iso,
               COUNT(*) AS count,
               COALESCE(SUM(amount),0) AS amount,
               COALESCE(SUM(fee),0) AS fee
        FROM slips
        WHERE {where_clause}
        GROUP BY bot_key, company_name, raw_name, bank, account, slip_date_display, slip_date_iso
        ORDER BY amount DESC, count DESC, raw_name ASC
        """,
        params,
    ).fetchall()
    groups: List[Dict[str, Any]] = []
    for row in rows:
        display_name = display_transferor_name(row["raw_name"]) or "(ไม่ทราบชื่อผู้โอน)"
        bot_key = clean_display(row["bot_key"]) or "default"
        company_name = clean_company_name(row["company_name"], bot_key)
        bank = display_bank(row["bank"])
        account = clean_display(row["account"])
        key = transferor_key(display_name)
        account_key = bank_key(account)
        date_key, date_label, sort_date = date_bucket(row["slip_date_display"], row["slip_date_iso"])
        target = None
        for group in groups:
            if group["bot_key"] != bot_key or group["company_name"] != company_name:
                continue
            if account_key or group["account_key"]:
                if account_key and account_key == group["account_key"] and banks_compatible(bank, group["bank"]):
                    target = group
                    break
            elif banks_compatible(bank, group["bank"]) and (key == group["key"] or names_similar(key, group["key"])):
                target = group
                break
        if target is None:
            row_limit_key = limit_key_for(bank, account, key)
            target = {
                "display_name": display_name,
                "company_name": company_name,
                "bot_key": bot_key,
                "key": key,
                "bank": bank,
                "account": account,
                "account_key": account_key,
                "limit_key": row_limit_key,
                "count": 0,
                "amount": 0.0,
                "fee": 0.0,
                "daily_buckets": {},
                "aliases": [],
            }
            groups.append(target)
        row_count = int(row["count"] or 0)
        row_amount = float(row["amount"] or 0)
        row_fee = float(row["fee"] or 0)
        target["count"] += row_count
        target["amount"] += row_amount
        target["fee"] += row_fee
        day = target["daily_buckets"].setdefault(
            date_key,
            {"date_key": date_key, "date": date_label, "sort_date": sort_date, "count": 0, "amount": 0.0, "fee": 0.0},
        )
        day["count"] += row_count
        day["amount"] += row_amount
        day["fee"] += row_fee
        if bank and not target["bank"]:
            target["bank"] = bank
            target["limit_key"] = limit_key_for(bank, target["account"], target["key"])
        if display_name not in target["aliases"]:
            target["aliases"].append(display_name)
    out = []
    for group in sorted(groups, key=lambda g: (-g["amount"], -g["count"], g["display_name"]))[:limit]:
        name = f"{group['display_name']} ({group['bank']})" if group["bank"] else group["display_name"]
        limit_row = account_limit_for(account_limits, group["limit_key"], group.get("bot_key", ""))
        raw_limit = limit_row.get("limit_amount") if limit_row else bank_limit_amount(group["bank"])
        limit_amount = float(raw_limit or 0)
        daily_rows = sorted(
            group.get("daily_buckets", {}).values(),
            key=lambda d: (str(d.get("sort_date") or ""), float(d.get("amount") or 0)),
            reverse=True,
        )
        peak_day = max(daily_rows, key=lambda d: float(d.get("amount") or 0), default={})
        peak_daily_amount = float(peak_day.get("amount") or 0)
        peak_daily_count = int(peak_day.get("count") or 0)
        remaining = limit_amount - peak_daily_amount if limit_amount else 0.0
        out.append(
            {
                "name": name,
                "display_name": group["display_name"],
                "company_name": group.get("company_name", ""),
                "bot_key": group.get("bot_key", "default"),
                "bank": group["bank"],
                "account": group["account"],
                "limit_key": group["limit_key"],
                "limit_amount": limit_amount,
                "remaining_amount": remaining,
                "daily_remaining_amount": remaining,
                "over_limit": bool(limit_amount and peak_daily_amount > limit_amount),
                "daily_over_limit": bool(limit_amount and peak_daily_amount > limit_amount),
                "peak_daily_amount": peak_daily_amount,
                "peak_daily_count": peak_daily_count,
                "peak_daily_date": peak_day.get("date") or peak_day.get("date_key") or "",
                "count": group["count"],
                "amount": group["amount"],
                "range_amount": group["amount"],
                "fee": group["fee"],
                "daily_buckets": daily_rows,
                "aliases": group["aliases"],
            }
        )
    return out


def daily_account_totals(conn: sqlite3.Connection, where_clause: str, params: List[Any], account_limits: Dict[str, Dict[str, Any]] | None = None, limit: int = 120) -> List[Dict[str, Any]]:
    """Summarize counted slip usage per day and transferor account."""
    account_limits = account_limits or {}
    rows = conn.execute(
        f"""
        SELECT COALESCE(NULLIF(transferor_name,''), NULLIF(sender_name,''), '(ไม่ทราบชื่อผู้โอน)') AS raw_name,
               COALESCE(NULLIF(company_name,''), NULLIF(bot_key,''), '') AS company_name,
               COALESCE(NULLIF(bot_key,''),'default') AS bot_key,
               COALESCE(NULLIF(from_bank,''), NULLIF(issuer_bank,''), '') AS bank,
               COALESCE(NULLIF(from_account,''), '') AS account,
               slip_date_display,
               slip_date_iso,
               COUNT(*) AS count,
               COALESCE(SUM(amount),0) AS amount,
               COALESCE(SUM(fee),0) AS fee
        FROM slips
        WHERE {where_clause}
          AND NULLIF(TRIM(COALESCE(from_account,'')), '') IS NOT NULL
        GROUP BY bot_key, company_name, raw_name, bank, account, slip_date_display, slip_date_iso
        ORDER BY COALESCE(NULLIF(slip_date_iso,''), slip_date_display) DESC, amount DESC, raw_name ASC
        """,
        params,
    ).fetchall()
    groups: List[Dict[str, Any]] = []
    for row in rows:
        display_name = display_transferor_name(row["raw_name"]) or "(ไม่ทราบชื่อผู้โอน)"
        bot_key = clean_display(row["bot_key"]) or "default"
        company_name = clean_company_name(row["company_name"], bot_key)
        bank = display_bank(row["bank"])
        account = clean_display(row["account"])
        date_key, date_label, sort_date = date_bucket(row["slip_date_display"], row["slip_date_iso"])
        name_key = transferor_key(display_name)
        account_key = bank_key(account)
        target = None
        for group in groups:
            if group["date_key"] != date_key or group["bot_key"] != bot_key or group["company_name"] != company_name:
                continue
            if account_key or group["account_key"]:
                if account_key and account_key == group["account_key"] and banks_compatible(bank, group["bank"]):
                    target = group
                    break
            elif banks_compatible(bank, group["bank"]) and (name_key == group["name_key"] or names_similar(name_key, group["name_key"])):
                target = group
                break
        if target is None:
            target = {
                "date_key": date_key,
                "date": date_label,
                "sort_date": sort_date,
                "display_name": display_name,
                "company_name": company_name,
                "bot_key": bot_key,
                "name_key": name_key,
                "bank": bank,
                "account": account,
                "account_key": account_key,
                "limit_key": limit_key_for(bank, account, name_key),
                "count": 0,
                "amount": 0.0,
                "fee": 0.0,
                "aliases": [],
            }
            groups.append(target)
        target["count"] += int(row["count"] or 0)
        target["amount"] += float(row["amount"] or 0)
        target["fee"] += float(row["fee"] or 0)
        if bank and not target["bank"]:
            target["bank"] = bank
            target["limit_key"] = limit_key_for(bank, target["account"], target["name_key"])
        if display_name not in target["aliases"]:
            target["aliases"].append(display_name)
    out: List[Dict[str, Any]] = []
    sorted_groups = sorted(groups, key=lambda g: str(g["display_name"]))
    sorted_groups = sorted(sorted_groups, key=lambda g: float(g["amount"] or 0), reverse=True)
    sorted_groups = sorted(sorted_groups, key=dict_company_sort_key)
    sorted_groups = sorted(sorted_groups, key=lambda g: str(g["sort_date"] or ""), reverse=True)
    selected_groups = sorted_groups if not limit or limit <= 0 else sorted_groups[:limit]
    for group in selected_groups:
        limit_row = account_limit_for(account_limits, group["limit_key"], group.get("bot_key", ""))
        raw_limit = limit_row.get("limit_amount") if limit_row else bank_limit_amount(group["bank"])
        daily_limit = float(raw_limit or 0)
        amount = float(group["amount"] or 0)
        remaining = daily_limit - amount if daily_limit else 0.0
        name = f"{group['display_name']} ({group['bank']})" if group["bank"] else group["display_name"]
        out.append(
            {
                "date_key": group["date_key"],
                "date": group["date"],
                "name": name,
                "display_name": group["display_name"],
                "company_name": group.get("company_name", ""),
                "bot_key": group.get("bot_key", "default"),
                "bank": group["bank"],
                "account": group["account"],
                "limit_key": group["limit_key"],
                "daily_limit": daily_limit,
                "limit_amount": daily_limit,
                "remaining_amount": remaining,
                "over_limit": bool(daily_limit and amount > daily_limit),
                "count": int(group["count"] or 0),
                "amount": amount,
                "fee": float(group["fee"] or 0),
                "aliases": group["aliases"],
            }
        )
    return out


def withdraw_limit_usage_summary(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Aggregate daily withdrawal account rows into company-level amount vs capacity."""
    groups: Dict[str, Dict[str, Any]] = {}
    for row in rows or []:
        bot_key = clean_display(row.get("bot_key")) or "default"
        company_name = clean_company_name(row.get("company_name"), bot_key)
        group = groups.setdefault(
            bot_key,
            {
                "bot_key": bot_key,
                "company_name": company_name,
                "withdraw_count": 0,
                "withdraw_amount": 0.0,
                "limit_amount": 0.0,
                "remaining_amount": 0.0,
                "over_limit_amount": 0.0,
                "usage_percent": 0.0,
                "account_count": 0,
                "account_day_count": 0,
                "known_limit_account_day_count": 0,
                "no_limit_account_day_count": 0,
                "over_limit": False,
                "accounts": set(),
                "days": set(),
            },
        )
        amount = float(row.get("amount") or 0)
        count = int(row.get("count") or 0)
        limit_amount = float(row.get("daily_limit") or row.get("limit_amount") or 0)
        account_key = f"{bank_key(row.get('bank'))}|{bank_key(row.get('account'))}"
        date_key = clean_display(row.get("date_key") or row.get("date"))
        group["withdraw_count"] += count
        group["withdraw_amount"] += amount
        group["limit_amount"] += limit_amount
        group["account_day_count"] += 1
        if account_key.strip("|"):
            group["accounts"].add(account_key)
        if date_key:
            group["days"].add(date_key)
        if limit_amount > 0:
            group["known_limit_account_day_count"] += 1
            if amount > limit_amount:
                group["over_limit_amount"] += amount - limit_amount
        else:
            group["no_limit_account_day_count"] += 1
    out: List[Dict[str, Any]] = []
    for group in groups.values():
        limit_amount = float(group["limit_amount"] or 0)
        withdraw_amount = float(group["withdraw_amount"] or 0)
        remaining = limit_amount - withdraw_amount if limit_amount else 0.0
        group["remaining_amount"] = remaining
        group["usage_percent"] = (withdraw_amount / limit_amount * 100.0) if limit_amount else 0.0
        group["over_limit"] = bool(limit_amount and withdraw_amount > limit_amount)
        group["account_count"] = len(group.pop("accounts"))
        group["day_count"] = len(group.pop("days"))
        out.append(group)
    out.sort(key=dict_company_sort_key)
    return out


def withdraw_limit_usage_totals(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    withdraw_amount = sum(float(r.get("withdraw_amount") or 0) for r in rows or [])
    limit_amount = sum(float(r.get("limit_amount") or 0) for r in rows or [])
    over_amount = sum(float(r.get("over_limit_amount") or 0) for r in rows or [])
    account_count = sum(int(r.get("account_count") or 0) for r in rows or [])
    account_day_count = sum(int(r.get("account_day_count") or 0) for r in rows or [])
    return {
        "withdraw_limit_capacity_amount": limit_amount,
        "withdraw_limit_remaining_amount": (limit_amount - withdraw_amount) if limit_amount else 0.0,
        "withdraw_limit_over_amount": over_amount,
        "withdraw_limit_usage_percent": (withdraw_amount / limit_amount * 100.0) if limit_amount else 0.0,
        "withdraw_limit_account_count": account_count,
        "withdraw_limit_account_day_count": account_day_count,
    }


def company_account_row_identity(row: sqlite3.Row) -> Tuple[str, str, str, str]:
    flow = flow_type_for_title(row["chat_title"], row["bot_key"], row["chat_id"])
    if flow == "deposit":
        return clean_display(row["to_account"]), display_bank(row["to_bank"]), "บัญชีรับเงิน/ปลายทาง", flow
    if flow == "withdraw":
        return clean_display(row["from_account"]), display_bank(row["from_bank"]), "บัญชีผู้โอน/ต้นทาง", flow
    account = clean_display(row["to_account"] or row["from_account"])
    bank = display_bank(row["to_bank"] or row["from_bank"])
    return account, bank, "บัญชีจากสลิป", flow


def company_account_daily_totals(conn: sqlite3.Connection, where_clause: str, params: List[Any], search: str = "", limit: int = 300) -> List[Dict[str, Any]]:
    """Daily account rows grouped by company, account, date, and flow."""
    search_clause, search_params = slip_search_clause("slips", search)
    rows = conn.execute(
        f"""
        SELECT COALESCE(NULLIF(bot_key,''),'default') AS bot_key,
               COALESCE(NULLIF(company_name,''), COALESCE(NULLIF(bot_key,''),'default')) AS company_name,
               chat_id, chat_title, slip_date_display, slip_date_iso,
               from_bank, from_account, to_bank, to_account,
               COUNT(*) AS count,
               COALESCE(SUM(amount),0) AS amount,
               COALESCE(SUM(fee),0) AS fee
        FROM slips
        WHERE {where_clause}
          AND (NULLIF(TRIM(COALESCE(from_account,'')), '') IS NOT NULL OR NULLIF(TRIM(COALESCE(to_account,'')), '') IS NOT NULL)
          {search_clause}
        GROUP BY bot_key, company_name, chat_id, chat_title, slip_date_display, slip_date_iso, from_bank, from_account, to_bank, to_account
        """,
        [*params, *search_params],
    ).fetchall()
    groups: Dict[Tuple[str, str, str, str, str, str], Dict[str, Any]] = {}
    deposit_group_index: Dict[Tuple[str, str, str, str, str], List[Dict[str, Any]]] = {}
    for row in rows:
        bot_key = clean_display(row["bot_key"]) or "default"
        company_name = clean_company_name(row["company_name"], bot_key)
        account, bank, account_role, flow = company_account_row_identity(row)
        if not account:
            continue
        date_key, date_label, sort_date = date_bucket(row["slip_date_display"], row["slip_date_iso"])
        account_pattern = masked_account_pattern(account)
        base_key = (bot_key, flow, date_key, bank_key(bank), account_role)
        group = None
        if flow == "deposit":
            for candidate in deposit_group_index.get(base_key, []):
                if masked_account_patterns_compatible(candidate.get("account_pattern", ""), account_pattern):
                    group = candidate
                    break
        key = (bot_key, flow, date_key, bank_key(bank), bank_key(account), account_role)
        if group is None:
            group = groups.setdefault(
                key,
                {
                    "bot_key": bot_key,
                    "company_name": company_name,
                    "flow_type": flow,
                    "flow_label": flow_label(flow),
                    "date_key": date_key,
                    "date": date_label,
                    "sort_date": sort_date,
                    "bank": bank,
                    "account": account,
                    "account_role": account_role,
                    "account_pattern": account_pattern,
                    "account_aliases": [],
                    "count": 0,
                    "amount": 0.0,
                    "fee": 0.0,
                    "chat_titles": [],
                },
            )
            if flow == "deposit":
                deposit_group_index.setdefault(base_key, []).append(group)
        elif flow == "deposit":
            group["account_pattern"] = masked_account_merge_pattern(group.get("account_pattern", ""), account_pattern)
        if account and account not in group["account_aliases"]:
            group["account_aliases"].append(account)
        if account and better_account_display(account, group.get("account", "")):
            group["account"] = account
        group["count"] += int(row["count"] or 0)
        group["amount"] += float(row["amount"] or 0)
        group["fee"] += float(row["fee"] or 0)
        title = clean_display(row["chat_title"])
        if title and title not in group["chat_titles"]:
            group["chat_titles"].append(title)
    out = []
    for group in groups.values():
        item = dict(group)
        item.pop("account_pattern", None)
        out.append(item)
    out.sort(key=lambda r: (clean_display(r.get("account")), FLOW_TYPE_LABELS.get(r.get("flow_type"), "")))
    out.sort(key=lambda r: str(r.get("sort_date") or ""), reverse=True)
    out.sort(key=dict_company_sort_key)
    return out[:limit]


def cross_company_withdraw_flow_type(flow_type: str) -> str:
    """Cross-company account audits are for withdrawal/source accounts only."""
    flow = normalize_flow_type(flow_type)
    return "withdraw" if flow in {"all", "withdraw"} else ""


def cross_company_account_usage(conn: sqlite3.Connection, scope: str = "all", flow_type: str = "all", selected_bot_key: str = "", search: str = "", limit: int = 120) -> List[Dict[str, Any]]:
    """Find withdrawal accounts that appear under more than one company in the selected period."""
    effective_flow = cross_company_withdraw_flow_type(flow_type)
    if not effective_flow:
        return []
    clause, params, _ = global_scope_where(scope, success_only=True)
    clause, params = apply_flow_sql(clause, params, effective_flow)
    rows = conn.execute(
        f"""
        SELECT COALESCE(NULLIF(bot_key,''),'default') AS bot_key,
               COALESCE(NULLIF(company_name,''), COALESCE(NULLIF(bot_key,''),'default')) AS company_name,
               chat_id, chat_title, slip_date_display, slip_date_iso,
               from_bank, from_account, to_bank, to_account,
               COUNT(*) AS count,
               COALESCE(SUM(amount),0) AS amount
        FROM slips
        WHERE {clause}
          AND (NULLIF(TRIM(COALESCE(from_account,'')), '') IS NOT NULL OR NULLIF(TRIM(COALESCE(to_account,'')), '') IS NOT NULL)
        GROUP BY bot_key, company_name, chat_id, chat_title, slip_date_display, slip_date_iso, from_bank, from_account, to_bank, to_account
        """,
        params,
    ).fetchall()
    selected_bot = clean_display(selected_bot_key)
    groups: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        account, bank, account_role, flow = company_account_row_identity(row)
        if not account:
            continue
        key = f"{bank_key(bank)}|{bank_key(account)}"
        bot_key = clean_display(row["bot_key"]) or "default"
        company_name = clean_company_name(row["company_name"], bot_key)
        group = groups.setdefault(
            key,
            {
                "bank": bank,
                "account": account,
                "account_key": key,
                "roles": [],
                "flows": [],
                "companies_map": {},
                "days_map": {},
                "total_count": 0,
                "total_amount": 0.0,
            },
        )
        if account_role and account_role not in group["roles"]:
            group["roles"].append(account_role)
        if flow and flow not in group["flows"]:
            group["flows"].append(flow)
        comp = group["companies_map"].setdefault(
            bot_key,
            {
                "bot_key": bot_key,
                "company_name": company_name,
                "count": 0,
                "amount": 0.0,
                "deposit_amount": 0.0,
                "withdraw_amount": 0.0,
                "flows": [],
                "days_map": {},
            },
        )
        comp["count"] += int(row["count"] or 0)
        row_count = int(row["count"] or 0)
        row_amount = float(row["amount"] or 0)
        comp["amount"] += row_amount
        if flow == "deposit":
            comp["deposit_amount"] += row_amount
        elif flow == "withdraw":
            comp["withdraw_amount"] += row_amount
        if flow and flow not in comp["flows"]:
            comp["flows"].append(flow)
        date_key, date_label, sort_date = date_bucket(row["slip_date_display"], row["slip_date_iso"])
        comp_day = comp["days_map"].setdefault(
            date_key,
            {"date_key": date_key, "date": date_label, "sort_date": sort_date, "count": 0, "amount": 0.0, "flows": []},
        )
        comp_day["count"] += row_count
        comp_day["amount"] += row_amount
        if flow and flow not in comp_day["flows"]:
            comp_day["flows"].append(flow)
        group_day = group["days_map"].setdefault(
            date_key,
            {"date_key": date_key, "date": date_label, "sort_date": sort_date, "count": 0, "amount": 0.0},
        )
        group_day["count"] += row_count
        group_day["amount"] += row_amount
        group["total_count"] += row_count
        group["total_amount"] += row_amount
    needle = normalize_match_text(search)
    out: List[Dict[str, Any]] = []
    for group in groups.values():
        companies = sorted(group.pop("companies_map").values(), key=dict_company_sort_key)
        for company in companies:
            days = list(company.pop("days_map", {}).values())
            days.sort(key=lambda r: (str(r.get("sort_date") or ""), str(r.get("date") or "")), reverse=True)
            for day in days:
                day["flow_labels"] = [flow_label(f) for f in day.pop("flows", [])]
                day.pop("sort_date", None)
            company["days"] = days
        if len(companies) < 2:
            continue
        if selected_bot and selected_bot not in {"__all__", "all"} and selected_bot not in {c["bot_key"] for c in companies}:
            continue
        if needle:
            hay = normalize_match_text(" ".join([group.get("bank", ""), group.get("account", ""), *[c.get("company_name", "") for c in companies]]))
            if needle not in hay:
                continue
        first = companies[0] if companies else {}
        days = list(group.pop("days_map", {}).values())
        days.sort(key=lambda r: (str(r.get("sort_date") or ""), str(r.get("date") or "")), reverse=True)
        for day in days:
            day.pop("sort_date", None)
        item = {
            **group,
            "company_count": len(companies),
            "companies": companies,
            "days": days,
            "company_name": first.get("company_name", ""),
            "bot_key": first.get("bot_key", ""),
            "flow_labels": [flow_label(f) for f in group.get("flows", [])],
        }
        out.append(item)
    out.sort(key=lambda r: (dict_company_sort_key(r), clean_display(r.get("bank")), clean_display(r.get("account"))))
    return out[:limit]


@lru_cache(maxsize=8192)
def _date_bucket_cached(display_raw: str, iso_raw: str) -> Tuple[str, str, str]:
    display_label, display_iso = normalize_date_parts(display_raw)
    iso_label, iso_iso = normalize_date_parts(iso_raw)
    label = display_label or iso_label or display_raw or iso_raw or "(ไม่ทราบวันที่)"
    sort_key = display_iso or iso_iso or label
    group_key = display_iso or iso_iso or label
    return group_key, label, sort_key


def date_bucket(display_value: Any, iso_value: Any = "") -> Tuple[str, str, str]:
    display_raw = clean_display(display_value)
    iso_raw = clean_display(iso_value)
    return _date_bucket_cached(display_raw, iso_raw)


date_bucket.cache_clear = _date_bucket_cached.cache_clear  # type: ignore[attr-defined]


def date_totals(conn: sqlite3.Connection, where_clause: str, params: List[Any], limit: int = 50) -> List[Dict[str, Any]]:
    rows = conn.execute(
        f"""
        SELECT slip_date_display, slip_date_iso,
               COALESCE(amount,0) AS amount,
               COALESCE(fee,0) AS fee
        FROM slips
        WHERE {where_clause}
        """,
        params,
    ).fetchall()
    groups: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        key, label, sort_key = date_bucket(row["slip_date_display"], row["slip_date_iso"])
        group = groups.setdefault(key, {"date": label, "sort_date": sort_key, "count": 0, "amount": 0.0, "fee": 0.0})
        group["count"] += 1
        group["amount"] += float(row["amount"] or 0)
        group["fee"] += float(row["fee"] or 0)
        if sort_key > group["sort_date"]:
            group["sort_date"] = sort_key
    out = sorted(groups.values(), key=lambda r: (r["sort_date"], r["date"]), reverse=True)[:limit]
    return [{"date": r["date"], "count": r["count"], "amount": r["amount"], "fee": r["fee"]} for r in out]


def daily_flow_totals(conn: sqlite3.Connection, where_clause: str, params: List[Any], limit: int = 60) -> List[Dict[str, Any]]:
    """Daily counted totals split into withdraw vs deposit/top-up for the dashboard chart."""
    rows = conn.execute(
        f"""
        SELECT COALESCE(NULLIF(bot_key,''),'default') AS bot_key, chat_id,
               slip_date_display, slip_date_iso, chat_title,
               COALESCE(amount,0) AS amount
        FROM slips
        WHERE {where_clause}
        """,
        params,
    ).fetchall()
    groups: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        key, label, sort_key = date_bucket(row["slip_date_display"], row["slip_date_iso"])
        flow = flow_type_for_title(row["chat_title"], row["bot_key"], row["chat_id"])
        if flow not in {"deposit", "withdraw"}:
            flow = "other"
        group = groups.setdefault(
            key,
            {
                "date_key": key,
                "date": label,
                "sort_date": sort_key,
                "withdraw_count": 0,
                "withdraw_amount": 0.0,
                "deposit_count": 0,
                "deposit_amount": 0.0,
                "other_count": 0,
                "other_amount": 0.0,
                "total_count": 0,
                "total_amount": 0.0,
            },
        )
        amount = float(row["amount"] or 0)
        group[f"{flow}_count"] += 1
        group[f"{flow}_amount"] += amount
        group["total_count"] += 1
        group["total_amount"] += amount
        if sort_key > group["sort_date"]:
            group["sort_date"] = sort_key
    out = sorted(groups.values(), key=lambda r: (r["sort_date"], r["date"]), reverse=True)[:limit]
    for row in out:
        row.pop("sort_date", None)
    return out


def normalize_match_date(value: Any) -> str:
    display, iso = normalize_date_parts(value)
    return iso or display or clean_display(value)


def normalize_match_text(value: Any) -> str:
    return re.sub(r"[\s\*\u200b\u200c\u200d.\-_/|()]+", "", clean_display(value).lower())


def extract_time_text(*values: Any) -> str:
    """Return HH:MM from Excel time cells or combined date/time text."""
    for value in values:
        if value in (None, ""):
            continue
        if isinstance(value, dt.datetime):
            return value.strftime("%H:%M")
        if isinstance(value, dt.time):
            return value.strftime("%H:%M")
        text = clean_display(value)
        match = re.search(r"(?<!\d)(\d{1,2}:\d{2})(?::\d{2})?(?!\d)", text)
        if match:
            return match.group(1)
    return ""


def detect_column(headers: List[str], aliases: List[str]) -> int:
    normalized = [normalize_match_text(h) for h in headers]
    alias_norm = [normalize_match_text(a) for a in aliases]
    for idx, head in enumerate(normalized):
        if head and any(alias == head for alias in alias_norm if alias):
            return idx
    for idx, head in enumerate(normalized):
        if any(alias in head or head in alias for alias in alias_norm if alias):
            return idx
    return -1


def backend_flow_type_for(*values: Any) -> str:
    text = clean_display(" ".join(clean_display(v) for v in values)).lower()
    if not text:
        return ""
    # These rows are accounting adjustments, not money entering/leaving via slip groups.
    if any(token in text for token in ["ยกเลิก", "cancel", "void", "โบนัส", "bonus", "เก็บ"]):
        return "other"
    if any(token in text for token in ["ถอน", "withdraw", "withdrawal"]):
        return "withdraw"
    if any(token in text for token in ["ฝาก", "เติม", "deposit", "topup", "top-up", "ออโต้", "auto"]):
        return "deposit"
    return ""


def parse_backend_excel(path: Path) -> List[Dict[str, Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    assert ws is not None
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header_idx = 0
    for i, row in enumerate(rows[:10]):
        texts = [clean_display(c) for c in row]
        if any(t for t in texts) and detect_column(texts, ["amount", "ยอด", "ยอดเงิน", "จำนวนเงิน", "total"]) >= 0:
            header_idx = i
            break
    headers = [clean_display(c) for c in rows[header_idx]]
    amount_i = detect_column(headers, ["amount", "ยอด", "ยอดเงิน", "จำนวน", "จำนวนเงิน", "เงิน", "total", "deposit", "credit", "gross amount"])
    received_i = detect_column(headers, ["จำนวนที่ได้รับ", "ยอดที่ได้รับ", "ได้รับ", "received", "received amount", "net amount", "credit amount"])
    date_i = detect_column(headers, ["date", "วันที่", "เวลา", "transaction date", "วันเวลา", "วันที่เวลา", "วันที่/เวลา", "วันทำรายการ", "created date", "created at"])
    time_i = detect_column(headers, ["time", "เวลา", "transaction time", "เวลาทำรายการ", "created time"])
    type_i = detect_column(headers, ["ประเภท", "type", "transaction type", "รายการประเภท"])
    operation_i = detect_column(headers, ["ประเภทดำเนินการ", "ดำเนินการ", "operation", "operation type", "action", "method"])
    name_i = detect_column(headers, ["name", "ชื่อ", "ชื่อผู้โอน", "ผู้โอน", "ยูสเซอร์", "ยูสเซอร์เนม", "user", "username", "user name", "member", "ลูกค้า", "customer", "payer", "transferor"])
    bank_i = detect_column(headers, ["bank", "ธนาคาร", "ธนาคารผู้โอน", "ธนาคารลูกค้า", "customer bank", "bank name"])
    ref_i = detect_column(headers, ["ref", "reference", "reference no", "เลขอ้างอิง", "หมายเลขอ้างอิง", "รหัส", "รหัสรายการ", "เลขที่รายการ", "รายการ", "transaction", "transaction id", "transaction code", "code", "seq", "id", "order id"])
    note_i = detect_column(headers, ["หมายเหตุ", "remark", "remarks", "memo", "note", "description"])
    source_i = detect_column(headers, ["ไฟล์", "แหล่งที่มา", "ไฟล์/แหล่งที่มา", "source", "filename", "file", "import"])
    if amount_i < 0 and received_i < 0:
        raise ValueError("ไม่พบคอลัมน์ยอดเงินใน Excel หลังบ้าน")
    parsed: List[Dict[str, Any]] = []
    for row_no, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        values = list(row)
        gross_amount = parse_number(values[amount_i] if 0 <= amount_i < len(values) else 0)
        received_amount = parse_number(values[received_i] if 0 <= received_i < len(values) else 0)
        amount = received_amount or gross_amount
        if not amount:
            continue
        date_raw = values[date_i] if 0 <= date_i < len(values) else ""
        raw_time_value = values[time_i] if 0 <= time_i < len(values) else ""
        time_text = extract_time_text(raw_time_value, date_raw)
        name = clean_display(values[name_i]) if 0 <= name_i < len(values) else ""
        bank = display_bank(values[bank_i]) if 0 <= bank_i < len(values) else ""
        transaction_type = clean_display(values[type_i]) if 0 <= type_i < len(values) else ""
        operation_type = clean_display(values[operation_i]) if 0 <= operation_i < len(values) else ""
        backend_flow = backend_flow_type_for(transaction_type, operation_type)
        reference = clean_display(values[ref_i]) if 0 <= ref_i < len(values) else ""
        note = clean_display(values[note_i]) if 0 <= note_i < len(values) else ""
        if not reference:
            reference = note
        source = clean_display(values[source_i]) if 0 <= source_i < len(values) else Path(path).name
        date_label, date_iso = normalize_date_parts(date_raw)
        parsed.append(
            {
                "row": row_no,
                "date": date_label or clean_display(date_raw),
                "date_key": date_iso or normalize_match_date(date_raw),
                "time": time_text,
                "name": name,
                "name_key": normalize_match_text(name),
                "bank": bank,
                "bank_key": normalize_match_text(bank),
                "amount": float(amount),
                "gross_amount": float(gross_amount or amount),
                "received_amount": float(received_amount or amount),
                "reference": reference,
                "reference_key": normalize_match_text(reference),
                "transaction_type": transaction_type,
                "operation_type": operation_type,
                "flow_type": backend_flow,
                "flow_label": flow_label(backend_flow) if backend_flow else "",
                "note": note,
                "source": source,
            }
        )
    return parsed


def backend_date_matches_scope(row: Dict[str, Any], scope: str) -> bool:
    range_start, range_end, _ = scope_date_range(scope)
    normalized, _ = scope_to_date(scope)
    normalized = clean_display(normalized)
    date_key = normalize_match_date(row.get("date_key") or row.get("date"))
    if range_start or range_end:
        if not date_key:
            return False
        if range_start and date_key < range_start:
            return False
        if range_end and date_key > range_end:
            return False
        return True
    if normalized in {"", "all", "open"}:
        return True
    if re.match(r"^\d{4}-\d{2}-\d{2}$", normalized):
        return date_key == normalized
    return date_key == normalize_match_date(normalized)


def backend_flow_matches(row: Dict[str, Any], flow_type: str) -> bool:
    flow = normalize_flow_type(flow_type)
    row_flow = clean_display(row.get("flow_type"))
    if flow in {"deposit", "withdraw"}:
        # Legacy Excel files may not have ประเภท/ประเภทดำเนินการ columns; keep
        # blank-flow rows so older reconciliation still works, but honor explicit
        # backend classifications when they are present.
        return row_flow in {"", flow}
    if flow == "other":
        return row_flow == "other"
    # For all-flow reconciliation, keep legacy files with no flow columns, but
    # exclude rows explicitly classified as accounting adjustments.
    return row_flow != "other"


def filter_backend_reconcile_rows(rows: List[Dict[str, Any]], scope: str, flow_type: str) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    kept: List[Dict[str, Any]] = []
    filtered: List[Dict[str, Any]] = []
    for row in rows:
        if backend_date_matches_scope(row, scope) and backend_flow_matches(row, flow_type):
            kept.append(row)
        else:
            filtered.append(row)
    return kept, filtered


def statement_flow_type_for(*values: Any) -> str:
    text = clean_display(" ".join(clean_display(v) for v in values)).lower()
    if not text:
        return ""
    if any(token in text for token in ["โอนออก", "เงินออก", "ถอน", "withdraw", "withdrawal", "debit", "จ่าย", "paid", "payment"]):
        return "withdraw"
    if any(token in text for token in ["รับเงินคืน", "เงินคืน", "รับเงิน", "เงินเข้า", "โอนเข้า", "ฝาก", "deposit", "credit", "refund", "received"]):
        return "deposit"
    return ""


def statement_default_flow_for_headers(headers: List[str]) -> str:
    """Infer source-level flow for statement exports that contain only one direction.

    True Wallet Dashboard exports are receive-only CSV files, so rows without an
    explicit debit/credit column should still be treated as deposit rows.
    """
    if (
        detect_column(headers, ["sender_mobile", "เบอร์ผู้โอน", "ผู้โอน", "sender mobile", "sender"]) >= 0
        and detect_column(headers, ["receiver_mobile", "เบอร์ผู้รับ", "ผู้รับ", "receiver mobile", "receiver"]) >= 0
        and detect_column(headers, ["amount_baht", "จำนวน (บาท)", "amount", "จำนวน"]) >= 0
    ):
        return "deposit"
    if detect_column(headers, ["received_time", "received time"]) >= 0 and detect_column(headers, ["amount_baht"]) >= 0:
        return "deposit"
    return ""


def parse_statement_rows(rows: List[Any], source_name: str) -> List[Dict[str, Any]]:
    """Parse statement-like rows from XLSX/CSV into amount/time/date reconciliation rows."""
    if not rows:
        return []
    header_idx = 0
    for i, row in enumerate(rows[:10]):
        texts = [clean_display(c) for c in row]
        if any(t for t in texts) and (
            detect_column(texts, ["โอนออก", "เงินออก", "debit", "withdraw", "ถอน"]) >= 0
            or detect_column(texts, ["รับเงินคืน", "เงินเข้า", "credit", "deposit", "ฝาก"]) >= 0
            or detect_column(texts, ["amount", "amount_baht", "ยอด", "จำนวน", "จำนวนเงิน", "เงิน", "total"]) >= 0
        ):
            header_idx = i
            break
    headers = [clean_display(c) for c in rows[header_idx]]
    date_i = detect_column(headers, ["date", "วันที่", "วันเวลา", "วันที่เวลา", "วันที่/เวลา", "วันทำรายการ", "transaction date", "created at", "received_time", "received time"])
    time_i = detect_column(headers, ["time", "เวลา", "transaction time", "เวลาทำรายการ", "received_time", "received time"])
    desc_i = detect_column(headers, ["รายการ", "รายละเอียด", "description", "memo", "message", "note", "หมายเหตุ", "ประเภท", "type", "transaction type"])
    debit_i = detect_column(headers, ["โอนออก", "เงินออก", "ยอดถอน", "ถอน", "debit", "withdraw", "withdrawal", "paid out", "payment"])
    credit_i = detect_column(headers, ["รับเงินคืน", "เงินเข้า", "ยอดฝาก", "ฝาก", "credit", "deposit", "received", "refund"])
    amount_i = detect_column(headers, ["amount_baht", "amount", "ยอด", "ยอดเงิน", "จำนวน", "จำนวนเงิน", "เงิน", "total"])
    ref_i = detect_column(headers, ["ref", "reference", "เลขอ้างอิง", "รหัสรายการ", "transaction id", "transaction_id", "เลขที่รายการ"])
    sender_i = detect_column(headers, ["sender_mobile", "เบอร์ผู้โอน", "ผู้โอน", "sender mobile", "sender"])
    receiver_i = detect_column(headers, ["receiver_mobile", "เบอร์ผู้รับ", "ผู้รับ", "receiver mobile", "receiver"])
    default_flow = statement_default_flow_for_headers(headers)
    if amount_i < 0 and debit_i < 0 and credit_i < 0:
        raise ValueError("ไม่พบคอลัมน์ยอดเงินในรายการเดินบัญชี")
    parsed: List[Dict[str, Any]] = []
    for row_no, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        values = list(row)
        date_raw = values[date_i] if 0 <= date_i < len(values) else ""
        time_raw = values[time_i] if 0 <= time_i < len(values) else ""
        desc = clean_display(values[desc_i]) if 0 <= desc_i < len(values) else ""
        reference = clean_display(values[ref_i]) if 0 <= ref_i < len(values) else ""
        sender = clean_display(values[sender_i]) if 0 <= sender_i < len(values) else ""
        receiver = clean_display(values[receiver_i]) if 0 <= receiver_i < len(values) else ""
        debit = abs(parse_number(values[debit_i] if 0 <= debit_i < len(values) else 0))
        credit = abs(parse_number(values[credit_i] if 0 <= credit_i < len(values) else 0))
        raw_amount = parse_number(values[amount_i] if 0 <= amount_i < len(values) else 0)
        flow = ""
        amount = 0.0
        if debit:
            flow = "withdraw"
            amount = debit
        elif credit:
            flow = "deposit"
            amount = credit
        elif raw_amount:
            flow = statement_flow_type_for(desc)
            if not flow and raw_amount < 0:
                flow = "withdraw"
            if not flow and raw_amount > 0:
                flow = default_flow
            amount = abs(raw_amount)
        if not amount:
            continue
        if not flow:
            flow = statement_flow_type_for(desc, headers[amount_i] if 0 <= amount_i < len(headers) else "") or default_flow
        date_label, date_iso = normalize_date_parts(date_raw)
        parsed.append(
            {
                "row": row_no,
                "date": date_label or clean_display(date_raw),
                "date_key": date_iso or normalize_match_date(date_raw),
                "time": extract_time_text(time_raw, date_raw),
                "description": desc,
                "flow_type": flow,
                "flow_label": flow_label(flow) if flow else "",
                "amount": float(amount),
                "reference": reference,
                "reference_key": normalize_match_text(reference),
                "sender": sender,
                "receiver": receiver,
                "source": source_name,
            }
        )
    return parsed


def parse_statement_excel(path: Path) -> List[Dict[str, Any]]:
    """Parse bank statement XLSX rows. Direction mapping: โอนออก=withdraw, รับเงินคืน/เงินเข้า=deposit."""
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    assert ws is not None
    return parse_statement_rows(list(ws.iter_rows(values_only=True)), Path(path).name)


def parse_statement_csv(path: Path) -> List[Dict[str, Any]]:
    """Parse statement CSV rows, including True Wallet Dashboard receive-history exports."""
    last_exc: Exception | None = None
    text = ""
    for encoding in ("utf-8-sig", "utf-8", "cp874"):
        try:
            text = Path(path).read_text(encoding=encoding)
            break
        except UnicodeDecodeError as exc:
            last_exc = exc
    if not text and last_exc:
        raise last_exc
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample)
    except csv.Error:
        dialect = csv.excel
    return parse_statement_rows(list(csv.reader(text.splitlines(), dialect)), Path(path).name)


def parse_statement_file(path: Path) -> List[Dict[str, Any]]:
    suffix = Path(path).suffix.lower()
    if suffix == ".csv":
        return parse_statement_csv(path)
    return parse_statement_excel(path)


BANK_LEDGER_TABLES = bank_ledger_component.BANK_LEDGER_TABLES
bank_ledger_account_identity = bank_ledger_component.bank_ledger_account_identity
ensure_bank_ledger_tables = bank_ledger_component.ensure_bank_ledger_tables
bank_ledger_tables_exist = bank_ledger_component.bank_ledger_tables_exist
bank_ledger_source_hash = bank_ledger_component.bank_ledger_source_hash
bank_ledger_entry_id = bank_ledger_component.bank_ledger_entry_id
bank_ledger_entries_from_statement = bank_ledger_component.bank_ledger_entries_from_statement
account_no_matches = bank_ledger_component.account_no_matches
slip_matches_bank_ledger_account = bank_ledger_component.slip_matches_bank_ledger_account
filter_slips_for_bank_ledger_account = bank_ledger_component.filter_slips_for_bank_ledger_account
match_ledger_entries_to_slips = bank_ledger_component.match_ledger_entries_to_slips
existing_bank_ledger_hashes = bank_ledger_component.existing_bank_ledger_hashes
bank_ledger_query_rows = bank_ledger_component.bank_ledger_query_rows
import_bank_ledger_statement = bank_ledger_component.import_bank_ledger_statement
preview_bank_ledger_import = bank_ledger_component.preview_bank_ledger_import
bank_ledger_snapshot = bank_ledger_component.bank_ledger_snapshot

def statement_flow_matches(row: Dict[str, Any], flow_type: str) -> bool:
    flow = normalize_flow_type(flow_type)
    row_flow = clean_display(row.get("flow_type"))
    if flow in {"deposit", "withdraw"}:
        return row_flow == flow
    if flow == "other":
        return row_flow == "other"
    return row_flow in {"deposit", "withdraw", ""}


def filter_statement_reconcile_rows(rows: List[Dict[str, Any]], scope: str, flow_type: str) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    kept: List[Dict[str, Any]] = []
    filtered: List[Dict[str, Any]] = []
    for row in rows:
        if backend_date_matches_scope(row, scope) and statement_flow_matches(row, flow_type):
            kept.append(row)
        else:
            filtered.append(row)
    return kept, filtered


RECONCILE_TIME_TOLERANCE_MIN = int(os.environ.get("AUDITSLIP_RECONCILE_TIME_TOL_MIN", "5"))


def _hhmm_to_minutes(value: Any) -> int:
    """Return minutes since midnight or -1 if unparseable."""
    t = clean_display(value)
    if not t:
        return -1
    parts = t.split(":")
    if len(parts) < 2:
        return -1
    try:
        h = int(parts[0]); m = int(parts[1])
    except (TypeError, ValueError):
        return -1
    if not (0 <= h < 24 and 0 <= m < 60):
        return -1
    return h * 60 + m


def amount_time_date_match(left: Dict[str, Any], right: Dict[str, Any]) -> bool:
    if abs(float(left.get("amount") or 0) - float(right.get("amount") or 0)) > 0.009:
        return False
    left_date = normalize_match_date(left.get("date_key") or left.get("date"))
    right_date = normalize_match_date(right.get("date_key") or right.get("date"))
    if left_date and right_date and left_date != right_date:
        return False
    # Time tolerance (default ±5 minutes) - configurable via env
    left_min = _hhmm_to_minutes(left.get("time"))
    right_min = _hhmm_to_minutes(right.get("time"))
    if left_min >= 0 and right_min >= 0:
        if abs(left_min - right_min) > RECONCILE_TIME_TOLERANCE_MIN:
            return False
    return True


def find_amount_time_match(target: Dict[str, Any], rows: List[Dict[str, Any]], used: set[int]) -> int:
    for idx, row in enumerate(rows):
        if idx in used:
            continue
        if amount_time_date_match(target, row):
            return idx
    return -1


def reconcile_daily_summary(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    groups: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        date_label = clean_display(row.get("date")) or "(ไม่ทราบวันที่)"
        key = clean_display(row.get("date_key")) or date_label
        group = groups.setdefault(key, {"date": date_label, "date_key": key, "count": 0, "amount": 0.0})
        group["count"] += 1
        group["amount"] += float(row.get("amount") or 0)
    return sorted(groups.values(), key=lambda r: (r["date_key"], r["date"]), reverse=True)


def slip_reconcile_rows(conn: sqlite3.Connection, chat_id: str = "", scope: str = "all", bot_key: str = "", flow_type: str = "all") -> List[Dict[str, Any]]:
    bot = clean_display(bot_key)
    if chat_id:
        where_clause, params, _ = scope_where(chat_id, scope, success_only=True, bot_key=bot)
    else:
        where_clause, params, _ = global_scope_where(scope, success_only=True, bot_key=bot)
    where_clause, params = apply_flow_sql(where_clause, params, flow_type)
    rows = conn.execute(
        f"""
        SELECT id, COALESCE(NULLIF(bot_key,''),'default') AS bot_key, company_name,
               chat_id, chat_title, message_id,
               slip_date_display, slip_date_iso, slip_time,
               transferor_name, sender_name, recipient_name,
               from_bank, from_account, issuer_bank, to_bank, to_account,
               amount, reference_no, seq, aid
        FROM slips
        WHERE {where_clause}
        ORDER BY created_at ASC
        """,
        params,
    ).fetchall()
    out = []
    for r in rows:
        date_value = r["slip_date_iso"] or r["slip_date_display"]
        bank = display_bank(r["from_bank"] or r["issuer_bank"] or r["to_bank"])
        reference = clean_display(r["reference_no"] or r["seq"] or r["aid"])
        name = display_transferor_name(r["transferor_name"] or r["sender_name"])
        message_id = clean_display(r["message_id"])
        out.append(
            {
                "id": r["id"],
                "bot_key": r["bot_key"],
                "company_name": clean_company_name(r["company_name"], r["bot_key"]),
                "chat_id": r["chat_id"],
                "chat_title": clean_display(r["chat_title"]),
                "message_id": message_id,
                "date": r["slip_date_display"] or r["slip_date_iso"],
                "date_key": normalize_match_date(date_value),
                "time": clean_display(r["slip_time"]),
                "name": name,
                "name_key": transferor_key(name),
                "recipient_name": clean_display(r["recipient_name"]),
                "bank": bank,
                "bank_key": normalize_match_text(bank),
                "from_bank": display_bank(r["from_bank"]),
                "from_account": clean_display(r["from_account"]),
                "to_bank": display_bank(r["to_bank"]),
                "to_account": clean_display(r["to_account"]),
                "issuer_bank": display_bank(r["issuer_bank"]),
                "amount": float(r["amount"] or 0),
                "reference": reference,
                "reference_key": normalize_match_text(reference),
                "source": f"msg {message_id}" if message_id else clean_display(r["chat_title"]),
            }
        )
    return out


def reconcile_backend_excel(db_path: Path, excel_path: Path, chat_id: str = "", scope: str = "all", bot_key: str = "", flow_type: str = "all") -> Dict[str, Any]:
    all_backend_rows = parse_backend_excel(excel_path)
    flow = normalize_flow_type(flow_type)
    bot = clean_display(bot_key)
    backend_rows, backend_filtered_out = filter_backend_reconcile_rows(all_backend_rows, scope, flow)
    with connect(db_path) as conn:
        slip_rows = slip_reconcile_rows(conn, chat_id=chat_id, scope=scope, bot_key=bot, flow_type=flow)
    used_slips: set[int] = set()
    matches: List[Dict[str, Any]] = []
    missing: List[Dict[str, Any]] = []
    for backend in backend_rows:
        best_idx = -1
        best_score = -1
        for idx, slip in enumerate(slip_rows):
            if idx in used_slips:
                continue
            if abs(float(backend["amount"]) - float(slip["amount"])) > 0.009:
                continue
            date_match = not backend["date_key"] or not slip["date_key"] or backend["date_key"] == slip["date_key"]
            if not date_match:
                continue
            score = 10
            if backend["reference_key"] and backend["reference_key"] == slip["reference_key"]:
                score += 100
            if backend["bank_key"] and backend["bank_key"] == slip["bank_key"]:
                score += 20
            if backend.get("time") and slip.get("time") and clean_display(backend.get("time")) == clean_display(slip.get("time")):
                score += 10
            if backend["name_key"] and (backend["name_key"] == slip["name_key"] or names_similar(backend["name_key"], slip["name_key"])):
                score += 20
            if backend["reference_key"] and slip["reference_key"] and backend["reference_key"] != slip["reference_key"]:
                score -= 50
            if score > best_score:
                best_score = score
                best_idx = idx
        if best_idx >= 0 and best_score >= 10:
            used_slips.add(best_idx)
            matches.append({"backend": backend, "slip": slip_rows[best_idx], "score": best_score})
        else:
            missing.append(backend)
    extra = [slip for idx, slip in enumerate(slip_rows) if idx not in used_slips]
    backend_amount = sum(float(r["amount"] or 0) for r in backend_rows)
    slip_amount = sum(float(r["amount"] or 0) for r in slip_rows)
    matched_amount = sum(float(m["backend"]["amount"] or 0) for m in matches)
    return {
        "ok": True,
        "excel_path": str(excel_path),
        "scope": {"chat_id": clean_display(chat_id), "bot_key": bot or "__all__", "flow_type": flow, "flow_label": flow_label(flow), "date_scope": clean_display(scope)},
        "backend": {"count": len(backend_rows), "amount": backend_amount},
        "backend_filtered_out": {"count": len(backend_filtered_out), "amount": sum(float(r["amount"] or 0) for r in backend_filtered_out)},
        "slips": {"count": len(slip_rows), "amount": slip_amount},
        "matched": {"count": len(matches), "amount": matched_amount, "rows": matches[:100]},
        "missing_in_slips": missing[:100],
        "extra_slips": extra[:100],
        "missing": {"count": len(missing), "amount": sum(float(r["amount"] or 0) for r in missing)},
        "extra": {"count": len(extra), "amount": sum(float(r["amount"] or 0) for r in extra)},
        "daily": {"backend": reconcile_daily_summary(backend_rows), "slips": reconcile_daily_summary(slip_rows)},
        "diff_amount": backend_amount - slip_amount,
    }


def reconcile_backend_slips_statement(db_path: Path, excel_path: Path, statement_path: Path, chat_id: str = "", scope: str = "all", bot_key: str = "", flow_type: str = "all") -> Dict[str, Any]:
    """Compare three sources by amount/time/date: backend Excel, OCR slips, and bank statement."""
    all_backend_rows = parse_backend_excel(excel_path)
    all_statement_rows = parse_statement_file(statement_path)
    flow = normalize_flow_type(flow_type)
    bot = clean_display(bot_key)
    backend_rows, backend_filtered_out = filter_backend_reconcile_rows(all_backend_rows, scope, flow)
    statement_rows, statement_filtered_out = filter_statement_reconcile_rows(all_statement_rows, scope, flow)
    with connect(db_path) as conn:
        slip_rows = slip_reconcile_rows(conn, chat_id=chat_id, scope=scope, bot_key=bot, flow_type=flow)

    used_slips: set[int] = set()
    used_statements: set[int] = set()
    matched: List[Dict[str, Any]] = []
    backend_slip_matched: List[Dict[str, Any]] = []
    backend_statement_matched: List[Dict[str, Any]] = []
    backend_missing_slip: List[Dict[str, Any]] = []
    backend_missing_statement: List[Dict[str, Any]] = []

    for backend in backend_rows:
        slip_idx = find_amount_time_match(backend, slip_rows, used_slips)
        statement_idx = find_amount_time_match(backend, statement_rows, used_statements)
        slip = slip_rows[slip_idx] if slip_idx >= 0 else None
        statement = statement_rows[statement_idx] if statement_idx >= 0 else None
        if slip_idx >= 0:
            used_slips.add(slip_idx)
            backend_slip_matched.append({"backend": backend, "slip": slip})
        else:
            backend_missing_slip.append(backend)
        if statement_idx >= 0:
            used_statements.add(statement_idx)
            backend_statement_matched.append({"backend": backend, "statement": statement})
        else:
            backend_missing_statement.append(backend)
        if slip is not None and statement is not None:
            matched.append({"backend": backend, "slip": slip, "statement": statement})

    slip_extra = [slip for idx, slip in enumerate(slip_rows) if idx not in used_slips]
    statement_extra = [row for idx, row in enumerate(statement_rows) if idx not in used_statements]
    backend_amount = sum(float(r["amount"] or 0) for r in backend_rows)
    slip_amount = sum(float(r["amount"] or 0) for r in slip_rows)
    statement_amount = sum(float(r["amount"] or 0) for r in statement_rows)
    matched_amount = sum(float(m["backend"]["amount"] or 0) for m in matched)
    return {
        "ok": True,
        "mode": "backend_slips_statement",
        "excel_path": str(excel_path),
        "statement_path": str(statement_path),
        "scope": {"chat_id": clean_display(chat_id), "bot_key": bot or "__all__", "flow_type": flow, "flow_label": flow_label(flow), "date_scope": clean_display(scope)},
        "backend": {"count": len(backend_rows), "amount": backend_amount},
        "backend_filtered_out": {"count": len(backend_filtered_out), "amount": sum(float(r["amount"] or 0) for r in backend_filtered_out)},
        "slips": {"count": len(slip_rows), "amount": slip_amount},
        "statement": {"count": len(statement_rows), "amount": statement_amount},
        "statement_filtered_out": {"count": len(statement_filtered_out), "amount": sum(float(r["amount"] or 0) for r in statement_filtered_out)},
        "matched": {"count": len(matched), "amount": matched_amount, "rows": matched[:100]},
        "backend_slip_matched": {"count": len(backend_slip_matched), "amount": sum(float(m["backend"]["amount"] or 0) for m in backend_slip_matched)},
        "backend_statement_matched": {"count": len(backend_statement_matched), "amount": sum(float(m["backend"]["amount"] or 0) for m in backend_statement_matched)},
        "backend_missing_slip": {"count": len(backend_missing_slip), "amount": sum(float(r["amount"] or 0) for r in backend_missing_slip), "rows": backend_missing_slip[:100]},
        "backend_missing_statement": {"count": len(backend_missing_statement), "amount": sum(float(r["amount"] or 0) for r in backend_missing_statement), "rows": backend_missing_statement[:100]},
        "slip_extra": {"count": len(slip_extra), "amount": sum(float(r["amount"] or 0) for r in slip_extra), "rows": slip_extra[:100]},
        "statement_extra": {"count": len(statement_extra), "amount": sum(float(r["amount"] or 0) for r in statement_extra), "rows": statement_extra[:100]},
        "daily": {"backend": reconcile_daily_summary(backend_rows), "slips": reconcile_daily_summary(slip_rows), "statement": reconcile_daily_summary(statement_rows)},
        "diff_amounts": {"backend_minus_slips": backend_amount - slip_amount, "backend_minus_statement": backend_amount - statement_amount, "slips_minus_statement": slip_amount - statement_amount},
    }


def scope_where(chat_id: str, scope: str = "open", success_only: bool = True, bot_key: str = "") -> Tuple[str, List[Any], str]:
    """Build the same success/open/date scope semantics used by the Telegram bot."""
    clause, params, label = global_scope_where(scope, success_only=success_only, bot_key=bot_key)
    return "chat_id=? AND " + clause, [str(chat_id), *params], label


def normalize_scope_bound(value: str) -> str:
    display, iso = normalize_date_parts(clean_display(value))
    return iso or clean_display(value)


def scope_date_range(scope: str) -> Tuple[str, str, str]:
    raw = clean_display(scope)
    text = raw[6:] if raw.lower().startswith("range:") else raw
    if ".." not in text:
        return "", "", ""
    start_text, end_text = text.split("..", 1)
    start = normalize_scope_bound(start_text)
    end = normalize_scope_bound(end_text)
    if not start and not end:
        return "", "", ""
    if start and end and start > end:
        start, end = end, start
    if start and end:
        label = f"{start} ถึง {end}"
    elif start:
        label = f"ตั้งแต่ {start}"
    else:
        label = f"ถึง {end}"
    return start, end, label


def append_scope_date_range(clause: str, params: List[Any], start: str, end: str, prefix: str = "") -> Tuple[str, List[Any]]:
    date_expr = f"{prefix}slip_date_iso"
    out = list(params)
    if start and end:
        clause += f" AND {date_expr} BETWEEN ? AND ?"
        out.extend([start, end])
    elif start:
        clause += f" AND {date_expr} >= ?"
        out.append(start)
    elif end:
        clause += f" AND {date_expr} <= ?"
        out.append(end)
    return clause, out


def global_scope_where(scope: str = "open", success_only: bool = True, bot_key: str = "") -> Tuple[str, List[Any], str]:
    """Build scope semantics across every chat, optionally limited to one bot/company."""
    range_start, range_end, range_label = scope_date_range(scope)
    normalized, label = scope_to_date(scope)
    clause = "1=1"
    params: List[Any] = []
    if bot_key and clean_display(bot_key) not in {"__all__", "all"}:
        clause += " AND COALESCE(bot_key,'default')=?"
        params.append(clean_display(bot_key) or "default")
    if success_only:
        clause += " AND status='success' AND COALESCE(is_duplicate,0)=0"
    if range_label:
        clause, params = append_scope_date_range(clause, params, range_start, range_end)
    elif normalized == "open":
        clause += " AND settlement_id IS NULL"
    elif normalized == "all":
        pass
    elif re.match(r"^\d{4}-\d{2}-\d{2}$", normalized):
        clause += " AND slip_date_iso=?"
        params.append(normalized)
    else:
        clause += " AND (slip_date_display=? OR slip_date_iso=?)"
        params.extend([normalized, normalized])
    return clause, params, range_label or label



FLOW_TYPE_LABELS = {"all": "รวมทุกกลุ่ม", "deposit": "ฝาก/เติมมือ", "withdraw": "ถอน", "other": "อื่นๆ"}
DEPOSIT_TITLE_TOKENS = ["ฝาก", "deposit", "deposits", "รับฝาก", "เติมเงิน", "เติมมือ", "topup", "top-up"]
WITHDRAW_TITLE_TOKENS = ["ถอน", "withdraw", "withdrawal", "withdrawals"]


def normalize_flow_type(value: Any) -> str:
    raw = clean_display(value).lower()
    if raw in {"all", "ทุกกลุ่ม", "รวมทุกกลุ่ม", ""}:
        return "all"
    if raw in {"deposit", "deposits", "ฝาก", "ฝาก/เติมมือ", "เติมมือ", "เติมเงิน", "รับฝาก", "in", "credit", "topup", "top-up", "manual topup"} or "เติมมือ" in raw or "เติมเงิน" in raw:
        return "deposit"
    if raw in {"withdraw", "withdrawal", "withdrawals", "ถอน", "out", "debit"}:
        return "withdraw"
    if raw in {"other", "อื่น", "อื่นๆ"}:
        return "other"
    return "all"


def configured_flow_map() -> Dict[Tuple[str, str], str]:
    """Explicit bot/chat -> flow mapping from AUDITSLIP_FLOW_MAP.

    Accepts JSON objects such as {"bot1|CHAT_ID":"deposit"} or CSV entries
    like bot1|CHAT_ID=deposit,bot1|OTHER=withdraw. A blank/* bot maps by chat_id.
    """
    raw = clean_display(os.environ.get("AUDITSLIP_FLOW_MAP") or os.environ.get("AUDITSLIP_GROUP_FLOW_MAP") or "")
    if not raw:
        return {}
    items: List[Tuple[str, Any]] = []
    try:
        parsed = json.loads(raw)
        if isinstance(parsed, dict):
            items = [(str(k), v) for k, v in parsed.items()]
        elif isinstance(parsed, list):
            for entry in parsed:
                if isinstance(entry, dict):
                    key = "|".join([clean_display(entry.get("bot_key") or entry.get("bot") or "*"), clean_display(entry.get("chat_id") or entry.get("chat") or "")])
                    items.append((key, entry.get("flow_type") or entry.get("flow") or ""))
    except Exception:
        for part in raw.split(","):
            if "=" in part:
                k, v = part.split("=", 1)
                items.append((k.strip(), v.strip()))
    result: Dict[Tuple[str, str], str] = {}
    for raw_key, raw_flow in items:
        flow = normalize_flow_type(raw_flow)
        if flow not in {"deposit", "withdraw", "other"}:
            continue
        key = clean_display(raw_key)
        if "|" in key:
            bot_key, chat_id = key.split("|", 1)
        elif ":" in key:
            bot_key, chat_id = key.split(":", 1)
        else:
            bot_key, chat_id = "*", key
        chat_id = clean_display(chat_id)
        if not chat_id:
            continue
        result[(clean_display(bot_key) or "*", chat_id)] = flow
    return result


def configured_flow_for(bot_key: Any = "", chat_id: Any = "") -> str:
    mapping = configured_flow_map()
    bot = clean_display(bot_key) or "default"
    chat = clean_display(chat_id)
    if not chat:
        return ""
    return mapping.get((bot, chat)) or mapping.get(("*", chat)) or mapping.get(("__all__", chat)) or ""


def title_tokens_expr(field: str, tokens: List[str]) -> Tuple[str, List[Any]]:
    clauses: List[str] = []
    params: List[Any] = []
    for token in tokens:
        if re.fullmatch(r"[a-z0-9_-]+", token):
            clauses.append(f"LOWER({field}) LIKE ?")
            params.append(f"%{token.lower()}%")
        else:
            clauses.append(f"{field} LIKE ?")
            params.append(f"%{token}%")
    return "(" + " OR ".join(clauses) + ")", params


def flow_map_expr(flow: str | None = None, alias: str = "") -> Tuple[str, List[Any]]:
    prefix = f"{alias}." if alias else ""
    bot_field = f"COALESCE(NULLIF({prefix}bot_key,''),'default')"
    chat_field = f"COALESCE({prefix}chat_id,'')"
    clauses: List[str] = []
    params: List[Any] = []
    for (bot_key, chat_id), mapped_flow in configured_flow_map().items():
        if flow and mapped_flow != flow:
            continue
        if bot_key in {"", "*", "__all__", "all"}:
            clauses.append(f"({chat_field}=?)")
            params.append(chat_id)
        else:
            clauses.append(f"({bot_field}=? AND {chat_field}=?)")
            params.extend([bot_key, chat_id])
    if not clauses:
        return "", []
    return "(" + " OR ".join(clauses) + ")", params


def flow_type_for_title(title: Any, bot_key: Any = "", chat_id: Any = "") -> str:
    mapped = configured_flow_for(bot_key, chat_id)
    if mapped:
        return mapped
    text = clean_display(title).lower()
    deposit_signal = any(token in text for token in DEPOSIT_TITLE_TOKENS)
    withdraw_signal = any(token in text for token in WITHDRAW_TITLE_TOKENS)
    if deposit_signal and not withdraw_signal:
        return "deposit"
    if withdraw_signal:
        return "withdraw"
    # Legacy/current groups that do not explicitly say ฝาก/เติมมือ are treated as ถอน,
    # so selecting บริษัท + ถอน never hides historical data just because the old
    # chat title was generic (for example a DM or an older audit room name).
    return "withdraw"


def flow_sql_clause(flow_type: str, alias: str = "") -> Tuple[str, List[Any]]:
    flow = normalize_flow_type(flow_type)
    if flow == "all":
        return "", []
    prefix = f"{alias}." if alias else ""
    field = f"COALESCE({prefix}chat_title,'')"
    deposit_expr, deposit_params = title_tokens_expr(field, DEPOSIT_TITLE_TOKENS)
    withdraw_expr, withdraw_params = title_tokens_expr(field, WITHDRAW_TITLE_TOKENS)
    deposit_only_expr = f"({deposit_expr} AND NOT {withdraw_expr})"
    deposit_only_params = [*deposit_params, *withdraw_params]
    mapped_any_expr, mapped_any_params = flow_map_expr(None, alias=alias)
    mapped_flow_expr, mapped_flow_params = flow_map_expr(flow, alias=alias)
    not_mapped_expr = f"NOT {mapped_any_expr}" if mapped_any_expr else "1=1"
    pieces: List[str] = []
    params: List[Any] = []
    if mapped_flow_expr:
        pieces.append(mapped_flow_expr)
        params.extend(mapped_flow_params)
    if flow == "deposit":
        pieces.append(f"({not_mapped_expr} AND {deposit_only_expr})")
        params.extend([*mapped_any_params, *deposit_only_params])
    elif flow == "withdraw":
        pieces.append(f"({not_mapped_expr} AND NOT {deposit_only_expr})")
        params.extend([*mapped_any_params, *deposit_only_params])
    else:
        pieces.append(f"({not_mapped_expr} AND 1=0)")
        params.extend(mapped_any_params)
    return " AND (" + " OR ".join(pieces) + ")", params


def apply_flow_sql(clause: str, params: List[Any], flow_type: str, alias: str = "") -> Tuple[str, List[Any]]:
    extra, extra_params = flow_sql_clause(flow_type, alias=alias)
    return clause + extra, [*params, *extra_params]


def limit_scope_clause(where_clause: str, params: List[Any], flow_type: str) -> Tuple[str, List[Any]]:
    """Scope transferor/limit panels to withdrawal slips only when the dashboard shows all flows."""
    flow = normalize_flow_type(flow_type)
    if flow == "deposit":
        return "1=0", []
    if flow == "all":
        return apply_flow_sql(where_clause, params, "withdraw")
    return where_clause, list(params)


def deposit_customer_scope_clause(where_clause: str, params: List[Any], flow_type: str) -> Tuple[str, List[Any]]:
    """Scope the customer-slip panel to deposit/top-up rooms only."""
    flow = normalize_flow_type(flow_type)
    if flow == "withdraw":
        return "1=0", []
    if flow == "deposit":
        return where_clause, list(params)
    if flow == "all":
        return apply_flow_sql(where_clause, params, "deposit")
    return "1=0", []


def flow_label(flow_type: str) -> str:
    return FLOW_TYPE_LABELS.get(normalize_flow_type(flow_type), "ทุกกลุ่ม")


def first_chat_selection(conn: sqlite3.Connection) -> Tuple[str, str]:
    row = conn.execute(
        """
        SELECT COALESCE(NULLIF(bot_key,''),'default') AS bot_key,
               chat_id,
               COALESCE(SUM(CASE WHEN status='success' AND COALESCE(is_duplicate,0)=0 AND settlement_id IS NULL THEN amount ELSE 0 END),0) AS open_amount,
               COUNT(*) AS total_rows
        FROM slips
        GROUP BY COALESCE(NULLIF(bot_key,''),'default'), chat_id
        ORDER BY open_amount DESC, total_rows DESC
        LIMIT 1
        """
    ).fetchone()
    return (str(row["bot_key"]), str(row["chat_id"])) if row else ("", "")



def grouped_totals(conn: sqlite3.Connection, where_clause: str, params: List[Any], name_expr: str, limit: int = 50) -> List[sqlite3.Row]:
    return conn.execute(
        f"""
        SELECT {name_expr} AS name,
               COUNT(*) AS count,
               COALESCE(SUM(amount),0) AS amount,
               COALESCE(SUM(fee),0) AS fee
        FROM slips
        WHERE {where_clause}
        GROUP BY name
        ORDER BY amount DESC, count DESC, name ASC
        LIMIT ?
        """,
        [*params, limit],
    ).fetchall()


def bank_totals(conn: sqlite3.Connection, where_clause: str, params: List[Any], name_expr: str, unknown_label: str, limit: int = 50, include_limit: bool = False) -> List[Dict[str, Any]]:
    rows = conn.execute(
        f"""
        SELECT {name_expr} AS raw_name,
               COUNT(*) AS count,
               COALESCE(SUM(amount),0) AS amount,
               COALESCE(SUM(fee),0) AS fee
        FROM slips
        WHERE {where_clause}
        GROUP BY raw_name
        ORDER BY amount DESC, count DESC, raw_name ASC
        """,
        params,
    ).fetchall()
    groups: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        name = display_bank(row["raw_name"]) or unknown_label
        key = bank_key(name) or name
        initial = {"name": name, "count": 0, "amount": 0.0, "fee": 0.0}
        if include_limit:
            initial["limit_amount"] = bank_limit_amount(name)
        group = groups.setdefault(key, initial)
        group["count"] += int(row["count"] or 0)
        group["amount"] += float(row["amount"] or 0)
        group["fee"] += float(row["fee"] or 0)
    return sorted(groups.values(), key=lambda g: (-g["amount"], -g["count"], g["name"]))[:limit]


def scoped_counts(conn: sqlite3.Connection, chat_id: str, scope: str, bot_key: str = "") -> sqlite3.Row:
    where_clause, params, _ = scope_where(chat_id, scope, success_only=True, bot_key=bot_key)
    return conn.execute(
        f"SELECT COUNT(*) AS count, COALESCE(SUM(amount),0) AS amount FROM slips WHERE {where_clause}",
        params,
    ).fetchone()


def _scope_clause_for_alias(alias: str, chat_id: str, scope: str = "open", success_only: bool = False, bot_key: str = "") -> Tuple[str, List[Any]]:
    clause, params = _scope_clause_for_alias_optional(alias, scope, success_only=success_only, bot_key=bot_key)
    prefix = f"{alias}."
    return f"{prefix}chat_id=? AND {clause}", [str(chat_id), *params]


def _scope_clause_for_alias_optional(alias: str, scope: str = "open", success_only: bool = False, bot_key: str = "") -> Tuple[str, List[Any]]:
    range_start, range_end, range_label = scope_date_range(scope)
    normalized, _ = scope_to_date(scope)
    prefix = f"{alias}."
    clause = "1=1"
    params: List[Any] = []
    if bot_key and clean_display(bot_key) not in {"__all__", "all"}:
        clause += f" AND COALESCE({prefix}bot_key,'default')=?"
        params.append(clean_display(bot_key) or "default")
    if success_only:
        clause += f" AND {prefix}status='success' AND COALESCE({prefix}is_duplicate,0)=0"
    if range_label:
        clause, params = append_scope_date_range(clause, params, range_start, range_end, prefix=prefix)
    elif normalized == "open":
        clause += f" AND {prefix}settlement_id IS NULL"
    elif normalized == "all":
        pass
    elif re.match(r"^\d{4}-\d{2}-\d{2}$", normalized):
        clause += f" AND {prefix}slip_date_iso=?"
        params.append(normalized)
    else:
        clause += f" AND ({prefix}slip_date_display=? OR {prefix}slip_date_iso=?)"
        params.extend([normalized, normalized])
    return clause, params


def image_url_for(file_id: Any, slip_id: Any) -> str:
    return "/api/slip-image?" + urlencode({"id": clean_display(slip_id)}) if clean_display(file_id) else ""


def slip_search_clause(alias: str, query: str) -> Tuple[str, List[Any]]:
    query = clean_display(query)
    if not query:
        return "", []
    prefix = f"{alias}."
    fields = [
        "id",
        "duplicate_of",
        "message_id",
        "transferor_name",
        "recipient_name",
        "sender_name",
        "username",
        "amount",
        "reference_no",
        "seq",
        "aid",
        "from_bank",
        "from_account",
        "to_bank",
        "to_account",
        "issuer_bank",
        "slip_date_display",
        "slip_date_iso",
        "slip_time",
    ]
    pattern = f"%{query}%"
    clause = " OR ".join([f"CAST({prefix}{field} AS TEXT) LIKE ?" for field in fields])
    return f" AND ({clause})", [pattern] * len(fields)


def duplicate_pair_search_clause(query: str) -> Tuple[str, List[Any]]:
    query = clean_display(query)
    if not query:
        return "", []
    fields = [
        "d.id", "d.duplicate_of", "d.message_id", "d.transferor_name", "d.recipient_name", "d.sender_name", "d.username", "d.amount", "d.reference_no", "d.seq", "d.aid", "d.from_bank", "d.from_account", "d.to_bank", "d.to_account", "d.issuer_bank", "d.slip_date_display", "d.slip_date_iso", "d.slip_time",
        "o.id", "o.message_id", "o.transferor_name", "o.recipient_name", "o.sender_name", "o.username", "o.amount", "o.reference_no", "o.seq", "o.aid", "o.from_bank", "o.from_account", "o.to_bank", "o.to_account", "o.issuer_bank", "o.slip_date_display", "o.slip_date_iso", "o.slip_time",
    ]
    pattern = f"%{query}%"
    clause = " OR ".join([f"CAST({field} AS TEXT) LIKE ?" for field in fields])
    return f" AND ({clause})", [pattern] * len(fields)


def duplicate_pair_rows(conn: sqlite3.Connection, chat_id: str, scope: str = "open", bot_key: str = "", search: str = "", limit: int = 40, flow_type: str = "all") -> List[Dict[str, Any]]:
    if chat_id:
        clause, params = _scope_clause_for_alias("d", chat_id, scope, success_only=False, bot_key=bot_key)
    else:
        clause, params = _scope_clause_for_alias_optional("d", scope, success_only=False, bot_key=bot_key)
    clause, params = apply_flow_sql(clause, params, flow_type, alias="d")
    clause += " AND d.status='success' AND COALESCE(d.is_duplicate,0)=1"
    search_clause, search_params = duplicate_pair_search_clause(search)
    clause += search_clause
    params.extend(search_params)
    rows = conn.execute(
        f"""
        SELECT d.id AS duplicate_id,
               COALESCE(NULLIF(d.bot_key,''),'default') AS duplicate_bot_key,
               d.company_name AS duplicate_company_name,
               d.chat_id AS duplicate_chat_id,
               d.chat_title AS duplicate_chat_title,
               d.file_id AS duplicate_file_id,
               d.message_id AS duplicate_message_id,
               d.sender_name AS duplicate_sender_name,
               d.username AS duplicate_username,
               d.reference_no AS reference_no,
               d.seq AS seq,
               d.aid AS aid,
               d.created_at_iso AS duplicate_created_at_iso,
               (SELECT MIN(j.created_at) FROM ocr_jobs j WHERE j.slip_id=d.id AND COALESCE(j.bot_key,'default')=COALESCE(d.bot_key,'default')) AS duplicate_submitted_at,
               d.duplicate_of AS original_id,
               d.slip_date_display AS slip_date_display,
               d.slip_date_iso AS slip_date_iso,
               d.slip_time AS slip_time,
               d.transferor_name AS transferor_name,
               d.recipient_name AS recipient_name,
               d.amount AS amount,
               d.from_bank AS from_bank,
               d.to_bank AS to_bank,
               d.issuer_bank AS issuer_bank,
               COALESCE(NULLIF(o.bot_key,''),'default') AS original_bot_key,
               o.company_name AS original_company_name,
               o.chat_id AS original_chat_id,
               o.chat_title AS original_chat_title,
               o.file_id AS original_file_id,
               o.message_id AS original_message_id,
               o.sender_name AS original_sender_name,
               o.username AS original_username,
               o.reference_no AS original_reference_no,
               o.seq AS original_seq,
               o.aid AS original_aid,
               o.created_at_iso AS original_created_at_iso,
               (SELECT MIN(j.created_at) FROM ocr_jobs j WHERE j.slip_id=o.id AND COALESCE(j.bot_key,'default')=COALESCE(o.bot_key,'default')) AS original_submitted_at,
               o.slip_date_display AS original_slip_date_display,
               o.slip_date_iso AS original_slip_date_iso,
               o.slip_time AS original_slip_time,
               o.transferor_name AS original_transferor_name,
               o.recipient_name AS original_recipient_name,
               o.amount AS original_amount,
               o.from_bank AS original_from_bank,
               o.to_bank AS original_to_bank,
               o.issuer_bank AS original_issuer_bank
        FROM slips d
        LEFT JOIN slips o ON o.id=d.duplicate_of AND COALESCE(o.bot_key,'default')=COALESCE(d.bot_key,'default')
        WHERE {clause}
        ORDER BY d.created_at DESC
        LIMIT ?
        """,
        [*params, limit],
    ).fetchall()
    out = rows_to_dicts(rows)
    for row in out:
        row["duplicate_image_url"] = image_url_for(row.get("duplicate_file_id"), row.get("duplicate_id"))
        row["original_image_url"] = image_url_for(row.get("original_file_id"), row.get("original_id"))
        dup_bot = clean_display(row.get("duplicate_bot_key")) or "default"
        orig_bot = clean_display(row.get("original_bot_key")) or dup_bot
        row["duplicate_company_name"] = clean_company_name(row.get("duplicate_company_name"), dup_bot)
        row["original_company_name"] = clean_company_name(row.get("original_company_name"), orig_bot)
        row["duplicate_flow_type"] = flow_type_for_title(row.get("duplicate_chat_title"), dup_bot, row.get("duplicate_chat_id"))
        row["original_flow_type"] = flow_type_for_title(row.get("original_chat_title"), orig_bot, row.get("original_chat_id"))
        row["duplicate_flow_label"] = flow_label(row["duplicate_flow_type"])
        row["original_flow_label"] = flow_label(row["original_flow_type"])
        row["duplicate_submitted_at_iso"] = bkk_iso_from_ms(row.get("duplicate_submitted_at")) or clean_display(row.get("duplicate_created_at_iso"))
        row["original_submitted_at_iso"] = bkk_iso_from_ms(row.get("original_submitted_at")) or clean_display(row.get("original_created_at_iso"))
        row["from_bank"] = display_bank(row.get("from_bank"))
        row["to_bank"] = display_bank(row.get("to_bank"))
        row["issuer_bank"] = display_bank(row.get("issuer_bank"))
        row["original_from_bank"] = display_bank(row.get("original_from_bank"))
        row["original_to_bank"] = display_bank(row.get("original_to_bank"))
        row["original_issuer_bank"] = display_bank(row.get("original_issuer_bank"))
    return out


def source_bank_review_condition() -> str:
    missing_from_bank = """
      (
        NULLIF(TRIM(COALESCE(from_bank,'')), '') IS NULL
        OR LOWER(TRIM(COALESCE(from_bank,''))) IN ('unknown','n/a','na','none','null','-','xxx','xxxx','masked')
        OR COALESCE(from_bank,'') LIKE '%ไม่ทราบ%'
      )
    """
    missing_issuer_bank = """
      (
        NULLIF(TRIM(COALESCE(issuer_bank,'')), '') IS NULL
        OR LOWER(TRIM(COALESCE(issuer_bank,''))) IN ('unknown','n/a','na','none','null','-','xxx','xxxx','masked')
        OR COALESCE(issuer_bank,'') LIKE '%ไม่ทราบ%'
      )
    """
    return f"({missing_from_bank} AND {missing_issuer_bank})"


def pending_delete_slip_ids(conn: sqlite3.Connection) -> List[str]:
    """Slip ids that already have a live delete request.

    These rows are still `success` until approval/execution, so financial totals remain
    unchanged. But they should not keep showing in the source-bank-review queue as if
    the operator had never requested deletion.
    """
    try:
        rows = conn.execute(
            """
            SELECT payload_json
            FROM pending_actions
            WHERE action='slip.delete' AND status IN ('pending','approved')
            """
        ).fetchall()
    except sqlite3.OperationalError:
        return []
    out: List[str] = []
    seen = set()
    for row in rows:
        try:
            payload = json.loads(row["payload_json"] or "{}")
        except Exception:
            payload = {}
        slip_id = clean_display(payload.get("id") or payload.get("slip_id") or "") if isinstance(payload, dict) else ""
        if slip_id and slip_id not in seen:
            seen.add(slip_id)
            out.append(slip_id)
    return out


def _exclude_slip_ids_clause(exclude_ids: List[str] | None, column: str = "id") -> Tuple[str, List[Any]]:
    ids: List[str] = []
    seen = set()
    for value in exclude_ids or []:
        slip_id = clean_display(value)
        if slip_id and slip_id not in seen:
            seen.add(slip_id)
            ids.append(slip_id)
    if not ids:
        return "", []
    placeholders = ",".join("?" for _ in ids)
    return f" AND {column} NOT IN ({placeholders})", ids


def source_bank_review_count(conn: sqlite3.Connection, where_clause: str, params: List[Any], search: str = "", exclude_ids: List[str] | None = None) -> int:
    search_clause, search_params = slip_search_clause("slips", search)
    exclude_clause, exclude_params = _exclude_slip_ids_clause(exclude_ids)
    row = conn.execute(
        f"SELECT COUNT(*) AS count FROM slips WHERE {where_clause} AND {source_bank_review_condition()}{search_clause}{exclude_clause}",
        [*params, *search_params, *exclude_params],
    ).fetchone()
    return int(row["count"] or 0)


def source_bank_review_ids(conn: sqlite3.Connection, where_clause: str, params: List[Any], search: str = "", exclude_ids: List[str] | None = None) -> List[str]:
    search_clause, search_params = slip_search_clause("slips", search)
    exclude_clause, exclude_params = _exclude_slip_ids_clause(exclude_ids)
    rows = conn.execute(
        f"""
        SELECT id
        FROM slips
        WHERE {where_clause} AND {source_bank_review_condition()}{search_clause}{exclude_clause}
        ORDER BY created_at DESC
        """,
        [*params, *search_params, *exclude_params],
    ).fetchall()
    return [str(r["id"]) for r in rows]


def source_bank_review_rows(conn: sqlite3.Connection, where_clause: str, params: List[Any], search: str = "", limit: int = 40, exclude_ids: List[str] | None = None) -> List[Dict[str, Any]]:
    search_clause, search_params = slip_search_clause("slips", search)
    exclude_clause, exclude_params = _exclude_slip_ids_clause(exclude_ids)
    limit_clause = "LIMIT ?" if limit and limit > 0 else ""
    query_params: List[Any] = [*params, *search_params, *exclude_params]
    if limit_clause:
        query_params.append(limit)
    rows = conn.execute(
        f"""
        SELECT id, COALESCE(NULLIF(bot_key,''),'default') AS bot_key, company_name, chat_id, chat_title,
               file_id, message_id, sender_name, username, status, error,
               slip_date_display, slip_date_iso, slip_time,
               TRIM(COALESCE(NULLIF(slip_date_display,''), NULLIF(slip_date_iso,''), '') || ' ' || COALESCE(NULLIF(slip_time,''), '')) AS slip_date_text,
               transferor_name, recipient_name, issuer_bank, from_bank, from_account, to_bank, to_account,
               amount, confidence, created_at_iso
        FROM slips
        WHERE {where_clause} AND {source_bank_review_condition()}{search_clause}{exclude_clause}
        ORDER BY created_at DESC
        {limit_clause}
        """,
        query_params,
    ).fetchall()
    out = rows_to_dicts(rows)
    for row in out:
        for field in ["issuer_bank", "from_bank", "to_bank"]:
            row[field] = display_bank(row.get(field))
        row["image_url"] = image_url_for(row.get("file_id"), row.get("id"))
    return out


def slip_list_rows(conn: sqlite3.Connection, where_clause: str, params: List[Any], search: str = "", limit: int = 40) -> List[Dict[str, Any]]:
    search_clause, search_params = slip_search_clause("slips", search)
    rows = conn.execute(
        f"""
        SELECT id, COALESCE(NULLIF(bot_key,''),'default') AS bot_key, company_name, chat_id, chat_title,
               file_id, message_id, sender_name, username, status, error,
               slip_date_display, slip_date_iso, slip_time,
               TRIM(COALESCE(NULLIF(slip_date_display,''), NULLIF(slip_date_iso,''), '') || ' ' || COALESCE(NULLIF(slip_time,''), '')) AS slip_date_text,
               transferor_name, recipient_name, issuer_bank, from_bank, from_account, to_bank, to_account,
               amount, confidence, created_at_iso, settlement_id, is_duplicate, duplicate_of
        FROM slips
        WHERE {where_clause} AND COALESCE(status,'')!='deleted'{search_clause}
        ORDER BY created_at DESC
        LIMIT ?
        """,
        [*params, *search_params, limit],
    ).fetchall()
    out = rows_to_dicts(rows)
    for row in out:
        for field in ["issuer_bank", "from_bank", "to_bank"]:
            row[field] = display_bank(row.get(field))
        row["image_url"] = image_url_for(row.get("file_id"), row.get("id"))
        row["flow_type"] = flow_type_for_title(row.get("chat_title"), row.get("bot_key"), row.get("chat_id"))
        row["flow_label"] = flow_label(row["flow_type"])
    return out


def empty_account_slip_search(search: str = "") -> Dict[str, Any]:
    return {"query": clean_display(search), "count": 0, "amount": 0.0, "rows": [], "truncated": False}


DASHBOARD_LITE_EMPTY_ARRAY_KEYS = (
    "recent",
    "duplicate_pairs",
    "source_bank_review",
    "deposit_customer_slips",
    "issues",
    "jobs_recent",
    "provider_usage",
    "company_account_daily",
    "withdraw_limit_usage",
    "account_cross_company",
    "by_transferor",
    "by_account_day",
    "by_date",
    "daily_flow_summary",
    "by_from_bank",
    "by_to_bank",
    "by_sender",
)


def lite_dashboard_snapshot(snapshot: Dict[str, Any]) -> Dict[str, Any]:
    """Return a polling-safe dashboard payload with totals/navigation only.

    The browser auto-refresh runs every 10 seconds; it should not re-download slip-card
    image URLs and detail tables unless the operator explicitly refreshes or changes a
    filter.  Keep counts/totals/company navigation fresh, but strip detail-heavy arrays.
    """
    out = dict(snapshot)
    out["detail_level"] = "lite"
    out["detail_omitted"] = list(DASHBOARD_LITE_EMPTY_ARRAY_KEYS) + ["account_slip_search", "cross_company_account_slip_search"]
    for key in DASHBOARD_LITE_EMPTY_ARRAY_KEYS:
        out[key] = []
    query = clean_display(snapshot.get("slip_search"))
    out["account_slip_search"] = empty_account_slip_search(query)
    cross_empty = empty_account_slip_search(query)
    cross_empty.update({"company_count": 0, "companies": [], "is_cross_company": False})
    out["cross_company_account_slip_search"] = cross_empty
    return out


def account_slip_match(row: sqlite3.Row, search: str) -> Tuple[str, str]:
    """Return matched account side for an operator account/name/bank search."""
    query = clean_display(search)
    if not query:
        return "", ""
    needle = normalize_match_text(query)
    raw_needle = query.lower()
    query_account_pattern = masked_account_pattern(query)
    sides: List[Tuple[str, str]] = []
    candidates = [
        ("from", "บัญชีต้นทาง/ผู้โอน", row["from_account"], [row["from_account"], row["from_bank"], row["transferor_name"], row["sender_name"]]),
        ("to", "บัญชีปลายทาง/ผู้รับ", row["to_account"], [row["to_account"], row["to_bank"], row["recipient_name"], row["account_name"]]),
    ]
    for side, label, account_value, values in candidates:
        account_pattern = masked_account_pattern(account_value)
        if query_account_pattern and masked_account_patterns_compatible(query_account_pattern, account_pattern):
            sides.append((side, label))
            continue
        compact_values = [normalize_match_text(v) for v in values if clean_display(v)]
        raw_values = [clean_display(v).lower() for v in values if clean_display(v)]
        haystack = normalize_match_text(" ".join(clean_display(v) for v in values))
        if (
            (needle and (needle in haystack or any(value and (needle in value or value in needle) for value in compact_values)))
            or any(raw_needle and raw_needle in value for value in raw_values)
        ):
            sides.append((side, label))
    if not sides:
        return "", ""
    if len({side for side, _ in sides}) > 1:
        return "both", "ต้นทางและปลายทาง"
    return sides[0]


def account_slip_search_company_summary(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    companies: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        bot = clean_display(row.get("bot_key")) or "default"
        company = clean_company_name(row.get("company_name"), bot)
        group = companies.setdefault(bot, {"bot_key": bot, "company_name": company, "count": 0, "amount": 0.0})
        group["count"] += 1
        group["amount"] += float(row.get("amount") or 0)
    return sorted(companies.values(), key=dict_company_sort_key)


ACCOUNT_SLIP_SEARCH_SQL_FIELDS = (
    "from_account",
    "to_account",
    "transferor_name",
    "recipient_name",
    "account_name",
    "sender_name",
    "username",
    "from_bank",
    "to_bank",
    "issuer_bank",
    "reference_no",
    "seq",
    "aid",
)
SQL_COMPACT_REMOVE_CHARS = (" ", "\t", "\n", "\r", "*", "\u200b", "\u200c", "\u200d", ".", "-", "_", "/", "|", "(", ")")


def sql_literal(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def sql_compact_text_expr(column: str) -> str:
    expr = f"LOWER(COALESCE({column},''))"
    for ch in SQL_COMPACT_REMOVE_CHARS:
        expr = f"REPLACE({expr}, {sql_literal(ch)}, '')"
    return expr


def account_search_sql_predicate(search: str) -> Tuple[str, List[Any]]:
    """Broad SQL-side prefilter mirroring account_slip_match normalization."""
    query = clean_display(search)
    if not query:
        return "", []
    raw_needle = query.lower()
    compact_needle = normalize_match_text(query)
    account_pattern = masked_account_pattern(query)
    masked_token = masked_account_search_token(account_pattern)
    clauses: List[str] = []
    params: List[Any] = []
    for field in ACCOUNT_SLIP_SEARCH_SQL_FIELDS:
        raw_expr = f"LOWER(COALESCE({field},''))"
        compact_expr = sql_compact_text_expr(field)
        if raw_needle:
            clauses.append(f"{raw_expr} LIKE ?")
            params.append(f"%{raw_needle}%")
        if compact_needle:
            clauses.append(f"{compact_expr} LIKE ?")
            params.append(f"%{compact_needle}%")
            # Preserve account_slip_match's conservative reverse containment
            # behavior, e.g. a query like "SCB 123456" may match bank+account.
            clauses.append(f"({compact_expr} <> '' AND ? LIKE '%' || {compact_expr} || '%')")
            params.append(compact_needle)
        if masked_token and field in {"from_account", "to_account"}:
            clauses.append(f"{compact_expr} LIKE ?")
            params.append(f"%{masked_token}%")
    return "(" + " OR ".join(clauses) + ")", params


def account_slip_search_rows(conn: sqlite3.Connection, where_clause: str, params: List[Any], search: str = "", limit: int = 80) -> Dict[str, Any]:
    """Individual counted slip rows for auditing one account within the selected day/company/flow scope."""
    query = clean_display(search)
    if not query:
        return empty_account_slip_search(query)
    search_where, search_params = account_search_sql_predicate(query)
    effective_where = f"({where_clause}) AND {search_where}" if search_where else where_clause
    effective_params = list(params or []) + search_params
    rows = conn.execute(
        f"""
        SELECT id, COALESCE(NULLIF(bot_key,''),'default') AS bot_key, company_name, chat_id, chat_title,
               file_id, message_id, sender_name, username, status, error,
               slip_date_display, slip_date_iso, slip_time,
               TRIM(COALESCE(NULLIF(slip_date_display,''), NULLIF(slip_date_iso,''), '') || ' ' || COALESCE(NULLIF(slip_time,''), '')) AS slip_date_text,
               transferor_name, recipient_name, issuer_bank, from_bank, from_account, to_bank, to_account,
               account_name, amount, fee, reference_no, seq, aid, confidence, created_at, created_at_iso, settlement_id, is_duplicate, duplicate_of
        FROM slips
        WHERE {effective_where}
          AND (NULLIF(TRIM(COALESCE(from_account,'')), '') IS NOT NULL
               OR NULLIF(TRIM(COALESCE(to_account,'')), '') IS NOT NULL
               OR NULLIF(TRIM(COALESCE(transferor_name,'')), '') IS NOT NULL
               OR NULLIF(TRIM(COALESCE(recipient_name,'')), '') IS NOT NULL)
        ORDER BY COALESCE(NULLIF(slip_date_iso,''), '') DESC,
                 COALESCE(NULLIF(slip_time,''), '') DESC,
                 CAST(COALESCE(message_id,0) AS INTEGER) DESC,
                 created_at DESC
        """,
        effective_params,
    ).fetchall()
    matched: List[Dict[str, Any]] = []
    total_amount = 0.0
    for row in rows:
        matched_side, matched_label = account_slip_match(row, query)
        if not matched_side:
            continue
        date_key, date_label, _ = date_bucket(row["slip_date_display"], row["slip_date_iso"])
        amount = float(row["amount"] or 0)
        total_amount += amount
        item = dict(row)
        clean_company_fields(item)
        for field in ["issuer_bank", "from_bank", "to_bank"]:
            item[field] = display_bank(item.get(field))
        item.update(
            {
                "date_key": date_key,
                "date": date_label,
                "amount": amount,
                "fee": float(row["fee"] or 0),
                "reference": clean_display(row["reference_no"] or row["seq"] or row["aid"]),
                "matched_side": matched_side,
                "matched_label": matched_label,
                "flow_type": flow_type_for_title(row["chat_title"], row["bot_key"], row["chat_id"]),
                "flow_label": flow_label(flow_type_for_title(row["chat_title"], row["bot_key"], row["chat_id"])),
                "image_url": image_url_for(row["file_id"], row["id"]),
            }
        )
        matched.append(item)
    company_rows = account_slip_search_company_summary(matched)
    row_limit = int(limit or 0)
    visible_rows = matched if row_limit <= 0 else matched[:row_limit]
    return {"query": query, "count": len(matched), "amount": total_amount, "rows": visible_rows, "truncated": row_limit > 0 and len(matched) > row_limit, "company_count": len(company_rows), "companies": company_rows}


def cross_company_account_slip_search_rows(conn: sqlite3.Connection, scope: str = "all", flow_type: str = "all", search: str = "", limit: int = 120) -> Dict[str, Any]:
    """Individual counted withdrawal slip rows for one account across every company in the selected date scope."""
    query = clean_display(search)
    if not query:
        result = empty_account_slip_search(query)
        result.update({"company_count": 0, "companies": [], "is_cross_company": False})
        return result
    effective_flow = cross_company_withdraw_flow_type(flow_type)
    if not effective_flow:
        result = empty_account_slip_search(query)
        result.update({"company_count": 0, "companies": [], "is_cross_company": False})
        return result
    where_clause, params, _ = global_scope_where(scope, success_only=True)
    where_clause, params = apply_flow_sql(where_clause, params, effective_flow)
    result = account_slip_search_rows(conn, where_clause, params, query, limit=limit)
    company_rows = result.get("companies") or []
    company_count = len(company_rows)
    if company_count < 2:
        empty = empty_account_slip_search(query)
        empty.update({"company_count": company_count, "companies": company_rows, "is_cross_company": False})
        return empty
    result.update({"company_count": company_count, "companies": company_rows, "is_cross_company": True})
    return result


def count_amount_for_where(conn: sqlite3.Connection, where_clause: str, params: List[Any]) -> Dict[str, Any]:
    row = conn.execute(f"SELECT COUNT(*) AS count, COALESCE(SUM(amount),0) AS amount FROM slips WHERE {where_clause}", params).fetchone()
    return {"count": int(row["count"] or 0), "amount": float(row["amount"] or 0)}


def flow_split_panels(
    conn: sqlite3.Connection,
    selected_where: str,
    selected_params: List[Any],
    flow_type: str,
    account_limits: Dict[str, Dict[str, Any]],
    search: str = "",
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, Any], Dict[str, Any], List[Dict[str, Any]]]:
    limit_where, limit_params = limit_scope_clause(selected_where, selected_params, flow_type)
    deposit_where, deposit_params = deposit_customer_scope_clause(selected_where, selected_params, flow_type)
    by_transferor = transferor_totals(conn, limit_where, limit_params, account_limits=account_limits)
    by_account_day = daily_account_totals(conn, limit_where, limit_params, account_limits=account_limits)
    withdraw_limit_usage = withdraw_limit_usage_summary(daily_account_totals(conn, limit_where, limit_params, account_limits=account_limits, limit=0))
    deposit_customer_slips = slip_list_rows(conn, deposit_where, deposit_params, search)
    withdraw_limit_totals = count_amount_for_where(conn, limit_where, limit_params)
    deposit_customer_totals = count_amount_for_where(conn, deposit_where, deposit_params)
    return by_transferor, by_account_day, deposit_customer_slips, withdraw_limit_totals, deposit_customer_totals, withdraw_limit_usage


def ensure_bank_review_log_table(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS bank_review_logs (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          slip_id TEXT NOT NULL,
          bot_key TEXT,
          chat_id TEXT,
          provider TEXT,
          model TEXT,
          applied INTEGER DEFAULT 0,
          previous_json TEXT,
          suggested_json TEXT,
          created_at INTEGER NOT NULL
        )
        """
    )


_DASHBOARD_TOKENS_READY = False
_DASHBOARD_TOKENS_BOOTSTRAPPED = False


def _hash_token(token: str) -> str:
    return hashlib.sha256((token or "").encode("utf-8", "replace")).hexdigest()


def dashboard_owner_login_enabled() -> bool:
    return bool(DASHBOARD_OWNER_USER and DASHBOARD_OWNER_PASSWORD)


def dashboard_owner_session_token() -> str:
    if not dashboard_owner_login_enabled():
        return ""
    material = f"auditslip-owner-session-v1|{DASHBOARD_OWNER_USER}|{DASHBOARD_OWNER_PASSWORD}"
    return hashlib.sha256(material.encode("utf-8", "replace")).hexdigest()


def dashboard_owner_credentials_valid(username: str, password: str) -> bool:
    if not dashboard_owner_login_enabled():
        return False
    import hmac as _hmac
    return _hmac.compare_digest(clean_display(username), DASHBOARD_OWNER_USER) and _hmac.compare_digest(str(password or ""), DASHBOARD_OWNER_PASSWORD)


def dashboard_owner_session_role(token: str) -> str:
    expected = dashboard_owner_session_token()
    if not expected or not token:
        return ""
    import hmac as _hmac
    return "admin" if _hmac.compare_digest(token, expected) else ""


def ensure_dashboard_tokens_table(db_path: Path) -> None:
    """Create dashboard_tokens table (idempotent). Stores only sha256 hashes, never raw tokens."""
    global _DASHBOARD_TOKENS_READY
    if _DASHBOARD_TOKENS_READY:
        return
    with connect(db_path) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS dashboard_tokens (
              token_hash TEXT PRIMARY KEY,
              role TEXT NOT NULL DEFAULT 'viewer',
              label TEXT NOT NULL DEFAULT '',
              created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
              last_used_at TEXT DEFAULT NULL,
              revoked_at TEXT DEFAULT NULL
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS idx_dashboard_tokens_role ON dashboard_tokens(role)")
        conn.commit()
    _DASHBOARD_TOKENS_READY = True


def bootstrap_dashboard_admin_token(db_path: Path, legacy_token: str) -> None:
    """If the dashboard_tokens table is empty AND a legacy token is set, register it as admin."""
    global _DASHBOARD_TOKENS_BOOTSTRAPPED
    if _DASHBOARD_TOKENS_BOOTSTRAPPED:
        return
    if not legacy_token:
        _DASHBOARD_TOKENS_BOOTSTRAPPED = True
        return
    ensure_dashboard_tokens_table(db_path)
    with connect(db_path) as conn:
        row = conn.execute("SELECT COUNT(*) AS c FROM dashboard_tokens").fetchone()
        count = int(row["c"]) if row else 0
        if count == 0:
            conn.execute(
                "INSERT INTO dashboard_tokens(token_hash, role, label) VALUES (?,?,?)",
                (_hash_token(legacy_token), "admin", "legacy-bootstrap"),
            )
            conn.commit()
            print("Dashboard token registered as admin (legacy). Use /api/tokens to manage.", flush=True)
    _DASHBOARD_TOKENS_BOOTSTRAPPED = True


def lookup_dashboard_token_role(db_path: Path, token: str) -> str:
    """Return the role for a token (sha256 lookup). Empty string if missing/revoked.

    Updates last_used_at on a successful (non-revoked) lookup.
    """
    if not token:
        return ""
    ensure_dashboard_tokens_table(db_path)
    th = _hash_token(token)
    with connect(db_path) as conn:
        row = conn.execute(
            "SELECT role, revoked_at FROM dashboard_tokens WHERE token_hash=?",
            (th,),
        ).fetchone()
        if not row:
            return ""
        if row["revoked_at"]:
            return ""
        try:
            conn.execute(
                "UPDATE dashboard_tokens SET last_used_at=? WHERE token_hash=?",
                (dt.datetime.now(dt.timezone.utc).isoformat(), th),
            )
            conn.commit()
        except Exception:
            pass
        return str(row["role"] or "")


def dashboard_token_is_registered(db_path: Path, token: str) -> bool:
    """Return True if the token hash exists in dashboard_tokens (revoked or not).

    Used to decide whether the legacy DASHBOARD_TOKEN env fallback applies: once a
    token is registered, the DB is the source of truth — even for the legacy admin.
    This closes the revoke bypass where env-token holders kept admin after the
    bootstrap row was revoked.
    """
    if not token:
        return False
    ensure_dashboard_tokens_table(db_path)
    th = _hash_token(token)
    try:
        with connect(db_path) as conn:
            row = conn.execute(
                "SELECT 1 FROM dashboard_tokens WHERE token_hash=? LIMIT 1",
                (th,),
            ).fetchone()
        return bool(row)
    except Exception:
        return False


def list_dashboard_tokens(db_path: Path) -> List[Dict[str, Any]]:
    ensure_dashboard_tokens_table(db_path)
    with connect(db_path) as conn:
        rows = conn.execute(
            "SELECT token_hash, role, label, created_at, last_used_at, revoked_at "
            "FROM dashboard_tokens ORDER BY created_at DESC"
        ).fetchall()
    out: List[Dict[str, Any]] = []
    for r in rows:
        out.append({
            "token_hash_prefix": str(r["token_hash"])[:12],
            "role": r["role"],
            "label": r["label"],
            "created_at": r["created_at"],
            "last_used_at": r["last_used_at"],
            "revoked_at": r["revoked_at"],
        })
    return out


def create_dashboard_token(db_path: Path, role: str, label: str) -> Dict[str, Any]:
    role = (role or "").strip().lower()
    if role not in {"admin", "auditor", "operator", "viewer"}:
        return {"ok": False, "error": "invalid role"}
    label = (label or "").strip()[:120]
    ensure_dashboard_tokens_table(db_path)
    import secrets as _secrets
    token = _secrets.token_hex(32)
    th = _hash_token(token)
    with connect(db_path) as conn:
        conn.execute(
            "INSERT INTO dashboard_tokens(token_hash, role, label) VALUES (?,?,?)",
            (th, role, label),
        )
        conn.commit()
    return {"ok": True, "token": token, "token_hash_prefix": th[:12], "role": role, "label": label}


def revoke_dashboard_token(db_path: Path, token_hash_prefix: str) -> Dict[str, Any]:
    prefix = (token_hash_prefix or "").strip().lower()
    if len(prefix) < 6:
        return {"ok": False, "error": "prefix too short"}
    ensure_dashboard_tokens_table(db_path)
    with connect(db_path) as conn:
        rows = conn.execute(
            "SELECT token_hash, role, revoked_at FROM dashboard_tokens WHERE token_hash LIKE ?",
            (prefix + "%",),
        ).fetchall()
        if not rows:
            return {"ok": False, "error": "token not found"}
        if len(rows) > 1:
            return {"ok": False, "error": "prefix is ambiguous"}
        target = rows[0]
        if target["revoked_at"]:
            return {"ok": False, "error": "already revoked"}
        if (target["role"] or "") == "admin":
            active_admin_row = conn.execute(
                "SELECT COUNT(*) AS c FROM dashboard_tokens WHERE role='admin' AND revoked_at IS NULL"
            ).fetchone()
            if int(active_admin_row["c"] if active_admin_row else 0) <= 1:
                return {"ok": False, "error": "cannot revoke last admin"}
        conn.execute(
            "UPDATE dashboard_tokens SET revoked_at=? WHERE token_hash=?",
            (dt.datetime.now(dt.timezone.utc).isoformat(), target["token_hash"]),
        )
        conn.commit()
    return {"ok": True, "token_hash_prefix": target["token_hash"][:12]}


def active_dashboard_token_count(db_path: Path, roles: set[str] | None = None) -> int:
    ensure_dashboard_tokens_table(db_path)
    with connect(db_path) as conn:
        if roles:
            placeholders = ",".join(["?"] * len(roles))
            row = conn.execute(
                f"SELECT COUNT(*) AS c FROM dashboard_tokens WHERE revoked_at IS NULL AND role IN ({placeholders})",
                list(roles),
            ).fetchone()
        else:
            row = conn.execute("SELECT COUNT(*) AS c FROM dashboard_tokens WHERE revoked_at IS NULL").fetchone()
    return int(row["c"] if row else 0)


def simple_approval_enabled(db_path: Path, actor_fp: str = "", actor_role: str = "") -> bool:
    """Allow a one-token production setup to approve+execute its own pending action.

    This keeps the pending/audit-chain record, but removes the impossible two-person
    requirement when the live dashboard has only one active token. Set
    AUDITSLIP_SIMPLE_APPROVAL=0 to force strict two-person approval.
    """
    if os.environ.get("AUDITSLIP_SIMPLE_APPROVAL", "1") in {"0", "false", "False", "no"}:
        return False
    if (actor_role or "").strip().lower() != "admin":
        return False
    return active_dashboard_token_count(db_path) == 1


_MUTATION_LOG_READY = False


def ensure_dashboard_mutation_log_table(db_path: Path) -> None:
    global _MUTATION_LOG_READY
    if _MUTATION_LOG_READY:
        return
    with connect(db_path) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS dashboard_mutation_log (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              ts_iso TEXT NOT NULL,
              action TEXT NOT NULL,
              actor TEXT,
              chat_id TEXT,
              bot_key TEXT,
              slip_id TEXT,
              payload_json TEXT,
              result_status TEXT NOT NULL,
              result_summary TEXT
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS idx_dashboard_mutation_log_ts ON dashboard_mutation_log(ts_iso)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_dashboard_mutation_log_action ON dashboard_mutation_log(action)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_dashboard_mutation_log_slip ON dashboard_mutation_log(slip_id)")
        # Phase B1: hash-chain columns (idempotent on re-run; columns may already exist).
        try:
            conn.execute("ALTER TABLE dashboard_mutation_log ADD COLUMN prev_hash TEXT DEFAULT ''")
        except sqlite3.OperationalError:
            pass
        try:
            conn.execute("ALTER TABLE dashboard_mutation_log ADD COLUMN entry_hash TEXT DEFAULT ''")
        except sqlite3.OperationalError:
            pass
        conn.execute("CREATE INDEX IF NOT EXISTS idx_mutation_entry_hash ON dashboard_mutation_log(entry_hash)")
        _backfill_legacy_mutation_hashes(conn)
        conn.commit()
    _MUTATION_LOG_READY = True


# Phase B1: canonical hash for mutation rows.
# IMPORTANT: keep this byte-identical to tools/verify_audit_chain.py::compute_mutation_hash.
# 'id' is AUTOINCREMENT (unknown when the writer computes the hash), and 'entry_hash' is the
# field being computed; both must be excluded so the writer's hash matches the verifier's.
_MUTATION_HASH_EXCLUDE_KEYS = ("id", "entry_hash")


def compute_mutation_hash(prev_hash: str, row: Dict[str, Any]) -> str:
    canonical_obj = {k: row[k] for k in sorted(row.keys()) if k not in _MUTATION_HASH_EXCLUDE_KEYS}
    canonical = json.dumps(canonical_obj, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(((prev_hash or "") + "|" + canonical).encode("utf-8")).hexdigest()


def _backfill_legacy_mutation_hashes(conn: sqlite3.Connection) -> None:
    """Populate hashes for pre-Phase-B mutation rows whose hash fields were empty.

    This is a one-time migration for production rows created before `prev_hash` and
    `entry_hash` existed. Never overwrite non-empty hashes: if a row already carries
    a hash, the verifier must be allowed to catch any mismatch instead of silently
    repairing possible tampering.
    """
    rows = conn.execute(
        "SELECT id, ts_iso, action, actor, chat_id, bot_key, slip_id, payload_json, "
        "       result_status, result_summary, prev_hash, entry_hash "
        "FROM dashboard_mutation_log ORDER BY id ASC"
    ).fetchall()
    last_entry_hash = ""
    for row in rows:
        row_dict = dict(row)
        stored_prev = row_dict.get("prev_hash") or ""
        stored_entry = row_dict.get("entry_hash") or ""
        if not stored_prev and not stored_entry:
            row_dict["prev_hash"] = last_entry_hash
            entry_hash = compute_mutation_hash(last_entry_hash, row_dict)
            conn.execute(
                "UPDATE dashboard_mutation_log SET prev_hash=?, entry_hash=? WHERE id=?",
                (last_entry_hash, entry_hash, row_dict["id"]),
            )
            last_entry_hash = entry_hash
        else:
            last_entry_hash = stored_entry


_MUTATION_PAYLOAD_REDACT_KEYS = {"token", "cookie", "authorization", "auth", "password", "secret"}


def _sanitize_mutation_payload(payload: Any) -> Dict[str, Any]:
    if not isinstance(payload, dict):
        return {"_raw_type": type(payload).__name__}
    out: Dict[str, Any] = {}
    for k, v in payload.items():
        if k.lower() in _MUTATION_PAYLOAD_REDACT_KEYS:
            out[k] = "***"
        elif isinstance(v, (bytes, bytearray, memoryview)):
            out[k] = f"<bytes:{len(v)}>"
        elif isinstance(v, str) and len(v) > 500:
            out[k] = v[:500] + "...(truncated)"
        else:
            out[k] = v
    return out


def record_mutation(
    db_path: Path,
    action: str,
    *,
    actor: str = "",
    chat_id: str = "",
    bot_key: str = "",
    slip_id: str = "",
    payload: Any = None,
    result_status: str = "ok",
    result_summary: str = "",
) -> None:
    """Append row to dashboard_mutation_log. Failure must not break the original mutation.

    Phase B1: each row is hash-chained to its predecessor. Fetch-prev-then-insert runs in a
    single BEGIN IMMEDIATE transaction so concurrent writers cannot race on prev_hash. If hash
    computation or insert fails we still swallow the exception (matches existing posture: do not
    break the caller's mutation), which means a missing row is invisible to the verifier --
    deliberate tradeoff -- see commit message.
    """
    try:
        ensure_dashboard_mutation_log_table(db_path)
        row_data = {
            "ts_iso": dt.datetime.now(dt.timezone.utc).isoformat(),
            "action": action,
            "actor": actor or "",
            "chat_id": chat_id or "",
            "bot_key": bot_key or "",
            "slip_id": slip_id or "",
            "payload_json": json.dumps(_sanitize_mutation_payload(payload), ensure_ascii=False, default=str),
            "result_status": result_status,
            "result_summary": (result_summary or "")[:500],
        }
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        # Manual transaction control so BEGIN IMMEDIATE actually takes the reserved lock.
        conn.isolation_level = None
        try:
            conn.execute("BEGIN IMMEDIATE")
            prev_row = conn.execute(
                "SELECT entry_hash FROM dashboard_mutation_log ORDER BY id DESC LIMIT 1"
            ).fetchone()
            prev_hash = (prev_row["entry_hash"] if prev_row else "") or ""
            row_for_hash = dict(row_data)
            row_for_hash["prev_hash"] = prev_hash
            entry_hash = compute_mutation_hash(prev_hash, row_for_hash)
            conn.execute(
                "INSERT INTO dashboard_mutation_log "
                "(ts_iso, action, actor, chat_id, bot_key, slip_id, payload_json, result_status, result_summary, prev_hash, entry_hash) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                (
                    row_data["ts_iso"],
                    row_data["action"],
                    row_data["actor"],
                    row_data["chat_id"],
                    row_data["bot_key"],
                    row_data["slip_id"],
                    row_data["payload_json"],
                    row_data["result_status"],
                    row_data["result_summary"],
                    prev_hash,
                    entry_hash,
                ),
            )
            conn.execute("COMMIT")
        except Exception:
            try:
                conn.execute("ROLLBACK")
            except Exception:
                pass
            raise
        finally:
            conn.close()
    except Exception:
        logger.exception("record_mutation failed action=%s", action)


# Phase C2: high-risk action names (mutation-log convention -- short tokens, not dotted) that
# warrant a real-time Telegram alert to the owner. Lower-risk actions (reprocess, unmark_dup,
# bank_review*) are intentionally excluded.
_ALERT_ACTIONS = {
    "delete",
    "close",
    "clear",
    "account_limit",
    "company_account",
    "token.create",
    "token.revoke",
}


def send_owner_alert(message: str, action: str = "", request_id: str = "") -> None:
    """Best-effort Telegram alert to the watchdog/owner chat.

    Reads same env vars as auditslip_watchdog. Network failure is swallowed (logged at WARNING).
    Returns immediately if token/chat are unset so dev/test environments do not 'fail'.
    """
    try:
        token = (
            os.environ.get("AUDITSLIP_WATCHDOG_BOT_TOKEN")
            or os.environ.get("BOT_TOKEN")
            or os.environ.get("TELEGRAM_BOT_TOKEN")
            or ""
        )
        chat_id = os.environ.get("AUDITSLIP_WATCHDOG_ALERT_CHAT_ID") or next(
            (x.strip() for x in os.environ.get("AUDITSLIP_ADMIN_IDS", "").split(",") if x.strip()),
            "",
        )
        if not token or not chat_id:
            return
        url = f"https://api.telegram.org/bot{token}/sendMessage"
        resp = requests.post(url, data={"chat_id": chat_id, "text": message}, timeout=10)
        # Touch status_code to surface obvious 4xx in logs without raising.
        if resp.status_code >= 400:
            logger.warning("send_owner_alert non-2xx status=%s action=%s req=%s", resp.status_code, action, request_id)
    except Exception:
        logger.warning("send_owner_alert failed action=%s req=%s", action, request_id, exc_info=True)


def record_endpoint_mutation(db_path: Path, action: str, *, actor: str = "", request_id: str = "", payload: Any = None, result_status: str = "ok", result_summary: str = "", chat_id: str = "", bot_key: str = "", slip_id: str = "") -> None:
    """Thin wrapper around record_mutation that prefixes the per-request request_id into the result_summary.

    Phase C2: after logging, if `action` is in _ALERT_ACTIONS and AUDITSLIP_ALERT_ON_MUTATION!=0,
    fire a best-effort Telegram alert. Alert is fire-and-forget; failures never break the mutation.
    """
    prefix = f"req={request_id} " if request_id else ""
    record_mutation(db_path, action, actor=actor, chat_id=chat_id, bot_key=bot_key, slip_id=slip_id, payload=payload, result_status=result_status, result_summary=(prefix + (result_summary or "")))
    if action in _ALERT_ACTIONS and os.environ.get("AUDITSLIP_ALERT_ON_MUTATION", "1") != "0":
        try:
            actor_short = (actor or "")[:8]
            summary_snip = (result_summary or "")[:200]
            text = (
                f"🚨 Auditslip mutation: {action} by {actor_short} "
                f"req={request_id or '-'} result={result_status}"
            )
            if summary_snip:
                text = f"{text}\n{summary_snip}"
            send_owner_alert(text, action=action, request_id=request_id)
        except Exception:
            logger.warning("record_endpoint_mutation alert dispatch failed action=%s", action)


# Phase B1: parse a 'request_id' (12-hex) prefix stamped by record_endpoint_mutation back out
# of result_summary, so the tail endpoint can surface it without leaking the full summary text.
_REQUEST_ID_PREFIX_RE = re.compile(r"^req=([0-9a-fA-F]{8,32})\s")


def _extract_request_id(result_summary: str) -> str:
    if not result_summary:
        return ""
    m = _REQUEST_ID_PREFIX_RE.match(result_summary)
    return m.group(1) if m else ""


# ---------------------------------------------------------------------------
# Phase B3: two-person approval workflow for destructive operations
# ---------------------------------------------------------------------------

PENDING_ACTION_TTL_HOURS = 24
APPROVAL_REQUIRED_ACTIONS = {"slip.delete", "period.close", "company.account", "reconcile.run", "ledger.import"}
_PENDING_ACTIONS_READY = False


def ensure_pending_actions_table(db_path: Path) -> None:
    """Create pending_actions table if missing. Idempotent."""
    global _PENDING_ACTIONS_READY
    if _PENDING_ACTIONS_READY:
        return
    with connect(db_path) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS pending_actions (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              action TEXT NOT NULL,
              payload_json TEXT NOT NULL,
              requested_by TEXT NOT NULL,
              requested_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
              approved_by TEXT DEFAULT NULL,
              approved_at TEXT DEFAULT NULL,
              executed_at TEXT DEFAULT NULL,
              executed_result TEXT DEFAULT NULL,
              status TEXT NOT NULL DEFAULT 'pending',
              expires_at TEXT NOT NULL,
              request_id TEXT NOT NULL
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS idx_pending_status ON pending_actions(status, expires_at)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_pending_action ON pending_actions(action)")
        conn.commit()
    _PENDING_ACTIONS_READY = True


def _utc_now_iso() -> str:
    return dt.datetime.now(dt.timezone.utc).isoformat()


def _utc_expiry_iso(hours: int = PENDING_ACTION_TTL_HOURS) -> str:
    return (dt.datetime.now(dt.timezone.utc) + dt.timedelta(hours=hours)).isoformat()


def create_pending_action(
    db_path: Path,
    *,
    action: str,
    payload: Any,
    requested_by: str,
    request_id: str = "",
    ttl_hours: int = PENDING_ACTION_TTL_HOURS,
) -> int:
    """Insert a pending action row. Returns the new row id."""
    ensure_pending_actions_table(db_path)
    now_iso = _utc_now_iso()
    expires_iso = _utc_expiry_iso(ttl_hours)
    payload_json = json.dumps(_sanitize_mutation_payload(payload), ensure_ascii=False, default=str)
    with connect(db_path) as conn:
        cur = conn.execute(
            "INSERT INTO pending_actions "
            "(action, payload_json, requested_by, requested_at, status, expires_at, request_id) "
            "VALUES (?,?,?,?,?,?,?)",
            (action, payload_json, requested_by or "", now_iso, "pending", expires_iso, request_id or ""),
        )
        conn.commit()
        return int(cur.lastrowid)


def request_pending_action_once(
    db_path: Path,
    *,
    action: str,
    payload: Any,
    requested_by: str,
    request_id: str = "",
    ttl_hours: int = PENDING_ACTION_TTL_HOURS,
) -> Dict[str, Any]:
    """Create one pending action, reusing an identical pending request if it exists.

    Dashboard delete buttons are easy to tap repeatedly on mobile.  Because risky
    mutations require two-person approval, the first tap must create a pending row;
    subsequent taps for the same action/payload should explain the existing pending
    request instead of creating duplicate approvals that still do not remove the card.
    """
    ensure_pending_actions_table(db_path)
    expire_old_pending_actions(db_path)
    payload_json = json.dumps(_sanitize_mutation_payload(payload), ensure_ascii=False, default=str)
    with connect(db_path) as conn:
        existing = conn.execute(
            """
            SELECT id, request_id, expires_at
            FROM pending_actions
            WHERE action=? AND payload_json=? AND status='pending'
            ORDER BY id DESC
            LIMIT 1
            """,
            (action, payload_json),
        ).fetchone()
    if existing:
        return {
            "ok": True,
            "status": "pending",
            "pending_id": int(existing["id"]),
            "request_id": str(existing["request_id"] or ""),
            "expires_at": existing["expires_at"],
            "expires_in_hours": ttl_hours,
            "already_pending": True,
        }
    new_request_id = request_id or uuid.uuid4().hex[:12]
    pending_id = create_pending_action(
        db_path,
        action=action,
        payload=payload,
        requested_by=requested_by,
        request_id=new_request_id,
        ttl_hours=ttl_hours,
    )
    return {
        "ok": True,
        "status": "pending",
        "pending_id": pending_id,
        "request_id": new_request_id,
        "expires_in_hours": ttl_hours,
        "already_pending": False,
    }


def load_pending_action(db_path: Path, pending_id: int) -> Dict[str, Any]:
    """Return pending row as dict or {} if missing."""
    ensure_pending_actions_table(db_path)
    with connect(db_path) as conn:
        row = conn.execute("SELECT * FROM pending_actions WHERE id=?", (int(pending_id),)).fetchone()
    return dict(row) if row else {}


def list_pending_actions(db_path: Path, *, status: str = "", limit: int = 200) -> List[Dict[str, Any]]:
    """List pending actions newest first. Filter by status if non-empty."""
    ensure_pending_actions_table(db_path)
    with connect(db_path) as conn:
        if status:
            rows = conn.execute(
                "SELECT id, action, payload_json, requested_by, requested_at, approved_by, approved_at, "
                "executed_at, status, expires_at, request_id FROM pending_actions WHERE status=? "
                "ORDER BY id DESC LIMIT ?",
                (status, int(limit)),
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT id, action, payload_json, requested_by, requested_at, approved_by, approved_at, "
                "executed_at, status, expires_at, request_id FROM pending_actions "
                "ORDER BY id DESC LIMIT ?",
                (int(limit),),
            ).fetchall()
    out: List[Dict[str, Any]] = []
    for r in rows:
        d = dict(r)
        try:
            payload = json.loads(d.get("payload_json") or "{}")
        except Exception:
            payload = {}
        d["payload_summary"] = _pending_payload_summary(payload)
        d.pop("payload_json", None)
        out.append(d)
    return out


def _pending_payload_summary(payload: Any) -> Dict[str, Any]:
    """Compact, PII-light view of payload for /api/pending listing."""
    if not isinstance(payload, dict):
        return {}
    summary: Dict[str, Any] = {}
    for key in ("id", "slip_id", "bot_key", "chat_id", "company_name", "reason", "note"):
        if key in payload:
            value = payload.get(key)
            if isinstance(value, str) and len(value) > 80:
                value = value[:80] + "..."
            summary[key] = value
    return summary


def expire_old_pending_actions(db_path: Path) -> int:
    """Mark pending rows whose expires_at < now as 'expired'. Returns rowcount."""
    ensure_pending_actions_table(db_path)
    now_iso = _utc_now_iso()
    with connect(db_path) as conn:
        cur = conn.execute(
            "UPDATE pending_actions SET status='expired' WHERE status='pending' AND expires_at < ?",
            (now_iso,),
        )
        conn.commit()
        return int(cur.rowcount or 0)


def approve_pending_action(db_path: Path, pending_id: int, approver_fp: str, allow_self_approval: bool = False) -> Dict[str, Any]:
    """Approve a pending action. Rejects self-approval unless simple mode explicitly allows it."""
    ensure_pending_actions_table(db_path)
    expire_old_pending_actions(db_path)
    row = load_pending_action(db_path, pending_id)
    if not row:
        return {"ok": False, "error": "pending action not found", "status": 404}
    if row.get("status") != "pending":
        return {"ok": False, "error": f"cannot approve status={row.get('status')}", "status": 409}
    if not approver_fp:
        return {"ok": False, "error": "approver fingerprint required", "status": 403}
    if row.get("requested_by") == approver_fp and not allow_self_approval:
        return {"ok": False, "error": "self-approval not allowed", "status": 403}
    now_iso = _utc_now_iso()
    with connect(db_path) as conn:
        conn.execute(
            "UPDATE pending_actions SET status='approved', approved_by=?, approved_at=? "
            "WHERE id=? AND status='pending'",
            (approver_fp, now_iso, int(pending_id)),
        )
        conn.commit()
    return {"ok": True, "pending_id": int(pending_id), "approved_by": approver_fp, "approved_at": now_iso}


def reject_pending_action(db_path: Path, pending_id: int, rejecter_fp: str, reason: str = "") -> Dict[str, Any]:
    ensure_pending_actions_table(db_path)
    row = load_pending_action(db_path, pending_id)
    if not row:
        return {"ok": False, "error": "pending action not found", "status": 404}
    if row.get("status") not in {"pending", "approved"}:
        return {"ok": False, "error": f"cannot reject status={row.get('status')}", "status": 409}
    now_iso = _utc_now_iso()
    summary = (reason or "")[:480]
    with connect(db_path) as conn:
        conn.execute(
            "UPDATE pending_actions SET status='rejected', approved_by=?, approved_at=?, executed_result=? "
            "WHERE id=?",
            (rejecter_fp or "", now_iso, summary, int(pending_id)),
        )
        conn.commit()
    return {"ok": True, "pending_id": int(pending_id), "rejected_by": rejecter_fp, "reason": summary}


def cancel_pending_action(db_path: Path, pending_id: int, requester_fp: str) -> Dict[str, Any]:
    """Only the original requester may cancel."""
    ensure_pending_actions_table(db_path)
    row = load_pending_action(db_path, pending_id)
    if not row:
        return {"ok": False, "error": "pending action not found", "status": 404}
    if row.get("status") not in {"pending", "approved"}:
        return {"ok": False, "error": f"cannot cancel status={row.get('status')}", "status": 409}
    if not requester_fp or row.get("requested_by") != requester_fp:
        return {"ok": False, "error": "only requester may cancel", "status": 403}
    now_iso = _utc_now_iso()
    with connect(db_path) as conn:
        conn.execute(
            "UPDATE pending_actions SET status='cancelled', executed_at=? WHERE id=?",
            (now_iso, int(pending_id)),
        )
        conn.commit()
    return {"ok": True, "pending_id": int(pending_id), "cancelled_by": requester_fp}


def mark_pending_executed(db_path: Path, pending_id: int, executed_result: str = "") -> None:
    ensure_pending_actions_table(db_path)
    with connect(db_path) as conn:
        conn.execute(
            "UPDATE pending_actions SET status='executed', executed_at=?, executed_result=? WHERE id=?",
            (_utc_now_iso(), (executed_result or "")[:480], int(pending_id)),
        )
        conn.commit()


def pending_action_payload(row: Dict[str, Any]) -> Dict[str, Any]:
    """Parse the stored payload_json of a pending row back into a dict."""
    if not row:
        return {}
    try:
        data = json.loads(row.get("payload_json") or "{}")
    except Exception:
        return {}
    return data if isinstance(data, dict) else {}


def execute_pending_action(db_path: Path, pending_id: int, actor_fp: str) -> Dict[str, Any]:
    """Execute an already-approved pending action using its stored payload."""
    ensure_pending_actions_table(db_path)
    expire_old_pending_actions(db_path)
    row = load_pending_action(db_path, pending_id)
    if not row:
        return {"ok": False, "error": "pending action not found", "status_code": 404}
    if row.get("status") != "approved":
        return {"ok": False, "error": f"not approved (status={row.get('status')})", "pending_id": int(pending_id), "status_code": 409}
    action = str(row.get("action") or "")
    payload = pending_action_payload(row)
    request_id = str(row.get("request_id") or uuid.uuid4().hex[:12])
    result: Dict[str, Any]
    mutation_action = action.replace(".", "_")
    chat_id = str(payload.get("chat_id") or "")
    bot_key = str(payload.get("bot_key") or "")
    try:
        if action == "account.limit":
            result = save_account_limit(
                db_path,
                chat_id,
                str(payload.get("limit_key") or ""),
                str(payload.get("display_name") or ""),
                str(payload.get("bank") or ""),
                str(payload.get("account") or ""),
                parse_number(payload.get("limit_amount")),
            )
            mutation_action = "account_limit"
            summary = str(result.get("limit_key") or result.get("error") or "")
        elif action == "company.account":
            result = save_company_account(
                db_path,
                bot_key=bot_key or "default",
                chat_id=chat_id,
                company_name=str(payload.get("company_name") or ""),
                bank=str(payload.get("bank") or ""),
                account_no=str(payload.get("account_no") or ""),
                account_name=str(payload.get("account_name") or ""),
                daily_limit=parse_number(payload.get("daily_limit")),
            )
            mutation_action = "company_account"
            summary = str(result.get("account_key") or result.get("error") or "")
        elif action == "slip.delete":
            slip_id = str(payload.get("id") or payload.get("slip_id") or "")
            result = delete_dashboard_slip(db_path, slip_id, bot_key, str(payload.get("reason") or "dashboard operator delete"))
            mutation_action = "delete"
            summary = str(result.get("previous_status") or result.get("error") or "")
        elif action == "period.close":
            result = dashboard_close_period(db_path, chat_id, str(payload.get("note") or "dashboard close"), bot_key=bot_key or "default", company_name=str(payload.get("company_name") or ""))
            mutation_action = "close"
            summary = str(result.get("settlement_id") or result.get("error") or "")
        elif action == "reconcile.run":
            flow_type = str(payload.get("flow_type") or "all")
            scope = str(payload.get("scope") or "all")
            excel_path = safe_backend_excel_path(str(payload.get("excel_path") or ""))
            if not excel_path.exists():
                result = {"ok": False, "error": f"excel not found: {excel_path}", "status_code": 404}
                mutation_action = "reconcile"
                summary = "excel not found"
            else:
                result = reconcile_backend_excel(db_path, excel_path, chat_id=chat_id, scope=scope, bot_key=bot_key, flow_type=flow_type)
                mutation_action = "reconcile"
                summary = f"diff={result.get('diff_amount')} matched={result.get('matched', {}).get('count')} missing={result.get('missing', {}).get('count')} extra={result.get('extra', {}).get('count')}"
        elif action == "ledger.import":
            flow_type = str(payload.get("flow_type") or "all")
            scope = str(payload.get("scope") or "all")
            statement_path = safe_statement_file_path(str(payload.get("statement_path") or ""))
            mutation_action = "ledger_import"
            if not statement_path.exists():
                result = {"ok": False, "error": f"statement not found: {statement_path}", "status_code": 404}
                summary = "statement not found"
            else:
                result = import_bank_ledger_statement(
                    db_path,
                    statement_path,
                    bot_key=bot_key,
                    company_name=str(payload.get("company_name") or ""),
                    bank=str(payload.get("bank") or ""),
                    account_no=str(payload.get("account_no") or ""),
                    account_name=str(payload.get("account_name") or ""),
                    scope=scope,
                    flow_type=flow_type,
                    dry_run=False,
                )
                summary = f"inserted={result.get('inserted', {}).get('count')} duplicates={result.get('duplicates', {}).get('count')} matched={result.get('matched', {}).get('count')} ledger_extra={result.get('ledger_extra', {}).get('count')} slip_extra={result.get('slip_extra', {}).get('count')}"
        else:
            return {"ok": False, "error": f"unsupported pending action: {action}", "pending_id": int(pending_id), "status_code": 400}
    except Exception as exc:
        logger.exception("execute_pending_action failed action=%s pending_id=%s", action, pending_id)
        record_endpoint_mutation(db_path, mutation_action, actor=actor_fp, request_id=request_id, chat_id=chat_id, bot_key=bot_key, payload=payload, result_status="error", result_summary=safe_error(exc))
        return {"ok": False, "error": safe_error(exc), "pending_id": int(pending_id), "request_id": request_id, "status_code": 400}

    ok = bool(isinstance(result, dict) and result.get("ok"))
    record_endpoint_mutation(db_path, mutation_action, actor=actor_fp, request_id=request_id, chat_id=chat_id, bot_key=bot_key, slip_id=str(payload.get("id") or payload.get("slip_id") or ""), payload=payload, result_status=("ok" if ok else "error"), result_summary=summary)
    if ok:
        mark_pending_executed(db_path, pending_id, executed_result=summary)
    if isinstance(result, dict):
        result["pending_id"] = int(pending_id)
        result["request_id"] = request_id
        result["action"] = action
        if ok:
            result["status"] = "executed"
    return result


def simple_auto_execute_pending(db_path: Path, pending_id: int, actor_fp: str, actor_role: str) -> Dict[str, Any]:
    if not simple_approval_enabled(db_path, actor_fp, actor_role):
        return {}
    approval = approve_pending_action(db_path, pending_id, actor_fp, allow_self_approval=True)
    if not approval.get("ok"):
        return approval
    result = execute_pending_action(db_path, pending_id, actor_fp)
    if isinstance(result, dict):
        result["simple_approval"] = True
    return result


def verify_mutation_chain(db_path: Path) -> Dict[str, Any]:
    """Walk dashboard_mutation_log in id order; recompute each entry_hash from stored prev_hash
    and the row's content. Returns the first divergence (by id) or ok=True if all match.

    Cross-validates the chain links too: row N's prev_hash must equal row N-1's entry_hash.
    """
    ensure_dashboard_mutation_log_table(db_path)
    total = 0
    first_bad_id: Any = None
    first_bad_reason: Any = None
    with connect(db_path) as conn:
        cursor = conn.execute(
            "SELECT id, ts_iso, action, actor, chat_id, bot_key, slip_id, payload_json, "
            "       result_status, result_summary, prev_hash, entry_hash "
            "FROM dashboard_mutation_log ORDER BY id ASC"
        )
        last_entry_hash = ""
        for row in cursor:
            total += 1
            row_dict = dict(row)
            stored_entry_hash = row_dict.get("entry_hash") or ""
            stored_prev_hash = row_dict.get("prev_hash") or ""
            if first_bad_id is None and stored_prev_hash != last_entry_hash:
                first_bad_id = row_dict["id"]
                first_bad_reason = (
                    f"prev_hash mismatch: row {row_dict['id']} stores prev_hash="
                    f"{stored_prev_hash[:12]}... but prior entry_hash was {last_entry_hash[:12]}..."
                )
            recomputed = compute_mutation_hash(stored_prev_hash, row_dict)
            if first_bad_id is None and recomputed != stored_entry_hash:
                first_bad_id = row_dict["id"]
                first_bad_reason = (
                    f"entry_hash mismatch: row {row_dict['id']} stores {stored_entry_hash[:12]}..."
                    f" but recomputed {recomputed[:12]}..."
                )
            last_entry_hash = stored_entry_hash
    return {
        "ok": first_bad_id is None,
        "total_rows": total,
        "first_bad_id": first_bad_id,
        "first_bad_reason": first_bad_reason,
        "verified_at": dt.datetime.now(dt.timezone.utc).isoformat(),
    }


def mutation_chain_tail(db_path: Path, limit: int = 50) -> List[Dict[str, Any]]:
    """Return the last N entries of dashboard_mutation_log (newest first), WITHOUT the raw payload.

    result_summary is truncated to 200 chars. payload_json is intentionally omitted because it
    can carry account numbers in some endpoints; for full payload access use the standalone
    verifier with direct DB read (auditor scenario).
    """
    try:
        limit = int(limit)
    except (TypeError, ValueError):
        limit = 50
    if limit < 1:
        limit = 1
    if limit > 500:
        limit = 500
    ensure_dashboard_mutation_log_table(db_path)
    out: List[Dict[str, Any]] = []
    with connect(db_path) as conn:
        cursor = conn.execute(
            "SELECT id, ts_iso, actor, action, result_status, result_summary, prev_hash, entry_hash "
            "FROM dashboard_mutation_log ORDER BY id DESC LIMIT ?",
            (limit,),
        )
        for row in cursor:
            d = dict(row)
            summary = (d.get("result_summary") or "")[:200]
            out.append({
                "id": d["id"],
                "ts": d.get("ts_iso") or "",
                "actor_fingerprint": d.get("actor") or "",
                "action": d.get("action") or "",
                "request_id": _extract_request_id(d.get("result_summary") or ""),
                "result_status": d.get("result_status") or "",
                "result_summary": summary,
                "prev_hash": (d.get("prev_hash") or "")[:12],
                "entry_hash": (d.get("entry_hash") or "")[:12],
            })
    return out


def openai_bank_double_check_slip(db_path: Path, slip_id: str, apply: bool = True) -> Dict[str, Any]:
    slip_id = clean_display(slip_id)
    if not slip_id:
        return {"ok": False, "error": "missing slip id"}
    with connect(db_path) as conn:
        row = conn.execute(
            """
            SELECT id, COALESCE(NULLIF(bot_key,''),'default') AS bot_key, chat_id, file_id,
                   issuer_bank, from_bank, to_bank
            FROM slips
            WHERE id=? AND status='success'
            """,
            (slip_id,),
        ).fetchone()
    if not row:
        return {"ok": False, "error": "slip not found"}
    if not row["file_id"]:
        return {"ok": False, "error": "slip image not found"}

    body, mime = fetch_slip_image(db_path, slip_id)
    suffix = mimetypes.guess_extension(mime or "image/jpeg") or ".jpg"
    tmp_name = ""
    try:
        with tempfile.NamedTemporaryFile(prefix="auditslip-bank-review-", suffix=suffix, delete=False) as tmp:
            tmp.write(body)
            tmp_name = tmp.name
        data, meta = openai_extract(Path(tmp_name), mime)
    finally:
        if tmp_name:
            try:
                Path(tmp_name).unlink(missing_ok=True)
            except Exception:
                pass

    normalized = normalize_record(data)
    suggested = {
        "issuer_bank": display_bank(normalized.get("issuer_bank")),
        "from_bank": display_bank(normalized.get("from_bank")),
        "to_bank": display_bank(normalized.get("to_bank")),
        "confidence": float(normalized.get("confidence") or 0),
        "raw_text": clean_display(normalized.get("raw_text")),
    }
    if bank_needs_review(suggested.get("from_bank")) and is_known_bank(suggested.get("issuer_bank")):
        suggested["from_bank"] = suggested["issuer_bank"]
    previous = {field: display_bank(row[field]) for field in ["issuer_bank", "from_bank", "to_bank"]}
    applied: Dict[str, str] = {}
    for field in ["issuer_bank", "from_bank", "to_bank"]:
        value = suggested.get(field) or ""
        if value and bank_needs_review(previous.get(field)):
            applied[field] = value

    did_apply = bool(apply and applied)
    with connect(db_path) as conn:
        ensure_bank_review_log_table(conn)
        if did_apply:
            assignments = ", ".join([f"{field}=?" for field in applied])
            conn.execute(f"UPDATE slips SET {assignments} WHERE id=?", [*applied.values(), slip_id])
        conn.execute(
            """
            INSERT INTO bank_review_logs(slip_id, bot_key, chat_id, provider, model, applied, previous_json, suggested_json, created_at)
            VALUES (?,?,?,?,?,?,?,?,?)
            """,
            (
                slip_id,
                row["bot_key"],
                row["chat_id"],
                meta.get("provider", "openai"),
                meta.get("model", ""),
                1 if did_apply else 0,
                json.dumps(previous, ensure_ascii=False),
                json.dumps({"suggested": suggested, "applied": applied, "raw": data}, ensure_ascii=False),
                int(time.time()),
            ),
        )
        conn.commit()
    return {
        "ok": True,
        "id": slip_id,
        "provider": meta.get("provider", "openai"),
        "model": meta.get("model", ""),
        "previous": previous,
        "suggested": suggested,
        "applied": applied if did_apply else {},
    }


def dashboard_success_scope(chat_id: str = "", bot_key: str = "", scope: str = "open", flow_type: str = "all") -> Tuple[str, List[Any], str]:
    """Build the selected dashboard success/non-duplicate scope for bulk actions."""
    selected_bot = clean_display(bot_key)
    if chat_id:
        scoped_bot = selected_bot if selected_bot not in {"__all__", "all"} else ""
        where_clause, params, label = scope_where(chat_id, scope, success_only=True, bot_key=scoped_bot)
    else:
        scoped_bot = selected_bot if selected_bot not in {"__all__", "all"} else ""
        where_clause, params, label = global_scope_where(scope, success_only=True, bot_key=scoped_bot)
    where_clause, params = apply_flow_sql(where_clause, params, flow_type)
    return where_clause, params, label


def account_ledger_rows(
    db_path: Path,
    bot_key: str,
    chat_id: str,
    account_key: str,
    date_from: str = "",
    date_to: str = "",
    flow_type: str = "all",
    limit: int = 500,
) -> Dict[str, Any]:
    """Per-account ledger with running balance.

    Returns rows from `slips` matching the selected account, ordered by date+id,
    with running_balance calculated in Python (so duplicates don't double-count).
    """
    bot_key = clean_display(bot_key) or "default"
    chat_id = str(chat_id or "")
    account_key = clean_display(account_key)
    flow_type = normalize_flow_type(flow_type)
    if not account_key:
        return {"ok": False, "error": "missing account_key", "rows": []}

    with connect(db_path) as conn:
        ensure_company_account_table(conn)
        ensure_slip_reviews_table(conn)
        acc_row = conn.execute(
            "SELECT * FROM company_accounts WHERE bot_key=? AND chat_id=? AND account_key=?",
            (bot_key, chat_id, account_key),
        ).fetchone()
        if not acc_row:
            return {"ok": False, "error": "account not found", "rows": []}
        account = clean_company_fields(dict(acc_row))
        opening_balance = float(acc_row["opening_balance"] or 0) if "opening_balance" in acc_row.keys() else 0.0

        clauses = ["bot_key = ?", "chat_id = ?"]
        params: List[Any] = [bot_key, chat_id]
        bank = clean_display(acc_row["bank"])
        account_no = clean_display(acc_row["account_no"])
        if account_no:
            clauses.append("(COALESCE(to_account,'') LIKE ? OR COALESCE(to_bank,'') LIKE ? OR COALESCE(issuer_bank,'') LIKE ?)")
            params.extend([f"%{account_no}%", f"%{account_no}%", f"%{account_no}%"])
        elif bank:
            clauses.append("(COALESCE(to_bank,'') LIKE ? OR COALESCE(issuer_bank,'') LIKE ?)")
            params.extend([f"%{bank}%", f"%{bank}%"])
        if date_from:
            clauses.append("slip_date_text >= ?")
            params.append(date_from)
        if date_to:
            clauses.append("slip_date_text <= ?")
            params.append(date_to)
        clauses.append("status = 'success'")
        where = " AND ".join(clauses)
        where, params = apply_flow_sql(where, params, flow_type)

        rows = conn.execute(
            f"""
            SELECT s.*, sr.reviewed_at, sr.reviewed_by, sr.note AS review_note,
                   EXISTS(SELECT 1 FROM pending_actions pa
                          WHERE pa.action='slip.delete' AND pa.status IN ('pending','approved')
                          AND json_extract(pa.payload_json,'$.id')=s.id) AS pending_delete,
                   CASE WHEN s.is_duplicate=1 OR s.duplicate_of IS NOT NULL THEN 1 ELSE 0 END AS has_dup
            FROM slips s
            LEFT JOIN slip_reviews sr ON sr.slip_id = s.id
            WHERE {where}
            ORDER BY COALESCE(s.slip_date_iso, s.created_at_iso, ''), s.id
            LIMIT ?
            """,
            (*params, int(limit)),
        ).fetchall()

    running = opening_balance
    out_rows: List[Dict[str, Any]] = []
    total_in = 0.0
    total_out = 0.0
    issue_count = 0
    for r in rows:
        rec = dict(r)
        amount = float(rec.get("amount") or 0)
        flow = flow_type_for_title(rec.get("chat_title",""), rec.get("bot_key",""), rec.get("chat_id",""))
        is_dup = bool(rec.get("is_duplicate")) or bool(rec.get("duplicate_of")) or bool(rec.get("has_dup"))
        pending_del = bool(rec.get("pending_delete"))
        counted = not (is_dup or pending_del)
        delta = 0.0
        if counted and flow == "deposit":
            delta = amount
            total_in += amount
        elif counted and flow == "withdraw":
            delta = -amount
            total_out += amount
        running += delta
        flags: List[str] = []
        if is_dup:
            flags.append("duplicate")
        if pending_del:
            flags.append("pending_delete")
        if bank_needs_review(rec.get("from_bank")) and not is_dup:
            flags.append("bank_unclear")
        if float(rec.get("confidence") or 0) < 0.7 and float(rec.get("amount") or 0) > 0:
            flags.append("low_confidence")
        if rec.get("reviewed_at"):
            flags.append("reviewed")
        if any(f for f in flags if f not in ("reviewed",)):
            issue_count += 1
        rec["image_url"] = "/api/slip-image?" + urlencode({"id": rec.get("id", "")}) if rec.get("file_id") else ""
        rec["delta"] = delta
        rec["running_balance"] = running
        rec["counted"] = counted
        rec["flags"] = flags
        out_rows.append(clean_company_fields(rec))

    return {
        "ok": True,
        "account": account,
        "opening_balance": opening_balance,
        "rows": out_rows,
        "totals": {
            "in": total_in,
            "out": total_out,
            "net": total_in - total_out,
            "ending_balance": running,
            "row_count": len(out_rows),
            "issue_count": issue_count,
        },
    }


def mark_slip_reviewed(db_path: Path, slip_id: str, reviewed_by: str = "", note: str = "") -> Dict[str, Any]:
    slip_id = clean_display(slip_id)
    if not slip_id:
        return {"ok": False, "error": "missing slip id"}
    with connect(db_path) as conn:
        ensure_slip_reviews_table(conn)
        conn.execute(
            "INSERT OR REPLACE INTO slip_reviews(slip_id, reviewed_by, reviewed_at, note) VALUES (?,?,?,?)",
            (slip_id, clean_display(reviewed_by) or "owner", int(time.time()), clean_display(note)),
        )
        conn.commit()
    return {"ok": True, "id": slip_id}



def openai_bank_recheck_scope(
    db_path: Path,
    chat_id: str = "",
    bot_key: str = "",
    scope: str = "open",
    flow_type: str = "all",
    search: str = "",
    apply: bool = True,
) -> Dict[str, Any]:
    """Run bank recheck for every matching slip in the selected scope, not only visible cards."""
    where_clause, params, label = dashboard_success_scope(chat_id=chat_id, bot_key=bot_key, scope=scope, flow_type=flow_type)
    with connect(db_path) as conn:
        ids = source_bank_review_ids(conn, where_clause, params, search, exclude_ids=pending_delete_slip_ids(conn))
    ok_count = 0
    fail_count = 0
    errors: List[Dict[str, Any]] = []
    for slip_id in ids:
        result: Dict[str, Any] = {"ok": False, "error": "not attempted"}
        last_error = ""
        for attempt in range(1, 4):
            try:
                result = openai_bank_double_check_slip(db_path, slip_id, apply=apply)
                break
            except Exception as exc:
                last_error = str(exc)
                is_rate_limit = "429" in last_error or "rate limit" in last_error.lower()
                if is_rate_limit and attempt < 3:
                    time.sleep(1.0 * attempt)
                    continue
                result = {"ok": False, "error": last_error}
                break
        if result.get("ok"):
            ok_count += 1
        else:
            fail_count += 1
            if len(errors) < 10:
                errors.append({"id": slip_id, "error": result.get("error", "unknown error")})
    return {
        "ok": fail_count == 0,
        "scope": {"chat_id": chat_id, "bot_key": bot_key, "flow_type": normalize_flow_type(flow_type), "date_scope": label},
        "total_count": len(ids),
        "ok_count": ok_count,
        "fail_count": fail_count,
        "errors": errors,
    }



def reprocess_dashboard_slip(db_path: Path, slip_id: str, bot_key: str = "") -> Dict[str, Any]:
    """Re-run OCR for one unclear/error slip from the dashboard, preserving the same slip id."""
    slip_id = clean_display(slip_id)
    if not slip_id:
        return {"ok": False, "error": "missing slip id"}
    bot_filter = clean_display(bot_key)
    with connect(db_path) as conn:
        if bot_filter and bot_filter not in {"__all__", "all"}:
            row = conn.execute("SELECT * FROM slips WHERE id=? AND COALESCE(bot_key,'default')=?", (slip_id, bot_filter)).fetchone()
        else:
            row = conn.execute("SELECT * FROM slips WHERE id=?", (slip_id,)).fetchone()
    if not row:
        return {"ok": False, "error": "slip not found"}
    if not row["file_id"]:
        return {"ok": False, "error": "slip image not found"}

    previous = {"status": row["status"], "error": row["error"], "amount": float(row["amount"] or 0), "confidence": float(row["confidence"] or 0)}
    body, mime = fetch_slip_image(db_path, slip_id)
    suffix = mimetypes.guess_extension(mime or "image/jpeg") or ".jpg"
    tmp_name = ""
    try:
        with tempfile.NamedTemporaryFile(prefix="auditslip-reprocess-", suffix=suffix, delete=False) as tmp:
            tmp.write(body)
            tmp_name = tmp.name
        provider, data = ocr_extract(Path(tmp_name), mime)
    finally:
        if tmp_name:
            try:
                Path(tmp_name).unlink(missing_ok=True)
            except Exception:
                pass

    reason = unclear_reason(data)
    status = "unclear" if reason else "success"
    saved = dict(row)
    saved.update(data)
    saved["status"] = status
    saved["error"] = reason
    saved["ocr_provider"] = provider
    saved["raw_json"] = data.get("raw_json") or json.dumps(data, ensure_ascii=False)
    row_bot_key = clean_display(row["bot_key"]) or "default"
    bot = AuditslipBot(
        token=token_for_bot(row_bot_key),
        db_path=db_path,
        dry_run=True,
        bot_key=row_bot_key,
        company_name=clean_company_name(row["company_name"], row_bot_key),
    )
    bot.save_slip(saved)
    return {
        "ok": True,
        "id": slip_id,
        "bot_key": row_bot_key,
        "provider": provider,
        "previous": previous,
        "status": status,
        "error": reason,
        "amount": float(saved.get("amount") or 0),
        "confidence": float(saved.get("confidence") or 0),
    }



def company_overview_dicts(company_rows: List[sqlite3.Row], job_rows: List[sqlite3.Row]) -> List[Dict[str, Any]]:
    jobs_by_bot: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for r in job_rows:
        jobs_by_bot[str(r["bot_key"])][str(r["status"])] += int(r["count"] or 0)
    out = []
    for r in company_rows:
        item = clean_company_fields(dict(r))
        bot = str(item.get("bot_key") or "default")
        jobs = dict(jobs_by_bot.get(bot, {}))
        item["jobs"] = jobs
        item["queued"] = int(jobs.get("queued", 0))
        item["processing"] = int(jobs.get("processing", 0))
        item["failed"] = int(jobs.get("failed", 0))
        out.append(item)
    return sorted(out, key=dict_company_sort_key)

def dashboard_snapshot(db_path: Path = DB_PATH, chat_id: str = "", scope: str = "open", bot_key: str = "", slip_filter: str = "all", slip_search: str = "", flow_type: str = "all", account_search_mode: str = "", detail_level: str = "full") -> Dict[str, Any]:
    slip_filter_key = clean_display(slip_filter).lower() or "all"
    slip_search_key = clean_display(slip_search)
    flow_type_key = normalize_flow_type(flow_type)
    detail_level_key = clean_display(detail_level).lower() or "full"
    if detail_level_key not in {"full", "lite"}:
        detail_level_key = "full"
    include_detail_rows = detail_level_key == "full"
    account_search_mode_key = clean_display(account_search_mode).lower()
    if account_search_mode_key not in {"scoped", "cross", "both", ""}:
        account_search_mode_key = ""
    # Keep the company account index as the drill-down launcher. When the operator
    # clicks “ดูสลิปบัญชีนี้”, slip_search is reused for the image-heavy slip pane;
    # filtering the account table by that same query makes the list collapse to the
    # selected account only. Preserve the old combined filtering only for explicit
    # internal callers that pass account_search_mode="both".
    account_daily_search_key = slip_search_key if account_search_mode_key == "both" else ""
    with connect(db_path) as conn:
        chats = conn.execute(
            """
            SELECT COALESCE(NULLIF(bot_key,''),'default') AS bot_key,
                   COALESCE(MAX(NULLIF(company_name,'')), COALESCE(NULLIF(bot_key,''),'default')) AS company_name,
                   chat_id,
                   COALESCE(MAX(NULLIF(chat_title,'')), chat_id) AS chat_title,
                   COUNT(*) AS total_rows,
                   SUM(CASE WHEN status='success' AND COALESCE(is_duplicate,0)=0 AND settlement_id IS NULL THEN 1 ELSE 0 END) AS open_count,
                   COALESCE(SUM(CASE WHEN status='success' AND COALESCE(is_duplicate,0)=0 AND settlement_id IS NULL THEN amount ELSE 0 END),0) AS open_amount
            FROM slips
            GROUP BY COALESCE(NULLIF(bot_key,''),'default'), chat_id
            ORDER BY open_amount DESC, total_rows DESC
            LIMIT 100
            """
        ).fetchall()
        chat_dicts = rows_to_dicts(chats)
        for chat in chat_dicts:
            clean_company_fields(chat)
            chat["flow_type"] = flow_type_for_title(chat.get("chat_title"), chat.get("bot_key"), chat.get("chat_id"))
            chat["flow_label"] = flow_label(chat["flow_type"])
        chat_dicts.sort(key=lambda r: (*dict_company_sort_key(r), clean_display(r.get("chat_title")), clean_display(r.get("chat_id"))))
        requested_bot_key = clean_display(bot_key)
        overview_mode = (not chat_id) and (not requested_bot_key or requested_bot_key in {"__all__", "all"})
        if overview_mode:
            selected_bot_key = "__all__"
            selected_chat_id = ""
        elif chat_id:
            selected_chat_id = str(chat_id)
            match = next((c for c in chat_dicts if str(c.get("chat_id")) == selected_chat_id and (not requested_bot_key or str(c.get("bot_key")) == requested_bot_key)), None)
            selected_bot_key = requested_bot_key or (str(match.get("bot_key")) if match else "default")
        elif requested_bot_key:
            selected_bot_key = requested_bot_key
            selected_chat_id = ""
        else:
            selected_bot_key = "__all__"
            selected_chat_id = ""

        pending_delete_ids = pending_delete_slip_ids(conn)

        if selected_bot_key == "__all__":
            selected_where, selected_params, scope_label = global_scope_where(scope, success_only=True)
            selected_where, selected_params = apply_flow_sql(selected_where, selected_params, flow_type_key)
            selected_all_where, selected_all_params, _ = global_scope_where(scope, success_only=False)
            selected_all_where, selected_all_params = apply_flow_sql(selected_all_where, selected_all_params, flow_type_key)
            duplicate_where = selected_all_where + " AND status='success' AND COALESCE(is_duplicate,0)=1"
            duplicate_params = selected_all_params
            account_limits = load_bot_account_limits(conn)
            company_accounts = []
            open_where, open_params, _ = global_scope_where("open", success_only=True)
            open_where, open_params = apply_flow_sql(open_where, open_params, flow_type_key)
            all_where, all_params, _ = global_scope_where("all", success_only=True)
            all_where, all_params = apply_flow_sql(all_where, all_params, flow_type_key)
            open_totals = conn.execute(f"SELECT COUNT(*) AS count, COALESCE(SUM(amount),0) AS amount FROM slips WHERE {open_where}", open_params).fetchone()
            all_totals = conn.execute(f"SELECT COUNT(*) AS count, COALESCE(SUM(amount),0) AS amount FROM slips WHERE {all_where}", all_params).fetchone()
            selected_totals = conn.execute(f"SELECT COUNT(*) AS count, COALESCE(SUM(amount),0) AS amount FROM slips WHERE {selected_where}", selected_params).fetchone()
            duplicate_totals = conn.execute(f"SELECT COUNT(*) AS count, COALESCE(SUM(amount),0) AS amount FROM slips WHERE {duplicate_where}", duplicate_params).fetchone()
            status_where = "WHERE 1=1"
            status_params: List[Any] = []
            status_extra, status_extra_params = flow_sql_clause(flow_type_key)
            status_where += status_extra
            status_params.extend(status_extra_params)
            recent_where = f"WHERE {selected_all_where} AND COALESCE(status,'')!='deleted'"
            recent_params = list(selected_all_params)
            if slip_filter_key in {"duplicate", "duplicates", "dupe", "duplicateonly"}:
                recent_where += " AND status='success' AND COALESCE(is_duplicate,0)=1"
            elif slip_filter_key in {"nonduplicate", "normal", "success"}:
                recent_where += " AND status='success' AND COALESCE(is_duplicate,0)=0"
            elif slip_filter_key in {"issue", "issues", "error"}:
                recent_where += " AND status IN ('unclear','error')"
            recent_search_clause, recent_search_params = slip_search_clause("slips", slip_search_key)
            recent_where += recent_search_clause
            recent_params.extend(recent_search_params)
            issue_where = f"WHERE {selected_all_where} AND status IN ('unclear','error')"
            issue_params = list(selected_all_params)
            issue_extra, issue_extra_params = flow_sql_clause(flow_type_key)
            issue_where += issue_extra
            issue_params.extend(issue_extra_params)
            job_where = ""
            job_params = []
            job_recent_where = "WHERE status IN ('queued','processing','failed')"
            job_recent_params = []
            by_transferor, by_account_day, deposit_customer_slips, withdraw_limit_totals, deposit_customer_totals, withdraw_limit_usage = flow_split_panels(
                conn, selected_where, selected_params, flow_type_key, account_limits, slip_search_key
            )
            duplicate_pairs = duplicate_pair_rows(conn, "", scope, "", slip_search_key, flow_type=flow_type_key) if include_detail_rows else []
            source_bank_review = source_bank_review_rows(conn, selected_where, selected_params, slip_search_key, exclude_ids=pending_delete_ids) if include_detail_rows else []
            source_bank_review_total = source_bank_review_count(conn, selected_where, selected_params, slip_search_key, exclude_ids=pending_delete_ids)
            by_date = date_totals(conn, selected_where, selected_params) if include_detail_rows else []
            daily_flow_summary = daily_flow_totals(conn, selected_where, selected_params) if include_detail_rows else []
            by_from_bank = bank_totals(conn, selected_where, selected_params, "COALESCE(NULLIF(from_bank,''), '(ไม่ทราบธนาคารต้นทาง)')", "(ไม่ทราบธนาคารต้นทาง)") if include_detail_rows else []
            by_to_bank = bank_totals(conn, selected_where, selected_params, "COALESCE(NULLIF(to_bank,''), '(ไม่ทราบธนาคารปลายทาง)')", "(ไม่ทราบธนาคารปลายทาง)") if include_detail_rows else []
            by_sender = grouped_totals(conn, selected_where, selected_params, "COALESCE(NULLIF(sender_name,''), NULLIF(username,''), '(ไม่ทราบผู้ส่งรูป)')") if include_detail_rows else []
            company_account_daily = company_account_daily_totals(conn, selected_where, selected_params, account_daily_search_key)
            account_slip_search = account_slip_search_rows(conn, selected_where, selected_params, slip_search_key) if include_detail_rows else empty_account_slip_search(slip_search_key)
            account_cross_company = cross_company_account_usage(conn, scope, flow_type_key, selected_bot_key, slip_search_key) if include_detail_rows else []
        elif selected_chat_id:
            selected_where, selected_params, scope_label = scope_where(selected_chat_id, scope, success_only=True, bot_key=selected_bot_key)
            selected_where, selected_params = apply_flow_sql(selected_where, selected_params, flow_type_key)
            selected_all_where, selected_all_params, _ = scope_where(selected_chat_id, scope, success_only=False, bot_key=selected_bot_key)
            selected_all_where, selected_all_params = apply_flow_sql(selected_all_where, selected_all_params, flow_type_key)
            duplicate_where = selected_all_where + " AND status='success' AND COALESCE(is_duplicate,0)=1"
            duplicate_params = selected_all_params
            account_limits = load_account_limits(conn, selected_chat_id, selected_bot_key)
            company_accounts = load_company_accounts(conn, selected_bot_key, selected_chat_id)
            open_totals = scoped_counts(conn, selected_chat_id, "open", bot_key=selected_bot_key)
            all_totals = scoped_counts(conn, selected_chat_id, "all", bot_key=selected_bot_key)
            selected_totals = conn.execute(
                f"SELECT COUNT(*) AS count, COALESCE(SUM(amount),0) AS amount FROM slips WHERE {selected_where}",
                selected_params,
            ).fetchone()
            duplicate_totals = conn.execute(
                f"SELECT COUNT(*) AS count, COALESCE(SUM(amount),0) AS amount FROM slips WHERE {duplicate_where}",
                duplicate_params,
            ).fetchone()
            status_where = "WHERE chat_id=? AND COALESCE(bot_key,'default')=?"
            status_params = [selected_chat_id, selected_bot_key]
            status_extra, status_extra_params = flow_sql_clause(flow_type_key)
            status_where += status_extra
            status_params.extend(status_extra_params)
            recent_where = f"WHERE {selected_all_where} AND COALESCE(status,'')!='deleted'"
            recent_params = list(selected_all_params)
            if slip_filter_key in {"duplicate", "duplicates", "dupe", "duplicateonly"}:
                recent_where += " AND status='success' AND COALESCE(is_duplicate,0)=1"
            elif slip_filter_key in {"nonduplicate", "normal", "success"}:
                recent_where += " AND status='success' AND COALESCE(is_duplicate,0)=0"
            elif slip_filter_key in {"issue", "issues", "error"}:
                recent_where += " AND status IN ('unclear','error')"
            recent_search_clause, recent_search_params = slip_search_clause("slips", slip_search_key)
            recent_where += recent_search_clause
            recent_params.extend(recent_search_params)
            issue_where = f"WHERE {selected_all_where} AND status IN ('unclear','error')"
            issue_params = list(selected_all_params)
            issue_extra, issue_extra_params = flow_sql_clause(flow_type_key)
            issue_where += issue_extra
            issue_params.extend(issue_extra_params)
            job_where = "WHERE chat_id=? AND COALESCE(bot_key,'default')=?"
            job_params = [selected_chat_id, selected_bot_key]
            job_recent_where = "WHERE chat_id=? AND COALESCE(bot_key,'default')=? AND status IN ('queued','processing','failed')"
            job_recent_params = [selected_chat_id, selected_bot_key]
            by_transferor, by_account_day, deposit_customer_slips, withdraw_limit_totals, deposit_customer_totals, withdraw_limit_usage = flow_split_panels(
                conn, selected_where, selected_params, flow_type_key, account_limits, slip_search_key
            )
            duplicate_pairs = duplicate_pair_rows(conn, selected_chat_id, scope, selected_bot_key, slip_search_key, flow_type=flow_type_key) if include_detail_rows else []
            source_bank_review = source_bank_review_rows(conn, selected_where, selected_params, slip_search_key, exclude_ids=pending_delete_ids) if include_detail_rows else []
            source_bank_review_total = source_bank_review_count(conn, selected_where, selected_params, slip_search_key, exclude_ids=pending_delete_ids)
            by_date = date_totals(conn, selected_where, selected_params) if include_detail_rows else []
            daily_flow_summary = daily_flow_totals(conn, selected_where, selected_params) if include_detail_rows else []
            by_from_bank = bank_totals(conn, selected_where, selected_params, "COALESCE(NULLIF(from_bank,''), '(ไม่ทราบธนาคารต้นทาง)')", "(ไม่ทราบธนาคารต้นทาง)") if include_detail_rows else []
            by_to_bank = bank_totals(conn, selected_where, selected_params, "COALESCE(NULLIF(to_bank,''), '(ไม่ทราบธนาคารปลายทาง)')", "(ไม่ทราบธนาคารปลายทาง)") if include_detail_rows else []
            by_sender = grouped_totals(conn, selected_where, selected_params, "COALESCE(NULLIF(sender_name,''), NULLIF(username,''), '(ไม่ทราบผู้ส่งรูป)')") if include_detail_rows else []
            company_account_daily = company_account_daily_totals(conn, selected_where, selected_params, account_daily_search_key)
            account_slip_search = account_slip_search_rows(conn, selected_where, selected_params, slip_search_key) if include_detail_rows else empty_account_slip_search(slip_search_key)
            account_cross_company = cross_company_account_usage(conn, scope, flow_type_key, selected_bot_key, slip_search_key) if include_detail_rows else []
        else:
            selected_bot_key = selected_bot_key or requested_bot_key
            if selected_bot_key:
                selected_where, selected_params, scope_label = global_scope_where(scope, success_only=True, bot_key=selected_bot_key)
                selected_where, selected_params = apply_flow_sql(selected_where, selected_params, flow_type_key)
                selected_all_where, selected_all_params, _ = global_scope_where(scope, success_only=False, bot_key=selected_bot_key)
                selected_all_where, selected_all_params = apply_flow_sql(selected_all_where, selected_all_params, flow_type_key)
                duplicate_where = selected_all_where + " AND status='success' AND COALESCE(is_duplicate,0)=1"
                duplicate_params = selected_all_params
                open_where, open_params, _ = global_scope_where("open", success_only=True, bot_key=selected_bot_key)
                open_where, open_params = apply_flow_sql(open_where, open_params, flow_type_key)
                all_where, all_params, _ = global_scope_where("all", success_only=True, bot_key=selected_bot_key)
                all_where, all_params = apply_flow_sql(all_where, all_params, flow_type_key)
                open_totals = conn.execute(f"SELECT COUNT(*) AS count, COALESCE(SUM(amount),0) AS amount FROM slips WHERE {open_where}", open_params).fetchone()
                all_totals = conn.execute(f"SELECT COUNT(*) AS count, COALESCE(SUM(amount),0) AS amount FROM slips WHERE {all_where}", all_params).fetchone()
                selected_totals = conn.execute(f"SELECT COUNT(*) AS count, COALESCE(SUM(amount),0) AS amount FROM slips WHERE {selected_where}", selected_params).fetchone()
                duplicate_totals = conn.execute(f"SELECT COUNT(*) AS count, COALESCE(SUM(amount),0) AS amount FROM slips WHERE {duplicate_where}", duplicate_params).fetchone()
                account_limits = load_account_limits(conn, "", selected_bot_key)
                company_accounts = load_company_accounts(conn, selected_bot_key, "")
                status_where = "WHERE COALESCE(bot_key,'default')=?"
                status_params = [selected_bot_key]
                status_extra, status_extra_params = flow_sql_clause(flow_type_key)
                status_where += status_extra
                status_params.extend(status_extra_params)
                recent_where = f"WHERE {selected_all_where} AND COALESCE(status,'')!='deleted'"
                recent_params = list(selected_all_params)
                if slip_filter_key in {"duplicate", "duplicates", "dupe", "duplicateonly"}:
                    recent_where += " AND status='success' AND COALESCE(is_duplicate,0)=1"
                elif slip_filter_key in {"nonduplicate", "normal", "success"}:
                    recent_where += " AND status='success' AND COALESCE(is_duplicate,0)=0"
                elif slip_filter_key in {"issue", "issues", "error"}:
                    recent_where += " AND status IN ('unclear','error')"
                recent_search_clause, recent_search_params = slip_search_clause("slips", slip_search_key)
                recent_where += recent_search_clause
                recent_params.extend(recent_search_params)
                issue_where = f"WHERE {selected_all_where} AND status IN ('unclear','error')"
                issue_params = list(selected_all_params)
                issue_extra, issue_extra_params = flow_sql_clause(flow_type_key)
                issue_where += issue_extra
                issue_params.extend(issue_extra_params)
                job_where = "WHERE COALESCE(bot_key,'default')=?"
                job_params = [selected_bot_key]
                job_recent_where = "WHERE COALESCE(bot_key,'default')=? AND status IN ('queued','processing','failed')"
                job_recent_params = [selected_bot_key]
                by_transferor, by_account_day, deposit_customer_slips, withdraw_limit_totals, deposit_customer_totals, withdraw_limit_usage = flow_split_panels(
                    conn, selected_where, selected_params, flow_type_key, account_limits, slip_search_key
                )
                duplicate_pairs = duplicate_pair_rows(conn, "", scope, selected_bot_key, slip_search_key, flow_type=flow_type_key) if include_detail_rows else []
                source_bank_review = source_bank_review_rows(conn, selected_where, selected_params, slip_search_key, exclude_ids=pending_delete_ids) if include_detail_rows else []
                source_bank_review_total = source_bank_review_count(conn, selected_where, selected_params, slip_search_key, exclude_ids=pending_delete_ids)
                by_date = date_totals(conn, selected_where, selected_params) if include_detail_rows else []
                daily_flow_summary = daily_flow_totals(conn, selected_where, selected_params) if include_detail_rows else []
                by_from_bank = bank_totals(conn, selected_where, selected_params, "COALESCE(NULLIF(from_bank,''), '(ไม่ทราบธนาคารต้นทาง)')", "(ไม่ทราบธนาคารต้นทาง)") if include_detail_rows else []
                by_to_bank = bank_totals(conn, selected_where, selected_params, "COALESCE(NULLIF(to_bank,''), '(ไม่ทราบธนาคารปลายทาง)')", "(ไม่ทราบธนาคารปลายทาง)") if include_detail_rows else []
                by_sender = grouped_totals(conn, selected_where, selected_params, "COALESCE(NULLIF(sender_name,''), NULLIF(username,''), '(ไม่ทราบผู้ส่งรูป)')") if include_detail_rows else []
                company_account_daily = company_account_daily_totals(conn, selected_where, selected_params, account_daily_search_key)
                account_slip_search = account_slip_search_rows(conn, selected_where, selected_params, slip_search_key) if include_detail_rows else empty_account_slip_search(slip_search_key)
                account_cross_company = cross_company_account_usage(conn, scope, flow_type_key, selected_bot_key, slip_search_key) if include_detail_rows else []
            else:
                scope_label = "ไม่มีข้อมูล"
                open_totals = {"count": 0, "amount": 0}
                all_totals = {"count": 0, "amount": 0}
                selected_totals = {"count": 0, "amount": 0}
                duplicate_totals = {"count": 0, "amount": 0}
                status_where = ""
                status_params = []
                recent_where = ""
                recent_params = []
                issue_where = "WHERE status IN ('unclear','error')"
                issue_params = []
                job_where = ""
                job_params = []
                job_recent_where = "WHERE status IN ('queued','processing','failed')"
                job_recent_params = []
                company_accounts = []
                by_transferor = by_account_day = by_date = by_from_bank = by_to_bank = by_sender = []
                daily_flow_summary = []
                company_account_daily = []
                account_slip_search = empty_account_slip_search(slip_search_key)
                account_cross_company = []
                deposit_customer_slips = []
                withdraw_limit_totals = {"count": 0, "amount": 0.0}
                deposit_customer_totals = {"count": 0, "amount": 0.0}
                withdraw_limit_usage = []
                duplicate_pairs = []
                source_bank_review = []
                source_bank_review_total = 0

        if account_search_mode_key == "scoped" or not include_detail_rows:
            cross_company_account_slip_search = empty_account_slip_search(slip_search_key)
            cross_company_account_slip_search.update({"company_count": 0, "companies": [], "is_cross_company": False})
        else:
            cross_company_account_slip_search = cross_company_account_slip_search_rows(conn, scope, flow_type_key, slip_search_key)
        if account_search_mode_key == "cross":
            account_slip_search = empty_account_slip_search(slip_search_key)

        slip_status_rows = conn.execute(f"SELECT status, COUNT(*) AS count FROM slips {status_where} GROUP BY status", status_params).fetchall()
        job_status_rows = conn.execute(f"SELECT status, COUNT(*) AS count FROM ocr_jobs {job_where} GROUP BY status", job_params).fetchall()
        if include_detail_rows:
            recent = conn.execute(
                f"""
                SELECT id, COALESCE(NULLIF(bot_key,''),'default') AS bot_key, company_name, chat_id, chat_title, file_id, message_id,
                       status, error, slip_date_display, slip_date_iso, slip_time,
                       TRIM(COALESCE(NULLIF(slip_date_display,''), NULLIF(slip_date_iso,''), '') || ' ' || COALESCE(NULLIF(slip_time,''), '')) AS slip_date_text,
                       transferor_name, recipient_name, sender_name, username, issuer_bank, from_bank, to_bank,
                       amount, confidence, created_at_iso, settlement_id, is_duplicate, duplicate_of
                FROM slips
                {recent_where}
                -- Operator "recent slips" should follow the slip/message chronology, not OCR completion time.
                -- A delayed OCR batch can otherwise hide later withdrawal messages behind older slips that finished late.
                ORDER BY COALESCE(NULLIF(slip_date_iso,''), '') DESC,
                         COALESCE(NULLIF(slip_time,''), '') DESC,
                         CAST(COALESCE(message_id,0) AS INTEGER) DESC,
                         created_at DESC
                LIMIT 40
                """,
                recent_params,
            ).fetchall()
            issues = conn.execute(
                f"""
                SELECT id, COALESCE(NULLIF(bot_key,''),'default') AS bot_key, company_name, chat_id, chat_title, file_id,
                       status, error, message_id, created_at_iso,
                       TRIM(COALESCE(NULLIF(slip_date_display,''), NULLIF(slip_date_iso,''), '') || ' ' || COALESCE(NULLIF(slip_time,''), '')) AS slip_date_text,
                       transferor_name, recipient_name, sender_name, username, amount, confidence, raw_text
                FROM slips
                {issue_where}
                ORDER BY created_at DESC
                LIMIT 40
                """,
                issue_params,
            ).fetchall()
        else:
            recent = []
            issues = []
        issue_count_row = conn.execute(f"SELECT COUNT(*) AS count FROM slips {issue_where}", issue_params).fetchone()
        issue_queue_count_total = int(issue_count_row["count"] or 0) if issue_count_row else 0
        provider_usage = conn.execute(
            f"""
            SELECT COALESCE(NULLIF(ocr_provider,''), '(pending)') AS provider,
                   COALESCE(NULLIF(ocr_model,''), '') AS model,
                   status,
                   COUNT(*) AS count,
                   COALESCE(SUM(ocr_input_tokens),0) AS input_tokens,
                   COALESCE(SUM(ocr_output_tokens),0) AS output_tokens,
                   COALESCE(SUM(ocr_thought_tokens),0) AS thought_tokens,
                   COALESCE(SUM(ocr_total_tokens),0) AS total_tokens,
                   COALESCE(SUM(ocr_cost_usd),0) AS cost_usd
            FROM slips
            {status_where}
            GROUP BY provider, model, status
            ORDER BY provider, model, status
            """,
            status_params,
        ).fetchall() if include_detail_rows else []
        jobs_recent = conn.execute(
            f"""
            SELECT job_id, slip_id, COALESCE(NULLIF(bot_key,''),'default') AS bot_key, company_name, chat_id, message_id, status, attempts, max_attempts, locked_by, error, created_at, updated_at
            FROM ocr_jobs
            {job_recent_where}
            ORDER BY created_at ASC
            LIMIT 50
            """,
            job_recent_params,
        ).fetchall() if include_detail_rows else []

        company_where = "WHERE 1=1"
        company_where_params: List[Any] = []
        company_extra, company_extra_params = flow_sql_clause(flow_type_key)
        company_where += company_extra
        company_where_params.extend(company_extra_params)

        # The operator/company overview cards are labelled by the selected scope
        # (today / selected day / range / open).  Keep the historical rows only as
        # the company source list, but calculate every visible amount/count inside
        # the current dashboard scope.  Otherwise a new day still shows yesterday's
        # open-period totals and looks like the daily counter did not reset.
        company_scope_where, company_scope_params, _ = global_scope_where(scope, success_only=True)
        company_scope_where, company_scope_params = apply_flow_sql(company_scope_where, company_scope_params, flow_type_key)
        deposit_scope_where, deposit_scope_params, _ = global_scope_where(scope, success_only=True)
        deposit_scope_where, deposit_scope_params = apply_flow_sql(deposit_scope_where, deposit_scope_params, "deposit")
        withdraw_scope_where, withdraw_scope_params, _ = global_scope_where(scope, success_only=True)
        withdraw_scope_where, withdraw_scope_params = apply_flow_sql(withdraw_scope_where, withdraw_scope_params, "withdraw")
        duplicate_scope_where, duplicate_scope_params, _ = global_scope_where(scope, success_only=False)
        duplicate_scope_where += " AND status='success' AND COALESCE(is_duplicate,0)=1"
        duplicate_scope_where, duplicate_scope_params = apply_flow_sql(duplicate_scope_where, duplicate_scope_params, flow_type_key)
        issue_scope_where, issue_scope_params, _ = global_scope_where(scope, success_only=False)
        issue_scope_where += " AND status IN ('unclear','error')"
        issue_scope_where, issue_scope_params = apply_flow_sql(issue_scope_where, issue_scope_params, flow_type_key)
        company_params: List[Any] = [
            *company_scope_params,
            *company_scope_params,
            *deposit_scope_params,
            *deposit_scope_params,
            *withdraw_scope_params,
            *withdraw_scope_params,
            *duplicate_scope_params,
            *duplicate_scope_params,
            *issue_scope_params,
            *company_scope_params,
            *company_where_params,
        ]
        company_rows = conn.execute(
            f"""
            SELECT COALESCE(NULLIF(bot_key,''),'default') AS bot_key,
                   COALESCE(MAX(NULLIF(company_name,'')), COALESCE(NULLIF(bot_key,''),'default')) AS company_name,
                   COUNT(*) AS total_rows,
                   SUM(CASE WHEN {company_scope_where} THEN 1 ELSE 0 END) AS open_count,
                   COALESCE(SUM(CASE WHEN {company_scope_where} THEN amount ELSE 0 END),0) AS open_amount,
                   SUM(CASE WHEN {deposit_scope_where} THEN 1 ELSE 0 END) AS deposit_open_count,
                   COALESCE(SUM(CASE WHEN {deposit_scope_where} THEN amount ELSE 0 END),0) AS deposit_open_amount,
                   SUM(CASE WHEN {withdraw_scope_where} THEN 1 ELSE 0 END) AS withdraw_open_count,
                   COALESCE(SUM(CASE WHEN {withdraw_scope_where} THEN amount ELSE 0 END),0) AS withdraw_open_amount,
                   SUM(CASE WHEN {duplicate_scope_where} THEN 1 ELSE 0 END) AS duplicate_count,
                   COALESCE(SUM(CASE WHEN {duplicate_scope_where} THEN amount ELSE 0 END),0) AS duplicate_amount,
                   SUM(CASE WHEN {issue_scope_where} THEN 1 ELSE 0 END) AS issue_count,
                   MAX(CASE WHEN {company_scope_where} THEN created_at_iso ELSE NULL END) AS latest_slip_at
            FROM slips
            {company_where}
            GROUP BY COALESCE(NULLIF(bot_key,''),'default')
            ORDER BY open_amount DESC, total_rows DESC
            """,
            company_params,
        ).fetchall()
        company_job_rows = conn.execute(
            """
            SELECT COALESCE(NULLIF(bot_key,''),'default') AS bot_key, status, COUNT(*) AS count
            FROM ocr_jobs
            GROUP BY COALESCE(NULLIF(bot_key,''),'default'), status
            """
        ).fetchall()

    slip_statuses = defaultdict(int)
    for r in slip_status_rows:
        slip_statuses[str(r["status"])] += int(r["count"] or 0)
    jobs = defaultdict(int)
    for r in job_status_rows:
        jobs[str(r["status"])] += int(r["count"] or 0)
    recent_dicts = rows_to_dicts(recent)
    for row in recent_dicts:
        clean_company_fields(row)
        for field in ["issuer_bank", "from_bank", "to_bank"]:
            row[field] = display_bank(row.get(field))
        row["image_url"] = "/api/slip-image?" + urlencode({"id": row.get("id", "")}) if row.get("file_id") else ""
    issue_dicts = rows_to_dicts(issues)
    for row in issue_dicts:
        clean_company_fields(row)
        row["image_url"] = "/api/slip-image?" + urlencode({"id": row.get("id", "")}) if row.get("file_id") else ""
        row["raw_text"] = clean_display(row.get("raw_text"))[:240]

    limit_check_enabled = limit_check_enabled_for_flow(flow_type_key)
    if not limit_check_enabled:
        by_transferor = []
        by_account_day = []
        withdraw_limit_usage = []

    withdraw_limit_totals_extra = withdraw_limit_usage_totals(withdraw_limit_usage)
    company_menu_rows = company_overview_dicts(company_rows, company_job_rows)
    selected_company_summary = company_menu_rows
    if selected_bot_key and selected_bot_key not in {"__all__", "all"}:
        selected_company_summary = [row for row in company_menu_rows if str(row.get("bot_key")) == str(selected_bot_key)]

    duplicate_count_total = int(duplicate_totals["count"] or 0)
    issue_count_total = int(issue_queue_count_total or 0)
    queue_attention_count = int(jobs.get("queued", 0) or 0) + int(jobs.get("processing", 0) or 0) + int(jobs.get("failed", 0) or 0)
    over_limit_count = sum(1 for row in (by_account_day or []) if row.get("over_limit")) if limit_check_enabled else 0
    bank_ledger_summary = bank_ledger_snapshot(db_path, bot_key=selected_bot_key, scope=scope or "open", flow_type=flow_type_key)
    exception_summary = {
        "over_limit_count": int(over_limit_count),
        "issue_count": int(issue_count_total),
        "bank_review_count": int(source_bank_review_total or 0),
        "duplicate_count": int(duplicate_count_total),
        "queue_attention_count": int(queue_attention_count),
        "ledger_unmatched_count": int((bank_ledger_summary.get("unmatched_ledger") or {}).get("count") or 0),
    }
    exception_summary["total_count"] = sum(int(v or 0) for v in exception_summary.values())
    twallet_summary = fetch_twallet_summary()

    snapshot = {
        "app": APP_NAME,
        "generated_at": int(time.time()),
        "detail_level": detail_level_key,
        "selected_chat_id": selected_chat_id,
        "selected_bot_key": selected_bot_key,
        "scope": scope or "open",
        "slip_filter": slip_filter_key,
        "slip_search": slip_search_key,
        "flow_type": flow_type_key,
        "flow_label": flow_label(flow_type_key),
        "limit_check_enabled": limit_check_enabled,
        "scope_label": scope_label,
        "telegram_bots": public_telegram_bots(),
        "company_accounts": company_accounts,
        "totals": {
            "open_success_count": int(open_totals["count"] or 0),
            "open_success_amount": float(open_totals["amount"] or 0),
            "all_success_count": int(all_totals["count"] or 0),
            "all_success_amount": float(all_totals["amount"] or 0),
            "selected_success_count": int(selected_totals["count"] or 0),
            "selected_success_amount": float(selected_totals["amount"] or 0),
            "selected_duplicate_count": int(duplicate_totals["count"] or 0),
            "selected_duplicate_amount": float(duplicate_totals["amount"] or 0),
            "withdraw_limit_count": int(withdraw_limit_totals.get("count", 0) or 0),
            "withdraw_limit_amount": float(withdraw_limit_totals.get("amount", 0) or 0),
            **withdraw_limit_totals_extra,
            "deposit_customer_count": int(deposit_customer_totals.get("count", 0) or 0),
            "deposit_customer_amount": float(deposit_customer_totals.get("amount", 0) or 0),
            "source_bank_review_count": int(source_bank_review_total or 0),
        },
        "exception_summary": exception_summary,
        "bank_ledger_summary": bank_ledger_summary,
        "twallet_summary": twallet_summary,
        "slip_statuses": dict(slip_statuses),
        "jobs": dict(jobs),
        "chats": chat_dicts,
        "company_summary": selected_company_summary,
        "company_menu": company_menu_rows,
        "company_account_daily": aggregate_dicts(company_account_daily),
        "withdraw_limit_usage": aggregate_dicts(withdraw_limit_usage),
        "account_slip_search": account_slip_search,
        "cross_company_account_slip_search": cross_company_account_slip_search,
        "account_cross_company": aggregate_dicts(account_cross_company),
        "recent": recent_dicts,
        "duplicate_pairs": duplicate_pairs,
        "source_bank_review": source_bank_review,
        "deposit_customer_slips": deposit_customer_slips,
        "issues": issue_dicts,
        "provider_usage": rows_to_dicts(provider_usage),
        "jobs_recent": rows_to_dicts(jobs_recent),
        "by_transferor": aggregate_dicts(by_transferor),
        "by_account_day": aggregate_dicts(by_account_day),
        "by_date": aggregate_dicts(by_date),
        "daily_flow_summary": aggregate_dicts(daily_flow_summary),
        "by_from_bank": aggregate_dicts(by_from_bank),
        "by_to_bank": aggregate_dicts(by_to_bank),
        "by_sender": rows_to_dicts(by_sender),
    }
    if detail_level_key == "lite":
        return lite_dashboard_snapshot(snapshot)
    return snapshot


def safe_backend_excel_path(path_text: str = "") -> Path:
    base = Path(os.environ.get("AUDITSLIP_BACKEND_IMPORT_DIR", "/root/projects/auditslip/imports/backend"))
    base.mkdir(parents=True, exist_ok=True)
    base_resolved = base.resolve()
    if path_text:
        path = Path(path_text).expanduser()
        if not path.is_absolute():
            path = base / path
        resolved = path.resolve()
        try:
            resolved.relative_to(base_resolved)
        except ValueError:
            raise PermissionError("excel path outside backend import dir")
        if resolved.suffix.lower() not in {".xlsx", ".xlsm"}:
            raise PermissionError("excel path must be .xlsx or .xlsm")
        return resolved
    files = sorted([p for p in base.glob("*.xlsx") if p.is_file()], key=lambda p: p.stat().st_mtime, reverse=True)
    if not files:
        raise FileNotFoundError(f"ไม่พบไฟล์ .xlsx ใน {base}")
    return files[0]


def safe_statement_file_path(path_text: str = "") -> Path:
    base = Path(os.environ.get("AUDITSLIP_BACKEND_IMPORT_DIR", "/root/projects/auditslip/imports/backend"))
    base.mkdir(parents=True, exist_ok=True)
    base_resolved = base.resolve()
    allowed = {".xlsx", ".xlsm", ".csv"}
    if path_text:
        path = Path(path_text).expanduser()
        if not path.is_absolute():
            path = base / path
        resolved = path.resolve()
        try:
            resolved.relative_to(base_resolved)
        except ValueError:
            raise PermissionError("statement path outside backend import dir")
        if resolved.suffix.lower() not in allowed:
            raise PermissionError("statement path must be .xlsx, .xlsm, or .csv")
        return resolved
    files = sorted(
        [p for suffix in allowed for p in base.glob(f"*{suffix}") if p.is_file()],
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not files:
        raise FileNotFoundError(f"ไม่พบไฟล์ statement .xlsx/.xlsm/.csv ใน {base}")
    return files[0]


DASHBOARD_EXPORT_HEADERS = [
    "company_name",
    "chat_title",
    "username",
    "sender_name",
    "message_id",
    "caption",
    "error",
    "slip_date_display",
    "slip_date_iso",
    "slip_time",
    "issuer_bank",
    "seq",
    "location",
    "transaction_type",
    "transferor_name",
    "recipient_name",
    "from_bank",
    "from_account",
    "to_bank",
    "to_account",
    "account_name",
    "amount",
    "fee",
    "reference_no",
    "aid",
    "label",
    "raw_text",
    "confidence",
    "is_duplicate",
    "duplicate_of",
    "settlement_id",
    "created_at_iso",
]

DASHBOARD_DUPLICATE_HEADERS = [
    "duplicate_message_id",
    "matched_message_id",
    "slip_date_display",
    "slip_time",
    "transferor_name",
    "from_bank",
    "from_account",
    "to_bank",
    "to_account",
    "amount",
    "reference_no",
    "matched_reference_no",
    "sender_name",
    "matched_sender_name",
    "duplicate_created_at_iso",
    "matched_created_at_iso",
    "created_at_iso",
]

CROSS_COMPANY_ACCOUNT_EXPORT_HEADERS = [
    "company_name",
    "chat_title",
    "message_id",
    "slip_date_display",
    "slip_date_iso",
    "slip_time",
    "flow_label",
    "matched_label",
    "sender_name",
    "transferor_name",
    "recipient_name",
    "from_bank",
    "from_account",
    "to_bank",
    "to_account",
    "amount",
    "fee",
    "reference_no",
    "slip_image_url",
    "created_at_iso",
]

BANK_EXPORT_FIELDS = {"issuer_bank", "from_bank", "to_bank"}


def normalize_export_date(value: str) -> str:
    display, iso = normalize_date_parts(value)
    return iso or clean_display(value)


def export_date_bounds(start_date: str = "", end_date: str = "") -> Tuple[str, str]:
    start = normalize_export_date(start_date)
    end = normalize_export_date(end_date)
    if start and end and start > end:
        start, end = end, start
    return start, end


def export_row_date(row: sqlite3.Row) -> str:
    row_keys = set(row.keys())
    iso_value = row["slip_date_iso"] if "slip_date_iso" in row_keys else ""
    display_value = row["slip_date_display"] if "slip_date_display" in row_keys else ""
    return normalize_match_date(iso_value or display_value)


def export_row_in_date_range(row: sqlite3.Row, start_date: str = "", end_date: str = "") -> bool:
    start, end = export_date_bounds(start_date, end_date)
    if not start and not end:
        return True
    date_key = export_row_date(row)
    if not re.match(r"^\d{4}-\d{2}-\d{2}$", date_key):
        return False
    if start and date_key < start:
        return False
    if end and date_key > end:
        return False
    return True


def export_date_range_clause(start_date: str = "", end_date: str = "") -> Tuple[str, List[Any]]:
    # Kept for older tests/callers, but export_dashboard_excel filters date ranges
    # in Python so visible dates without slip_date_iso (for example 22/05/26 or Thai
    # Buddhist-era labels) are not silently excluded by lexical SQL comparisons.
    start, end = export_date_bounds(start_date, end_date)
    clause = ""
    params: List[Any] = []
    date_expr = "COALESCE(NULLIF(slip_date_iso,''), slip_date_display)"
    if start and end:
        clause += f" AND {date_expr} BETWEEN ? AND ?"
        params.extend([start, end])
    elif start:
        clause += f" AND {date_expr} >= ?"
        params.append(start)
    elif end:
        clause += f" AND {date_expr} <= ?"
        params.append(end)
    return clause, params


def export_where_clause(bot_key: str = "", chat_id: str = "", flow_type: str = "all", scope: str = "all", start_date: str = "", end_date: str = "") -> Tuple[str, List[Any]]:
    requested_bot = clean_display(bot_key)
    effective_scope = scope if not (start_date or end_date) else "all"
    if chat_id:
        clause, params, _ = scope_where(str(chat_id), effective_scope, success_only=False, bot_key=requested_bot)
    else:
        clause, params, _ = global_scope_where(effective_scope, success_only=False, bot_key=requested_bot)
    clause, params = apply_flow_sql(clause, params, flow_type)
    clause += " AND COALESCE(status,'')!='deleted'"
    if not (start_date or end_date):
        date_clause, date_params = export_date_range_clause(start_date, end_date)
        clause += date_clause
        params = [*params, *date_params]
    return clause, params


def safe_export_name(value: Any) -> str:
    text = clean_display(value) or "export"
    return re.sub(r"[^A-Za-z0-9ก-๙_.-]+", "-", text).strip("-._") or "export"


def export_cell_value(row: sqlite3.Row, header: str) -> Any:
    if header not in row.keys():
        return ""
    value = row[header]
    return display_bank(value) if header in BANK_EXPORT_FIELDS else value


def export_row_get(row: Any, key: str, default: Any = "") -> Any:
    if isinstance(row, dict):
        return row.get(key, default)
    try:
        if key in row.keys():
            return row[key]
    except Exception:
        pass
    return default


def export_row_flow_type(row: Any) -> str:
    explicit = normalize_flow_type(export_row_get(row, "flow_type", ""))
    if explicit in {"deposit", "withdraw", "other"}:
        return explicit
    return flow_type_for_title(
        export_row_get(row, "chat_title", ""),
        export_row_get(row, "bot_key", ""),
        export_row_get(row, "chat_id", ""),
    )


def export_row_is_duplicate(row: Any) -> bool:
    try:
        return bool(int(export_row_get(row, "is_duplicate", 0) or 0))
    except Exception:
        return False


def counted_export_rows(rows: List[sqlite3.Row]) -> List[sqlite3.Row]:
    return [r for r in rows if str(r["status"] or "") == "success" and not int(r["is_duplicate"] or 0)]


def dashboard_slip_export_rows(rows: List[sqlite3.Row], flow_type: str = "all") -> List[List[Any]]:
    target = normalize_flow_type(flow_type)
    out: List[List[Any]] = []
    for row in rows:
        if export_row_is_duplicate(row):
            continue
        if target != "all" and export_row_flow_type(row) != target:
            continue
        out.append([export_cell_value(row, header) for header in DASHBOARD_EXPORT_HEADERS])
    return out


def cross_company_account_slip_export_rows(rows: List[Dict[str, Any]], flow_type: str = "all") -> List[List[Any]]:
    target = normalize_flow_type(flow_type)
    out: List[List[Any]] = []
    for row in rows:
        if export_row_is_duplicate(row):
            continue
        if target != "all" and export_row_flow_type(row) != target:
            continue
        out.append([cross_company_account_export_cell(row, header) for header in CROSS_COMPANY_ACCOUNT_EXPORT_HEADERS])
    return out


def export_group_summary(rows: List[sqlite3.Row], key_field: str, label: str) -> List[List[Any]]:
    groups: Dict[str, Dict[str, Any]] = {}
    for row in counted_export_rows(rows):
        key = clean_company_name(row[key_field], row["bot_key"]) if key_field == "company_name" else clean_display(row[key_field])
        key = key or label
        group = groups.setdefault(key, {label: key, "count": 0, "amount": 0.0, "fee": 0.0})
        group["count"] += 1
        group["amount"] += float(row["amount"] or 0)
        group["fee"] += float(row["fee"] or 0)
    return [[g[label], g["count"], g["amount"], g["fee"]] for g in sorted(groups.values(), key=lambda g: (-float(g["amount"] or 0), str(g[label])))]


def export_daily_summary(rows: List[sqlite3.Row]) -> List[List[Any]]:
    groups: Dict[str, Dict[str, Any]] = {}
    for row in counted_export_rows(rows):
        date_key, date_label, sort_date = date_bucket(row["slip_date_display"], row["slip_date_iso"])
        group = groups.setdefault(date_key, {"date": date_label, "sort_date": sort_date, "count": 0, "amount": 0.0, "fee": 0.0})
        group["count"] += 1
        group["amount"] += float(row["amount"] or 0)
        group["fee"] += float(row["fee"] or 0)
    ordered = sorted(groups.values(), key=lambda g: str(g.get("sort_date") or ""))
    return [[g["date"], g["count"], g["amount"], g["fee"]] for g in ordered]


def write_export_sheet(ws: Any, headers: List[str], rows: List[List[Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)
    for cell in ws[1]:
        font = copy(cell.font)
        font.bold = True
        cell.font = font


def autofit_workbook(wb: Workbook) -> None:
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            sheet.column_dimensions[col[0].column_letter].width = min(max(12, max_len + 2), 48)


def export_rows_for_selection(conn: sqlite3.Connection, bot_key: str = "", chat_id: str = "", flow_type: str = "all", scope: str = "all", start_date: str = "", end_date: str = "") -> List[sqlite3.Row]:
    where, params = export_where_clause(bot_key=bot_key, chat_id=chat_id, flow_type=flow_type, scope=scope, start_date=start_date, end_date=end_date)
    rows = conn.execute(
        f"SELECT * FROM slips WHERE {where} ORDER BY COALESCE(NULLIF(slip_date_iso,''), slip_date_display), slip_time, created_at",
        params,
    ).fetchall()
    if start_date or end_date:
        rows = [row for row in rows if export_row_in_date_range(row, start_date, end_date)]
    return rows


def duplicate_export_rows(conn: sqlite3.Connection, rows: List[sqlite3.Row]) -> List[List[Any]]:
    duplicate_rows = [row for row in rows if int(row["is_duplicate"] or 0)]
    original_ids = [str(row["duplicate_of"] or "") for row in duplicate_rows if row["duplicate_of"]]
    original_by_id: Dict[str, sqlite3.Row] = {}
    if original_ids:
        placeholders = ",".join("?" for _ in original_ids)
        originals = conn.execute(f"SELECT * FROM slips WHERE id IN ({placeholders})", original_ids).fetchall()
        original_by_id = {str(row["id"]): row for row in originals}
    submitted_by_id: Dict[str, str] = {}
    submitted_ids = [str(row["id"]) for row in duplicate_rows] + original_ids
    if submitted_ids:
        placeholders = ",".join("?" for _ in submitted_ids)
        submitted_rows = conn.execute(
            f"SELECT slip_id, MIN(created_at) AS submitted_at FROM ocr_jobs WHERE slip_id IN ({placeholders}) GROUP BY slip_id",
            submitted_ids,
        ).fetchall()
        submitted_by_id = {str(row["slip_id"]): bkk_iso_from_ms(row["submitted_at"]) for row in submitted_rows}
    out = []
    for row in duplicate_rows:
        original = original_by_id.get(str(row["duplicate_of"] or ""))
        out.append([
            row["message_id"],
            original["message_id"] if original else "",
            row["slip_date_display"],
            row["slip_time"],
            row["transferor_name"],
            display_bank(row["from_bank"]),
            row["from_account"],
            display_bank(row["to_bank"]),
            row["to_account"],
            row["amount"],
            row["reference_no"],
            original["reference_no"] if original else "",
            row["sender_name"],
            original["sender_name"] if original else "",
            submitted_by_id.get(str(row["id"])) or row["created_at_iso"],
            submitted_by_id.get(str(row["duplicate_of"] or "")) or (original["created_at_iso"] if original else ""),
            row["created_at_iso"],
        ])
    return out


def export_dashboard_excel(db_path: Path, bot_key: str = "", chat_id: str = "", flow_type: str = "all", scope: str = "all", start_date: str = "", end_date: str = "", company_name: str = "") -> Path:
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    with connect(db_path) as conn:
        rows = export_rows_for_selection(conn, bot_key=bot_key, chat_id=chat_id, flow_type=flow_type, scope=scope, start_date=start_date, end_date=end_date)
        duplicate_rows = duplicate_export_rows(conn, rows)
    wb = Workbook()
    summary_ws = wb.active
    assert summary_ws is not None
    summary_ws.title = "SummaryByCompany"
    write_export_sheet(summary_ws, ["company_name", "count", "amount", "fee"], export_group_summary(rows, "company_name", "company_name"))
    transferor_ws = wb.create_sheet("SummaryByTransferor")
    write_export_sheet(transferor_ws, ["transferor_name", "count", "amount", "fee"], export_group_summary(rows, "transferor_name", "transferor_name"))
    daily_ws = wb.create_sheet("DailySummary")
    write_export_sheet(daily_ws, ["date", "count", "amount", "fee"], export_daily_summary(rows))
    slips_ws = wb.create_sheet("Slips")
    write_export_sheet(slips_ws, DASHBOARD_EXPORT_HEADERS, dashboard_slip_export_rows(rows, "all"))
    deposit_ws = wb.create_sheet("DepositSlips")
    write_export_sheet(deposit_ws, DASHBOARD_EXPORT_HEADERS, dashboard_slip_export_rows(rows, "deposit"))
    withdraw_ws = wb.create_sheet("WithdrawSlips")
    write_export_sheet(withdraw_ws, DASHBOARD_EXPORT_HEADERS, dashboard_slip_export_rows(rows, "withdraw"))
    dup_ws = wb.create_sheet("DuplicateSlips")
    write_export_sheet(dup_ws, DASHBOARD_DUPLICATE_HEADERS, duplicate_rows)
    autofit_workbook(wb)
    label_parts = [safe_export_name(company_name or bot_key or "all"), normalize_flow_type(flow_type), normalize_export_date(start_date) or clean_display(scope or "all")]
    if end_date:
        label_parts.append(normalize_export_date(end_date))
    out = EXPORT_DIR / ("auditslip-" + "-".join([p for p in label_parts if p]) + f"-{int(time.time())}.xlsx")
    wb.save(out)
    return out


def export_dashboard_zip_by_company(db_path: Path, bot_key: str = "__all__", flow_type: str = "all", scope: str = "all", start_date: str = "", end_date: str = "") -> Path:
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    requested_bot = clean_display(bot_key)
    with connect(db_path) as conn:
        rows = export_rows_for_selection(conn, bot_key="" if requested_bot in {"__all__", "all", ""} else requested_bot, chat_id="", flow_type=flow_type, scope=scope, start_date=start_date, end_date=end_date)
    companies: Dict[str, str] = {}
    for row in rows:
        row_bot = clean_display(row["bot_key"]) or "default"
        companies[row_bot] = clean_company_name(row["company_name"], row_bot) or row_bot
    if not companies:
        raise FileNotFoundError("no slips for export selection")
    zip_path = EXPORT_DIR / f"auditslip-by-company-{normalize_flow_type(flow_type)}-{int(time.time())}.zip"
    generated: List[Path] = []
    try:
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for row_bot in sorted(companies, key=lambda b: company_sort_key(companies[b], b)):
                path = export_dashboard_excel(db_path, bot_key=row_bot, chat_id="", flow_type=flow_type, scope=scope, start_date=start_date, end_date=end_date, company_name=companies[row_bot])
                generated.append(path)
                zf.write(path, arcname=f"{safe_export_name(row_bot)}-{safe_export_name(companies[row_bot])}.xlsx")
    finally:
        for path in generated:
            try:
                path.unlink(missing_ok=True)
            except Exception:
                pass
    return zip_path


def cross_company_account_export_cell(row: Dict[str, Any], header: str) -> Any:
    if header == "reference_no":
        return row.get("reference_no") or row.get("reference") or ""
    if header == "slip_image_url":
        return row.get("image_url") or ""
    value = row.get(header, "")
    return display_bank(value) if header in BANK_EXPORT_FIELDS else value


def export_cross_company_account_slips_excel(db_path: Path, flow_type: str = "all", scope: str = "all", search: str = "") -> Path:
    """Export one account's counted slips across every company as a single workbook."""
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    query = clean_display(search)
    if not query:
        raise ValueError("missing cross-company account search")
    with connect(db_path) as conn:
        result = cross_company_account_slip_search_rows(conn, scope=scope, flow_type=flow_type, search=query, limit=0)
    rows = list(result.get("rows") or [])
    wb = Workbook()
    summary_ws = wb.active
    assert summary_ws is not None
    summary_ws.title = "SummaryByCompany"
    summary_rows = [[row.get("company_name") or row.get("bot_key") or "", int(row.get("count") or 0), float(row.get("amount") or 0)] for row in (result.get("companies") or [])]
    write_export_sheet(summary_ws, ["company_name", "count", "amount"], summary_rows)
    slips_ws = wb.create_sheet("CrossCompanyAccountSlips")
    write_export_sheet(slips_ws, CROSS_COMPANY_ACCOUNT_EXPORT_HEADERS, cross_company_account_slip_export_rows(rows, "all"))
    deposit_ws = wb.create_sheet("DepositSlips")
    write_export_sheet(deposit_ws, CROSS_COMPANY_ACCOUNT_EXPORT_HEADERS, cross_company_account_slip_export_rows(rows, "deposit"))
    withdraw_ws = wb.create_sheet("WithdrawSlips")
    write_export_sheet(withdraw_ws, CROSS_COMPANY_ACCOUNT_EXPORT_HEADERS, cross_company_account_slip_export_rows(rows, "withdraw"))
    autofit_workbook(wb)
    out = EXPORT_DIR / f"auditslip-cross-company-account-{normalize_flow_type(flow_type)}-{safe_export_name(query)}-{safe_export_name(clean_display(scope or 'all'))}-{int(time.time())}.xlsx"
    wb.save(out)
    return out


def export_preview_filename(company_name: str = "", bot_key: str = "", flow_type: str = "all", scope: str = "all", start_date: str = "", end_date: str = "", *, ext: str = "xlsx", prefix: str = "auditslip") -> str:
    label_parts = [safe_export_name(company_name or bot_key or "all"), normalize_flow_type(flow_type), normalize_export_date(start_date) or clean_display(scope or "all")]
    if end_date:
        label_parts.append(normalize_export_date(end_date))
    label = "-".join([p for p in label_parts if p]) or "all"
    return f"{prefix}-{label}-preview.{ext}"


def export_workbook_preview(rows: List[sqlite3.Row], *, bot_key: str = "", chat_id: str = "", flow_type: str = "all", scope: str = "all", start_date: str = "", end_date: str = "", company_name: str = "") -> Dict[str, Any]:
    duplicate_rows = duplicate_export_rows_for_preview(rows)
    counted_rows = counted_export_rows(rows)
    sheets = {
        "SummaryByCompany": len(export_group_summary(rows, "company_name", "company_name")),
        "SummaryByTransferor": len(export_group_summary(rows, "transferor_name", "transferor_name")),
        "DailySummary": len(export_daily_summary(rows)),
        "Slips": len(dashboard_slip_export_rows(rows, "all")),
        "DepositSlips": len(dashboard_slip_export_rows(rows, "deposit")),
        "WithdrawSlips": len(dashboard_slip_export_rows(rows, "withdraw")),
        "DuplicateSlips": len(duplicate_rows),
    }
    flow = normalize_flow_type(flow_type)
    return {
        "ok": True,
        "dry_run": True,
        "format": "xlsx",
        "mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "filename": export_preview_filename(company_name=company_name, bot_key=bot_key, flow_type=flow, scope=scope, start_date=start_date, end_date=end_date, ext="xlsx"),
        "scope": {"bot_key": clean_display(bot_key) or "default", "chat_id": str(chat_id or ""), "flow_type": flow, "date_scope": clean_display(scope or "all"), "start_date": normalize_export_date(start_date), "end_date": normalize_export_date(end_date)},
        "rows": {"selected": len(rows), "counted": len(counted_rows), "duplicates": len(duplicate_rows)},
        "sheets": sheets,
    }


def duplicate_export_rows_for_preview(rows: List[sqlite3.Row]) -> List[sqlite3.Row]:
    """Count duplicate evidence rows without reading linked originals or creating workbooks."""
    return [row for row in rows if int(row["is_duplicate"] or 0)]


def export_dashboard_preview(db_path: Path, bot_key: str = "", chat_id: str = "", flow_type: str = "all", scope: str = "all", start_date: str = "", end_date: str = "", cross_account_search: str = "", company_name: str = "") -> Dict[str, Any]:
    """Read-only export smoke metadata. Does not create XLSX/ZIP files or mutation log rows."""
    flow = normalize_flow_type(flow_type)
    query = clean_display(cross_account_search)
    if query:
        with connect(db_path) as conn:
            result = cross_company_account_slip_search_rows(conn, scope=scope, flow_type=flow, search=query, limit=0)
        rows = list(result.get("rows") or [])
        sheets = {
            "SummaryByCompany": len(result.get("companies") or []),
            "CrossCompanyAccountSlips": len(cross_company_account_slip_export_rows(rows, "all")),
            "DepositSlips": len(cross_company_account_slip_export_rows(rows, "deposit")),
            "WithdrawSlips": len(cross_company_account_slip_export_rows(rows, "withdraw")),
        }
        return {
            "ok": True,
            "dry_run": True,
            "format": "xlsx",
            "mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "filename": f"auditslip-cross-company-account-{flow}-{safe_export_name(query)}-{safe_export_name(clean_display(scope or 'all'))}-preview.xlsx",
            "scope": {"bot_key": "__all__", "chat_id": "", "flow_type": flow, "date_scope": clean_display(scope or "all"), "start_date": normalize_export_date(start_date), "end_date": normalize_export_date(end_date)},
            "rows": {"selected": len(rows), "counted": len(cross_company_account_slip_export_rows(rows, "all")), "duplicates": sum(1 for row in rows if export_row_is_duplicate(row))},
            "company_count": int(result.get("company_count") or 0),
            "is_cross_company": bool(result.get("is_cross_company")),
            "sheets": sheets,
        }

    requested_all = clean_display(bot_key) in {"__all__", "all"} and not chat_id
    with connect(db_path) as conn:
        rows = export_rows_for_selection(conn, bot_key="" if requested_all else bot_key, chat_id=chat_id, flow_type=flow, scope=scope, start_date=start_date, end_date=end_date)
    if requested_all:
        companies: Dict[str, Dict[str, Any]] = {}
        for row in rows:
            row_bot = clean_display(row["bot_key"]) or "default"
            company = clean_company_name(row["company_name"], row_bot) or row_bot
            entry = companies.setdefault(row_bot, {"bot_key": row_bot, "company_name": company, "selected": 0, "counted": 0, "duplicates": 0})
            entry["selected"] += 1
            if int(row["is_duplicate"] or 0):
                entry["duplicates"] += 1
            elif str(row["status"] or "") == "success":
                entry["counted"] += 1
        return {
            "ok": True,
            "dry_run": True,
            "format": "zip",
            "mime": "application/zip",
            "filename": export_preview_filename(company_name="by-company", bot_key=bot_key or "all", flow_type=flow, scope=scope, start_date=start_date, end_date=end_date, ext="zip", prefix="auditslip"),
            "scope": {"bot_key": clean_display(bot_key) or "__all__", "chat_id": "", "flow_type": flow, "date_scope": clean_display(scope or "all"), "start_date": normalize_export_date(start_date), "end_date": normalize_export_date(end_date)},
            "rows": {"selected": len(rows), "counted": len(counted_export_rows(rows)), "duplicates": len(duplicate_export_rows_for_preview(rows))},
            "company_count": len(companies),
            "companies": sorted(companies.values(), key=lambda r: company_sort_key(str(r.get("company_name") or ""), str(r.get("bot_key") or ""))),
            "sheets": {"zip_members": len(companies), "per_workbook": ["SummaryByCompany", "SummaryByTransferor", "DailySummary", "Slips", "DepositSlips", "WithdrawSlips", "DuplicateSlips"]},
        }
    return export_workbook_preview(rows, bot_key=bot_key, chat_id=chat_id, flow_type=flow, scope=scope, start_date=start_date, end_date=end_date, company_name=company_name)


def resolve_export_selection(db_path: Path, chat_id: str = "", bot_key: str = "", flow_type: str = "all") -> Dict[str, Any]:
    """Resolve an export request to a chat that belongs to the selected company/bot.

    The dashboard UI can briefly have a stale chat dropdown after changing companies.
    Never let that stale chat override the selected bot_key, otherwise Export Excel can
    download a file for the previous company.
    """
    requested_chat = str(chat_id or "")
    requested_bot = clean_display(bot_key) or "default"
    flow_type_key = normalize_flow_type(flow_type)
    flow_extra, flow_extra_params = flow_sql_clause(flow_type_key)
    all_bots = requested_bot in {"__all__", "all"}

    def row_payload(row: sqlite3.Row, stale: bool = False) -> Dict[str, Any]:
        return {
            "ok": True,
            "bot_key": str(row["bot_key"] or "default"),
            "company_name": str(row["company_name"] or row["bot_key"] or ""),
            "chat_id": str(row["chat_id"] or ""),
            "chat_title": str(row["chat_title"] or row["chat_id"] or ""),
            "stale_chat_replaced": bool(stale),
            "flow_type": flow_type_for_title(row["chat_title"], row["bot_key"], row["chat_id"]),
        }

    base_select = """
        SELECT COALESCE(NULLIF(bot_key,''),'default') AS bot_key,
               COALESCE(MAX(NULLIF(company_name,'')), COALESCE(NULLIF(bot_key,''),'default')) AS company_name,
               chat_id,
               COALESCE(MAX(NULLIF(chat_title,'')), chat_id) AS chat_title,
               COUNT(*) AS total_rows,
               COALESCE(SUM(CASE WHEN status='success' AND COALESCE(is_duplicate,0)=0 AND settlement_id IS NULL THEN amount ELSE 0 END),0) AS open_amount
        FROM slips
    """
    with connect(db_path) as conn:
        if requested_chat:
            if not all_bots:
                row = conn.execute(
                    base_select + f"""
                    WHERE chat_id=? AND COALESCE(bot_key,'default')=? {flow_extra}
                    GROUP BY COALESCE(NULLIF(bot_key,''),'default'), chat_id
                    LIMIT 1
                    """,
                    (requested_chat, requested_bot, *flow_extra_params),
                ).fetchone()
                if row:
                    return row_payload(row)
            else:
                row = conn.execute(
                    base_select + f"""
                    WHERE chat_id=? {flow_extra}
                    GROUP BY COALESCE(NULLIF(bot_key,''),'default'), chat_id
                    ORDER BY open_amount DESC, total_rows DESC
                    LIMIT 1
                    """,
                    (requested_chat, *flow_extra_params),
                ).fetchone()
                if row:
                    return row_payload(row)

        where = ""
        params: List[Any] = []
        stale = bool(requested_chat and not all_bots)
        if not all_bots:
            where = "WHERE COALESCE(bot_key,'default')=?"
            params.append(requested_bot)
        else:
            where = "WHERE 1=1"
        where += flow_extra
        params.extend(flow_extra_params)
        row = conn.execute(
            base_select + f"""
            {where}
            GROUP BY COALESCE(NULLIF(bot_key,''),'default'), chat_id
            ORDER BY open_amount DESC, total_rows DESC
            LIMIT 1
            """,
            params,
        ).fetchone()
        if row:
            return row_payload(row, stale=stale)
    return {"ok": False, "error": "no chat data for selected company", "bot_key": requested_bot, "chat_id": requested_chat}

def dashboard_close_period(db_path: Path, chat_id: str, note: str = "", bot_key: str = "default", company_name: str = "") -> Dict[str, Any]:
    if not chat_id:
        return {"ok": False, "error": "missing chat_id"}
    bot = AuditslipBot(token="", db_path=db_path, dry_run=True, bot_key=bot_key or "default", company_name=company_name or APP_NAME)
    bot.init_db()
    result = bot.close_period(chat_id, closed_by="dashboard", note=note or "dashboard close")
    return {"ok": True, **result}


def unmark_duplicate_slip(db_path: Path, slip_id: str, bot_key: str = "") -> Dict[str, Any]:
    slip_id = clean_display(slip_id)
    bot_key = clean_display(bot_key)
    if not slip_id:
        return {"ok": False, "error": "missing slip id"}
    with connect(db_path) as conn:
        clause = "id=? AND status='success' AND COALESCE(is_duplicate,0)=1"
        params: List[Any] = [slip_id]
        if bot_key and bot_key not in {"__all__", "all"}:
            clause += " AND COALESCE(bot_key,'default')=?"
            params.append(bot_key)
        before = conn.execute(f"SELECT id, duplicate_of FROM slips WHERE {clause}", params).fetchone()
        if not before:
            return {"ok": False, "error": "duplicate slip not found"}
        conn.execute(f"UPDATE slips SET is_duplicate=0, duplicate_of=NULL WHERE {clause}", params)
        conn.commit()
    return {"ok": True, "id": slip_id, "previous_duplicate_of": before["duplicate_of"]}


def delete_dashboard_slip(db_path: Path, slip_id: str, bot_key: str = "", reason: str = "dashboard delete") -> Dict[str, Any]:
    """Soft-delete one slip so it disappears from normal dashboard totals/lists.

    The amount is kept as evidence in the row, but status becomes `deleted`, so
    financial totals that count success/non-duplicate slips subtract it.
    """
    slip_id = clean_display(slip_id)
    bot_key = clean_display(bot_key)
    if not slip_id:
        return {"ok": False, "error": "missing slip id"}
    with connect(db_path) as conn:
        clause = "id=?"
        params: List[Any] = [slip_id]
        if bot_key and bot_key not in {"__all__", "all"}:
            clause += " AND COALESCE(bot_key,'default')=?"
            params.append(bot_key)
        row = conn.execute(
            f"""
            SELECT id, COALESCE(NULLIF(bot_key,''),'default') AS bot_key, chat_id,
                   status, COALESCE(is_duplicate,0) AS is_duplicate,
                   COALESCE(amount,0) AS amount, error
            FROM slips
            WHERE {clause}
            """,
            params,
        ).fetchone()
        if not row:
            return {"ok": False, "error": "slip not found"}
        amount = float(row["amount"] or 0)
        previous_status = str(row["status"] or "")
        was_counted = previous_status == "success" and int(row["is_duplicate"] or 0) == 0
        if previous_status == "deleted":
            return {
                "ok": True,
                "id": slip_id,
                "bot_key": row["bot_key"],
                "chat_id": row["chat_id"],
                "previous_status": previous_status,
                "removed_amount": 0.0,
                "was_counted": False,
                "already_deleted": True,
            }
        note = clean_display(reason) or "dashboard delete"
        previous_error = clean_display(row["error"])
        deleted_note = f"ลบจาก dashboard: {note}"
        error_text = deleted_note if not previous_error else f"{previous_error} | {deleted_note}"
        conn.execute(
            f"UPDATE slips SET status='deleted', error=? WHERE {clause}",
            [error_text, *params],
        )
        conn.commit()
    return {
        "ok": True,
        "id": slip_id,
        "bot_key": row["bot_key"],
        "chat_id": row["chat_id"],
        "previous_status": previous_status,
        "removed_amount": amount,
        "was_counted": was_counted,
        "already_deleted": False,
    }


def render_dashboard_html(token: str = "") -> str:
    return f"""<!doctype html>
<html lang="th">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>Auditslip Dashboard</title>
  <link rel="icon" href="data:," />
  <script>if('scrollRestoration' in window.history)window.history.scrollRestoration='manual';window.scrollTo(0,0);</script>
  <style>
    :root {{ color-scheme: dark; --bg:#0b1020; --panel:#121a2f; --panel-2:#0f172a; --muted:#94a3b8; --text:#e5e7eb; --good:#22c55e; --warn:#f59e0b; --bad:#ef4444; --line:#26324d; --accent:#3b82f6; --accent-2:#8b5cf6; --soft:#17223a; }}
    * {{ box-sizing:border-box; }}
    body {{ margin:0; font-family: ui-sans-serif, system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif; background:radial-gradient(circle at top left,#1e293b,#0b1020 42%); color:var(--text); overflow-x:hidden; }}
    .topbar {{ min-height:72px; padding:14px 18px; display:flex; justify-content:space-between; gap:12px; align-items:center; border-bottom:1px solid var(--line); position:sticky; top:0; z-index:30; background:rgba(11,16,32,.92); backdrop-filter: blur(14px); }}
    .header-title {{ min-width:0; flex:1; }}
    .top-actions {{ display:flex; flex-wrap:wrap; gap:8px; align-items:center; justify-content:flex-end; min-width:260px; }}
    .scope-chip {{ display:inline-flex; align-items:center; min-height:36px; max-width:360px; padding:7px 10px; border:1px solid rgba(59,130,246,.35); border-radius:999px; background:rgba(37,99,235,.18); color:#dbeafe; font-size:12px; font-weight:850; overflow:hidden; text-overflow:ellipsis; white-space:nowrap; }}
    .last-updated {{ color:var(--muted); font-size:11px; white-space:nowrap; }}
    .admin-toggle {{ background:#334155; }}
    h1 {{ margin:0; font-size:22px; letter-spacing:.01em; }}
    .app-shell {{ display:flex; align-items:flex-start; }}
    .icon-button {{ width:44px; height:44px; display:inline-flex; align-items:center; justify-content:center; border:1px solid var(--line); border-radius:14px; background:#111a2d; color:#dbeafe; box-shadow:0 8px 20px rgba(0,0,0,.16); }}
    .side-menu {{ width:318px; max-width:318px; min-height:calc(100vh - 72px); max-height:calc(100vh - 72px); position:sticky; top:72px; align-self:flex-start; padding:14px; border-right:1px solid rgba(148,163,184,.18); background:linear-gradient(180deg,rgba(15,23,42,.96),rgba(11,16,32,.92)); box-shadow:20px 0 45px rgba(0,0,0,.20); overflow-y:auto; overscroll-behavior:contain; transition:width .18s ease, max-width .18s ease, padding .18s ease, transform .22s ease; }}
    .side-scroll {{ display:grid; gap:12px; }}
    .side-brand {{ display:flex; gap:10px; align-items:center; padding:12px; margin-bottom:12px; border:1px solid rgba(59,130,246,.22); border-radius:18px; background:linear-gradient(135deg,rgba(59,130,246,.18),rgba(139,92,246,.08)); }}
    .side-close {{ margin-left:auto; display:none; width:38px; height:38px; border-radius:12px; }}
    .side-scrim {{ display:none; position:fixed; inset:0; z-index:50; background:rgba(2,6,23,.62); backdrop-filter:blur(2px); }}
    body.side-open .side-scrim {{ display:block; }}
    .brand-mark {{ width:40px; height:40px; border-radius:14px; display:flex; align-items:center; justify-content:center; font-weight:900; background:linear-gradient(135deg,var(--accent),var(--accent-2)); box-shadow:0 10px 25px rgba(59,130,246,.24); }}
    .brand-title {{ font-weight:850; line-height:1.15; }}
    .brand-subtitle {{ color:var(--muted); font-size:12px; margin-top:2px; }}
    .side-panel {{ border:1px solid rgba(148,163,184,.18); border-radius:18px; padding:12px; background:rgba(18,26,47,.72); box-shadow:0 12px 30px rgba(0,0,0,.14); }}
    .side-heading {{ display:flex; justify-content:space-between; align-items:flex-end; gap:8px; margin:0 0 10px; padding-bottom:8px; border-bottom:1px solid rgba(148,163,184,.12); }}
    .side-heading span {{ font-size:14px; font-weight:850; color:#f8fafc; }}
    .side-heading small {{ color:var(--muted); font-size:11px; white-space:nowrap; }}
    .side-field {{ display:grid; gap:5px; margin:8px 0; }}
    .side-field span {{ color:#cbd5e1; font-size:12px; font-weight:700; }}
    .side-actions {{ display:grid; grid-template-columns:1fr 1fr; gap:8px; margin-top:10px; }}
    .side-actions.single {{ grid-template-columns:1fr; }}
    .side-nav {{ display:grid; gap:7px; }}
    .side-menu-item {{ width:100%; display:flex; gap:10px; align-items:center; text-align:left; padding:10px; border-radius:14px; background:rgba(15,23,42,.72); border:1px solid transparent; color:var(--text); }}
    .side-menu-item:hover, .side-menu-item.active {{ border-color:rgba(59,130,246,.45); background:linear-gradient(135deg,rgba(37,99,235,.28),rgba(15,23,42,.8)); }}
    .side-menu-icon {{ width:28px; height:28px; border-radius:10px; display:flex; align-items:center; justify-content:center; flex:0 0 auto; background:#1f2a44; color:#bfdbfe; font-size:14px; }}
    .side-menu-text {{ min-width:0; }}
    .side-menu-title {{ font-weight:800; font-size:13px; }}
    .side-menu-desc {{ color:var(--muted); font-size:11px; margin-top:1px; }}
    .side-company {{ width:100%; display:block; border:1px solid rgba(148,163,184,.15); border-radius:14px; padding:10px; margin-top:8px; background:rgba(15,23,42,.75); color:var(--text); text-align:left; }}
    .side-company:hover {{ border-color:rgba(59,130,246,.45); background:rgba(30,41,59,.86); }}
    .bot-list {{ display:grid; gap:8px; }}
    .bot-row {{ border:1px solid rgba(148,163,184,.14); border-radius:13px; padding:9px; background:rgba(15,23,42,.62); }}
    .bot-row .top {{ display:flex; justify-content:space-between; gap:8px; }}
    .wrap {{ padding:18px; max-width:1280px; margin:0 auto; flex:1; min-width:0; }}
    .grid {{ display:grid; grid-template-columns:repeat(4,minmax(0,1fr)); gap:14px; }}
    .operator-home-grid {{ display:grid; grid-template-columns:repeat(3,minmax(0,1fr)); gap:12px; }}
    .operator-card {{ border:1px solid rgba(148,163,184,.16); border-radius:16px; padding:12px; background:linear-gradient(180deg,rgba(15,23,42,.92),rgba(15,23,42,.66)); min-width:0; }}
    .operator-card .head {{ display:flex; justify-content:space-between; gap:8px; align-items:flex-start; }}
    .operator-stats {{ display:grid; grid-template-columns:1fr 1fr; gap:8px; margin-top:10px; }}
    .operator-stat {{ border:1px solid rgba(148,163,184,.12); border-radius:12px; padding:8px; background:rgba(2,6,23,.24); }}
    .exception-list {{ display:grid; gap:8px; }}
    .exception-item {{ display:flex; justify-content:space-between; align-items:center; gap:10px; border:1px solid rgba(148,163,184,.16); border-radius:14px; padding:10px; background:rgba(15,23,42,.70); }}
    .exception-item strong {{ font-size:15px; }}
    .flow-chart {{ display:grid; gap:9px; margin-top:10px; }}
    .flow-chart-row {{ display:grid; grid-template-columns:minmax(72px,.65fr) minmax(0,2fr) minmax(112px,.75fr); gap:10px; align-items:center; padding:8px 0; border-bottom:1px solid rgba(148,163,184,.10); }}
    .flow-chart-row:last-child {{ border-bottom:0; }}
    .flow-bars {{ display:grid; gap:5px; min-width:0; }}
    .flow-bar {{ height:16px; min-width:2px; border-radius:999px; box-shadow:inset 0 0 0 1px rgba(255,255,255,.08); }}
    .flow-bar.withdraw {{ background:linear-gradient(90deg,#60a5fa,#2563eb); }}
    .flow-bar.deposit {{ background:linear-gradient(90deg,#34d399,#059669); }}
    .flow-chart-values {{ text-align:right; font-size:12px; color:#cbd5e1; font-variant-numeric:tabular-nums; }}
    .flow-legend {{ display:flex; gap:10px; flex-wrap:wrap; margin-top:6px; color:#cbd5e1; font-size:12px; }}
    .legend-dot {{ width:10px; height:10px; display:inline-block; border-radius:999px; margin-right:4px; }}
    .legend-dot.withdraw {{ background:#3b82f6; }} .legend-dot.deposit {{ background:#10b981; }}
    .limit-usage-chart {{ display:grid; gap:10px; margin-top:10px; }}
    .limit-usage-summary {{ display:grid; grid-template-columns:repeat(3,minmax(0,1fr)); gap:8px; margin:8px 0 12px; }}
    .limit-usage-stat {{ border:1px solid rgba(148,163,184,.12); border-radius:12px; padding:9px; background:rgba(2,6,23,.22); }}
    .limit-usage-row {{ display:grid; grid-template-columns:minmax(92px,.7fr) minmax(0,2fr) minmax(130px,.8fr); gap:10px; align-items:center; padding:9px 0; border-bottom:1px solid rgba(148,163,184,.10); }}
    .limit-usage-row:last-child {{ border-bottom:0; }}
    .limit-usage-track {{ height:18px; border-radius:999px; overflow:hidden; background:rgba(148,163,184,.18); box-shadow:inset 0 0 0 1px rgba(255,255,255,.08); }}
    .limit-usage-fill {{ height:100%; border-radius:999px; background:linear-gradient(90deg,#38bdf8,#2563eb); min-width:2px; }}
    .limit-usage-fill.over {{ background:linear-gradient(90deg,#f59e0b,#dc2626); }}
    .limit-usage-values {{ text-align:right; font-size:12px; color:#cbd5e1; font-variant-numeric:tabular-nums; }}
    body:not(.admin-mode) [data-admin-only="true"] {{ display:none !important; }}
    .card {{ background:rgba(18,26,47,.94); border:1px solid var(--line); border-radius:16px; padding:16px; box-shadow:0 10px 30px rgba(0,0,0,.18); min-width:0; overflow:hidden; }}
    .label {{ color:var(--muted); font-size:13px; }}
    .value {{ font-size:28px; font-weight:800; margin-top:6px; }}
    .good {{ color:var(--good); }} .warn {{ color:var(--warn); }} .bad {{ color:var(--bad); }}
    .sections {{ display:grid; grid-template-columns: 1.2fr .8fr; gap:14px; margin-top:14px; }}
    .menu-section[hidden] {{ display:none !important; }}
    .pending-badge {{ display:inline-block; min-width:18px; padding:1px 6px; margin-left:6px; border-radius:999px; background:#dc2626; color:#fff; font-size:11px; font-weight:800; line-height:16px; text-align:center; vertical-align:middle; }}
    .pending-badge[hidden] {{ display:none !important; }}
    .pending-toolbar {{ display:flex; gap:8px; flex-wrap:wrap; align-items:center; margin:8px 0; }}
    .pending-toast-host {{ position:fixed; right:16px; bottom:16px; display:flex; flex-direction:column; gap:8px; z-index:9999; }}
    .pending-toast {{ background:rgba(15,23,42,.95); border:1px solid var(--line); color:#e2e8f0; padding:10px 14px; border-radius:10px; box-shadow:0 12px 30px rgba(0,0,0,.35); font-size:13px; max-width:320px; }}
    .pending-toast.success {{ border-color:rgba(34,197,94,.55); }}
    .pending-toast.error {{ border-color:rgba(220,38,38,.55); }}
    .responsive-table {{ width:100%; overflow-x:auto; -webkit-overflow-scrolling:touch; }}
    table {{ width:100%; border-collapse:collapse; font-size:13px; min-width:620px; }}
    th,td {{ text-align:left; padding:9px 8px; border-bottom:1px solid var(--line); vertical-align:top; }}
    th {{ color:#cbd5e1; font-weight:700; }}
    td {{ overflow-wrap:anywhere; word-break:break-word; }}
    th.action-col, td.action-cell {{ width:120px; min-width:120px; white-space:nowrap; overflow-wrap:normal; word-break:normal; text-align:center; }}
    td.action-cell button {{ min-width:96px; white-space:nowrap; padding:9px 12px; border-radius:12px; }}
    .num, .money-cell {{ white-space:nowrap; font-variant-numeric:tabular-nums; }}
    .muted {{ color:var(--muted); }}
    .pill {{ display:inline-block; padding:3px 8px; border-radius:999px; background:#1f2a44; font-size:12px; }}
    .slip-cards {{ display:grid; gap:10px; }}
    .slip-card {{ border:1px solid var(--line); border-radius:12px; padding:10px; background:#0f172a; display:grid; grid-template-columns:92px 1fr; gap:10px; align-items:start; }}
    .slip-card.dupe-card {{ grid-template-columns:minmax(192px, auto) 1fr; }}
    .dupe-thumbs {{ display:flex; gap:8px; align-items:flex-start; }}
    .slip-card .top {{ display:flex; justify-content:space-between; gap:8px; align-items:center; }}
    .slip-thumb {{ width:92px; height:122px; object-fit:cover; border-radius:10px; border:1px solid var(--line); background:#111827; }}
    .slip-body {{ min-width:0; }}
    .mini {{ font-size:12px; color:var(--muted); margin-top:4px; }}
    .cross-company-list {{ display:grid; gap:12px; margin-top:10px; }}
    .cross-company-card {{ display:grid; gap:12px; border:1px solid rgba(148,163,184,.18); border-radius:16px; padding:13px; background:linear-gradient(180deg,rgba(15,23,42,.95),rgba(15,23,42,.72)); box-shadow:0 8px 22px rgba(0,0,0,.12); min-width:0; }}
    .cross-company-head {{ display:flex; justify-content:space-between; gap:12px; align-items:flex-start; padding-bottom:10px; border-bottom:1px solid rgba(148,163,184,.14); }}
    .cross-company-account {{ font-size:15px; font-weight:900; color:#e5eefc; word-break:break-word; }}
    .cross-company-bank {{ color:#cbd5e1; font-size:13px; margin-top:3px; }}
    .cross-company-chips {{ display:flex; gap:6px; flex-wrap:wrap; justify-content:flex-end; }}
    .cross-company-summary {{ display:grid; grid-template-columns:repeat(2,minmax(0,1fr)); gap:8px; }}
    .cross-company-stat {{ border:1px solid rgba(148,163,184,.12); border-radius:12px; padding:9px; background:rgba(2,6,23,.22); }}
    .cross-company-stat .value {{ font-size:18px; margin-top:3px; }}
    .cross-company-section-title {{ color:#cbd5e1; font-size:12px; font-weight:900; letter-spacing:.02em; margin-bottom:7px; }}
    .cross-company-companies {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(220px,1fr)); gap:10px; }}
    .cross-company-company {{ border:1px solid rgba(59,130,246,.18); border-radius:14px; padding:10px; background:rgba(30,41,59,.36); min-width:0; }}
    .cross-company-company-head {{ display:flex; justify-content:space-between; gap:10px; align-items:flex-start; }}
    .cross-company-company-name {{ font-weight:900; color:#f8fafc; }}
    .cross-company-total {{ text-align:right; white-space:nowrap; font-variant-numeric:tabular-nums; color:#e2e8f0; }}
    .cross-company-days {{ display:grid; gap:2px; margin-top:8px; }}
    .cross-company-day-row {{ display:grid; grid-template-columns:minmax(74px,.8fr) minmax(0,1.2fr); gap:8px; padding:5px 0; border-top:1px solid rgba(148,163,184,.10); color:#cbd5e1; }}
    .cross-company-day-row .day-amount {{ text-align:right; white-space:nowrap; font-variant-numeric:tabular-nums; color:#e5e7eb; }}
    .cross-company-overall-days {{ border:1px dashed rgba(148,163,184,.18); border-radius:12px; padding:8px 10px; background:rgba(2,6,23,.18); }}
    .reconcile-controls {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(180px,1fr)); gap:10px; align-items:end; margin-top:10px; }}
    .reconcile-step {{ display:grid; gap:6px; color:#cbd5e1; font-size:12px; font-weight:850; }}
    .reconcile-step span {{ color:#cbd5e1; }}
    .reconcile-upload, .reconcile-preview, .reconcile-actions {{ grid-column:1/-1; }}
    .reconcile-preview {{ border:1px solid rgba(59,130,246,.22); border-radius:12px; padding:10px; background:rgba(37,99,235,.10); color:#dbeafe; font-size:13px; }}
    .reconcile-actions {{ display:flex; gap:8px; flex-wrap:wrap; align-items:center; }}
    a.button, button {{ color:white; text-decoration:none; background:#2563eb; border:0; border-radius:10px; padding:9px 12px; font-weight:700; cursor:pointer; }}
    .toolbar {{ display:flex; gap:8px; flex-wrap:wrap; align-items:center; }}
    .danger {{ background:#dc2626; }}
    .statusline {{ min-width:180px; color:var(--muted); font-size:13px; }}
    .modal-backdrop {{ position:fixed; inset:0; z-index:100; display:none; align-items:center; justify-content:center; padding:18px; background:rgba(2,6,23,.72); backdrop-filter:blur(5px); }}
    .modal-backdrop.open {{ display:flex; }}
    .modal-card {{ width:min(440px,100%); border:1px solid rgba(148,163,184,.24); border-radius:18px; background:linear-gradient(180deg,#121a2f,#0f172a); box-shadow:0 30px 80px rgba(0,0,0,.42); padding:18px; }}
    .modal-title {{ font-size:18px; font-weight:900; margin:0 0 8px; color:#f8fafc; }}
    .modal-message {{ color:#cbd5e1; line-height:1.55; white-space:pre-wrap; }}
    .modal-input {{ width:100%; min-height:86px; margin-top:12px; resize:vertical; background:#0f172a; color:var(--text); border:1px solid var(--line); border-radius:12px; padding:10px; font:inherit; }}
    .modal-input[hidden] {{ display:none !important; }}
    .modal-actions {{ display:flex; justify-content:flex-end; gap:8px; margin-top:16px; }}
    .modal-cancel {{ background:#334155; }}
    .modal-primary.danger {{ background:#dc2626; }}
    select, input {{ width:100%; background:#0f172a; color:var(--text); border:1px solid var(--line); border-radius:12px; padding:9px 10px; max-width:100%; }}
    details.side-panel summary {{ list-style:none; display:flex; justify-content:space-between; align-items:center; cursor:pointer; }}
    details.side-panel summary::-webkit-details-marker {{ display:none; }}
    details.side-panel summary span {{ font-size:14px; font-weight:850; }}
    details.side-panel summary small {{ color:var(--muted); font-size:11px; }}
    body.side-collapsed .side-menu {{ width:82px; padding:10px; }}
    body.side-collapsed .brand-title, body.side-collapsed .brand-subtitle, body.side-collapsed .side-panel:not(.side-nav-panel):not(.side-menu-compact), body.side-collapsed .side-heading, body.side-collapsed .side-menu-desc, body.side-collapsed .side-menu-title {{ display:none; }}
    body.side-collapsed .side-brand {{ justify-content:center; padding:10px; }}
    body.side-collapsed .side-menu-item {{ justify-content:center; padding:10px 8px; }}
    @media (max-width: 1100px) {{ .grid {{ grid-template-columns:repeat(2,minmax(0,1fr)); }} .operator-home-grid {{ grid-template-columns:repeat(2,minmax(0,1fr)); }} .sections {{ grid-template-columns:1fr; }} .side-menu {{ width:292px; }} }}
    @media (max-width: 780px) {{ .topbar {{ align-items:flex-start; flex-wrap:wrap; }} .top-actions {{ width:100%; justify-content:flex-start; min-width:0; }} .scope-chip {{ max-width:100%; }} .app-shell {{ display:block; }} .side-close {{ display:inline-flex; }} .side-menu {{ position:fixed; top:0; left:0; bottom:0; width:min(88vw,340px); max-width:min(88vw,340px); min-height:100dvh; max-height:100dvh; z-index:60; border-right:1px solid rgba(148,163,184,.22); border-bottom:0; border-radius:0 22px 22px 0; transform:translateX(-105%); }} body.side-open .side-menu {{ transform:translateX(0); }} body.side-collapsed .side-menu {{ width:min(88vw,340px); max-width:min(88vw,340px); padding:14px; }} body.side-collapsed .side-scroll {{ display:grid; }} body.side-collapsed .side-brand {{ margin-bottom:12px; justify-content:flex-start; padding:12px; }} body.side-collapsed .side-menu-item {{ justify-content:flex-start; padding:10px; }} .side-actions {{ grid-template-columns:1fr; }} }}
    @media (max-width: 720px) {{ .slip-card {{ grid-template-columns:76px 1fr; }} .slip-card.dupe-card {{ grid-template-columns:1fr; }} .dupe-thumbs .slip-thumb {{ width:calc(50vw - 30px); max-width:120px; }} .slip-thumb {{ width:76px; height:104px; }} table {{ min-width:560px; }} }}
    @media (max-width: 640px) {{
      .responsive-table {{ overflow-x:visible; }}
      .responsive-table table {{ min-width:0; border-collapse:separate; border-spacing:0 10px; }}
      .responsive-table thead {{ display:none; }}
      .responsive-table tbody, .responsive-table tr, .responsive-table td {{ display:block; width:100%; }}
      .responsive-table tr {{ border:1px solid rgba(148,163,184,.18); border-radius:14px; padding:8px; margin-bottom:10px; background:#0f172a; box-shadow:0 8px 18px rgba(0,0,0,.10); }}
      .responsive-table td {{ display:grid; grid-template-columns:minmax(104px,42%) 1fr; gap:8px; align-items:start; padding:7px 6px; border-bottom:1px solid rgba(148,163,184,.12); text-align:right; }}
      .responsive-table td::before {{ content:attr(data-label); color:var(--muted); font-weight:800; text-align:left; }}
      .responsive-table td:last-child {{ border-bottom:0; }}
      .responsive-table td button {{ width:100%; }}
      .responsive-table td.action-cell {{ width:100%; min-width:0; display:block; text-align:left; }}
      .responsive-table td.action-cell::before {{ display:none; }}
    }}
    @media (max-width: 420px) {{ .responsive-table td {{ grid-template-columns:1fr; text-align:left; }} .responsive-table td::before {{ margin-bottom:2px; }} }}
    @media (max-width: 560px) {{ .topbar {{ position:static; }} .wrap {{ padding:10px; }} .grid {{ grid-template-columns:1fr; }} .operator-home-grid {{ grid-template-columns:1fr; }} .value {{ font-size:24px; }} .card {{ padding:12px; border-radius:14px; }} h1 {{ font-size:20px; }} .limit-usage-summary {{ grid-template-columns:1fr; }} .limit-usage-row {{ grid-template-columns:1fr; gap:6px; }} .limit-usage-values {{ text-align:left; }} .cross-company-head, .cross-company-company-head {{ display:grid; }} .cross-company-chips {{ justify-content:flex-start; }} .cross-company-summary {{ grid-template-columns:1fr; }} .cross-company-companies {{ grid-template-columns:1fr; }} .cross-company-total {{ text-align:left; white-space:normal; }} .cross-company-day-row {{ grid-template-columns:1fr; gap:2px; }} .cross-company-day-row .day-amount {{ text-align:left; white-space:normal; }} }}
  </style>
</head>
<body>
  <header class="topbar">
    <button id="sideMenuToggle" class="icon-button" type="button" onclick="toggleSideMenu()" aria-label="เปิด/ปิด side menu" aria-expanded="true">☰</button>
    <div class="header-title"><h1>Auditslip Dashboard</h1><div class="muted">บริษัท · ฝาก/ถอน · รอบ · รายการที่ต้องจัดการ</div></div>
    <div id="mobileTopActions" class="top-actions">
      <span id="activeScopeChip" class="scope-chip">กำลังโหลด scope...</span>
      <span id="lastUpdatedLabel" class="last-updated">ยังไม่อัปเดต</span>
      <button id="topRefreshButton" type="button" onclick="refreshDashboardHome()">รีเฟรช</button>
      <button id="adminModeToggle" class="admin-toggle" type="button" onclick="toggleAdminMode()">ตั้งค่า/Admin</button>
    </div>
  </header>
  <div id="sideScrim" class="side-scrim" onclick="closeSideMenu()" aria-hidden="true"></div>
  <div class="app-shell">
    <aside id="sideMenu" class="side-menu" aria-label="ระบบ">
      <div class="side-brand">
        <div class="brand-mark">AS</div>
        <div><div class="brand-title">ระบบ Auditslip</div><div class="brand-subtitle">ควบคุมบริษัท · ฝาก/ถอน · Export</div></div>
        <button id="sideMenuClose" class="icon-button side-close" type="button" onclick="closeSideMenu()" aria-label="ปิด side menu">×</button>
      </div>
      <div class="side-scroll">
        <section class="side-panel side-controls">
          <div class="side-heading"><span>ควบคุมมุมมอง</span><small>เลือกข้อมูล</small></div>
          <label class="side-field"><span>บริษัท</span><select id="botFilter" title="เลือกบริษัท"></select></label>
          <label class="side-field"><span>ฝาก/ถอน</span><select id="flowFilter" title="เลือกกลุ่มฝาก/เติมมือหรือถอน"><option value="all">รวมทุกกลุ่ม</option><option value="withdraw">ทุกกลุ่มถอน</option><option value="deposit">ทุกกลุ่มฝาก/เติมมือ</option></select></label>
          <label class="side-field"><span>กลุ่ม</span><select id="chat" title="เลือกกลุ่มของบริษัท"></select></label>
          <label class="side-field"><span>รอบ</span><select id="scope"><option value="today" selected>วันนี้</option><option value="open">รอบเปิด</option><option value="all">ทั้งหมด</option></select></label>
          <label class="side-field"><span>เลือกวันที่เดียว</span><input id="customDateFilter" type="date" title="เลือกวันที่เพื่อดูทุกตารางเฉพาะวันนี้" /></label>
          <label class="side-field"><span>ช่วงวันที่เริ่ม</span><input id="summaryStartDate" type="date" title="วันที่เริ่มสำหรับสรุปแดชบอร์ด" /></label>
          <label class="side-field"><span>ช่วงวันที่สิ้นสุด</span><input id="summaryEndDate" type="date" title="วันที่สิ้นสุดสำหรับสรุปแดชบอร์ด" /></label>
          <label class="side-field"><span>สถานะสลิป</span><select id="slipFilter"><option value="all">สลิปทั้งหมด</option><option value="success">เฉพาะไม่ซ้ำ</option><option id="duplicateOnly" value="duplicate">สลิปซ้ำที่จับแล้ว</option><option value="issues">อ่านไม่ชัด/ error</option></select></label>
          <label class="side-field"><span>ค้นหา</span><input id="slipSearch" placeholder="ค้นหาสลิป / ค้นหาสลิปซ้ำ" /></label>
          <div class="side-actions"><button onclick="load({{scrollTop:true}})">ค้นหา</button><button onclick="clearSlipSearch()">ล้างค้นหา</button><button onclick="load({{scrollTop:true}})">Refresh</button></div>
        </section>

        <details class="side-panel side-admin-login-panel" data-admin-only="true" open>
          <summary><span>เข้าสู่ระบบ Admin</span><small>ไม่ใช้ token URL</small></summary>
          <div class="side-field" style="margin-top:10px"><span>User</span><input id="adminUsername" autocomplete="username" value="owner" /></div>
          <div class="side-field"><span>Password</span><input id="adminPassword" type="password" autocomplete="current-password" placeholder="รหัสผ่าน" /></div>
          <div class="side-actions"><button type="button" onclick="adminLogin()">Login Admin</button><button type="button" onclick="adminLogout()">Logout</button></div>
          <div id="adminLoginStatus" class="mini muted">เปิดดูข้อมูลได้เลยโดยไม่ต้อง login · login เฉพาะตอนตั้งค่า/อนุมัติ/แก้ไข</div>
        </details>

        <details class="side-panel side-admin-panel" data-admin-only="true">
          <summary><span>ตั้งค่าบอท Telegram</span><small>Admin/debug</small></summary>
          <div id="botSettings" style="margin-top:10px"><div id="telegramBots" class="muted">โหลดข้อมูลบอท...</div></div>
        </details>

        <section class="side-panel side-nav-panel" data-legacy-label="เมนูฟังก์ชั่น">
          <div class="side-heading"><span>เมนูหลัก</span><small>จัดหมวดตามงาน</small></div>
          <div class="side-nav">
            <button class="side-menu-item active" type="button" data-menu-target="section-operator-home" onclick="showMenuSection('section-operator-home')"><span class="side-menu-icon">★</span><span class="side-menu-text"><span class="side-menu-title">งานวันนี้</span><span class="side-menu-desc">ภาพรวม operator และรายการที่ต้องจัดการ</span></span></button>
            <button class="side-menu-item" type="button" data-menu-target="section-overview" onclick="showMenuSection('section-overview')"><span class="side-menu-icon">▦</span><span class="side-menu-text"><span class="side-menu-title">ภาพรวม</span><span class="side-menu-desc">ยอดรวมแยกบริษัท</span></span></button>
            <button class="side-menu-item" type="button" data-menu-target="all" onclick="showAllMenuSections()"><span class="side-menu-icon">⌘</span><span class="side-menu-text"><span class="side-menu-title">แสดงทั้งหมด</span><span class="side-menu-desc">ทุกการ์ดและทุกตาราง</span></span></button>
          </div>
          <div class="side-heading" style="margin-top:14px"><span>ตรวจเงิน</span><small>ledger · ธนาคาร · กระทบยอด</small></div>
          <div class="side-nav">
            <button class="side-menu-item" type="button" data-menu-target="section-account-ledger" onclick="showMenuSection('section-account-ledger')"><span class="side-menu-icon">≡</span><span class="side-menu-text"><span class="side-menu-title">เดินบัญชีรายบัญชี</span><span class="side-menu-desc">timeline + running balance</span></span></button>
            <button class="side-menu-item" type="button" data-menu-target="section-employee-audit" onclick="showMenuSection('section-employee-audit')"><span class="side-menu-icon">◫</span><span class="side-menu-text"><span class="side-menu-title">ออดิตพนักงาน</span><span class="side-menu-desc">1 เทียบ ledger · 2 รายวัน · 3 ซ้ำข้ามบริษัท</span></span></button>
            <button class="side-menu-item" type="button" data-menu-target="section-bank-ledger" onclick="showMenuSection('section-bank-ledger')"><span class="side-menu-icon">▤</span><span class="side-menu-text"><span class="side-menu-title">Preview Statement</span><span class="side-menu-desc">อัปโหลด statement เทียบสลิป</span></span></button>
            <button class="side-menu-item" type="button" data-menu-target="section-banks" onclick="showMenuSection('section-banks')"><span class="side-menu-icon">⇄</span><span class="side-menu-text"><span class="side-menu-title">ยอดธนาคาร</span><span class="side-menu-desc">ยอดแยกต้นทาง/ปลายทาง</span></span></button>
            <button class="side-menu-item" type="button" data-menu-target="section-reconcile" onclick="showMenuSection('section-reconcile')"><span class="side-menu-icon">✓</span><span class="side-menu-text"><span class="side-menu-title">กระทบยอด Statement</span><span class="side-menu-desc">หลังบ้าน vs สลิป</span></span></button>
            <button class="side-menu-item" type="button" data-menu-target="section-bank-review" onclick="showMenuSection('section-bank-review')"><span class="side-menu-icon">◎</span><span class="side-menu-text"><span class="side-menu-title">รีเช็คธนาคาร</span><span class="side-menu-desc">OpenAI เช็กต้นทางที่ไม่ชัด</span></span></button>
          </div>
          <div class="side-heading" style="margin-top:14px"><span>จัดการสลิป</span><small>สลิปล่าสุด · ฝาก/ถอน · ซ้ำ</small></div>
          <div class="side-nav">
            <button class="side-menu-item" type="button" data-menu-target="section-recent" onclick="showMenuSection('section-recent')"><span class="side-menu-icon">●</span><span class="side-menu-text"><span class="side-menu-title">Recent / Queue</span><span class="side-menu-desc">รายการล่าสุดและคิว OCR</span></span></button>
            <button class="side-menu-item" type="button" data-menu-target="section-deposit-slips" onclick="showMenuSection('section-deposit-slips')"><span class="side-menu-icon">＋</span><span class="side-menu-text"><span class="side-menu-title">ฝาก/เติมมือ</span><span class="side-menu-desc">สลิปลูกค้า ไม่มีวงเงิน</span></span></button>
            <button class="side-menu-item" type="button" data-menu-target="limitSection" onclick="showMenuSection('limitSection')"><span class="side-menu-icon">↯</span><span class="side-menu-text"><span class="side-menu-title">ฝั่งถอน/วงเงิน</span><span class="side-menu-desc">วงเงินรายวัน/ผู้โอนถอน</span></span></button>
            <button class="side-menu-item" type="button" data-menu-target="section-date-sender" onclick="showMenuSection('section-date-sender')"><span class="side-menu-icon">◷</span><span class="side-menu-text"><span class="side-menu-title">วันที่/ผู้ส่งรูป</span><span class="side-menu-desc">ยอดรายวันและคนส่งรูป</span></span></button>
            <button class="side-menu-item" type="button" data-menu-target="section-duplicates" onclick="showMenuSection('section-duplicates')"><span class="side-menu-icon">⧉</span><span class="side-menu-text"><span class="side-menu-title">สลิปซ้ำ</span><span class="side-menu-desc">ตรวจคู่ซ้ำ/ยกเลิกซ้ำ</span></span></button>
          </div>
          <div class="side-heading" style="margin-top:14px"><span>ระบบ</span><small>บัญชี · พิจารณา</small></div>
          <div class="side-nav">
            <button class="side-menu-item" type="button" data-menu-target="section-pending" onclick="showMenuSection('section-pending')"><span class="side-menu-icon">⏳</span><span class="side-menu-text"><span class="side-menu-title">รออนุมัติ <span id="pendingBadge" class="pending-badge" hidden>0</span></span><span class="side-menu-desc">two-person approval · ยังไม่ได้อนุมัติ</span></span></button>
            <button class="side-menu-item" type="button" data-menu-target="section-company-accounts" onclick="showMenuSection('section-company-accounts')"><span class="side-menu-icon">🏦</span><span class="side-menu-text"><span class="side-menu-title">บัญชี/ค้นสลิป</span><span class="side-menu-desc">บัญชีรับเงินและดูรูปสลิปต่อบัญชี</span></span></button>
          </div>
        </section>

        <section class="side-panel side-export-panel">
          <div class="side-heading"><span>ส่งออก Excel</span><small>แยกบริษัท/ตามวัน</small></div>
          <label class="side-field"><span>บริษัทที่ส่งออก</span><select id="exportCompanyFilter" title="เลือกบริษัทสำหรับ export และกรองทั้งหน้า"></select></label>
          <label class="side-field"><span>วันที่เริ่ม</span><input id="exportStartDate" type="date" title="วันที่เริ่ม export" /></label>
          <label class="side-field"><span>วันที่สิ้นสุด</span><input id="exportEndDate" type="date" title="วันที่สิ้นสุด export" /></label>
          <div class="side-actions single"><a id="excel" class="button" href="#" target="exportDownloadFrame" onclick="return exportExcel()">Export Excel</a></div>
          <iframe id="exportDownloadFrame" name="exportDownloadFrame" title="export download" hidden></iframe>
          <!-- export_dashboard_zip_by_company -->
        </section>

        <section class="side-panel side-danger-panel" data-admin-only="true">
          <div class="side-heading"><span>ปิดรอบ</span><small>เฉพาะกลุ่มที่เลือก</small></div>
          <label class="side-field"><span>note ปิดรอบ</span><input id="closeNote" placeholder="note ปิดรอบ" /></label>
          <div id="closeGuardHint" class="mini muted">ต้องเลือกบริษัทและกลุ่มเดียวก่อนปิดรอบ</div>
          <div class="side-actions single"><button id="closePeriodButton" class="danger" onclick="closeOpenPeriod()">Close / เคลียร์ยอด</button></div>
          <span id="statusline" class="statusline"></span>
        </section>

        <section class="side-panel side-company-panel" aria-label="submenu บริษัท">
          <div class="side-heading"><span>บริษัท</span><small>เลือกบริษัท</small></div>
          <div id="sideCompanies" class="muted">โหลดบริษัท...</div>
        </section>
      </div>
    </aside>
    <main class="wrap">
    <section id="section-metrics" class="grid menu-section" data-always-visible="true">
      <div class="card"><div class="label">ยอดถอนวันนี้/ช่วงที่เลือก</div><div id="withdrawAmount" class="value good">-</div></div>
      <div class="card"><div class="label">สลิปถอน</div><div id="withdrawCount" class="value">-</div></div>
      <div class="card"><div class="label">ถอนรวม / วงเงินรวม</div><div id="withdrawLimitUsageRatio" class="value good">-</div><div id="withdrawLimitUsageMeta" class="mini">วงเงินรวมทุกบัญชีถอน</div></div>
      <div class="card"><div class="label">เหลือ/เกินวงเงินถอน</div><div id="withdrawLimitRemaining" class="value good">-</div><div id="withdrawLimitRemainingMeta" class="mini">เทียบกับวงเงินรวม</div></div>
      <div class="card"><div class="label">ยอดฝาก/เติมมือวันนี้/ช่วงที่เลือก</div><div id="depositAmount" class="value good">-</div></div>
      <div class="card"><div class="label">สลิปฝาก/เติมมือ</div><div id="depositCount" class="value">-</div></div>
      <div class="card"><div class="label">คิวรอ OCR</div><div id="queued" class="value warn">-</div></div>
      <div class="card"><div class="label">processing / failed</div><div id="processing" class="value">-</div></div>
      <div class="card"><div class="label">สลิปซ้ำตามช่วง</div><div id="duplicateCount" class="value warn">-</div><div id="duplicateAmount" class="mini"></div></div>
      <div class="card"><div class="label">รีเช็คธนาคารต้นทาง</div><div id="sourceBankReviewCount" class="value warn">-</div><div class="mini">success ไม่ซ้ำ แต่ยังไม่เจอธนาคารต้นทางจริง ๆ · ปลายทางว่างไม่นับเป็น issue</div></div>
      <div class="card"><div class="label">True Wallet วันนี้</div><div id="twalletTodayAmount" class="value good">-</div><div id="twalletTodayMeta" class="mini">โหลดข้อมูล TWallet...</div></div>
    </section>
    <section id="section-operator-home" class="sections menu-section" hidden>
      <div class="card"><h3>True Wallet</h3><div class="mini">ยอดจาก dashboard TWallet · แสดงเฉพาะยอด/จำนวน ไม่แสดง token หรือเบอร์เต็ม</div><div id="twalletSummary"></div></div>
      <div class="card"><h3>งานวันนี้</h3><div class="mini">ภาพรวมบริษัท · ฝาก/ถอน · รอบ สำหรับ operator บนมือถือ</div><div id="operatorHome"></div></div>
      <div class="card"><h3>รายการที่ต้องจัดการ</h3><div class="mini">รวมเฉพาะ exception ที่ควรเปิดดูต่อ: เกินวงเงิน, OCR issue, สลิปซ้ำ, รีเช็คธนาคาร</div><div id="exceptionQueue"></div></div>
    </section>
    <section id="section-overview" class="sections menu-section" hidden>
      <div class="card"><h3>ภาพรวม / บริษัทที่เลือก</h3><div class="mini">เลือกทุกบริษัทเพื่อดูรวมทุกบริษัท หรือเลือกบริษัทเดียวเพื่อให้ทุกเมนูแสดงเฉพาะบริษัทนั้น</div><div id="companyOverview"></div></div>
      <div class="card"><h3>ตัวกรองที่ใช้อยู่</h3><div id="activeSelectionSummary" class="muted">เลือกบริษัท/บอทจาก side menu</div></div>
    </section>
    <section id="section-company-accounts" class="sections menu-section" hidden>
      <div class="card"><h3>บริษัทย่อย/บัญชีรับเงิน</h3><div class="toolbar"><input id="companyName" placeholder="ชื่อบริษัท" /><input id="accountBank" placeholder="ธนาคาร" /><input id="accountNo" placeholder="เลขบัญชีที่ใช้" /><input id="accountName" placeholder="ชื่อบัญชีที่ใช้" /><input id="accountDailyLimit" type="number" step="0.01" placeholder="วงเงิน/วัน" /><button onclick="saveCompanyAccount()">บันทึกบัญชีบริษัท</button></div><div id="companyAccounts"></div></div>
      <div class="card"><h3>ฝั่งถอน · รายบัญชีตามวันที่</h3><div class="mini">บัญชีของบริษัท/ผู้โอน [กลุ่มถอน] · บัญชีผู้โอน/ต้นทางจากกลุ่มถอน แยกตามวันและบริษัท วางไว้ก่อนเพื่อไม่ต้องเลื่อนผ่านรายการฝาก/เติมมือ</div><div id="companyAccountDailyWithdraw"></div></div>
      <div class="card"><h3>ฝั่งฝาก/เติมมือ · รายบัญชีตามวันที่</h3><div class="mini">บัญชีของบริษัท/ผู้โอน [กลุ่มฝาก/เติมมือ] · บัญชีรับเงิน/ปลายทางจากกลุ่มฝาก/เติมมือ แยกไว้คนละส่วนกับฝั่งถอน</div><div id="companyAccountDailyDeposit"></div></div>
      <div class="card"><h3>ค้นหารายการสลิปตามบัญชี</h3><div class="mini">พิมพ์เลขบัญชี/ชื่อ/ธนาคารในช่องค้นหา ระบบจะแสดงจำนวนสลิป ยอดรวม และรูปสลิปทีละใบใน scope ที่เลือก</div><div id="accountSlipSearch"></div><div id="accountCrossCompanyBlock" class="cross-conditional" hidden><h3>บัญชีถอนที่พบข้ามบริษัท</h3><div class="mini">เฉพาะสลิปถอนเท่านั้น ไม่เอาชื่อ/บัญชีจากสลิปฝากหรือเติมมือมาแสดง</div><div id="accountCrossCompany"></div></div><div id="crossCompanyAccountSlipSearchBlock" class="cross-conditional" hidden><h3>ค้นหาสลิปถอนข้ามบริษัท</h3><div class="mini">ค้นทุกบริษัทเฉพาะสลิปถอน เพื่อดูรูปสลิปของบัญชีถอนที่ซ้ำข้ามบริษัท</div><div id="crossCompanyAccountSlipSearch"></div></div></div>
    </section>
    <section id="section-date-sender" class="sections menu-section" hidden>
      <div class="card"><h3>กราฟรายวัน ฝาก/ถอน</h3><div class="mini">นับเฉพาะสลิป success ไม่ซ้ำ แยกกลุ่มฝาก/เติมมือและถอนตามวันที่ของสลิป</div><div id="dailyFlowChart"></div></div>
      <div class="card"><h3>ยอดแยกตามวันที่</h3><div id="byDate"></div></div>
      <div class="card"><h3>ยอดแยกตามผู้ส่งรูป</h3><div id="bySender"></div></div>
    </section>
    <section id="limitSection" class="sections menu-section" hidden>
      <div class="card"><h3>ยอดถอนรวม / วงเงินรวมทุกบัญชี</h3><div class="mini">รวมทุกบัญชีถอนในบริษัท/วันที่เลือก · วงเงินเป็นวงเงินรายวันรวมตามบัญชีและวันที่ใน scope</div><div id="withdrawLimitUsageChart"></div></div>
      <div class="card"><h3>ฝั่งถอน · วงเงินรายวันต่อบัญชี</h3><div id="withdrawLimitSummary" class="mini">นับเฉพาะกลุ่มถอน ไม่รวมฝาก/เติมมือ</div><div class="mini">แยกตามวันที่ของสลิปและเลขบัญชีผู้โอน วงเงิน/วันจะ reset ทุกวัน และไม่นับสลิปซ้ำ</div><div id="byAccountDay"></div></div>
      <div class="card"><h3>ฝั่งถอน · ตั้งวงเงินจากยอดรายวัน</h3><div class="mini">ตั้งวงเงินบัญชี: กด “ตั้งวงเงิน” จากแถวบัญชี ระบบจะจำบริษัทของแถวนั้นให้เอง แม้ตอนดูรวมทุกบริษัท</div><div class="toolbar limit-edit-form"><input id="limitKey" placeholder="เลือกบัญชีจากปุ่มตั้งวงเงิน" readonly /><input id="limitScopeValue" placeholder="บริษัท/กลุ่มที่จะบันทึก" readonly /><input id="limitName" placeholder="ชื่อบัญชี" /><input id="limitBank" placeholder="ธนาคาร" /><input id="limitAccount" placeholder="เลขบัญชี" /><input id="limitAmount" type="number" step="0.01" placeholder="วงเงินต่อวัน" /><button onclick="saveAccountLimit()">บันทึกวงเงิน</button></div><div id="limitScopeHint" class="mini muted">เลือกบัญชีจากตารางด้านล่างก่อนแก้ไขวงเงิน</div><div id="byTransferor"></div></div>
    </section>
    <section id="section-deposit-slips" class="sections menu-section" hidden>
      <div class="card"><h3>ฝั่งฝาก/เติมมือ · สลิปลูกค้าฝาก/เติมมือ</h3><div id="depositCustomerSummary" class="mini">สลิปลูกค้า ไม่มีวงเงิน และไม่เอาไปรวมกับถอน/วงเงิน</div><div id="depositCustomerSlips"></div></div>
    </section>
    <section id="section-account-ledger" class="sections menu-section" hidden>
      <div class="card">
        <div style="display:flex;gap:10px;align-items:center;flex-wrap:wrap;margin-bottom:10px">
          <b>เดินบัญชีรายบัญชี</b>
          <select id="ledgerAccountKey" style="flex:1;min-width:180px" onchange="loadLedger()"><option value="">— เลือกบัญชี —</option></select>
          <select id="ledgerFlow" onchange="loadLedger()"><option value="all">ทั้งหมด</option><option value="deposit">ฝาก</option><option value="withdraw">ถอน</option></select>
          <input type="date" id="ledgerDateFrom" onchange="loadLedger()" style="width:130px">
          <input type="date" id="ledgerDateTo" onchange="loadLedger()" style="width:130px">
          <button type="button" onclick="loadLedger()" style="white-space:nowrap">🔄 โหลด</button>
          <button type="button" id="ledgerExportBtn" onclick="exportLedgerExcel()" style="white-space:nowrap" disabled>⬇ Excel</button>
        </div>
        <div id="ledgerKpi" class="mini muted" style="margin-bottom:8px"></div>
      </div>
      <div class="card">
        <div id="ledgerBody"><div class="muted mini">เลือกบัญชีและกด "โหลด"</div></div>
      </div>
    </section>
    <section id="section-banks" class="sections menu-section" hidden>
      <div class="card"><h3>ยอดแยกตามธนาคารต้นทาง/ผู้โอน</h3><div id="byFromBank"></div><h3>ธนาคารปลายทาง</h3><div id="byToBank"></div></div>
      <div class="card"><h3>หมายเหตุธนาคาร</h3><div class="muted">ตัดตารางยอดแยกตาม issuer ออกแล้ว เหลือเฉพาะต้นทางและปลายทางที่ใช้ตรวจยอดจริง</div></div>
    </section>
    <section id="section-bank-ledger" class="sections menu-section" hidden>
      <div class="card">
        <h3>เดินบัญชีรายบัญชี <span class="pill good">preview-first</span></h3>
        <div class="mini">อัปโหลด statement รายบัญชีเพื่อ preview ก่อน import · import จริงต้องผ่าน pending approval (ledger.import) และเก็บ audit-chain</div>
        <div class="reconcile-controls">
          <label class="reconcile-step"><span>บริษัท</span><select id="ledgerCompanyFilter" title="เลือกบริษัทของ ledger"></select></label>
          <label class="reconcile-step"><span>ฝาก/ถอน</span><select id="ledgerFlowFilter" title="เลือกทิศทาง statement"><option value="deposit">ฝาก/เงินเข้า</option><option value="withdraw">ถอน/เงินออก</option><option value="all">ทุกทิศทาง</option></select></label>
          <label class="reconcile-step"><span>วันที่</span><input id="ledgerDateScope" type="date" title="วันที่ของรายการเดินบัญชี ถ้าเว้นว่างจะใช้ scope หลัก" /></label>
          <label class="reconcile-step"><span>ธนาคาร</span><input id="statementBank" placeholder="เช่น KBANK" /></label>
          <label class="reconcile-step"><span>เลขบัญชี</span><input id="statementAccountNo" placeholder="เลขบัญชีที่ต้องการออดิต" /></label>
          <label class="reconcile-step"><span>ชื่อบัญชี</span><input id="statementAccountName" placeholder="ชื่อบัญชี" /></label>
          <label class="reconcile-step reconcile-upload"><span>statement Excel/CSV</span><input id="ledgerStatementFile" type="file" accept=".xlsx,.xlsm,.csv" /></label>
          <label class="reconcile-step reconcile-upload" data-admin-only="true"><span>หรือ path บน server</span><input id="ledgerStatementPath" placeholder="statement_path" /></label>
          <div class="reconcile-actions"><button onclick="runBankLedgerPreview()">Preview เดินบัญชีรายบัญชี</button><button data-admin-only="true" onclick="requestBankLedgerImport()">ขอ Import หลัง approval</button></div>
        </div>
        <div id="bankLedgerSummary" style="margin-top:12px"></div>
      </div>
    </section>

    <section id="section-employee-audit" class="sections menu-section" hidden>
      <div class="card">
        <h3>ออดิตยอดพนักงาน</h3>
        <div class="mini">ใช้ตัวกรองบริษัท/กลุ่ม/ฝากถอน/วันที่ด้านซ้ายร่วมกัน · ข้อ 1 เทียบสลิปกับ ledger, ข้อ 2 สรุปยอดรายวันต่อพนักงาน, ข้อ 3 จับสลิปซ้ำข้ามบอทหรือบริษัท</div>
        <div class="reconcile-controls">
          <label class="reconcile-step"><span>เลขบัญชี/Account key สำหรับข้อ 1</span><input id="auditAccountKey" placeholder="เว้นว่างเพื่อดูทั้ง scope" /></label>
          <label class="reconcile-step"><span>Threshold variance</span><input id="auditVarianceThreshold" type="number" step="0.01" value="100" /></label>
          <div class="reconcile-actions"><button onclick="runEmployeeAudit()">โหลดออดิต 1-2-3</button></div>
        </div>
        <div id="employeeAuditSummary" style="margin-top:12px"><div class="muted">กดโหลดออดิตเพื่อเริ่มตรวจ</div></div>
      </div>
      <div class="card">
        <h3>ผลลัพธ์ออดิต</h3>
        <div id="employeeAuditDetails"><div class="muted">ยังไม่มีข้อมูล</div></div>
      </div>
    </section>

    <section id="section-reconcile" class="sections menu-section" hidden>
      <div class="card"><h3>เทียบ Excel หลังบ้าน</h3><div class="mini">ขั้นตอน: เลือกบริษัท → เลือกยอดฝาก/ถอน → เลือกวันที่เทียบ → อัปโหลด Excel ของบริษัทนั้น ระบบจะเทียบเฉพาะ scope ที่เลือก ไม่ปนบริษัทอื่น</div><div class="reconcile-controls"><label class="reconcile-step"><span>1 เลือกบริษัทก่อน</span><select id="reconcileCompanyFilter" title="เลือกบริษัทสำหรับเทียบหลังบ้าน"></select></label><label class="reconcile-step"><span>2 เลือกยอดฝาก/ถอน</span><select id="reconcileFlowFilter" title="เลือกยอดฝากหรือถอนสำหรับไฟล์หลังบ้าน"><option value="">เลือกยอดฝาก/ถอน</option><option value="withdraw">ยอดถอน</option><option value="deposit">ยอดฝาก/เติมมือ</option></select></label><label class="reconcile-step"><span>3 เลือกวันที่เทียบ</span><input id="reconcileDateScope" type="date" title="เลือกวันที่ของไฟล์หลังบ้าน ถ้าเว้นว่างจะใช้ช่วงวันที่หลักของ dashboard" /></label><label class="reconcile-step reconcile-upload"><span>4 อัปโหลด Excel ของบริษัทนี้</span><input id="backendExcelFile" type="file" accept=".xlsx,.xlsm" /></label><label class="reconcile-step reconcile-upload" data-admin-only="true"><span>หรือ path ไฟล์บน server</span><input id="backendExcel" placeholder="path หรือเว้นว่างเพื่อใช้ไฟล์ล่าสุด" /></label><div id="reconcileScopePreview" class="reconcile-preview">ไฟล์นี้จะถูกเทียบเฉพาะบริษัทและฝาก/ถอนที่เลือก และวันที่เลือก</div><div class="reconcile-actions"><button onclick="runReconcile()">เทียบยอดตามบริษัท/ฝากถอน/วันที่เลือก</button></div></div><div id="reconcile"></div></div>
      <div class="card"><h3>เทียบ 3 ฝั่ง: หลังบ้าน + สลิป + รายการเดินบัญชี</h3><div class="mini">ใช้เมื่อไฟล์หลังบ้านและสลิปมีแค่ยอด/เวลา แล้วอัปโหลดรายการเดินบัญชีมาเป็นตัวกลาง เทียบด้วยยอด+เวลา+วันที่ใน scope เดียวกัน · รองรับ Excel และ CSV จาก True Wallet Dashboard · รายการเดินบัญชี: โอนออก=ถอน, รับเงินคืน/เงินเข้า=ฝาก</div><div class="reconcile-controls"><label class="reconcile-step reconcile-upload"><span>รายการเดินบัญชี Excel/CSV</span><input id="statementExcelFile" type="file" accept=".xlsx,.xlsm,.csv" /></label><label class="reconcile-step reconcile-upload" data-admin-only="true"><span>หรือ path รายการเดินบัญชีบน server</span><input id="statementExcel" placeholder="statement_path หรือไฟล์รายการเดินบัญชี" /></label><div class="reconcile-actions"><button onclick="runStatementReconcile()">เทียบ 3 ฝั่งตามบริษัท/ฝากถอน/วันที่เลือก</button></div></div><div id="statementReconcile"></div></div>
      <div class="card"><h3>สถานะข้อมูล</h3><div class="muted">ยอดรวมทุกแผงนับเฉพาะ success ที่ไม่ซ้ำ ยกเว้นการ์ดสลิปซ้ำที่แสดงเพื่อเช็กซ้ำโดยเฉพาะ</div></div>
    </section>
    <section id="section-duplicates" class="sections menu-section" hidden>
      <div class="card"><h3>คู่สลิปซ้ำ</h3><div class="mini">แสดงใบที่ถูกจับว่าซ้ำ และซ้ำกับใบไหนเพื่อให้ตรวจย้อนหลังได้</div><div id="duplicatePairs"></div></div>
    </section>
    <section id="section-bank-review" class="sections menu-section" hidden>
      <div class="card"><h3>รีเช็คธนาคารต้นทาง</h3><div class="mini">รายการ success ที่ยังหาธนาคารต้นทางไม่ได้จริง ๆ ใช้ OpenAI รีเช็คจากรูปสลิปได้</div><div class="toolbar" style="margin:8px 0" data-admin-only="true"><button onclick="openaiBankRecheckAll(this)">OpenAI รีเช็คทั้งหมดในขอบเขตนี้</button></div><div id="sourceBankReview"></div></div>
    </section>
    <section id="section-recent" class="sections menu-section" hidden>
      <div class="card"><h3>Recent slips</h3><div id="recent"></div></div>
      <div class="card"><h3>Queue / Issues</h3><div id="issues"></div><h3>Provider usage</h3><div id="usage"></div></div>
    </section>
    <section id="section-pending" class="sections menu-section" hidden>
      <div class="card">
        <h3>รออนุมัติ (Two-person approval) / โหมดง่าย</h3>
        <div class="mini">ถ้ามี token เดียว ระบบจะใช้โหมดง่าย: ยังบันทึก pending/audit เหมือนเดิม แต่ admin สามารถอนุมัติ+ทำรายการได้ทันที</div>
        <div class="pending-toolbar">
          <label class="mini">สถานะ
            <select id="pendingStatusFilter" title="กรองสถานะคำขอ">
              <option value="pending">pending</option>
              <option value="approved">approved</option>
              <option value="executed">executed</option>
              <option value="rejected">rejected</option>
              <option value="cancelled">cancelled</option>
              <option value="expired">expired</option>
              <option value="all">ทั้งหมด</option>
            </select>
          </label>
          <button id="pendingRefreshBtn" type="button" onclick="loadPendingActions({{scrollTop:false}})">รีเฟรช</button>
          <span id="pendingMeta" class="mini muted">โหลด...</span>
        </div>
        <div id="pendingTableContainer"></div>
        <div class="mini">approve/reject ทำได้เฉพาะ role admin หรือ auditor และห้าม self-approve</div>
      </div>
    </section>
    </main>
    <div id="pendingToastHost" class="pending-toast-host" aria-live="polite"></div>
  </div>
  <div id="dashboardModal" class="modal-backdrop" role="dialog" aria-modal="true" aria-labelledby="dashboardModalTitle" aria-hidden="true">
    <div class="modal-card">
      <div id="dashboardModalTitle" class="modal-title">แจ้งเตือน</div>
      <div id="dashboardModalMessage" class="modal-message"></div>
      <textarea id="dashboardModalInput" class="modal-input" hidden></textarea>
      <div class="modal-actions">
        <button id="dashboardModalCancel" type="button" class="modal-cancel">ยกเลิก</button>
        <button id="dashboardModalOk" type="button" class="modal-primary">ตกลง</button>
      </div>
    </div>
  </div>
<script>
let transferorRows = [];
let currentSnapshot = null;
const SIDE_MENU_KEY = 'auditslip.sideMenuCollapsed';
const ADMIN_MODE_KEY = 'auditslip.adminMode';
const ACTION_HEADER = {{'X-Auditslip-Action':'dashboard'}};
let dashboardModalResolver = null;
function hdr() {{ return ACTION_HEADER; }}
function dashboardModalElements() {{
  return {{
    root: document.getElementById('dashboardModal'),
    title: document.getElementById('dashboardModalTitle'),
    message: document.getElementById('dashboardModalMessage'),
    input: document.getElementById('dashboardModalInput'),
    ok: document.getElementById('dashboardModalOk'),
    cancel: document.getElementById('dashboardModalCancel')
  }};
}}
function closeDashboardModal(result=false) {{
  const el = dashboardModalElements();
  if (el.root) {{ el.root.classList.remove('open'); el.root.setAttribute('aria-hidden', 'true'); }}
  const resolver = dashboardModalResolver;
  dashboardModalResolver = null;
  if (resolver) resolver(result);
}}
function showDashboardModal(options={{}}) {{
  return new Promise(resolve => {{
    const el = dashboardModalElements();
    if (!el.root) {{ resolve(options.input ? {{ok:true, value:''}} : true); return; }}
    dashboardModalResolver = resolve;
    if (el.title) el.title.textContent = options.title || 'แจ้งเตือน';
    if (el.message) el.message.textContent = options.message || '';
    const wantsInput = Boolean(options.input);
    if (el.input) {{
      el.input.hidden = !wantsInput;
      el.input.value = wantsInput ? String(options.inputValue ?? '') : '';
      el.input.placeholder = wantsInput ? String(options.inputPlaceholder || '') : '';
      el.input.onkeydown = (event) => {{
        if ((event.ctrlKey || event.metaKey) && event.key === 'Enter') {{
          event.preventDefault();
          closeDashboardModal({{ok:true, value:el.input.value || ''}});
        }}
      }};
    }}
    if (el.ok) {{
      el.ok.textContent = options.confirmText || 'ตกลง';
      el.ok.classList.toggle('danger', Boolean(options.danger));
      el.ok.onclick = () => closeDashboardModal(wantsInput ? {{ok:true, value:el.input ? (el.input.value || '') : ''}} : true);
    }}
    if (el.cancel) {{
      el.cancel.textContent = options.cancelText || 'ยกเลิก';
      el.cancel.style.display = options.showCancel === false ? 'none' : '';
      el.cancel.onclick = () => closeDashboardModal(wantsInput ? {{ok:false, value:null}} : false);
    }}
    el.root.classList.add('open');
    el.root.setAttribute('aria-hidden', 'false');
    if (wantsInput && el.input) el.input.focus();
    else if (el.ok) el.ok.focus();
  }});
}}
function dashboardNotify(message, title='แจ้งเตือน') {{ return showDashboardModal({{title, message, showCancel:false, confirmText:'ตกลง'}}); }}
function dashboardConfirm(message, title='ยืนยัน', danger=false) {{ return showDashboardModal({{title, message, showCancel:true, confirmText:'ยืนยัน', cancelText:'ยกเลิก', danger}}); }}
async function dashboardInput(message, title='กรอกข้อมูล', defaultValue='', options={{}}) {{
  const result = await showDashboardModal({{title, message, input:true, inputValue:defaultValue, inputPlaceholder:options.placeholder || '', showCancel:true, confirmText:options.confirmText || 'ตกลง', cancelText:options.cancelText || 'ยกเลิก', danger:Boolean(options.danger)}});
  if (!result || result.ok !== true) return null;
  return String(result.value || '');
}}
document.addEventListener('keydown', (event) => {{ if (event.key === 'Escape' && dashboardModalResolver) closeDashboardModal(false); }});
if ('scrollRestoration' in window.history) window.history.scrollRestoration = 'manual';
window.scrollTo(0, 0);
function safeStorageGet(key) {{ try {{ return window.localStorage.getItem(key); }} catch (e) {{ return null; }} }}
function safeStorageSet(key, value) {{ try {{ window.localStorage.setItem(key, value); }} catch (e) {{}} }}
function isMobileSideMenu() {{ return window.matchMedia('(max-width: 780px)').matches; }}
function applySideMenuState() {{
  const mobile = isMobileSideMenu();
  const collapsed = safeStorageGet(SIDE_MENU_KEY) === '1';
  if (mobile) {{
    document.body.classList.remove('side-collapsed');
  }} else {{
    document.body.classList.remove('side-open');
    document.body.classList.toggle('side-collapsed', collapsed);
  }}
  const open = document.body.classList.contains('side-open');
  const btn = document.getElementById('sideMenuToggle');
  if (btn) {{
    btn.textContent = mobile ? (open ? '×' : '☰') : (collapsed ? '☰' : '×');
    btn.setAttribute('aria-expanded', String(mobile ? open : !collapsed));
  }}
}}
function openSideMenu() {{
  document.body.classList.add('side-open');
  applySideMenuState();
}}
function closeSideMenu() {{
  document.body.classList.remove('side-open');
  applySideMenuState();
}}
function toggleSideMenu() {{
  if (isMobileSideMenu()) {{
    if (document.body.classList.contains('side-open')) closeSideMenu(); else openSideMenu();
    return;
  }}
  const next = document.body.classList.contains('side-collapsed') ? '0' : '1';
  safeStorageSet(SIDE_MENU_KEY, next);
  applySideMenuState();
}}
window.addEventListener('resize', applySideMenuState);
function closeSideMenuIfMobile() {{
  if (isMobileSideMenu()) closeSideMenu();
}}
function scrollDashboardTop(smooth=true) {{
  const target = document.getElementById('section-metrics') || document.body;
  if (target && target.scrollIntoView) target.scrollIntoView({{behavior: smooth ? 'smooth' : 'auto', block:'start'}});
  window.scrollTo({{top:0, left:0, behavior: smooth ? 'smooth' : 'auto'}});
}}
function scrollElementIntoView(id, smooth=true) {{
  const target = document.getElementById(id);
  if (target && target.scrollIntoView) target.scrollIntoView({{behavior: smooth ? 'smooth' : 'auto', block:'start'}});
}}
function refreshDashboardHome() {{
  showMenuSection('section-operator-home', {{scroll:false, persist:false}});
  return load({{home:true, scrollTop:true, smooth:false}});
}}
function setActiveMenu(target) {{
  document.querySelectorAll('[data-menu-target]').forEach(btn => btn.classList.toggle('active', String(btn.dataset.menuTarget || '') === String(target || 'all')));
}}
function showAllMenuSections(options={{}}) {{
  document.querySelectorAll('.menu-section').forEach(section => section.hidden = false);
  setActiveMenu('all');
  closeSideMenuIfMobile();
  if (options && options.scrollTop) scrollDashboardTop();
}}
function showMenuSection(target, options={{}}) {{
  if (!target || target === 'all') return showAllMenuSections(options);
  document.querySelectorAll('.menu-section').forEach(section => {{
    section.hidden = !(section.dataset.alwaysVisible === 'true' || section.id === target);
  }});
  setActiveMenu(target);
  const section = document.getElementById(target);
  if (options.scroll !== false && section) section.scrollIntoView({{behavior:'smooth', block:'start'}});
  if (target === 'section-account-ledger') populateLedgerAccounts();
  closeSideMenuIfMobile();
}}
applySideMenuState();
applyAdminMode();
let initialMenuApplied = false;
function splitChatValue(value) {{
  const raw = String(value || '');
  const parts = raw.split('|');
  if (parts.length < 2) return {{bot_key:'', chat_id:raw, flow_type:'all'}};
  return {{bot_key: parts[0] || '', chat_id: parts[1] || '', flow_type: parts[2] || 'all'}};
}}
function selectedChatParts() {{ return splitChatValue(document.getElementById('chat').value || ''); }}
function selectedBotKey() {{ return document.getElementById('botFilter').value || selectedChatParts().bot_key || ''; }}
function query(params={{}}) {{
  const p = new URLSearchParams();
  Object.entries(params).forEach(([k,v]) => {{ if (v !== undefined && v !== null && String(v) !== '') p.set(k, v); }});
  const qs = p.toString();
  return qs ? '?' + qs : '';
}}
function postHeaders(extra={{}}) {{ return {{...ACTION_HEADER, 'Content-Type':'application/json', ...extra}}; }}
async function adminLogin() {{
  const status = document.getElementById('adminLoginStatus');
  const usernameEl = document.getElementById('adminUsername');
  const passwordEl = document.getElementById('adminPassword');
  const username = usernameEl ? usernameEl.value : '';
  const password = passwordEl ? passwordEl.value : '';
  if (status) status.textContent = 'กำลัง login admin...';
  const res = await fetch('/api/login', {{method:'POST', headers:postHeaders(), body: JSON.stringify({{username, password}})}});
  let data = {{}};
  try {{ data = await res.json(); }} catch (e) {{ data = {{error:'อ่านผล login ไม่ได้'}}; }}
  if (!res.ok || !data.ok) {{
    if (status) status.textContent = 'Login ไม่สำเร็จ';
    return await dashboardNotify(data.error || 'Login ไม่สำเร็จ');
  }}
  if (passwordEl) passwordEl.value = '';
  if (status) status.textContent = 'Login แล้ว · สิทธิ '+(data.role || 'admin');
  await loadPendingActions({{scrollTop:false}});
  refreshPendingBadge();
  await load({{scrollTop:false}});
}}
async function adminLogout() {{
  const status = document.getElementById('adminLoginStatus');
  const res = await fetch('/api/logout', {{method:'POST', headers:ACTION_HEADER}});
  if (status) status.textContent = res.ok ? 'Logout แล้ว · ยังดู dashboard ได้ตามปกติ' : 'Logout ไม่สำเร็จ';
  await loadPendingActions({{scrollTop:false}});
}}
function scopeName(value) {{ return value === 'open' ? 'รอบเปิด' : (value === 'today' ? 'วันนี้' : (value === 'all' ? 'ทั้งหมด' : value || '-')); }}
function isSingleChatSelected() {{
  const parts = selectedChatParts();
  return Boolean(parts.chat_id && parts.chat_id !== '__all__' && selectedBotKey() !== '__all__');
}}
function closeOpenPeriodGuard() {{
  const ok = isSingleChatSelected();
  const btn = document.getElementById('closePeriodButton');
  const hint = document.getElementById('closeGuardHint');
  if (btn) btn.disabled = !ok;
  if (hint) hint.textContent = ok ? 'พร้อมปิดรอบเฉพาะกลุ่มที่เลือก' : 'ต้องเลือกบริษัทและกลุ่มเดียวก่อนปิดรอบ';
  return ok;
}}
function updateTopStatus(data, selectedBot, activeFlow) {{
  const chip = document.getElementById('activeScopeChip');
  const updated = document.getElementById('lastUpdatedLabel');
  const botRow = (data.telegram_bots || []).find(b => b.bot_key === selectedBot) || {{}};
  const company = selectedBot === '__all__' ? 'ทุกบริษัท' : (botRow.company_name || selectedBot || '-');
  const scope = data.scope_label || scopeName(data.scope || document.getElementById('scope').value || 'today');
  if (chip) chip.textContent = company + ' · ' + flowName(data.flow_type || activeFlow) + ' · ' + scope;
  const now = new Date();
  if (updated) updated.textContent = 'อัปเดตล่าสุด ' + now.toLocaleTimeString('th-TH', {{hour:'2-digit', minute:'2-digit', second:'2-digit'}});
}}
function applyAdminMode() {{
  const enabled = safeStorageGet(ADMIN_MODE_KEY) === '1';
  document.body.classList.toggle('admin-mode', enabled);
  const btn = document.getElementById('adminModeToggle');
  if (btn) btn.textContent = enabled ? 'ปิด Admin' : 'ตั้งค่า/Admin';
}}
function applyAdminVisibility() {{ applyAdminMode(); }}
function toggleAdminMode() {{
  const enabled = safeStorageGet(ADMIN_MODE_KEY) === '1';
  safeStorageSet(ADMIN_MODE_KEY, enabled ? '0' : '1');
  applyAdminMode();
}}
function money(n) {{ return Number(n || 0).toLocaleString('th-TH', {{minimumFractionDigits:2, maximumFractionDigits:2}}); }}
function formatIsoTime(value) {{
  const raw = String(value ?? '').trim();
  if (!raw) return '-';
  const d = new Date(raw);
  if (!Number.isNaN(d.getTime())) return d.toLocaleString('th-TH', {{year:'2-digit', month:'2-digit', day:'2-digit', hour:'2-digit', minute:'2-digit', second:'2-digit', hour12:false}});
  return raw;
}}
function limitMoney(n) {{ return Number(n || 0) > 0 ? money(n) : '-'; }}
function nameWithCompany(r) {{
  const name = r.name || r.display_name || '-';
  const company = r.company_name || '';
  return company ? (name + ' (' + company + ')') : name;
}}
function esc(s) {{ return String(s ?? '').replace(/[&<>"']/g, c => ({{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}}[c])); }}
function table(rows, cols) {{
  if (!rows || !rows.length) return '<div class="muted">ไม่มีข้อมูล</div>';
  return '<div class="responsive-table"><table><thead><tr>'+cols.map(c=>'<th>'+esc(c[0])+'</th>').join('')+'</tr></thead><tbody>'+
    rows.map(r=>'<tr>'+cols.map(c=>'<td>'+esc(typeof c[1]==='function'?c[1](r):r[c[1]])+'</td>').join('')+'</tr>').join('')+'</tbody></table></div>';
}}
function enhanceResponsiveTables(root=document) {{
  (root || document).querySelectorAll('.responsive-table table').forEach(tbl => {{
    const headers = Array.from(tbl.querySelectorAll('thead th')).map(th => (th.textContent || '').trim());
    tbl.querySelectorAll('tbody tr').forEach(tr => {{
      Array.from(tr.children).forEach((td, i) => {{ if (headers[i]) td.setAttribute('data-label', headers[i]); }});
    }});
  }});
}}
function aggregateTable(rows) {{ return table(rows, [['ชื่อ','name'], ['สลิป','count'], ['ยอด', r => money(r.amount)]]); }}
function sourceBankTable(rows) {{ return aggregateTable(rows); }}
function dailyAccountLimitTable(rows) {{
  if (!rows || !rows.length) return '<div class="muted">ไม่มีข้อมูล</div>';
  return '<div class="responsive-table"><table><thead><tr><th>วันที่</th><th>บัญชีผู้โอน (บริษัท)</th><th>เลขบัญชี</th><th>สลิป</th><th>ยอดวันนี้</th><th>วงเงิน/วัน</th><th>เหลือ/เกิน</th></tr></thead><tbody>'+
    rows.map(r => '<tr>'+[
      '<td>'+esc(r.date || r.date_key || '-')+'</td>',
      '<td>'+esc(nameWithCompany(r))+'</td>',
      '<td>'+esc(r.account || '-')+'</td>',
      '<td>'+esc(r.count || 0)+'</td>',
      '<td>'+money(r.amount)+'</td>',
      '<td>'+limitMoney(r.daily_limit ?? r.limit_amount)+'</td>',
      '<td class="'+(r.over_limit?'bad':'')+'">'+(Number((r.daily_limit ?? r.limit_amount)||0)>0 ? money(r.remaining_amount) : '-')+'</td>'
    ].join('')+'</tr>').join('')+'</tbody></table></div>';
}}
function transferorLimitTable(rows) {{
  transferorRows = rows || [];
  if (!transferorRows.length) return '<div class="muted">ไม่มีข้อมูล</div>';
  return '<div class="responsive-table"><table><thead><tr><th>วันที่</th><th>บัญชีผู้โอน (บริษัท)</th><th>บัญชี</th><th>สลิป</th><th>ยอดวันนี้</th><th>วงเงิน/วัน</th><th>เหลือ/เกินวันนี้</th><th class="action-col">จัดการ</th></tr></thead><tbody>'+
    transferorRows.map((r, i) => '<tr>'+[
      '<td>'+esc(r.date || r.date_key || '-')+'</td>',
      '<td>'+esc(nameWithCompany(r))+'</td>',
      '<td>'+esc(r.account || '-')+'</td>',
      '<td>'+esc(r.count || 0)+'</td>',
      '<td class="money-cell">'+money(r.amount)+'</td>',
      '<td class="money-cell">'+limitMoney(r.daily_limit ?? r.limit_amount)+'</td>',
      '<td class="money-cell '+(r.over_limit?'bad':'')+'">'+(Number((r.daily_limit ?? r.limit_amount)||0)>0 ? money(r.remaining_amount) : '-')+'</td>',
      '<td class="action-cell"><button onclick="pickLimitIndex('+i+')">ตั้งวงเงิน</button></td>'
    ].join('')+'</tr>').join('')+'</tbody></table></div>';
}}
function dateTable(rows) {{ return table(rows, [['วันที่','date'], ['สลิป','count'], ['ยอด', r => money(r.amount)]]); }}
function renderDailyFlowChart(rows) {{
  if (!rows || !rows.length) return '<div class="muted">ยังไม่มีข้อมูลรายวันใน scope นี้</div>';
  const maxAmount = Math.max(1, ...rows.map(r => Math.max(Number(r.withdraw_amount || 0), Number(r.deposit_amount || 0), Number(r.other_amount || 0))));
  const bar = (cls, amount) => '<div class="flow-bar bar '+cls+'" style="width:'+Math.max(2, Math.round((Number(amount || 0) / maxAmount) * 100))+'%"></div>';
  return '<div class="flow-legend"><span><i class="legend-dot withdraw"></i>ถอน</span><span><i class="legend-dot deposit"></i>ฝาก/เติมมือ</span></div>'+
    '<div class="flow-chart">'+rows.map(r => {{
      const other = Number(r.other_amount || 0) ? '<div class="mini">อื่นๆ '+money(r.other_amount || 0)+'</div>' : '';
      return '<div class="flow-chart-row">'
        + '<div><b>'+esc(r.date || r.date_key || '-')+'</b><div class="mini">รวม '+esc(r.total_count || 0)+' สลิป</div></div>'
        + '<div class="flow-bars">'+bar('withdraw', r.withdraw_amount || 0)+bar('deposit', r.deposit_amount || 0)+'</div>'
        + '<div class="flow-chart-values"><div>ถอน '+money(r.withdraw_amount || 0)+' / '+esc(r.withdraw_count || 0)+'</div><div>ฝาก '+money(r.deposit_amount || 0)+' / '+esc(r.deposit_count || 0)+'</div>'+other+'</div>'
        + '</div>';
    }}).join('')+'</div>';
}}
function renderWithdrawLimitUsageChart(rows, totals={{}}) {{
  const items = rows || [];
  const withdrawn = Number((totals && totals.withdraw_limit_amount) || 0);
  const capacity = Number((totals && totals.withdraw_limit_capacity_amount) || 0);
  const remaining = Number((totals && totals.withdraw_limit_remaining_amount) || 0);
  const percent = Number((totals && totals.withdraw_limit_usage_percent) || 0);
  const over = Number((totals && totals.withdraw_limit_over_amount) || 0);
  const summary = '<div class="limit-usage-summary">'
    + '<div class="limit-usage-stat"><div class="label">ถอนรวม</div><div class="value good">'+money(withdrawn)+'</div></div>'
    + '<div class="limit-usage-stat"><div class="label">วงเงินรวม</div><div class="value">'+(capacity ? money(capacity) : '-')+'</div><div class="mini">'+esc((totals && totals.withdraw_limit_account_day_count) || 0)+' บัญชี/วัน</div></div>'
    + '<div class="limit-usage-stat"><div class="label">เหลือ/เกิน</div><div class="value '+(remaining < 0 || over > 0 ? 'bad' : 'good')+'">'+(capacity ? money(remaining) : '-')+'</div><div class="mini">'+(capacity ? percent.toFixed(1)+'%' : 'ยังไม่มีวงเงิน')+'</div></div>'
    + '</div>';
  if (!items.length) return summary + '<div class="muted">ยังไม่มีข้อมูลวงเงินถอนใน scope นี้</div>';
  return summary + '<div class="limit-usage-chart">'+items.map(r => {{
    const rowPercent = Number(r.usage_percent || 0);
    const width = Math.max(2, Math.min(100, Math.round(rowPercent)));
    const isOver = Boolean(r.over_limit);
    const remainingText = Number(r.limit_amount || 0) ? money(r.remaining_amount || 0) : '-';
    const noLimit = Number(r.no_limit_account_day_count || 0) ? ' · ไม่มีวงเงิน '+esc(r.no_limit_account_day_count)+' บัญชี/วัน' : '';
    return '<div class="limit-usage-row '+(isOver ? 'over' : '')+'">'
      + '<div><b>'+esc(r.company_name || r.bot_key || '-')+'</b><div class="mini">'+esc(r.account_count || 0)+' บัญชี · '+esc(r.withdraw_count || 0)+' สลิป</div></div>'
      + '<div><div class="limit-usage-track"><div class="limit-usage-fill '+(isOver ? 'over' : '')+'" style="width:'+width+'%"></div></div><div class="mini">ใช้ไป '+rowPercent.toFixed(1)+'%'+noLimit+'</div></div>'
      + '<div class="limit-usage-values"><div>ถอน '+money(r.withdraw_amount || 0)+'</div><div>วงเงิน '+(Number(r.limit_amount || 0) ? money(r.limit_amount || 0) : '-')+'</div><div class="'+(isOver ? 'bad' : 'good')+'">เหลือ/เกิน '+remainingText+'</div></div>'
      + '</div>';
  }}).join('')+'</div>';
}}
function flowName(value) {{ return value === 'deposit' ? 'ฝาก/เติมมือ' : (value === 'withdraw' ? 'ถอน' : (value === 'other' ? 'อื่นๆ' : 'รวมทุกกลุ่ม')); }}
function flowChipName(value) {{ return value === 'deposit' ? 'กลุ่มฝาก/เติมมือ' : (value === 'withdraw' ? 'กลุ่มถอน' : (value === 'other' ? 'กลุ่มอื่นๆ' : 'รวมทุกกลุ่ม')); }}
function renderCompanyAccountDaily(rows) {{
  if (!rows || !rows.length) return '<div class="muted">ไม่มีรายการรายบัญชีตามวันที่</div>';
  return '<div class="responsive-table"><table><thead><tr><th>บริษัท</th><th>วันที่</th><th>กลุ่ม</th><th>บทบาทบัญชี</th><th>ธนาคาร</th><th>เลขบัญชี</th><th>สลิป</th><th>ยอด</th><th>กลุ่ม Telegram</th><th class="action-col">ดูรายการ</th></tr></thead><tbody>'+
    rows.map(r => {{
      const accountAliases = (r.account_aliases || []).filter(a => a && a !== r.account);
      const accountCell = esc(r.account || '-') + (accountAliases.length ? '<div class="mini">รวมรูปแบบ: '+esc(accountAliases.join(', '))+'</div>' : '');
      return '<tr>'+[
        '<td>'+esc(r.company_name || r.bot_key || '-')+'</td>',
        '<td>'+esc(r.date || r.date_key || '-')+'</td>',
        '<td>['+esc(flowChipName(r.flow_type))+']</td>',
        '<td>'+esc(r.account_role || '-')+'</td>',
        '<td>'+esc(r.bank || '-')+'</td>',
        '<td>'+accountCell+'</td>',
        '<td>'+esc(r.count || 0)+'</td>',
        '<td>'+money(r.amount)+'</td>',
        '<td>'+esc((r.chat_titles || []).join(', '))+'</td>',
        '<td class="action-cell"><button type="button" data-account-search="'+esc(String(r.account || ''))+'">ดูสลิปบัญชีนี้</button></td>'
      ].join('')+'</tr>';
    }}).join('')+'</tbody></table></div>';
}}
function pickAccountSearch(queryText, crossCompany=false) {{
  const input = document.getElementById('slipSearch');
  if (!input || !queryText) return;
  input.value = queryText;
  window.__accountSearchMode = crossCompany ? 'cross' : 'scoped';
  if (crossCompany) {{
    const flow = document.getElementById('flowFilter').value || 'all';
    document.getElementById('botFilter').value = '__all__';
    const exportCompany = document.getElementById('exportCompanyFilter');
    if (exportCompany) exportCompany.value = '__all__';
    document.getElementById('chat').value = '__all__||' + flow;
  }}
  showMenuSection('section-company-accounts', {{scroll:false}});
  load({{scrollTarget:'accountSlipSearch', smooth:true}});
}}
function wireAccountSearchButtons() {{
  document.querySelectorAll('[data-account-search]').forEach(btn => {{
    btn.onclick = function() {{ pickAccountSearch(this.dataset.accountSearch || ''); }};
  }});
  document.querySelectorAll('[data-cross-account-search]').forEach(btn => {{
    btn.onclick = function() {{ pickAccountSearch(this.dataset.crossAccountSearch || '', true); }};
  }});
}}
function renderAccountSlipSearch(result) {{
  const queryText = result && result.query ? String(result.query) : '';
  if (!queryText) return '<div class="muted">กรอกเลขบัญชี/ชื่อ/ธนาคารในช่องค้นหา หรือกด “ดูสลิปบัญชีนี้” จากตารางด้านบน</div>';
  const rows = (result && result.rows) || [];
  const summary = '<div class="good"><b>ผลค้นหา: '+esc(queryText)+'</b></div><div class="mini">พบ '+esc((result && result.count) || 0)+' สลิป · ยอดรวม '+money((result && result.amount) || 0)+(result && result.truncated ? ' · แสดงเฉพาะรายการแรก ๆ' : '')+'</div>';
  if (!rows.length) return summary + '<div class="muted">ไม่พบสลิปของบัญชีนี้ในบริษัท/วัน/ฝากถอนที่เลือก</div>';
  return summary + '<div class="slip-cards">'+rows.map(r => {{
    const image = thumb(r.image_url, 'account slip');
    const banks = [r.from_bank, r.to_bank].filter(Boolean).join(' → ') || r.issuer_bank || '-';
    const accounts = [r.from_account, r.to_account].filter(Boolean).join(' → ') || '-';
    const names = [r.transferor_name, r.recipient_name].filter(Boolean).join(' → ') || r.sender_name || '-';
    return '<div class="slip-card">'+image+'<div class="slip-body"><div class="top"><b>'+esc(r.company_name || r.bot_key || '-')+'</b><span class="pill">'+esc(r.flow_label || flowName(r.flow_type))+'</span></div>'
      + '<div class="mini">'+esc(r.slip_date_text || ((r.date || r.date_key || '')+' '+(r.slip_time || '')))+' · msg '+esc(r.message_id || '-')+' · '+esc(r.matched_label || '-')+'</div>'
      + '<div>'+esc(names)+'</div><div class="mini">'+esc(banks)+' · '+esc(accounts)+'</div>'
      + '<div>ยอด <b>'+money(r.amount || 0)+'</b></div>'
      + '<div class="mini">ref '+esc(r.reference || r.reference_no || '-')+'</div>'
      + '</div></div>';
  }}).join('')+'</div>';
}}
function renderCrossCompanyAccountSlipSearch(result) {{
  const queryText = result && result.query ? String(result.query) : '';
  if (!queryText) return '<div class="muted">กรอกเลขบัญชี/ชื่อ/ธนาคาร หรือกด “ดูสลิปข้ามบริษัท” เพื่อค้นทุกบริษัทในวันที่/ฝากถอนที่เลือก</div>';
  const rows = (result && result.rows) || [];
  const companies = (result && result.companies) || [];
  const companyLine = companies.length ? '<div class="mini">บริษัทที่พบ: '+companies.map(c => esc(c.company_name || c.bot_key || '-')+' '+money(c.amount || 0)+' / '+esc(c.count || 0)+' สลิป').join(' · ')+'</div>' : '';
  const summary = '<div class="good"><b>ค้นข้ามบริษัท: '+esc(queryText)+'</b></div><div class="mini">พบ '+esc((result && result.count) || 0)+' สลิป · '+esc((result && result.company_count) || 0)+' บริษัท · ยอดรวม '+money((result && result.amount) || 0)+(result && result.truncated ? ' · แสดงเฉพาะรายการแรก ๆ' : '')+'</div>'+companyLine;
  const exportLink = Number((result && result.count) || 0) > 0 ? '<div class="toolbar" style="margin:8px 0"><a class="button" target="exportDownloadFrame" href="'+esc(buildCrossCompanyAccountExcelUrl(queryText))+'">ส่งออก Excel ข้ามบริษัท</a></div>' : '';
  if (!rows.length) return summary + exportLink + '<div class="muted">ไม่พบสลิปของบัญชีนี้ข้ามบริษัทในวันที่/ฝากถอนที่เลือก</div>';
  return summary + exportLink + '<div class="slip-cards">'+rows.map(r => {{
    const image = thumb(r.image_url, 'cross company account slip');
    const banks = [r.from_bank, r.to_bank].filter(Boolean).join(' → ') || r.issuer_bank || '-';
    const accounts = [r.from_account, r.to_account].filter(Boolean).join(' → ') || '-';
    const names = [r.transferor_name, r.recipient_name].filter(Boolean).join(' → ') || r.sender_name || '-';
    return '<div class="slip-card">'+image+'<div class="slip-body"><div class="top"><b>'+esc(r.company_name || r.bot_key || '-')+'</b><span class="pill">'+esc(r.flow_label || flowName(r.flow_type))+'</span></div>'
      + '<div class="mini">'+esc(r.slip_date_text || ((r.date || r.date_key || '')+' '+(r.slip_time || '')))+' · msg '+esc(r.message_id || '-')+' · '+esc(r.matched_label || '-')+'</div>'
      + '<div>'+esc(names)+'</div><div class="mini">'+esc(banks)+' · '+esc(accounts)+'</div>'
      + '<div>ยอด <b>'+money(r.amount || 0)+'</b></div>'
      + '<div class="mini">ref '+esc(r.reference || r.reference_no || '-')+'</div>'
      + '</div></div>';
  }}).join('')+'</div>';
}}
function renderAccountCrossCompany(rows) {{
  if (!rows || !rows.length) return '<div class="muted">ยังไม่พบบัญชีเดียวกันข้ามบริษัทในช่วงนี้</div>';
  const dayRows = (days) => (days || []).map(d => '<div class="cross-company-day-row"><span>'+esc(d.date || d.date_key || '-')+'</span><span class="day-amount">'+money(d.amount || 0)+' / '+esc(d.count || 0)+' สลิป</span></div>').join('') || '<div class="muted">ไม่มีรายวัน</div>';
  const companyCards = (companies) => (companies || []).map(c => '<div class="cross-company-company">'
    + '<div class="cross-company-company-head"><div class="cross-company-company-name">'+esc(c.company_name || c.bot_key || '-')+'</div><div class="cross-company-total">'+money(c.amount || 0)+' / '+esc(c.count || 0)+' สลิป</div></div>'
    + '<div class="cross-company-days">'+dayRows(c.days)+'</div>'
    + '</div>').join('') || '<div class="muted">ไม่มีบริษัท</div>';
  return '<div class="cross-company-list">'+rows.map(r => '<article class="cross-company-card">'
    + '<div class="cross-company-head"><div><div class="label">บัญชี</div><div class="cross-company-account">'+esc(r.account || '-')+'</div><div class="cross-company-bank">ธนาคาร '+esc(r.bank || '-')+'</div></div><div class="cross-company-chips">'+esc((r.flow_labels || []).map(v => '['+v+']').join(' '))+'</div></div>'
    + '<div class="cross-company-summary"><div class="cross-company-stat"><div class="label">สลิปรวมบัญชีนี้</div><div class="value">'+esc(r.total_count || 0)+'</div></div><div class="cross-company-stat"><div class="label">รวมบัญชีนี้</div><div class="value good">'+money(r.total_amount || 0)+'</div></div></div>'
    + '<div class="toolbar"><button type="button" data-cross-account-search="'+esc(String(r.account || ''))+'">ดูสลิปข้ามบริษัท</button></div>'
    + '<div><div class="cross-company-section-title">แยกตามบริษัท · ไปอยู่บริษัทไหน / ยอดเท่าไหร่</div><div class="cross-company-companies">'+companyCards(r.companies)+'</div></div>'
    + '<div class="cross-company-overall-days"><div class="cross-company-section-title">ยอดรายวันรวม</div><div class="cross-company-days">'+dayRows(r.days)+'</div></div>'
    + '</article>').join('')+'</div>';
}}
function renderCompanyOverview(rows) {{
  if (!rows || !rows.length) return '<div class="muted">ยังไม่มีข้อมูลบริษัท</div>';
  return '<div class="responsive-table"><table><thead><tr><th>บริษัท</th><th>ยอดรวมเปิด</th><th>สลิปรวม</th><th>ยอดถอน</th><th>สลิปถอน</th><th>ยอดฝาก</th><th>สลิปฝาก</th><th>ซ้ำ</th><th>คิว</th><th></th></tr></thead><tbody>'+
    rows.map(r => '<tr>'+[
      '<td>'+esc(r.company_name || r.bot_key || '-')+'</td>',
      '<td>'+money(r.open_amount)+'</td>',
      '<td>'+esc(r.open_count || 0)+'</td>',
      '<td>'+money(r.withdraw_open_amount || 0)+'</td>',
      '<td>'+esc(r.withdraw_open_count || 0)+'</td>',
      '<td>'+money(r.deposit_open_amount || 0)+'</td>',
      '<td>'+esc(r.deposit_open_count || 0)+'</td>',
      '<td>'+esc(r.duplicate_count || 0)+'</td>',
      '<td>'+esc((r.queued||0)+' / '+(r.processing||0)+' / '+(r.failed||0))+'</td>',
      '<td><button data-company-bot-key="'+esc(String(r.bot_key || 'default'))+'">เปิด</button></td>'
    ].join('')+'</tr>').join('')+'</tbody></table></div>';
}}
function renderSideCompanies(rows) {{
  if (!rows || !rows.length) return '<div class="muted">ยังไม่มีข้อมูลบริษัท</div>';
  return rows.map(r => '<button type="button" class="side-company" data-company-bot-key="'+esc(String(r.bot_key || 'default'))+'"><b>'+esc(r.company_name || r.bot_key || '-')+'</b><div class="mini">ถอน '+money(r.withdraw_open_amount || 0)+' / '+esc(r.withdraw_open_count || 0)+' สลิป · ฝาก '+money(r.deposit_open_amount || 0)+' / '+esc(r.deposit_open_count || 0)+' สลิป · ซ้ำ '+esc(r.duplicate_count || 0)+'</div></button>').join('');
}}
function renderTWalletSummary(t) {{
  if (!t || t.enabled === false) return '<div class="muted">ยังไม่ได้ตั้งค่า True Wallet</div>';
  if (!t.ok) return '<div class="bad"><b>True Wallet ดึงข้อมูลไม่ได้</b></div><div class="mini">'+esc(t.error || 'offline')+'</div>';
  const last = t.last_receive || {{}};
  const lastLine = last.amount ? 'ล่าสุด '+money(last.amount)+' · '+esc(last.sender_mobile || '-')+' → '+esc(last.receiver_mobile || '-')+' · '+esc(last.received_time || '-') : 'ยังไม่มีรายการล่าสุด';
  return '<div class="operator-home-grid">'
    + '<div class="operator-stat"><div class="mini">ยอดรับวันนี้</div><b>'+money(t.today_total || 0)+'</b><div class="mini">'+esc(t.today_count || 0)+' รายการ · '+esc(t.today_date || '')+'</div></div>'
    + '<div class="operator-stat"><div class="mini">ยอดคงเหลือ</div><b>'+money(t.balance_amount || 0)+'</b><div class="mini">เบอร์ '+esc(t.mobile_masked || '-')+' · '+esc(t.balance_updated_at || '')+'</div></div>'
    + '<div class="operator-stat"><div class="mini">รายการทั้งหมด</div><b>'+esc(t.tx_count || 0)+'</b><div class="mini">'+esc(lastLine)+'</div></div>'
    + '</div>';
}}
function renderOperatorHome(rows) {{
  if (!rows || !rows.length) return '<div class="muted">ยังไม่มีข้อมูลบริษัท</div>';
  return '<div class="operator-home-grid">'+rows.map(r => {{
    const issueCount = Number(r.failed || 0) + Number(r.processing || 0) + Number(r.queued || 0);
    const dupe = Number(r.duplicate_count || 0);
    const warn = issueCount || dupe;
    return '<button type="button" class="operator-card" data-company-bot-key="'+esc(String(r.bot_key || 'default'))+'">'
      + '<div class="head"><b>'+esc(r.company_name || r.bot_key || '-')+'</b><span class="pill '+(warn?'warn':'good')+'">'+(warn?'ต้องดู':'ปกติ')+'</span></div>'
      + '<div class="operator-stats"><div class="operator-stat"><div class="mini">ถอน</div><b>'+money(r.withdraw_open_amount || 0)+'</b><div class="mini">'+esc(r.withdraw_open_count || 0)+' สลิป</div></div>'
      + '<div class="operator-stat"><div class="mini">ฝาก</div><b>'+money(r.deposit_open_amount || 0)+'</b><div class="mini">'+esc(r.deposit_open_count || 0)+' สลิป</div></div></div>'
      + '<div class="mini">ซ้ำ '+esc(dupe)+' · คิว/ประมวลผล/failed '+esc((r.queued||0)+'/'+(r.processing||0)+'/'+(r.failed||0))+'</div>'
      + '</button>';
  }}).join('')+'</div>';
}}
function renderExceptionQueue(data) {{
  const items = [];
  const summary = data.exception_summary || {{}};
  const overLimitRows = (data.by_account_day || []).filter(r => r.over_limit);
  const overLimitCount = Number(summary.over_limit_count || overLimitRows.length || 0);
  if (overLimitCount) items.push({{title:'เกินวงเงินรายวัน', count:overLimitCount, target:'limitSection', detail:overLimitRows.slice(0,3).map(r => (r.company_name || r.name || '-')+' '+(r.date || '')+' '+money(Math.abs(r.remaining_amount || 0))).join(' · ') || 'เปิดดูบัญชีที่เกินวงเงิน'}});
  const issueRows = data.issues || [];
  const issueCount = Number(summary.issue_count || issueRows.length || 0);
  if (issueCount) items.push({{title:'OCR issue / อ่านไม่ชัด', count:issueCount, target:'section-recent', detail:issueRows.slice(0,3).map(r => (r.company_name || r.bot_key || '-')+' msg '+(r.message_id || '-')).join(' · ') || 'มีรายการอ่านไม่ชัดใน scope นี้'}});
  const bankRows = data.source_bank_review || [];
  const bankReviewCount = Number(summary.bank_review_count || data.totals.source_bank_review_count || bankRows.length || 0);
  if (bankReviewCount) items.push({{title:'รอรีเช็คธนาคารต้นทาง', count:bankReviewCount, target:'section-bank-review', detail:bankRows.slice(0,3).map(r => (r.company_name || r.bot_key || '-')+' '+money(r.amount || 0)).join(' · ') || 'มีสลิปที่ต้องเติมธนาคารต้นทาง'}});
  const dupeRows = data.duplicate_pairs || [];
  const dupeCount = Number(summary.duplicate_count || data.totals.selected_duplicate_count || dupeRows.length || 0);
  if (dupeCount) items.push({{title:'สลิปซ้ำ', count:dupeCount, target:'section-duplicates', detail:dupeRows.slice(0,3).map(r => (r.company_name || r.bot_key || '-')+' '+money(r.amount || r.original_amount || 0)).join(' · ') || 'มีสลิปซ้ำใน scope นี้'}});
  const queueCount = Number(summary.queue_attention_count || 0);
  if (queueCount) items.push({{title:'คิว OCR / processing / failed', count:queueCount, target:'section-recent', detail:'รอ '+(data.jobs.queued || 0)+' · processing '+(data.jobs.processing || 0)+' · failed '+(data.jobs.failed || 0)}});
  const ledgerSummary = data.bank_ledger_summary || {{}};
  const ledgerUnmatched = Number(summary.ledger_unmatched_count || ((ledgerSummary.unmatched_ledger || {{}}).count) || 0);
  if (ledgerUnmatched) items.push({{title:'เดินบัญชียังไม่ match สลิป', count:ledgerUnmatched, target:'section-bank-ledger', detail:'เปิดดู statement รายบัญชีและรายการที่ยังค้างตรวจ'}});
  if (!items.length) return '<div class="good"><b>ไม่มีรายการที่ต้องจัดการใน scope นี้</b></div><div class="mini">ถ้าต้องการดูรายละเอียดทั้งหมด เปิดเมนูแสดงทั้งหมด</div>';
  return '<div class="exception-list">'+items.map(item => '<button type="button" class="exception-item" data-menu-jump="'+esc(item.target)+'"><span><strong>'+esc(item.title)+'</strong><div class="mini">'+esc(item.detail || '-')+'</div></span><span class="pill warn">'+esc(item.count)+'</span></button>').join('')+'</div>';
}}
function wireExceptionButtons() {{
  document.querySelectorAll('[data-menu-jump]').forEach(btn => {{
    btn.onclick = function() {{ showMenuSection(this.dataset.menuJump || 'all'); }};
  }});
}}
function wireCompanyButtons() {{
  document.querySelectorAll('[data-company-bot-key]').forEach(btn => {{
    btn.onclick = function() {{ pickCompany(this.dataset.companyBotKey || '__all__'); }};
  }});
}}
function pickCompany(botKey) {{
  document.getElementById('botFilter').value = botKey;
  const exportCompany = document.getElementById('exportCompanyFilter');
  if (exportCompany) exportCompany.value = botKey;
  document.getElementById('chat').value = botKey + '||' + (document.getElementById('flowFilter').value || 'all');
  showAllMenuSections();
  load({{scrollTop:true}});
}}
function recentCards(rows) {{
  if (!rows || !rows.length) return '<div class="muted">ไม่มีข้อมูล</div>';
  return '<div class="slip-cards">'+rows.slice(0,24).map(r => {{
    const image = r.image_url ? '<a href="'+esc(r.image_url)+'" target="_blank" rel="noopener"><img class="slip-thumb" loading="lazy" src="'+esc(r.image_url)+'" alt="slip image" /></a>' : '<div class="slip-thumb muted" style="display:flex;align-items:center;justify-content:center">ไม่มีรูป</div>';
    const dup = Number(r.is_duplicate||0) ? '<div class="mini bad">สลิปซ้ำที่จับแล้ว · ไม่นับในยอดรวม</div>' : '';
    const company = r.company_name ? '<div class="mini">บริษัท: '+esc(r.company_name)+' · Bot: '+esc(r.bot_key || '')+'</div>' : '';
    return '<div class="slip-card">'+image+'<div class="slip-body"><div class="top"><b>'+esc(r.transferor_name || r.sender_name || '(ไม่ทราบชื่อ)')+'</b><span class="pill">'+esc(r.status)+(Number(r.is_duplicate||0)?' · ซ้ำ':'')+'</span></div><div class="mini">'+esc(r.slip_date_text || '')+' · '+esc([r.from_bank, r.to_bank].filter(Boolean).join(' → ') || r.issuer_bank || '')+'</div>'+company+'<div>ยอด <b>'+money(r.amount)+'</b></div>'+dup+'<div class="toolbar" style="margin-top:8px" data-admin-only="true"><button class="danger" data-delete-slip-id="'+esc(String(r.id || ''))+'" onclick="deleteSlip(this.dataset.deleteSlipId)">ขอลบรายการนี้</button></div></div></div>';
  }}).join('')+'</div>';
}}
function thumb(url, label) {{
  return url ? '<a href="'+esc(url)+'" target="_blank" rel="noopener"><img class="slip-thumb" loading="lazy" src="'+esc(url)+'" alt="'+esc(label||'slip')+'" /></a>' : '<div class="slip-thumb muted" style="display:flex;align-items:center;justify-content:center">ไม่มีรูป</div>';
}}
function renderQueueIssues(jobs, issues) {{
  const jobRows = (jobs || []).slice(0,20);
  const issueRows = (issues || []).slice(0,40);
  let html = '';
  if (jobRows.length) {{
    html += '<h4>คิว OCR</h4>' + table(jobRows, [['สถานะ','status'], ['msg', 'message_id'], ['attempts', r => (r.attempts||0)+'/'+(r.max_attempts||0)], ['error','error']]);
  }}
  if (!issueRows.length) {{
    html += '<div class="muted">ไม่มีรายการ error/อ่านไม่ชัด</div>';
    return html;
  }}
  html += '<div class="toolbar" style="margin:8px 0" data-admin-only="true"><button onclick="reprocessAllIssues(this)">รี OCR Issues ในขอบเขตนี้</button></div>';
  html += '<div class="slip-cards">'+issueRows.map(r => {{
    const image = thumb(r.image_url, 'issue slip');
    const who = r.transferor_name || r.sender_name || r.username || '(ไม่ทราบชื่อ)';
    const detail = 'msg '+esc(r.message_id || '-')+' · '+esc(r.slip_date_text || r.created_at_iso || '')+' · '+esc(r.company_name || r.bot_key || '');
    const amount = Number(r.amount || 0) ? money(r.amount) : '-';
    const conf = r.confidence !== undefined && r.confidence !== null ? Number(r.confidence || 0).toFixed(2) : '-';
    return '<div class="slip-card">'+image+'<div class="slip-body"><div class="top"><b>'+esc(who)+'</b><span class="pill">'+esc(r.status || 'issue')+'</span></div><div class="mini">'+detail+'</div><div class="bad">'+esc(r.error || '-')+'</div><div class="mini">ยอด: '+amount+' · confidence: '+esc(conf)+'</div>'+(r.raw_text ? '<div class="mini">OCR: '+esc(r.raw_text)+'</div>' : '')+'<div class="toolbar" style="margin-top:8px" data-admin-only="true"><button data-slip-id="'+esc(String(r.id || ''))+'" onclick="reprocessIssue(this.dataset.slipId, this)">รี OCR</button><button class="danger" data-delete-slip-id="'+esc(String(r.id || ''))+'" onclick="deleteSlip(this.dataset.deleteSlipId)">ขอลบรายการนี้</button></div></div></div>';
  }}).join('')+'</div>';
  return html;
}}
async function reprocessIssue(slipId, btn, reload=true) {{
  if (!slipId) return;
  const status = document.getElementById('statusline');
  const oldText = btn ? btn.textContent : '';
  if (btn) {{ btn.disabled = true; btn.textContent = 'กำลังรี OCR...'; }}
  if (status) status.textContent = 'กำลังรี OCR issue...';
  try {{
    const res = await fetch('/api/slip/reprocess'+query(), {{method:'POST', headers:postHeaders(), body: JSON.stringify({{id: slipId, bot_key: selectedBotKey()}})}});
    const data = await res.json();
    if (!res.ok || !data.ok) throw new Error(data.error || res.status);
    if (status) status.textContent = 'รี OCR แล้ว: '+data.status+(data.error ? ' · '+data.error : '')+' · '+money(data.amount || 0);
    if (reload) await load();
    return data;
  }} catch (e) {{
    const message = 'รี OCR ไม่สำเร็จ: '+(e && e.message ? e.message : e);
    if (status) status.textContent = message;
    await dashboardNotify(message);
    return null;
  }} finally {{
    if (btn) {{ btn.disabled = false; btn.textContent = oldText || 'รี OCR'; }}
  }}
}}
async function reprocessAllIssues(btn) {{
  const rows = (currentSnapshot && currentSnapshot.issues) || [];
  if (!rows.length) return await dashboardNotify('ไม่มี Issues ให้รี OCR');
  if (!await dashboardConfirm('รี OCR Issues ในขอบเขตนี้ '+rows.length+' รายการที่โหลดมา?')) return;
  const oldText = btn ? btn.textContent : '';
  if (btn) {{ btn.disabled = true; btn.textContent = 'กำลังรี OCR ทั้งหมด...'; }}
  let ok = 0, fail = 0;
  for (const row of rows) {{
    const result = await reprocessIssue(row.id, null, false);
    if (result && result.ok) ok += 1; else fail += 1;
  }}
  if (btn) {{ btn.disabled = false; btn.textContent = oldText || 'รี OCR Issues ในขอบเขตนี้'; }}
  const status = document.getElementById('statusline');
  if (status) status.textContent = 'รี OCR Issues เสร็จ: สำเร็จ '+ok+' / fail '+fail;
  await load();
}}
function duplicateContextLine(prefix, company, bot, chatTitle, flowLabel, senderName, username) {{
  const sender = senderName || (username ? '@'+username : '');
  const senderPart = sender ? ' · ชื่อผู้ส่ง '+esc(sender) : '';
  return '<div class="mini dupe-meta"><b>'+esc(prefix)+'</b> · บริษัท/Bot '+esc(company || bot || '-')+' ('+esc(bot || '-')+') · กลุ่ม Telegram '+esc(chatTitle || '-')+senderPart+' · '+esc(flowLabel || '-').replace('กลุ่ม','กลุ่ม')+'</div>';
}}
function renderDuplicatePairs(rows) {{
  if (!rows || !rows.length) return '<div class="muted">ยังไม่พบสลิปซ้ำในช่วงนี้</div>';
  return '<div class="slip-cards">'+rows.map(r => {{
    const dupImg = thumb(r.duplicate_image_url, 'duplicate slip');
    const origImg = thumb(r.original_image_url, 'original slip');
    const banks = [r.from_bank, r.to_bank].filter(Boolean).join(' → ') || r.issuer_bank || '-';
    const originalBanks = [r.original_from_bank, r.original_to_bank].filter(Boolean).join(' → ') || r.original_issuer_bank || banks || '-';
    const ref = r.reference_no || r.seq || r.aid || '-';
    const originalRef = r.original_reference_no || r.original_seq || r.original_aid || '-';
    const duplicateSubmitted = formatIsoTime(r.duplicate_submitted_at_iso || r.duplicate_created_at_iso);
    const originalSubmitted = formatIsoTime(r.original_submitted_at_iso || r.original_created_at_iso);
    return '<div class="slip-card dupe-card">'
      + '<div class="dupe-thumbs">'+dupImg+origImg+'</div>'
      + '<div class="slip-body"><div class="top"><b>สลิปซ้ำ</b><span class="pill">ซ้ำกับใบต้นฉบับ</span></div>'
      + duplicateContextLine('ใบซ้ำ', r.duplicate_company_name, r.duplicate_bot_key, r.duplicate_chat_title, r.duplicate_flow_label || flowName(r.duplicate_flow_type), r.duplicate_sender_name, r.duplicate_username)
      + duplicateContextLine('ต้นฉบับ', r.original_company_name, r.original_bot_key, r.original_chat_title, r.original_flow_label || flowName(r.original_flow_type), r.original_sender_name, r.original_username)
      + '<div class="mini dupe-submitted"><b>เวลาส่งเข้ามา</b>: ต้นฉบับ '+esc(originalSubmitted)+' · ใบซ้ำ '+esc(duplicateSubmitted)+'</div>'
      + '<div><b>ข้อมูลใบซ้ำ</b>: '+esc((r.slip_date_display || r.slip_date_iso || '')+' '+(r.slip_time || ''))+' · '+esc(r.transferor_name || '(ไม่ทราบผู้โอน)')+' → '+esc(r.recipient_name || '-')+'</div>'
      + '<div class="mini">'+esc(banks)+' · ref '+esc(ref)+' · msg '+esc(r.duplicate_message_id || '-')+'</div>'
      + '<div><b>ข้อมูลต้นฉบับ</b>: '+esc((r.original_slip_date_display || r.original_slip_date_iso || '')+' '+(r.original_slip_time || ''))+' · '+esc(r.original_transferor_name || r.transferor_name || '(ไม่ทราบผู้โอน)')+' → '+esc(r.original_recipient_name || r.recipient_name || '-')+'</div>'
      + '<div class="mini">'+esc(originalBanks)+' · ref '+esc(originalRef)+' · msg '+esc(r.original_message_id || '-')+'</div>'
      + '<div>ยอด <b>'+money(r.amount || r.original_amount)+'</b></div>'
      + '<div class="toolbar" style="margin-top:8px"><button data-dupe-id="'+esc(String(r.duplicate_id || ''))+'">ยกเลิกการนับซ้ำใบนี้</button></div>'
      + '</div></div>';
  }}).join('')+'</div>';
}}
function sourceBankReviewCards(rows) {{
  if (!rows || !rows.length) return '<div class="muted">ไม่มีสลิปที่ต้องรีเช็คธนาคารต้นทาง</div>';
  return '<div class="slip-cards">'+rows.map(r => {{
    const image = thumb(r.image_url, 'source bank review slip');
    const hint = 'issuer: '+esc(r.issuer_bank || '-')+' · ปลายทาง: '+esc(r.to_bank || '-')+' · ต้นทาง: '+esc(r.from_bank || '(ยังไม่เจอ)');
    return '<div class="slip-card">'+image+'<div class="slip-body"><div class="top"><b>'+esc(r.transferor_name || r.sender_name || '(ไม่ทราบชื่อ)')+'</b><span class="pill">รีเช็คต้นทาง</span></div><div class="mini">ยอดนี้นับเข้าแดชบอร์ดแล้ว ถ้าเป็น success ไม่ซ้ำ · รีเช็คเพื่อเติมธนาคารต้นทางให้แผงธนาคารและวงเงินครบ</div><div class="mini">'+esc(r.slip_date_text || '')+' · msg '+esc(r.message_id || '-')+'</div><div>'+hint+'</div><div>ยอด <b>'+money(r.amount)+'</b></div><div class="toolbar" style="margin-top:8px" data-admin-only="true"><button data-slip-id="'+esc(String(r.id || ''))+'" onclick="openaiBankRecheck(this.dataset.slipId, this)">OpenAI รีเช็คธนาคาร</button><button class="danger" data-delete-slip-id="'+esc(String(r.id || ''))+'" onclick="deleteSlip(this.dataset.deleteSlipId)">ขอลบรายการนี้</button></div></div></div>';
  }}).join('')+'</div>';
}}

async function openaiBankRecheck(slipId, btn, ask=true, reload=true) {{
  if (!slipId) return;
  if (ask) {{
    const ok = await dashboardConfirm('ยืนยันให้ OpenAI รีเช็คธนาคารจากรูปสลิปใบนี้?\\n\\nระบบจะอัปเดตเฉพาะ field ที่ยังว่าง/ไม่ทราบ และอาจใช้เวลาสักครู่');
    if (!ok) return;
  }}
  const status = document.getElementById('statusline');
  const oldText = btn ? btn.textContent : '';
  if (btn) {{ btn.disabled = true; btn.textContent = 'กำลังรีเช็ค...'; }}
  if (status) status.textContent = 'OpenAI กำลังรีเช็คธนาคาร...';
  try {{
    const res = await fetch('/api/bank-review/openai'+query(), {{method:'POST', headers:postHeaders(), body: JSON.stringify({{id: slipId, apply: true}})}});
    let data = {{}};
    try {{ data = await res.json(); }} catch (e) {{ data = {{error: 'อ่านผลลัพธ์จาก server ไม่ได้'}}; }}
    if (!res.ok || !data.ok) {{
      const message = 'รีเช็คไม่สำเร็จ: ' + (data.error || res.status);
      if (status) status.textContent = message;
      await dashboardNotify(message);
      return;
    }}
    const applied = Object.entries(data.applied || {{}}).map(([k,v]) => k+'='+v).join(', ') || 'ไม่มี field ใหม่';
    const message = 'OpenAI รีเช็คแล้ว: ' + applied;
    if (status) status.textContent = message;
    if (ask) await dashboardNotify(message);
    if (reload) await load();
  }} catch (e) {{
    const message = 'รีเช็คไม่สำเร็จ: ' + (e && e.message ? e.message : e);
    if (status) status.textContent = message;
    await dashboardNotify(message);
  }} finally {{
    if (btn) {{ btn.disabled = false; btn.textContent = oldText || 'OpenAI รีเช็คธนาคาร'; }}
  }}
}}

async function openaiBankRecheckAllScope(btn) {{
  const summary = (currentSnapshot && currentSnapshot.exception_summary) || {{}};
  const total = Number(summary.bank_review_count || (currentSnapshot && currentSnapshot.totals && currentSnapshot.totals.source_bank_review_count) || 0);
  if (!total) return await dashboardNotify('ไม่มีรายการที่ต้องรีเช็คในขอบเขตนี้');
  if (!await dashboardConfirm('ยืนยันให้ OpenAI รีเช็คธนาคารต้นทางทั้งหมด '+total+' รายการในบริษัท/กลุ่ม/วันที่ที่เลือก?\\n\\nระบบจะอัปเดตเฉพาะ field ธนาคารที่ยังว่างหรือไม่ทราบ')) return;
  const status = document.getElementById('statusline');
  const oldText = btn ? btn.textContent : '';
  if (btn) {{ btn.disabled = true; btn.textContent = 'กำลังรีเช็คทั้ง scope...'; }}
  if (status) status.textContent = 'OpenAI กำลังรีเช็คธนาคารทั้ง scope...';
  try {{
    const parts = selectedChatParts();
    const payload = {{
      chat_id: (parts.bot_key === selectedBotKey()) ? (parts.chat_id || '') : '',
      bot_key: selectedBotKey(),
      flow_type: document.getElementById('flowFilter').value || parts.flow_type || 'all',
      scope: currentSnapshot ? (currentSnapshot.scope || document.getElementById('scope').value || 'open') : (document.getElementById('scope').value || 'open'),
      slip_search: document.getElementById('slipSearch').value || '',
      apply: true
    }};
    const res = await fetch('/api/bank-review/openai-all'+query(), {{method:'POST', headers:postHeaders(), body: JSON.stringify(payload)}});
    const data = await res.json();
    const message = 'รีเช็คธนาคารทั้ง scope เสร็จ: สำเร็จ '+(data.ok_count || 0)+' / fail '+(data.fail_count || 0)+' / ทั้งหมด '+(data.total_count || total);
    if (status) status.textContent = message;
    if (!res.ok && !data.ok) await dashboardNotify(message);
    await load();
  }} catch (e) {{
    const message = 'รีเช็คทั้ง scope ไม่สำเร็จ: ' + (e && e.message ? e.message : e);
    if (status) status.textContent = message;
    await dashboardNotify(message);
  }} finally {{
    if (btn) {{ btn.disabled = false; btn.textContent = oldText || 'OpenAI รีเช็คทั้งหมดในขอบเขตนี้'; }}
  }}
}}
const openaiBankRecheckAll = openaiBankRecheckAllScope;

async function unmarkDuplicate(slipId) {{
  if (!slipId) return;
  if (!await dashboardConfirm('ยืนยันยกเลิกการนับซ้ำของใบนี้? ใบนี้จะกลับไปนับในยอดรวม')) return;
  const res = await fetch('/api/duplicate/unmark'+query(), {{method:'POST', headers:postHeaders(), body: JSON.stringify({{id: slipId, bot_key: selectedBotKey()}})}});
  const data = await res.json();
  if (!res.ok || !data.ok) return await dashboardNotify(data.error || 'ยกเลิกไม่สำเร็จ');
  await load();
}}
async function deleteSlip(slipId) {{
  if (!slipId) return;
  if (!await dashboardConfirm('ยืนยันส่งคำขอลบรายการนี้? รายการจะยังไม่หายจนกว่าจะมีผู้อนุมัติอีกคน และระบบจะหักยอดเมื่อ execute แล้ว', 'ส่งคำขอลบรายการ', true)) return;
  const status = document.getElementById('statusline');
  if (status) status.textContent = 'กำลังส่งคำขอลบรายการ...';
  const res = await fetch('/api/slip/delete'+query({{approval:'request'}}), {{method:'POST', headers:postHeaders(), body: JSON.stringify({{id: slipId, bot_key: selectedBotKey(), reason: 'dashboard operator delete'}})}});
  const data = await res.json();
  if (!res.ok || !data.ok) {{ if (status) status.textContent = 'ส่งคำขอลบไม่สำเร็จ: ' + (data.error || res.status); return await dashboardNotify(data.error || 'ส่งคำขอลบไม่สำเร็จ'); }}
  if (data.status === 'pending') {{
    const pendingMessage = (data.already_pending ? 'มีคำขอลบค้างอยู่แล้ว' : 'ส่งคำขอลบแล้ว') + ' · รออนุมัติ #' + (data.pending_id || '-');
    if (status) status.textContent = pendingMessage;
    await dashboardNotify(pendingMessage + ' ซ่อนจากคิวรีเช็คแล้ว แต่ยอดจะถูกหักหลังอนุมัติและ execute');
    await load();
    return;
  }}
  if (status) status.textContent = 'ลบรายการแล้ว · หักยอดออก ' + money(data.removed_amount || 0);
  await load();
}}
function wireDuplicateButtons() {{
  document.querySelectorAll('[data-dupe-id]').forEach(function(btn) {{
    btn.onclick = function() {{ unmarkDuplicate(this.dataset.dupeId); }};
  }});
}}
function pickLimitIndex(index) {{
  const r = transferorRows[index] || {{}};
  pickLimit(r.limit_key || '', r.display_name || r.name || '', r.bank || '', r.account || '', Number((r.limit_amount ?? r.daily_limit) || 0), r.bot_key || '');
}}
function selectedLimitScopeKey(botKey='') {{
  const bot = botKey || '';
  if (bot && bot !== '__all__' && bot !== 'all') return 'bot:' + bot;
  return limitScopeKey();
}}
function pickLimit(limitKey, name, bank, account, amount, botKey='') {{
  document.getElementById('limitKey').value = limitKey;
  const scope = selectedLimitScopeKey(botKey);
  const scopeEl = document.getElementById('limitScopeValue');
  if (scopeEl) scopeEl.value = scope;
  const hint = document.getElementById('limitScopeHint');
  if (hint) hint.textContent = scope ? ('จะบันทึกวงเงินใน scope: ' + scope) : 'ไม่พบบริษัทของบัญชีนี้ กรุณาเลือกบริษัทก่อนบันทึก';
  document.getElementById('limitName').value = name;
  document.getElementById('limitBank').value = bank;
  document.getElementById('limitAccount').value = account;
  document.getElementById('limitAmount').value = amount || '';
}}
function limitScopeKey() {{
  const parts = selectedChatParts();
  if (parts.chat_id) return parts.chat_id;
  const bot = selectedBotKey() || parts.bot_key || (currentSnapshot && currentSnapshot.selected_bot_key) || '';
  if (bot && bot !== '__all__' && bot !== 'all') return 'bot:' + bot;
  return '';
}}
async function saveAccountLimit() {{
  const selectedScopeEl = document.getElementById('limitScopeValue');
  const scopeKey = (selectedScopeEl && selectedScopeEl.value) ? selectedScopeEl.value : limitScopeKey();
  const payload = {{chat_id: scopeKey, limit_key: document.getElementById('limitKey').value, display_name: document.getElementById('limitName').value, bank: document.getElementById('limitBank').value, account: document.getElementById('limitAccount').value, limit_amount: Number(document.getElementById('limitAmount').value || 0)}};
  if (!payload.chat_id) return await dashboardNotify('เลือกบริษัทหรือกลุ่มก่อนตั้งวงเงิน');
  if (!payload.limit_key) return await dashboardNotify('เลือกบัญชีจากปุ่มตั้งวงเงินก่อน');
  const res = await fetch('/api/account-limit'+query(), {{method:'POST', headers:postHeaders(), body: JSON.stringify(payload)}});
  const data = await res.json();
  if (!res.ok || !data.ok) return await dashboardNotify(data.error || 'บันทึกไม่สำเร็จ');
  await dashboardNotify('บันทึกวงเงินแล้ว');
  await load({{scrollTarget:'byTransferor', smooth:false}});
}}
function renderTelegramBots(rows) {{
  if (!rows || !rows.length) return '<div class="muted">ยังไม่มีข้อมูลบอท</div>';
  return '<div class="bot-list">'+rows.map(b => '<div class="bot-row"><div class="top"><b>'+esc(b.company_name || b.bot_key || '-')+'</b><span class="pill">'+(b.has_token ? 'พร้อมใช้' : 'ไม่มี token')+'</span></div><div class="mini">bot: '+esc(b.bot_key || '-')+' · token: '+(b.has_token ? 'ตั้งค่าแล้ว' : 'ยังไม่ตั้งค่า')+'</div></div>').join('')+'</div>';
}}
function renderCompanyAccounts(rows) {{
  return table(rows || [], [['บริษัท','company_name'], ['ธนาคาร','bank'], ['เลขบัญชี','account_no'], ['ชื่อบัญชี','account_name'], ['วงเงิน/วัน', r => limitMoney(r.daily_limit)]]);
}}
async function saveCompanyAccount() {{
  const parts = selectedChatParts();
  const payload = {{
    bot_key: selectedBotKey() || parts.bot_key || (currentSnapshot && currentSnapshot.selected_bot_key) || 'default',
    chat_id: parts.chat_id,
    company_name: document.getElementById('companyName').value || '',
    bank: document.getElementById('accountBank').value || '',
    account_no: document.getElementById('accountNo').value || '',
    account_name: document.getElementById('accountName').value || '',
    daily_limit: Number(document.getElementById('accountDailyLimit').value || 0)
  }};
  if (!payload.chat_id || !payload.company_name || !payload.account_no) return await dashboardNotify('กรอกบริษัทและเลขบัญชีก่อน');
  const res = await fetch('/api/company-account'+query(), {{method:'POST', headers:postHeaders(), body: JSON.stringify(payload)}});
  const data = await res.json();
  if (!res.ok || !data.ok) return await dashboardNotify(data.error || 'บันทึกบัญชีไม่สำเร็จ');
  await load();
}}
function reconcileSummary(data) {{
  if (!data) return '<div class="muted">ยังไม่ได้เทียบ</div>';
  if (!data.ok) return '<div class="bad">'+esc(data.error || 'reconcile failed')+'</div>';
  const complete = (data.missing.count||0) === 0 && (data.extra.count||0) === 0;
  const verdict = complete ? 'ผลเทียบยอด: ครบ' : 'ผลเทียบยอด: ไม่ครบ';
  const scope = data.scope || {{}};
  const cards = '<div class="'+(complete?'good':'warn')+'"><b>'+verdict+'</b></div>'
    + '<div id="reconcile_scope" class="mini">บริษัท/Bot: '+esc(scope.bot_key || '__all__')+' · กลุ่ม: '+esc(scope.flow_label || scope.flow_type || 'รวมทุกกลุ่ม')+' · ช่วง: '+esc(scope.date_scope || '-')+'</div>'
    + '<div class="muted">หลังบ้าน '+(data.backend.count||0)+' รายการ · '+money(data.backend.amount)+' | สลิป '+(data.slips.count||0)+' รายการ · '+money(data.slips.amount)+' | match '+(data.matched.count||0)+' | ขาด '+(data.missing.count||0)+' | เกิน '+(data.extra.count||0)+' | diff '+money(data.diff_amount)+'</div>';
  const matchRows = (data.matched.rows || []).map(m => ({{...m.backend, status:'ตรงกัน', slip_source:(m.slip||{{}}).source || '', slip_name:(m.slip||{{}}).name || '', slip_time:(m.slip||{{}}).time || '', score:m.score}}));
  const missing = table(data.missing_in_slips, [['สถานะ', r => 'หลังบ้านมี แต่ไม่พบในสลิป'], ['วันที่','date'], ['เวลา','time'], ['ชื่อ','name'], ['ธนาคาร','bank'], ['ไฟล์/แหล่งที่มา','source'], ['ref/รหัส','reference'], ['ยอด', r => money(r.amount)]]);
  const extra = table(data.extra_slips, [['สถานะ', r => 'สลิปมี แต่ไม่พบในหลังบ้าน'], ['วันที่','date'], ['เวลา','time'], ['ชื่อ','name'], ['ธนาคาร','bank'], ['ไฟล์/แหล่งที่มา','source'], ['ref/รหัส','reference'], ['ยอด', r => money(r.amount)]]);
  const matched = table(matchRows, [['สถานะ','status'], ['วันที่','date'], ['เวลา','time'], ['ชื่อ','name'], ['ธนาคาร','bank'], ['ไฟล์/แหล่งที่มา','source'], ['ref/รหัส','reference'], ['ยอด', r => money(r.amount)], ['score','score']]);
  const dailyRows = [
    ...(data.daily.backend || []).map(r => ({{...r, source:'หลังบ้าน'}})),
    ...(data.daily.slips || []).map(r => ({{...r, source:'สลิป'}})),
  ];
  const daily = table(dailyRows, [['แหล่ง','source'], ['วันที่','date'], ['รายการ','count'], ['ยอด', r => money(r.amount)]]);
  return cards + '<h4>สรุปรายวัน</h4>' + daily + '<h4>รายการที่ตรงกัน</h4>' + matched + '<h4>หลังบ้านมี แต่ไม่พบในสลิป</h4>' + missing + '<h4>สลิปมี แต่ไม่พบในหลังบ้าน</h4>' + extra;
}}
function statementReconcileSummary(data) {{
  if (!data) return '<div class="muted">ยังไม่ได้เทียบ 3 ฝั่ง</div>';
  if (!data.ok) return '<div class="bad">'+esc(data.error || 'statement reconcile failed')+'</div>';
  const complete = (data.backend_missing_slip.count||0) === 0 && (data.backend_missing_statement.count||0) === 0 && (data.slip_extra.count||0) === 0 && (data.statement_extra.count||0) === 0;
  const verdict = complete ? 'ผลเทียบ 3 ฝั่ง: ครบ' : 'ผลเทียบ 3 ฝั่ง: ไม่ครบ';
  const scope = data.scope || {{}};
  const diff = data.diff_amounts || {{}};
  const cards = '<div class="'+(complete?'good':'warn')+'"><b>'+verdict+'</b></div>'
    + '<div class="mini">บริษัท/Bot: '+esc(scope.bot_key || '__all__')+' · กลุ่ม: '+esc(scope.flow_label || scope.flow_type || '-')+' · ช่วง: '+esc(scope.date_scope || '-')+'</div>'
    + '<div class="muted">หลังบ้าน '+(data.backend.count||0)+' รายการ · '+money(data.backend.amount)+' | สลิป '+(data.slips.count||0)+' รายการ · '+money(data.slips.amount)+' | เดินบัญชี '+(data.statement.count||0)+' รายการ · '+money(data.statement.amount)+' | match 3 ฝั่ง '+(data.matched.count||0)+' | diff หลังบ้าน-สลิป '+money(diff.backend_minus_slips||0)+' | diff หลังบ้าน-เดินบัญชี '+money(diff.backend_minus_statement||0)+'</div>';
  const matchedRows = (data.matched.rows || []).map(m => ({{...m.backend, status:'ตรง 3 ฝั่ง', slip_time:(m.slip||{{}}).time || '', statement_time:(m.statement||{{}}).time || '', statement_desc:(m.statement||{{}}).description || ''}}));
  const matched = table(matchedRows, [['สถานะ','status'], ['วันที่','date'], ['เวลาหลังบ้าน','time'], ['เวลาสลิป','slip_time'], ['เวลาเดินบัญชี','statement_time'], ['รายการเดินบัญชี','statement_desc'], ['ยอด', r => money(r.amount)]]);
  const missingSlip = table((data.backend_missing_slip || {{}}).rows || [], [['สถานะ', r => 'หลังบ้านมี แต่ไม่พบในสลิป'], ['วันที่','date'], ['เวลา','time'], ['ยอด', r => money(r.amount)], ['แหล่ง','source']]);
  const missingStatement = table((data.backend_missing_statement || {{}}).rows || [], [['สถานะ', r => 'หลังบ้านมี แต่ไม่พบในเดินบัญชี'], ['วันที่','date'], ['เวลา','time'], ['ยอด', r => money(r.amount)], ['แหล่ง','source']]);
  const slipExtra = table((data.slip_extra || {{}}).rows || [], [['สถานะ', r => 'สลิปมี แต่ไม่พบในหลังบ้าน'], ['วันที่','date'], ['เวลา','time'], ['ยอด', r => money(r.amount)], ['แหล่ง','source']]);
  const statementExtra = table((data.statement_extra || {{}}).rows || [], [['สถานะ', r => 'เดินบัญชีมี แต่ไม่พบในหลังบ้าน'], ['วันที่','date'], ['เวลา','time'], ['รายการ','description'], ['ยอด', r => money(r.amount)], ['flow','flow_label']]);
  return cards + '<h4>ตรง 3 ฝั่ง</h4>' + matched + '<h4>หลังบ้านมี แต่ไม่พบในสลิป</h4>' + missingSlip + '<h4>หลังบ้านมี แต่ไม่พบในรายการเดินบัญชี</h4>' + missingStatement + '<h4>สลิปเกินจากหลังบ้าน</h4>' + slipExtra + '<h4>รายการเดินบัญชีเกินจากหลังบ้าน</h4>' + statementExtra;
}}
function renderBankLedgerSnapshot(data) {{
  data = data || {{}};
  const entries = data.entries || {{}};
  const matched = data.matched || {{}};
  const unmatchedLedger = data.unmatched_ledger || data.ledger_extra || {{}};
  const unmatchedSlips = data.unmatched_slips || data.slip_extra || {{}};
  const account = data.account || {{}};
  const accountLine = account.account_no ? ('บัญชี '+esc(account.account_no)+' · '+esc(account.bank || '-')+' · '+esc(account.account_name || '')) : 'ยังไม่มี ledger ใน scope นี้';
  const cards = '<div class="muted">'+accountLine+'</div>'
    + '<div class="mini">รายการเดินบัญชี '+esc(entries.count || 0)+' · '+money(entries.amount || 0)+' | match '+esc(matched.count || 0)+' | เดินบัญชีเกิน '+esc(unmatchedLedger.count || 0)+' | สลิปเกิน '+esc(unmatchedSlips.count || 0)+'</div>';
  const ledgerRows = (unmatchedLedger.rows || []).slice(0,50);
  const slipRows = (unmatchedSlips.rows || []).slice(0,50);
  return cards
    + '<h4>เดินบัญชีที่ยังไม่ match สลิป</h4>' + table(ledgerRows, [['วันที่','date'], ['เวลา','time'], ['รายการ','description'], ['flow','flow_label'], ['ref','reference'], ['ยอด', r => money(r.amount)]])
    + '<h4>สลิปที่ยังไม่ match เดินบัญชี</h4>' + table(slipRows, [['วันที่','date'], ['เวลา','time'], ['บัญชีต้นทาง','from_account'], ['บัญชีปลายทาง','to_account'], ['ref','reference'], ['ยอด', r => money(r.amount)]]);
}}
function bankLedgerPreviewSummary(data) {{
  if (!data) return '<div class="muted">ยังไม่ได้ preview</div>';
  if (!data.ok) return '<div class="bad">'+esc(data.error || 'ledger preview failed')+'</div>';
  const complete = (data.ledger_extra.count||0) === 0 && (data.slip_extra.count||0) === 0;
  const account = data.account || {{}};
  const cards = '<div class="'+(complete?'good':'warn')+'"><b>'+(complete?'Preview เดินบัญชี: ครบ':'Preview เดินบัญชี: มีรายการค้างตรวจ')+'</b></div>'
    + '<div class="mini">บริษัท/Bot: '+esc(account.bot_key || '-')+' · บัญชี '+esc(account.account_no || '-')+' · '+esc(account.bank || '-')+' · dry-run '+esc(data.dry_run ? 'yes' : 'no')+'</div>'
    + '<div class="muted">statement '+(data.incoming.count||0)+' รายการ · '+money(data.incoming.amount)+' | สลิป '+(data.slips.count||0)+' รายการ · '+money(data.slips.amount)+' | match '+(data.matched.count||0)+' | ledger เกิน '+(data.ledger_extra.count||0)+' | สลิปเกิน '+(data.slip_extra.count||0)+'</div>';
  const matchedRows = (data.matched.rows || []).map(m => ({{...(m.ledger || {{}}), status:'match', slip_source:(m.slip||{{}}).source || ''}}));
  const matched = table(matchedRows, [['สถานะ','status'], ['วันที่','date'], ['เวลา','time'], ['รายการ','description'], ['ref','reference'], ['ยอด', r => money(r.amount)], ['สลิป','slip_source']]);
  const ledgerExtra = table((data.ledger_extra || {{}}).rows || [], [['สถานะ', r => 'เดินบัญชีมี แต่ไม่พบสลิป'], ['วันที่','date'], ['เวลา','time'], ['รายการ','description'], ['ref','reference'], ['ยอด', r => money(r.amount)]]);
  const slipExtra = table((data.slip_extra || {{}}).rows || [], [['สถานะ', r => 'สลิปมี แต่ไม่พบเดินบัญชี'], ['วันที่','date'], ['เวลา','time'], ['ต้นทาง','from_account'], ['ปลายทาง','to_account'], ['ref','reference'], ['ยอด', r => money(r.amount)]]);
  return cards + '<h4>รายการที่ match</h4>' + matched + '<h4>เดินบัญชีเกิน</h4>' + ledgerExtra + '<h4>สลิปเกิน</h4>' + slipExtra;
}}
async function runBankLedgerPreview() {{
  const bot = document.getElementById('ledgerCompanyFilter').value || selectedBotKey() || '__all__';
  const flow = document.getElementById('ledgerFlowFilter').value || document.getElementById('flowFilter').value || 'all';
  const scope = document.getElementById('ledgerDateScope').value || selectedDashboardScope();
  const bank = document.getElementById('statementBank').value || '';
  const account_no = document.getElementById('statementAccountNo').value || '';
  const account_name = document.getElementById('statementAccountName').value || '';
  const statement_path = document.getElementById('ledgerStatementPath').value || '';
  const statementFile = document.getElementById('ledgerStatementFile').files[0];
  const box = document.getElementById('bankLedgerSummary');
  if (!account_no) {{
    const msg = 'ใส่เลขบัญชีของ statement ก่อน preview';
    box.innerHTML = '<div class="bad">'+esc(msg)+'</div>';
    await dashboardNotify(msg);
    return;
  }}
  if (!statementFile && !statement_path) {{
    const msg = 'อัปโหลด statement หรือใส่ path statement ก่อน';
    box.innerHTML = '<div class="bad">'+esc(msg)+'</div>';
    await dashboardNotify(msg);
    return;
  }}
  box.innerHTML = '<div class="muted">กำลัง preview ledger รายบัญชี...</div>';
  let res;
  if (statementFile) {{
    const form = new FormData();
    form.append('bot_key', bot);
    form.append('flow_type', flow);
    form.append('scope', scope);
    form.append('bank', bank);
    form.append('account_no', account_no);
    form.append('account_name', account_name);
    if (statement_path) form.append('statement_path', statement_path);
    form.append('statement', statementFile);
    res = await fetch('/api/ledger/preview'+query(), {{method:'POST', headers:ACTION_HEADER, body: form}});
  }} else {{
    res = await fetch('/api/ledger/preview'+query(), {{method:'POST', headers:postHeaders(), body: JSON.stringify({{bot_key:bot, flow_type:flow, scope, bank, account_no, account_name, statement_path}})}});
  }}
  const data = await res.json();
  box.innerHTML = bankLedgerPreviewSummary(data);
  enhanceResponsiveTables(box);
}}
async function requestBankLedgerImport() {{
  const ok = await dashboardConfirm('ขอ Import หลัง approval?', 'ระบบจะสร้าง pending action ledger.import ก่อนเขียนรายการเดินบัญชีเข้า ledger จริง');
  if (!ok) return;
  const bot = document.getElementById('ledgerCompanyFilter').value || selectedBotKey() || '__all__';
  const flow = document.getElementById('ledgerFlowFilter').value || document.getElementById('flowFilter').value || 'all';
  const scope = document.getElementById('ledgerDateScope').value || selectedDashboardScope();
  const bank = document.getElementById('statementBank').value || '';
  const account_no = document.getElementById('statementAccountNo').value || '';
  const account_name = document.getElementById('statementAccountName').value || '';
  const statement_path = document.getElementById('ledgerStatementPath').value || '';
  const statementFile = document.getElementById('ledgerStatementFile').files[0];
  const box = document.getElementById('bankLedgerSummary');
  if (!account_no) {{ await dashboardNotify('ใส่เลขบัญชีก่อนขอ import'); return; }}
  if (!statementFile && !statement_path) {{ await dashboardNotify('อัปโหลด statement หรือใส่ path statement ก่อน'); return; }}
  box.innerHTML = '<div class="muted">กำลังสร้างคำขอ import ledger...</div>';
  let res;
  if (statementFile) {{
    const form = new FormData();
    form.append('bot_key', bot);
    form.append('flow_type', flow);
    form.append('scope', scope);
    form.append('bank', bank);
    form.append('account_no', account_no);
    form.append('account_name', account_name);
    if (statement_path) form.append('statement_path', statement_path);
    form.append('statement', statementFile);
    res = await fetch('/api/ledger/import'+query(), {{method:'POST', headers:ACTION_HEADER, body: form}});
  }} else {{
    res = await fetch('/api/ledger/import'+query(), {{method:'POST', headers:postHeaders(), body: JSON.stringify({{bot_key:bot, flow_type:flow, scope, bank, account_no, account_name, statement_path}})}});
  }}
  const data = await res.json();
  if (!res.ok || !data.ok) {{
    box.innerHTML = '<div class="bad">'+esc(data.error || 'สร้างคำขอ import ไม่สำเร็จ')+'</div>';
    await dashboardNotify(data.error || 'สร้างคำขอ import ไม่สำเร็จ');
    return;
  }}
  if (data.status === 'executed') {{
    box.innerHTML = '<div class="good"><b>Import สำเร็จ</b></div><div class="mini">inserted '+esc((data.inserted||{{}}).count||0)+' · duplicates '+esc((data.duplicates||{{}}).count||0)+' · action '+esc(data.action||'ledger.import')+'</div>';
  }} else {{
    box.innerHTML = '<div class="warn"><b>สร้างคำขอ import แล้ว</b></div><div class="mini">pending_id '+esc(data.pending_id)+' · action '+esc(data.action||'ledger.import')+' · ไปที่เมนูรออนุมัติเพื่อ approve/execute</div>';
  }}
  await refreshPendingBadge();
  await loadPendingActions({{scrollTop:false}});
  enhanceResponsiveTables(box);
}}
function reconcileScopeValue() {{
  const dateEl = document.getElementById('reconcileDateScope');
  if (dateEl && dateEl.value) return dateEl.value;
  const scopeEl = document.getElementById('scope');
  return scopeEl ? (scopeEl.value || 'all') : 'all';
}}
function updateReconcileScopePreview() {{
  const botEl = document.getElementById('reconcileCompanyFilter');
  const flowEl = document.getElementById('reconcileFlowFilter');
  const box = document.getElementById('reconcileScopePreview');
  if (!box || !botEl || !flowEl) return;
  const bot = botEl.value || '__all__';
  const botLabel = botEl.selectedOptions && botEl.selectedOptions[0] ? botEl.selectedOptions[0].textContent : bot;
  const flow = flowEl.value || '';
  const ready = bot && bot !== '__all__' && (flow === 'deposit' || flow === 'withdraw');
  if (!ready) {{
    box.innerHTML = '<b>เลือกบริษัทและฝาก/ถอนก่อนอัปโหลดไฟล์หลังบ้าน</b><div class="mini">ป้องกันการเอาไฟล์บริษัทหนึ่งไปเทียบกับยอดอีกบริษัทหรืออีกฝั่ง</div>';
    return;
  }}
  box.innerHTML = 'ไฟล์นี้จะถูกเทียบเฉพาะ <b>'+esc(botLabel)+'</b> · <b>'+esc(flowName(flow))+'</b> · '+esc(scopeName(reconcileScopeValue()));
}}
async function runReconcile() {{
  const guardScopeMsg = 'เลือกบริษัทและฝาก/ถอนก่อนอัปโหลดไฟล์หลังบ้าน';
  const bot = document.getElementById('reconcileCompanyFilter').value || '__all__';
  const flow = document.getElementById('reconcileFlowFilter').value || '';
  const scope = reconcileScopeValue();
  const excel_path = document.getElementById('backendExcel').value || '';
  const file = document.getElementById('backendExcelFile').files[0];
  const box = document.getElementById('reconcile');
  if (bot === '__all__' || !bot) {{
    const msg = 'เลือกบริษัทก่อนอัปโหลด/เทียบไฟล์หลังบ้าน';
    box.innerHTML = '<div class="bad">'+esc(msg)+'</div>';
    await dashboardNotify(msg);
    return;
  }}
  if (flow !== 'deposit' && flow !== 'withdraw') {{
    const msg = 'เลือกยอดฝาก/ถอนก่อนอัปโหลดไฟล์หลังบ้าน';
    box.innerHTML = '<div class="bad">'+esc(msg)+'</div>';
    await dashboardNotify(msg);
    return;
  }}
  if (!file && !excel_path) {{
    const msg = 'อัปโหลด Excel ของบริษัทนี้ก่อน หรือใส่ path ไฟล์บน server';
    box.innerHTML = '<div class="bad">'+esc(msg)+'</div>';
    await dashboardNotify(msg);
    return;
  }}
  box.innerHTML = '<div class="muted">กำลังเทียบยอดเฉพาะบริษัท/ฝากถอนที่เลือก...</div>';
  let res;
  if (file) {{
    const form = new FormData();
    form.append('chat_id', '');
    form.append('bot_key', bot);
    form.append('flow_type', flow);
    form.append('scope', scope);
    form.append('excel', file);
    res = await fetch('/api/reconcile'+query(), {{method:'POST', headers:ACTION_HEADER, body: form}});
  }} else {{
    res = await fetch('/api/reconcile'+query(), {{method:'POST', headers:postHeaders(), body: JSON.stringify({{chat_id:'', bot_key:bot, flow_type:flow, scope, excel_path}})}});
  }}
  const data = await res.json();
  box.innerHTML = reconcileSummary(data);
  enhanceResponsiveTables(box);
}}
async function runStatementReconcile() {{
  const bot = document.getElementById('reconcileCompanyFilter').value || '__all__';
  const flow = document.getElementById('reconcileFlowFilter').value || '';
  const scope = reconcileScopeValue();
  const excel_path = document.getElementById('backendExcel').value || '';
  const file = document.getElementById('backendExcelFile').files[0];
  const statement_path = document.getElementById('statementExcel').value || '';
  const statementFile = document.getElementById('statementExcelFile').files[0];
  const box = document.getElementById('statementReconcile');
  if (bot === '__all__' || !bot) {{
    const msg = 'เลือกบริษัทก่อนเทียบ 3 ฝั่ง';
    box.innerHTML = '<div class="bad">'+esc(msg)+'</div>';
    await dashboardNotify(msg);
    return;
  }}
  if (flow !== 'deposit' && flow !== 'withdraw') {{
    const msg = 'เลือกยอดฝาก/ถอนก่อนเทียบ 3 ฝั่ง';
    box.innerHTML = '<div class="bad">'+esc(msg)+'</div>';
    await dashboardNotify(msg);
    return;
  }}
  if (!file && !excel_path) {{
    const msg = 'อัปโหลด Excel หลังบ้าน หรือใส่ path ไฟล์หลังบ้านก่อน';
    box.innerHTML = '<div class="bad">'+esc(msg)+'</div>';
    await dashboardNotify(msg);
    return;
  }}
  if (!statementFile && !statement_path) {{
    const msg = 'อัปโหลดรายการเดินบัญชี หรือใส่ statement_path ก่อน';
    box.innerHTML = '<div class="bad">'+esc(msg)+'</div>';
    await dashboardNotify(msg);
    return;
  }}
  box.innerHTML = '<div class="muted">กำลังเทียบ 3 ฝั่งด้วยยอด/เวลา...</div>';
  let res;
  if (file || statementFile) {{
    const form = new FormData();
    form.append('chat_id', '');
    form.append('bot_key', bot);
    form.append('flow_type', flow);
    form.append('scope', scope);
    if (excel_path) form.append('excel_path', excel_path);
    if (statement_path) form.append('statement_path', statement_path);
    if (file) form.append('excel', file);
    if (statementFile) form.append('statement', statementFile);
    res = await fetch('/api/reconcile/statement'+query(), {{method:'POST', headers:ACTION_HEADER, body: form}});
  }} else {{
    res = await fetch('/api/reconcile/statement'+query(), {{method:'POST', headers:postHeaders(), body: JSON.stringify({{chat_id:'', bot_key:bot, flow_type:flow, scope, excel_path, statement_path}})}});
  }}
  const data = await res.json();
  box.innerHTML = statementReconcileSummary(data);
  enhanceResponsiveTables(box);
}}
function employeeAuditParams() {{
  const parts = selectedChatParts();
  const bot = selectedBotKey() || parts.bot_key || '__all__';
  const flow = document.getElementById('flowFilter').value || parts.flow_type || 'all';
  const scope = (currentSnapshot && currentSnapshot.scope) || selectedDashboardScope();
  const chat = (parts.bot_key === bot && (flow === 'all' || parts.flow_type === flow)) ? (parts.chat_id || '') : '';
  const account = (document.getElementById('auditAccountKey') || {{}}).value || '';
  const threshold = (document.getElementById('auditVarianceThreshold') || {{}}).value || '100';
  return {{bot_key:bot, chat_id:chat, flow_type:flow, scope, account_key:account, threshold}};
}}
function renderEmployeeAuditSummary(reconcile, variance, crossDup) {{
  const rec = (reconcile && reconcile.summary) || {{}};
  const hasLedger = reconcile && reconcile.has_ledger;
  const scope = (reconcile && reconcile.scope) || {{}};
  const cards = [
    '<div class="operator-stat"><div class="label">1 สลิปไม่เจอ ledger</div><div class="value '+(Number(rec.slip_only_count||0) ? 'warn' : 'good')+'">'+esc(rec.slip_only_count || 0)+'</div><div class="mini">'+money(rec.slip_only_amount || 0)+'</div></div>',
    '<div class="operator-stat"><div class="label">1 ledger ไม่เจอสลิป</div><div class="value '+(Number(rec.ledger_only_count||0) ? 'warn' : 'good')+'">'+esc(rec.ledger_only_count || 0)+'</div><div class="mini">'+money(rec.ledger_only_amount || 0)+'</div></div>',
    '<div class="operator-stat"><div class="label">2 วันพนักงานที่ flagged</div><div class="value '+(Number((variance||{{}}).flagged_count||0) ? 'warn' : 'good')+'">'+esc((variance||{{}}).flagged_count || 0)+'</div><div class="mini">พนักงาน '+esc((variance||{{}}).employee_count || 0)+'</div></div>',
    '<div class="operator-stat"><div class="label">3 กลุ่มซ้ำข้ามบริษัท</div><div class="value '+(Number((crossDup||{{}}).group_count||0) ? 'bad' : 'good')+'">'+esc((crossDup||{{}}).group_count || 0)+'</div><div class="mini">ยอดเสี่ยง '+money((crossDup||{{}}).total_suspicious_amount || 0)+'</div></div>'
  ].join('');
  const note = hasLedger ? 'มี ledger สำหรับเทียบ' : 'ยังไม่มีตาราง ledger/import statement ใน scope นี้ ข้อ 1 จะแสดงสลิปเป็นหลัก';
  return '<div class="operator-home-grid">'+cards+'</div><div class="mini">บริษัท '+esc(scope.bot_key || '-')+' · กลุ่ม '+esc(scope.flow_type || '-')+' · scope '+esc(scope.scope || '-')+' · '+esc(note)+'</div>';
}}
function renderEmployeeVariance(data) {{
  if (!data || !data.ok) return '<div class="bad">'+esc((data && data.error) || 'โหลด variance ไม่สำเร็จ')+'</div>';
  const employees = (data.employees || []).slice(0,80);
  const rows = (data.rows || []).slice(0,120);
  const employeeTable = table(employees, [['พนักงาน','employee'], ['Employee ID','employee_id'], ['แหล่ง identity','identity_source'], ['สลิป','total_slips'], ['ยอดรวม', r => money(r.total_amount)], ['วันที่ flagged','flagged_days']]);
  const rowTable = table(rows, [['วันที่','date'], ['พนักงาน','employee'], ['Employee ID','employee_id'], ['identity','identity_source'], ['บริษัท','company_name'], ['สลิป','slip_count'], ['ยอดสลิป', r => money(r.slip_total)], ['ยอด ledger วันที่เดียวกัน', r => r.ledger_total === null || r.ledger_total === undefined ? '-' : money(r.ledger_total)], ['variance', r => r.variance === null || r.variance === undefined ? '-' : money(r.variance)], ['flag', r => r.flagged ? 'ต้องตรวจ' : '']]);
  return '<h4>2 สรุปพนักงาน</h4>'+employeeTable+'<h4>2 รายวันต่อพนักงาน</h4>'+rowTable;
}}
function renderEmployeeReconcile(data) {{
  if (!data || !data.ok) return '<div class="bad">'+esc((data && data.error) || 'โหลด reconcile ไม่สำเร็จ')+'</div>';
  const summary = data.summary || {{}};
  const top = '<div class="mini">สลิป '+esc(summary.slip_count || 0)+' · '+money(summary.slip_total || 0)+' | ledger '+esc(summary.ledger_count || 0)+' · '+money(summary.ledger_total || 0)+' | match '+esc(summary.matched_count || 0)+' · diff '+money(summary.diff_amount || 0)+'</div>';
  const matched = (data.matched || []).map(m => ({{...(m.slip || {{}}), score:m.score, ledger_ref:(m.entry||{{}}).reference || '', ledger_desc:(m.entry||{{}}).description || ''}}));
  const matchedTable = table(matched, [['วันที่','slip_date_iso'], ['ชื่อ','transferor_name'], ['บัญชีต้นทาง','from_account'], ['บัญชีปลายทาง','to_account'], ['ref สลิป','reference_no'], ['ref ledger','ledger_ref'], ['ยอด', r => money(r.amount)], ['score','score']]);
  const slipOnlyTable = table(data.slip_only || [], [['วันที่','slip_date_iso'], ['ชื่อ','transferor_name'], ['ต้นทาง','from_account'], ['ปลายทาง','to_account'], ['ref','reference_no'], ['ยอด', r => money(r.amount)]]);
  const ledgerOnlyTable = table(data.ledger_only || [], [['วันที่','date_key'], ['เวลา','time'], ['บัญชี','account_no'], ['รายการ','description'], ['ref','reference'], ['ยอด', r => money(r.amount)]]);
  return '<h4>1 เทียบสลิปกับ ledger</h4>'+top+'<h4>1 รายการที่ match</h4>'+matchedTable+'<h4>1 สลิปมี แต่ไม่พบ ledger</h4>'+slipOnlyTable+'<h4>1 ledger มี แต่ไม่พบสลิป</h4>'+ledgerOnlyTable;
}}
function renderCrossBotDuplicates(data) {{
  if (!data || !data.ok) return '<div class="bad">'+esc((data && data.error) || 'โหลด cross duplicate ไม่สำเร็จ')+'</div>';
  const groups = data.groups || [];
  if (!groups.length) return '<h4>3 สลิปซ้ำข้ามบอท/บริษัท</h4><div class="good">ไม่พบกลุ่มซ้ำข้าม source ใน scope นี้</div>';
  const html = groups.slice(0,80).map(g => {{
    const slipRows = (g.slips || []).map(s => ({{...s, source:(s.company_name || s.bot_key || '-') + ' · ' + (s.chat_title || s.chat_id || '-')}}));
    return '<article class="cross-company-card"><div class="cross-company-head"><div><div class="label">fingerprint</div><div class="cross-company-account">'+esc(g.fingerprint || '-')+'</div></div><div class="cross-company-total">'+esc(g.slip_count || 0)+' สลิป · '+money(g.total_amount || 0)+'</div></div>'
      + table(slipRows, [['source','source'], ['วันที่','slip_date_iso'], ['ชื่อ','transferor_name'], ['ref','reference_no'], ['ยอด', r => money(r.amount)], ['ซ้ำแล้ว', r => Number(r.is_duplicate || 0) ? 'ใช่' : 'ยัง']])
      + '</article>';
  }}).join('');
  return '<h4>3 สลิปซ้ำข้ามบอท/บริษัท</h4><div class="cross-company-list">'+html+'</div>';
}}
async function runEmployeeAudit() {{
  const summaryBox = document.getElementById('employeeAuditSummary');
  const detailsBox = document.getElementById('employeeAuditDetails');
  const p = employeeAuditParams();
  if (summaryBox) summaryBox.innerHTML = '<div class="muted">กำลังโหลดออดิต 1-2-3...</div>';
  if (detailsBox) detailsBox.innerHTML = '<div class="muted">รอผลตรวจ...</div>';
  const common = {{bot_key:p.bot_key, chat_id:p.chat_id, flow_type:p.flow_type, scope:p.scope}};
  try {{
    const [recRes, varRes, dupRes] = await Promise.all([
      fetch('/api/audit/reconcile'+query({{...common, account_key:p.account_key}}), {{cache:'no-store'}}),
      fetch('/api/audit/daily-variance'+query({{...common, threshold:p.threshold}}), {{cache:'no-store'}}),
      fetch('/api/audit/cross-dup'+query({{bot_key:p.bot_key, scope:p.scope}}), {{cache:'no-store'}})
    ]);
    const [reconcile, variance, crossDup] = await Promise.all([recRes.json(), varRes.json(), dupRes.json()]);
    if (summaryBox) summaryBox.innerHTML = renderEmployeeAuditSummary(reconcile, variance, crossDup);
    if (detailsBox) detailsBox.innerHTML = renderEmployeeReconcile(reconcile) + renderEmployeeVariance(variance) + renderCrossBotDuplicates(crossDup);
    enhanceResponsiveTables(detailsBox || document);
  }} catch (err) {{
    if (summaryBox) summaryBox.innerHTML = '<div class="bad">โหลดออดิตไม่สำเร็จ</div>';
    if (detailsBox) detailsBox.innerHTML = '<div class="bad">'+esc(err)+'</div>';
  }}
}}
function showToast(msg, type='info') {{
  try {{
    const host = document.getElementById('pendingToastHost');
    if (!host) {{ return; }}
    const div = document.createElement('div');
    div.className = 'pending-toast ' + (type === 'success' ? 'success' : (type === 'error' ? 'error' : ''));
    div.textContent = String(msg || '');
    host.appendChild(div);
    setTimeout(() => {{ try {{ div.remove(); }} catch (e) {{}} }}, 4200);
  }} catch (e) {{}}
}}
const PENDING_ACTION_ICONS = {{
  'slip.delete': '🗑️',
  'period.close': '🔒',
  'account_limit': '💰',
  'account_limit.delete': '💰',
  'company_account': '🏦',
  'company_account.delete': '🏦',
  'reconcile': '🔁',
  'token.create': '🔑',
  'token.revoke': '🔑',
  'token.update': '🔑'
}};
function pendingActionEmoji(action) {{
  const key = String(action || '');
  if (PENDING_ACTION_ICONS[key]) return PENDING_ACTION_ICONS[key];
  if (key.startsWith('slip.delete') || key.endsWith('.delete')) return '🗑️';
  if (key.startsWith('period.close') || key.endsWith('.close')) return '🔒';
  if (key.startsWith('account_limit')) return '💰';
  if (key.startsWith('company_account')) return '🏦';
  if (key.startsWith('reconcile')) return '🔁';
  if (key.startsWith('token.')) return '🔑';
  return '•';
}}
function pendingPayloadSummaryText(summary) {{
  if (!summary || typeof summary !== 'object') return '';
  const parts = [];
  ['id','slip_id','company_name','bot_key','chat_id','reason','note'].forEach(k => {{
    if (summary[k] !== undefined && summary[k] !== null && String(summary[k]) !== '') {{
      parts.push(esc(k) + ':' + esc(String(summary[k])));
    }}
  }});
  return parts.join(' · ');
}}
function pendingShortFp(fp) {{
  const s = String(fp || '');
  if (!s) return '-';
  return esc(s.substring(0, 8));
}}
function pendingStatusBadge(status) {{
  const s = String(status || '');
  const cls = (s === 'pending') ? 'warn' : (s === 'rejected' || s === 'expired' || s === 'cancelled') ? 'bad' : (s === 'executed' || s === 'approved') ? 'good' : '';
  return '<span class="'+cls+'">'+esc(s || '-')+'</span>';
}}
function pendingDisplayTime(value) {{
  const s = String(value || '');
  if (!s) return '-';
  return esc(s.replace('T', ' ').replace(/\\.[0-9]+Z?$/, '').replace(/Z$/, ''));
}}
// Backward marker for regression tests: pendingRowsTable(rows, currentActor)
function pendingRowsTable(rows, currentActor, simpleApproval=false) {{
  if (!rows || !rows.length) return '<div class="muted">ไม่มีคำขอในสถานะนี้</div>';
  const head = '<thead><tr>'+
    '<th>ID</th>'+
    '<th>Action</th>'+
    '<th>ผู้ขอ</th>'+
    '<th>เวลาขอ</th>'+
    '<th>expire</th>'+
    '<th>status</th>'+
    '<th class="action-col">จัดการ</th>'+
    '</tr></thead>';
  const body = rows.map(r => {{
    const id = Number(r.id || 0);
    const status = String(r.status || '');
    const action = String(r.action || '');
    const summary = pendingPayloadSummaryText(r.payload_summary || {{}});
    const actionCell = pendingActionEmoji(action) + ' ' + esc(action) + (summary ? '<div class="mini muted">'+summary+'</div>' : '');
    const isSelfRequest = currentActor && String(r.requested_by || '') === String(currentActor);
    const buttons = [];
    if (status === 'pending') {{
      if (isSelfRequest && simpleApproval) {{
        buttons.push('<button data-pending-id="'+id+'" data-admin-only="true" onclick="executePending('+id+')">อนุมัติ+ทำรายการ</button>');
        buttons.push('<span class="mini good">โหมดง่าย · token เดียว</span>');
      }} else if (isSelfRequest) {{
        buttons.push('<button disabled title="คำขอที่คุณสร้างเอง ต้องใช้ token อื่นอนุมัติ">ใช้ token อื่นอนุมัติ</button>');
        buttons.push('<span class="mini warn">คำขอที่คุณสร้างเอง · ห้าม self-approve</span>');
      }} else {{
        buttons.push('<button data-pending-id="'+id+'" data-admin-only="true" onclick="approvePending('+id+')">อนุมัติ</button>');
      }}
      buttons.push('<button class="danger" data-pending-id="'+id+'" data-admin-only="true" onclick="rejectPending('+id+')">ปฏิเสธ</button>');
      buttons.push('<button data-pending-id="'+id+'" onclick="cancelPending('+id+')">ยกเลิก</button>');
    }} else if (status === 'approved') {{
      buttons.push('<button data-pending-id="'+id+'" data-admin-only="true" onclick="executePending('+id+')">execute</button>');
    }}
    const actionsHtml = buttons.length ? buttons.join(' ') : '<span class="muted mini">-</span>';
    return '<tr>'+
      '<td>'+esc(String(id))+'</td>'+
      '<td>'+actionCell+'</td>'+
      '<td>'+pendingShortFp(r.requested_by)+'</td>'+
      '<td>'+pendingDisplayTime(r.requested_at)+'</td>'+
      '<td>'+pendingDisplayTime(r.expires_at)+'</td>'+
      '<td>'+pendingStatusBadge(status)+'</td>'+
      '<td class="action-cell">'+actionsHtml+'</td>'+
      '</tr>';
  }}).join('');
  return '<div class="responsive-table"><table>'+head+'<tbody>'+body+'</tbody></table></div>';
}}

// ========== Per-Account Ledger ===========
let _ledgerSnapshot = null;
async function populateLedgerAccounts() {{
  const parts = selectedChatParts();
  const chat_id = parts.chat_id || '';
  const bot_key = parts.bot_key || '';
  const url = '/api/summary?' + new URLSearchParams({{chat_id, bot_key, scope:'all', detail:'light'}}).toString();
  try {{
    const r = await fetch(url, {{headers: hdr()}});
    const data = await r.json();
    const accounts = (data.company_accounts || []);
    const sel = document.getElementById('ledgerAccountKey');
    if (!sel) return;
    const prev = sel.value;
    sel.innerHTML = '<option value="">— เลือกบัญชี —</option>';
    accounts.forEach(a => {{
      const key = a.account_key || '';
      const label = [a.company_name, a.bank, a.account_no, a.account_name].filter(Boolean).join(' · ');
      const opt = document.createElement('option');
      opt.value = key; opt.textContent = label || key;
      sel.appendChild(opt);
    }});
    if (prev) sel.value = prev;
  }} catch(e) {{ console.warn('populateLedgerAccounts', e); }}
}}
function ledgerFlag(flags) {{
  if (!flags || !flags.length) return '';
  const map = {{duplicate:'ซ้ำ', pending_delete:'รอลบ', bank_unclear:'ธนาคารไม่ชัด', low_confidence:'OCR ต่ำ', reviewed:'✓'}};
  return flags.map(f => `<span class="pill ${{f==='reviewed'?'':'warn'}}">${{map[f]||f}}</span>`).join(' ');
}}
async function loadLedger() {{
  const parts = selectedChatParts();
  const chat_id = parts.chat_id || '';
  const bot_key = parts.bot_key || '';
  const account_key = (document.getElementById('ledgerAccountKey')||{{}}).value || '';
  const flow = (document.getElementById('ledgerFlow')||{{}}).value || 'all';
  const date_from = (document.getElementById('ledgerDateFrom')||{{}}).value || '';
  const date_to   = (document.getElementById('ledgerDateTo')||{{}}).value || '';
  const body = document.getElementById('ledgerBody');
  const kpi  = document.getElementById('ledgerKpi');
  if (!body) return;
  if (!account_key) {{ body.innerHTML = '<div class="muted mini">เลือกบัญชีก่อน</div>'; return; }}
  body.innerHTML = '<div class="muted mini">กำลังโหลด…</div>';
  const q = new URLSearchParams({{chat_id, bot_key, account_key, flow_type:flow, date_from, date_to, limit:500}}).toString();
  try {{
    const r = await fetch('/api/ledger?' + q, {{headers: hdr()}});
    const data = await r.json();
    if (!data.ok) {{ body.innerHTML = `<div class="muted mini">Error: ${{esc(data.error||'ไม่พบข้อมูล')}}</div>`; return; }}
    _ledgerSnapshot = data;
    const totals = data.totals || {{}};
    const acc = data.account || {{}};
    if (kpi) kpi.innerHTML = `
      <span>บัญชี: <b>${{esc([acc.company_name,acc.bank,acc.account_no].filter(Boolean).join(' '))}}</b></span> &nbsp;·&nbsp;
      <span>ยอดเข้า: <b class="good">+${{money(totals.in||0)}}</b></span> &nbsp;·&nbsp;
      <span>ยอดออก: <b class="warn">-${{money(totals.out||0)}}</b></span> &nbsp;·&nbsp;
      <span>สุทธิ: <b>${{money(totals.net||0)}}</b></span> &nbsp;·&nbsp;
      <span>ยอดสะสม: <b>${{money(totals.ending_balance||0)}}</b></span> &nbsp;·&nbsp;
      <span>รายการ: <b>${{totals.row_count||0}}</b></span>
      ${{totals.issue_count ? ` &nbsp;·&nbsp;<span class="pill warn">${{totals.issue_count}} ปัญหา</span>` : ''}}
    `;
    const rows = data.rows || [];
    if (!rows.length) {{ body.innerHTML = '<div class="muted mini">ไม่พบรายการในช่วงที่เลือก</div>'; return; }}
    let html = `<div class="scroll-x"><table class="responsive-table"><thead><tr>
      <th>วันที่/เวลา</th><th>Flow</th><th>ผู้โอน</th><th>ต้นทาง</th><th>ยอดเข้า</th><th>ยอดออก</th><th>ยอดสะสม</th><th>Badge</th><th>Actions</th>
    </tr></thead><tbody>`;
    rows.forEach(r => {{
      const flow_label = r.chat_title||'';
      const is_deposit = flow_label.toLowerCase().includes('ฝาก')||flow_label.toLowerCase().includes('deposit');
      const amount = parseFloat(r.amount||0);
      const bal = parseFloat(r.running_balance||0);
      const flags = r.flags||[];
      const rowClass = flags.includes('pending_delete')?'opacity:0.5':flags.includes('reviewed')?'color:var(--muted)':'';
      html += `<tr style="${{rowClass}}">
        <td class="mini">${{esc(r.slip_date_iso||r.slip_date_display||'-')}}</td>
        <td>${{is_deposit ? '<span class="pill good">ฝาก</span>' : '<span class="pill warn">ถอน</span>'}}</td>
        <td class="mini">${{esc(r.transferor_name||r.sender_name||'-')}}</td>
        <td class="mini">${{esc(r.from_bank||r.issuer_bank||'-')}}</td>
        <td class="good">${{is_deposit ? money(amount) : ''}}</td>
        <td class="warn">${{!is_deposit ? money(amount) : ''}}</td>
        <td><b>${{money(bal)}}</b></td>
        <td>${{ledgerFlag(flags)}}</td>
        <td style="white-space:nowrap">
          ${{r.image_url ? `<a href="${{esc(r.image_url)}}" target="_blank" class="mini">📄</a> ` : ''}}
          ${{flags.includes('bank_unclear') ? `<button class="mini" data-admin-only="true" onclick="ledgerRecheckBank('${{esc(r.id)}}',this)">🔍</button> ` : ''}}
          ${{!flags.includes('reviewed') ? `<button class="mini" data-admin-only="true" onclick="ledgerMarkReviewed('${{esc(r.id)}}',this)">✓</button>` : ''}}
        </td>
      </tr>`;
    }});
    html += '</tbody></table></div>';
    body.innerHTML = html;
    applyAdminVisibility();
    const exportBtn = document.getElementById('ledgerExportBtn');
    if (exportBtn) exportBtn.disabled = false;
  }} catch(e) {{ body.innerHTML = `<div class="muted mini">โหลดผิดพลาด: ${{e}}</div>`; console.error(e); }}
}}
async function ledgerRecheckBank(slipId, btn) {{
  if (!await dashboardConfirm('ให้ OpenAI รีเช็คธนาคารสำหรับสลิปนี้?')) return;
  if (btn) {{ btn.disabled=true; btn.textContent='…'; }}
  try {{
    const r = await fetch('/api/bank-review/openai' + query(), {{method:'POST', headers:postHeaders(), body:JSON.stringify({{id:slipId,apply:true}})}});
    const d = await r.json();
    await loadLedger();
  }} catch(e) {{ console.error(e); }}
}}
async function ledgerMarkReviewed(slipId, btn) {{
  if (btn) {{ btn.disabled=true; btn.textContent='…'; }}
  try {{
    const r = await fetch('/api/slip/mark-reviewed', {{method:'POST', headers:postHeaders(), body:JSON.stringify({{id:slipId}})}});
    await loadLedger();
  }} catch(e) {{ console.error(e); }}
}}
function exportLedgerExcel() {{
  if (!_ledgerSnapshot) return;
  const acc = _ledgerSnapshot.account || {{}};
  const rows = _ledgerSnapshot.rows || [];
  const headers = ['วันที่','Flow','ผู้โอน','ต้นทาง','ยอดเข้า','ยอดออก','ยอดสะสม','Badge','Slip ID'];
  const csvRows = [headers.join(',')];
  rows.forEach(r => {{
    const flow_label = (r.chat_title||'').toLowerCase();
    const is_dep = flow_label.includes('ฝาก')||flow_label.includes('deposit');
    const amt = parseFloat(r.amount||0);
    csvRows.push([
      r.slip_date_iso||r.slip_date_display||'',
      is_dep?'ฝาก':'ถอน',
      r.transferor_name||r.sender_name||'',
      r.from_bank||r.issuer_bank||'',
      is_dep?amt:'',
      !is_dep?amt:'',
      parseFloat(r.running_balance||0),
      (r.flags||[]).join('|'),
      r.id||'',
    ].map(v=>`"${{String(v).replace(/"/g,'""')}}"`).join(','));
  }});
  const blob = new Blob([String.fromCharCode(0xFEFF)+csvRows.join(String.fromCharCode(10))], {{type:'text/csv;charset=utf-8'}});
  const a = document.createElement('a');
  a.href = URL.createObjectURL(blob);
  a.download = `ledger-${{acc.company_name||'account'}}-${{new Date().toISOString().slice(0,10)}}.csv`;
  a.click();
}}
// ========== End Ledger ===========

async function loadPendingActions(options={{}}) {{
  const filterEl = document.getElementById('pendingStatusFilter');
  const container = document.getElementById('pendingTableContainer');
  const meta = document.getElementById('pendingMeta');
  if (!container) return;
  const status = filterEl ? (filterEl.value || 'pending') : 'pending';
  const qs = (status && status !== 'all') ? ('?status=' + encodeURIComponent(status)) : '';
  try {{
    const res = await fetch('/api/pending' + qs, {{cache:'no-store'}});
    if (!res.ok) {{
      container.innerHTML = '<div class="bad">โหลดรายการรออนุมัติไม่สำเร็จ ('+res.status+')</div>';
      if (meta) meta.textContent = '';
      return;
    }}
    const data = await res.json();
    const rows = (data && data.items) || [];
    container.innerHTML = pendingRowsTable(rows, data.current_actor || '', Boolean(data.simple_approval_enabled));
    enhanceResponsiveTables(container);
    if (meta) meta.textContent = 'พบ ' + (data.count || rows.length || 0) + ' รายการ' + (data.simple_approval_enabled ? ' · โหมดง่าย token เดียว' : '');
    if (status === 'pending') {{ updatePendingBadge(rows.length); }}
  }} catch (err) {{
    container.innerHTML = '<div class="bad">โหลดรายการรออนุมัติไม่สำเร็จ</div>';
    if (meta) meta.textContent = '';
  }}
}}
function updatePendingBadge(count) {{
  const badge = document.getElementById('pendingBadge');
  if (!badge) return;
  const n = Number(count || 0);
  if (n > 0) {{
    badge.textContent = String(n);
    badge.hidden = false;
  }} else {{
    badge.textContent = '0';
    badge.hidden = true;
  }}
}}
async function refreshPendingBadge() {{
  try {{
    const res = await fetch('/api/pending?status=pending', {{cache:'no-store'}});
    if (!res.ok) return;
    const data = await res.json();
    updatePendingBadge((data && (data.count || (data.items || []).length)) || 0);
  }} catch (err) {{}}
}}
async function approvePending(pendingId) {{
  const ok = await dashboardConfirm('อนุมัติคำขอ #' + pendingId + '?', 'ยืนยันการอนุมัติ');
  if (!ok) return;
  try {{
    const res = await fetch('/api/pending/approve', {{method:'POST', headers:postHeaders(), body: JSON.stringify({{pending_id: pendingId}})}});
    const data = await res.json().catch(() => ({{}}));
    if (res.ok && data && data.ok) {{
      showToast('อนุมัติคำขอ #' + pendingId + ' แล้ว', 'success');
    }} else {{
      showToast('อนุมัติไม่สำเร็จ: ' + (data && data.error ? data.error : res.status), 'error');
    }}
  }} catch (err) {{
    showToast('อนุมัติไม่สำเร็จ', 'error');
  }}
  await loadPendingActions({{scrollTop:false}});
  refreshPendingBadge();
  await load({{scrollTop:false}});
}}
async function rejectPending(pendingId) {{
  const reason = await dashboardInput('เหตุผลที่ปฏิเสธคำขอ #' + pendingId + ' (ระบุก็ได้)', 'ปฏิเสธคำขอ', '', {{placeholder:'เหตุผล (ไม่บังคับ)', confirmText:'ปฏิเสธ', danger:true}});
  if (reason === null) return;
  try {{
    const res = await fetch('/api/pending/reject', {{method:'POST', headers:postHeaders(), body: JSON.stringify({{pending_id: pendingId, reason: reason}})}});
    const data = await res.json().catch(() => ({{}}));
    if (res.ok && data && data.ok) {{
      showToast('ปฏิเสธคำขอ #' + pendingId + ' แล้ว', 'success');
    }} else {{
      showToast('ปฏิเสธไม่สำเร็จ: ' + (data && data.error ? data.error : res.status), 'error');
    }}
  }} catch (err) {{
    showToast('ปฏิเสธไม่สำเร็จ', 'error');
  }}
  await loadPendingActions({{scrollTop:false}});
  refreshPendingBadge();
  await load({{scrollTop:false}});
}}
async function cancelPending(pendingId) {{
  const ok = await dashboardConfirm('ยกเลิกคำขอ #' + pendingId + '? (ใช้ได้เฉพาะผู้ที่ขอ)', 'ยืนยันยกเลิก', true);
  if (!ok) return;
  try {{
    const res = await fetch('/api/pending/cancel', {{method:'POST', headers:postHeaders(), body: JSON.stringify({{pending_id: pendingId}})}});
    const data = await res.json().catch(() => ({{}}));
    if (res.ok && data && data.ok) {{
      showToast('ยกเลิกคำขอ #' + pendingId + ' แล้ว', 'success');
    }} else {{
      showToast('ยกเลิกไม่สำเร็จ: ' + (data && data.error ? data.error : res.status), 'error');
    }}
  }} catch (err) {{
    showToast('ยกเลิกไม่สำเร็จ', 'error');
  }}
  await loadPendingActions({{scrollTop:false}});
  refreshPendingBadge();
  await load({{scrollTop:false}});
}}
async function executePending(pendingId) {{
  const ok = await dashboardConfirm('ดำเนินการคำขอที่อนุมัติแล้ว #' + pendingId + '?', 'ยืนยัน execute');
  if (!ok) return;
  try {{
    const res = await fetch('/api/pending/approve?approval=execute', {{method:'POST', headers:postHeaders(), body: JSON.stringify({{pending_id: pendingId, approval: 'execute'}})}});
    const data = await res.json().catch(() => ({{}}));
    if (res.ok && data && data.ok) {{
      showToast('execute สำเร็จ #' + pendingId, 'success');
    }} else {{
      showToast('execute ไม่สำเร็จ: ' + (data && data.error ? data.error : res.status), 'error');
    }}
  }} catch (err) {{
    showToast('execute ไม่สำเร็จ', 'error');
  }}
  await loadPendingActions({{scrollTop:false}});
  refreshPendingBadge();
  await load({{scrollTop:false}});
}}
async function load(options={{}}) {{
  const botEl = document.getElementById('botFilter');
  const exportCompanyEl = document.getElementById('exportCompanyFilter');
  const chatEl = document.getElementById('chat');
  const flowEl = document.getElementById('flowFilter');
  const scopeEl = document.getElementById('scope');
  const customDateEl = document.getElementById('customDateFilter');
  const summaryStartEl = document.getElementById('summaryStartDate');
  const summaryEndEl = document.getElementById('summaryEndDate');
  const slipFilterEl = document.getElementById('slipFilter');
  const slipSearchEl = document.getElementById('slipSearch');
  const current = splitChatValue(chatEl.value || '');
  const currentBot = botEl.value || current.bot_key || '';
  const currentFlow = flowEl.value || current.flow_type || 'all';
  const currentDate = customDateEl ? (customDateEl.value || '') : '';
  const summaryStart = summaryStartEl ? (summaryStartEl.value || '') : '';
  const summaryEnd = summaryEndEl ? (summaryEndEl.value || '') : '';
  const rangeScope = (summaryStart || summaryEnd) ? `range:${{summaryStart}}..${{summaryEnd}}` : '';
  const currentScope = rangeScope || currentDate || scopeEl.value || 'today';
  const currentFilter = slipFilterEl.value || 'all';
  const currentSearch = slipSearchEl.value || '';
  const requestChat = (current.bot_key === currentBot && (currentFlow === 'all' || current.flow_type === currentFlow)) ? current.chat_id : '';
  const accountSearchMode = (typeof window.__accountSearchMode === 'string') ? window.__accountSearchMode : '';
  const detailLevel = (options && options.lite) ? 'lite' : 'full';
  const res = await fetch('/api/summary'+query({{chat_id: requestChat, bot_key: currentBot, flow_type: currentFlow, scope: currentScope, slip_filter: currentFilter, slip_search: currentSearch, account_search_mode: accountSearchMode, detail: detailLevel}}), {{cache:'no-store'}});
  if (!res.ok) {{ document.body.innerHTML = '<main class="wrap"><div class="card bad">Unauthorized or dashboard unavailable</div></main>'; return; }}
  const data = await res.json();
  const isLiteSnapshot = data.detail_level === 'lite';
  if (isLiteSnapshot && currentSnapshot) {{
    currentSnapshot = Object.assign({{}}, currentSnapshot, data, {{
      recent: currentSnapshot.recent || [],
      duplicate_pairs: currentSnapshot.duplicate_pairs || [],
      source_bank_review: currentSnapshot.source_bank_review || [],
      deposit_customer_slips: currentSnapshot.deposit_customer_slips || [],
      issues: currentSnapshot.issues || [],
      jobs_recent: currentSnapshot.jobs_recent || [],
      provider_usage: currentSnapshot.provider_usage || [],
      company_account_daily: currentSnapshot.company_account_daily || [],
      withdraw_limit_usage: currentSnapshot.withdraw_limit_usage || [],
      account_slip_search: currentSnapshot.account_slip_search || data.account_slip_search,
      cross_company_account_slip_search: currentSnapshot.cross_company_account_slip_search || data.cross_company_account_slip_search,
      account_cross_company: currentSnapshot.account_cross_company || [],
      by_transferor: currentSnapshot.by_transferor || [],
      by_account_day: currentSnapshot.by_account_day || [],
      by_date: currentSnapshot.by_date || [],
      daily_flow_summary: currentSnapshot.daily_flow_summary || [],
      by_from_bank: currentSnapshot.by_from_bank || [],
      by_to_bank: currentSnapshot.by_to_bank || [],
      by_sender: currentSnapshot.by_sender || []
    }});
  }} else {{
    currentSnapshot = data;
  }}
  const selectedBot = data.selected_bot_key || currentBot || '__all__';
  const companyOptions = '<option value="__all__">ทุกบริษัท</option>' + (data.telegram_bots || []).map(b => '<option value="'+esc(b.bot_key)+'">'+esc(b.company_name || b.bot_key)+'</option>').join('');
  botEl.innerHTML = companyOptions;
  botEl.value = selectedBot || '__all__';
  if (exportCompanyEl) {{
    exportCompanyEl.innerHTML = companyOptions;
    exportCompanyEl.value = selectedBot || '__all__';
  }}
  const reconcileCompanyFilter = document.getElementById('reconcileCompanyFilter');
  const previousReconcileBot = reconcileCompanyFilter ? (reconcileCompanyFilter.value || '') : '';
  if (reconcileCompanyFilter) {{
    reconcileCompanyFilter.innerHTML = companyOptions;
    const hasPreviousReconcileBot = previousReconcileBot && Array.from(reconcileCompanyFilter.options || []).some(opt => opt.value === previousReconcileBot);
    reconcileCompanyFilter.value = hasPreviousReconcileBot ? previousReconcileBot : (selectedBot && selectedBot !== '__all__' ? selectedBot : '__all__');
  }}
  const ledgerCompanyFilter = document.getElementById('ledgerCompanyFilter');
  const previousLedgerBot = ledgerCompanyFilter ? (ledgerCompanyFilter.value || '') : '';
  if (ledgerCompanyFilter) {{
    ledgerCompanyFilter.innerHTML = companyOptions;
    const hasPreviousLedgerBot = previousLedgerBot && Array.from(ledgerCompanyFilter.options || []).some(opt => opt.value === previousLedgerBot);
    ledgerCompanyFilter.value = hasPreviousLedgerBot ? previousLedgerBot : (selectedBot && selectedBot !== '__all__' ? selectedBot : '__all__');
  }}
  flowEl.value = data.flow_type || currentFlow || 'all';
  const activeFlow = flowEl.value || 'all';
  const reconcileFlowFilter = document.getElementById('reconcileFlowFilter');
  if (reconcileFlowFilter && !reconcileFlowFilter.value && (activeFlow === 'deposit' || activeFlow === 'withdraw')) {{ reconcileFlowFilter.value = activeFlow; }}
  const ledgerFlowFilter = document.getElementById('ledgerFlowFilter');
  if (ledgerFlowFilter && (activeFlow === 'deposit' || activeFlow === 'withdraw')) {{ ledgerFlowFilter.value = activeFlow; }}
  const chatRows = (data.chats || []).filter(c => String(c.bot_key || 'default') === String(selectedBot || 'default') && (activeFlow === 'all' || String(c.flow_type || 'other') === activeFlow));
  if (selectedBot === '__all__') {{
    chatEl.disabled = true;
    chatEl.innerHTML = '<option value="__all__||'+esc(activeFlow)+'">รวมทุกบริษัท · '+flowName(activeFlow)+'</option>';
  }} else if (chatRows.length) {{
    chatEl.disabled = false;
    const allLabel = activeFlow === 'all' ? 'ทุกกลุ่มของบริษัทนี้' : ('ทุกกลุ่ม'+flowName(activeFlow)+'ของบริษัทนี้');
    chatEl.innerHTML = '<option value="'+esc((selectedBot || 'default')+'||'+activeFlow)+'">'+esc(allLabel)+'</option>' + chatRows.map(c => '<option value="'+esc((c.bot_key || 'default')+'|'+c.chat_id+'|'+(c.flow_type || 'other'))+'">'+esc('['+flowChipName(c.flow_type)+'] '+(c.chat_title || c.chat_id))+' ('+money(c.open_amount)+')</option>').join('');
    const selectedValue = (selectedBot || 'default') + '|' + (data.selected_chat_id || '') + '|' + (data.flow_type || activeFlow);
    chatEl.value = data.selected_chat_id ? selectedValue : ((selectedBot || 'default')+'||'+activeFlow);
  }} else {{
    chatEl.disabled = true;
    const noChatLabel = activeFlow === 'all' ? 'ยังไม่มีกลุ่มของบริษัทนี้' : ('ยังไม่มีกลุ่ม'+flowName(activeFlow)+'ของบริษัทนี้');
    chatEl.innerHTML = '<option value="'+esc((selectedBot || 'default')+'||'+activeFlow)+'">'+esc(noChatLabel)+'</option>';
  }}
  document.getElementById('withdrawAmount').textContent = money(data.totals.withdraw_limit_amount || 0);
  document.getElementById('withdrawCount').textContent = data.totals.withdraw_limit_count || 0;
  const withdrawCapacity = Number(data.totals.withdraw_limit_capacity_amount || 0);
  const withdrawRemaining = Number(data.totals.withdraw_limit_remaining_amount || 0);
  const withdrawUsagePct = Number(data.totals.withdraw_limit_usage_percent || 0);
  const withdrawOver = Number(data.totals.withdraw_limit_over_amount || 0);
  const usageRatioEl = document.getElementById('withdrawLimitUsageRatio');
  const usageMetaEl = document.getElementById('withdrawLimitUsageMeta');
  const remainingEl = document.getElementById('withdrawLimitRemaining');
  const remainingMetaEl = document.getElementById('withdrawLimitRemainingMeta');
  if (usageRatioEl) usageRatioEl.textContent = withdrawCapacity ? (money(data.totals.withdraw_limit_amount || 0) + ' / ' + money(withdrawCapacity)) : '-';
  if (usageMetaEl) usageMetaEl.textContent = withdrawCapacity ? (withdrawUsagePct.toFixed(1) + '% · ' + esc(data.totals.withdraw_limit_account_day_count || 0) + ' บัญชี/วัน') : 'ยังไม่มีวงเงินถอนใน scope นี้';
  if (remainingEl) {{
    remainingEl.textContent = withdrawCapacity ? money(withdrawRemaining) : '-';
    remainingEl.classList.toggle('bad', withdrawRemaining < 0 || withdrawOver > 0);
    remainingEl.classList.toggle('good', !(withdrawRemaining < 0 || withdrawOver > 0));
  }}
  if (remainingMetaEl) remainingMetaEl.textContent = withdrawOver > 0 ? ('เกินรวม ' + money(withdrawOver)) : 'เหลือจากวงเงินรวม';
  document.getElementById('depositAmount').textContent = money(data.totals.deposit_customer_amount || 0);
  document.getElementById('depositCount').textContent = data.totals.deposit_customer_count || 0;
  document.getElementById('queued').textContent = data.jobs.queued || 0;
  document.getElementById('processing').textContent = (data.jobs.processing || 0) + ' / ' + (data.jobs.failed || 0);
  document.getElementById('duplicateCount').textContent = data.totals.selected_duplicate_count || 0;
  document.getElementById('duplicateAmount').textContent = 'ยอดซ้ำ ' + money(data.totals.selected_duplicate_amount || 0);
  document.getElementById('sourceBankReviewCount').textContent = data.totals.source_bank_review_count || 0;
  const tw = data.twallet_summary || {{}};
  document.getElementById('twalletTodayAmount').textContent = tw.ok ? money(tw.today_total || 0) : '-';
  document.getElementById('twalletTodayMeta').textContent = tw.ok ? ((tw.today_count || 0)+' รายการ · คงเหลือ '+money(tw.balance_amount || 0)) : ('TWallet: '+(tw.error || 'offline'));
  if (data.scope && ['today','open','all'].includes(String(data.scope))) scopeEl.value = data.scope;
  if (customDateEl && /^\\d{{4}}-\\d{{2}}-\\d{{2}}$/.test(data.scope || '')) {{
    customDateEl.value = data.scope;
    if (summaryStartEl) summaryStartEl.value = '';
    if (summaryEndEl) summaryEndEl.value = '';
    scopeEl.value = 'all';
  }} else if (String(data.scope || '').startsWith('range:')) {{
    const rangeText = String(data.scope || '').slice(6);
    const rangeParts = rangeText.split('..');
    if (summaryStartEl) summaryStartEl.value = rangeParts[0] || '';
    if (summaryEndEl) summaryEndEl.value = rangeParts[1] || '';
    if (customDateEl) customDateEl.value = '';
    scopeEl.value = 'all';
  }} else if (customDateEl && !currentDate) {{
    customDateEl.value = '';
  }}
  if (data.slip_filter) slipFilterEl.value = data.slip_filter;
  if (document.activeElement !== slipSearchEl) slipSearchEl.value = data.slip_search || currentSearch || '';
  document.getElementById('companyOverview').innerHTML = renderCompanyOverview(data.company_summary);
  document.getElementById('twalletSummary').innerHTML = renderTWalletSummary(data.twallet_summary);
  document.getElementById('operatorHome').innerHTML = renderOperatorHome(data.company_summary);
  document.getElementById('exceptionQueue').innerHTML = renderExceptionQueue(data);
  wireExceptionButtons();
  const bankLedgerSummaryEl = document.getElementById('bankLedgerSummary');
  if (bankLedgerSummaryEl && data.bank_ledger_summary) bankLedgerSummaryEl.innerHTML = renderBankLedgerSnapshot(data.bank_ledger_summary);
  document.getElementById('sideCompanies').innerHTML = renderSideCompanies(data.company_menu || data.company_summary);
  wireCompanyButtons();
  document.getElementById('botSettings').innerHTML = renderTelegramBots(data.telegram_bots);
  updateTopStatus(data, selectedBot, activeFlow);
  closeOpenPeriodGuard();
  const activeSummary = document.getElementById('activeSelectionSummary');
  if (activeSummary) activeSummary.innerHTML = '<div><b>'+esc(selectedBot === '__all__' ? 'ทุกบริษัท' : (selectedBot || '-'))+'</b> · '+esc(flowName(data.flow_type || activeFlow))+' · '+esc(data.scope_label || data.scope || scopeName(data.scope || ''))+'</div><div class="mini">ใช้ตัวกรองและส่งออก Excel ได้จาก side menu ด้านซ้าย</div>';
  if (!isLiteSnapshot) {{
    document.getElementById('companyAccounts').innerHTML = renderCompanyAccounts(data.company_accounts);
    const accountDailyRows = data.company_account_daily || [];
    document.getElementById('companyAccountDailyWithdraw').innerHTML = renderCompanyAccountDaily(accountDailyRows.filter(r => r.flow_type === 'withdraw'));
    document.getElementById('companyAccountDailyDeposit').innerHTML = renderCompanyAccountDaily(accountDailyRows.filter(r => r.flow_type === 'deposit'));
    document.getElementById('accountSlipSearch').innerHTML = renderAccountSlipSearch(data.account_slip_search);
    document.getElementById('accountCrossCompany').innerHTML = renderAccountCrossCompany(data.account_cross_company);
    document.getElementById('crossCompanyAccountSlipSearch').innerHTML = renderCrossCompanyAccountSlipSearch(data.cross_company_account_slip_search);
    const accountCrossBlock = document.getElementById('accountCrossCompanyBlock');
    if (accountCrossBlock) accountCrossBlock.hidden = !((data.account_cross_company || []).length > 0);
    const crossSearchBlock = document.getElementById('crossCompanyAccountSlipSearchBlock');
    const crossSearchData = data.cross_company_account_slip_search || {{}};
    if (crossSearchBlock) crossSearchBlock.hidden = !(crossSearchData.is_cross_company === true && (crossSearchData.company_count || 0) > 0);
    wireAccountSearchButtons();
    const selectedBotRow = (data.telegram_bots || []).find(b => b.bot_key === selectedBot) || {{}};
    if (!document.getElementById('companyName').value) document.getElementById('companyName').value = selectedBotRow.company_name || selectedBot || '';
    document.getElementById('dailyFlowChart').innerHTML = renderDailyFlowChart(data.daily_flow_summary);
    document.getElementById('byDate').innerHTML = dateTable(data.by_date);
    const withdrawSummary = document.getElementById('withdrawLimitSummary');
    const withdrawUsageChart = document.getElementById('withdrawLimitUsageChart');
    const depositSummary = document.getElementById('depositCustomerSummary');
    if (withdrawUsageChart) withdrawUsageChart.innerHTML = renderWithdrawLimitUsageChart(data.withdraw_limit_usage || [], data.totals || {{}});
    if (withdrawSummary) withdrawSummary.textContent = 'ฝั่งถอน/วงเงิน: ' + (data.totals.withdraw_limit_count || 0) + ' สลิป · ' + money(data.totals.withdraw_limit_amount || 0) + ' / วงเงินรวม ' + money(data.totals.withdraw_limit_capacity_amount || 0) + ' · ไม่รวมฝาก/เติมมือ';
    if (depositSummary) depositSummary.textContent = 'ฝั่งฝาก/เติมมือ: ' + (data.totals.deposit_customer_count || 0) + ' สลิป · ' + money(data.totals.deposit_customer_amount || 0) + ' · สลิปลูกค้า ไม่มีวงเงิน';
    if (data.limit_check_enabled === false) {{
      const limitNote = '<div class="muted">กลุ่มฝาก/เติมมือ ไม่ต้องเช็กวงเงิน</div>';
      document.getElementById('byAccountDay').innerHTML = limitNote;
      document.getElementById('byTransferor').innerHTML = limitNote;
    }} else {{
      document.getElementById('byAccountDay').innerHTML = dailyAccountLimitTable(data.by_account_day);
      document.getElementById('byTransferor').innerHTML = transferorLimitTable(data.by_account_day);
    }}
    document.getElementById('depositCustomerSlips').innerHTML = recentCards(data.deposit_customer_slips || []);
    document.getElementById('bySender').innerHTML = aggregateTable(data.by_sender);
    document.getElementById('byFromBank').innerHTML = sourceBankTable(data.by_from_bank);
    document.getElementById('byToBank').innerHTML = aggregateTable(data.by_to_bank);
    document.getElementById('duplicatePairs').innerHTML = renderDuplicatePairs(data.duplicate_pairs);
    wireDuplicateButtons();
    document.getElementById('sourceBankReview').innerHTML = sourceBankReviewCards(data.source_bank_review);
    document.getElementById('recent').innerHTML = recentCards(data.recent);
    document.getElementById('issues').innerHTML = renderQueueIssues(data.jobs_recent || [], data.issues || []);
    document.getElementById('usage').innerHTML = table(data.provider_usage, [['provider','provider'], ['model','model'], ['status','status'], ['calls','count'], ['input','input_tokens'], ['output','output_tokens'], ['thought','thought_tokens'], ['total token','total_tokens'], ['cost $','cost_usd']]);
  }}
  enhanceResponsiveTables();
  updateReconcileScopePreview();
  updateExcel();
  if (!initialMenuApplied || (options && options.home)) {{
    initialMenuApplied = true;
    const initialHashTarget = String(location.hash || '').replace(/^#/, '');
    if (initialHashTarget && document.getElementById(initialHashTarget) && document.getElementById(initialHashTarget).classList.contains('menu-section')) {{
      showMenuSection(initialHashTarget, {{scroll:false, persist:false}});
    }} else if (!(options && options.scrollTarget)) {{
      showMenuSection('section-operator-home', {{scroll:false, persist:false}});
    }}
  }}
  if (options && options.scrollTarget) scrollElementIntoView(options.scrollTarget, options.smooth !== false);
  else if (options && options.scrollTop) scrollDashboardTop(options.smooth !== false);
  try {{ loadPendingActions({{scrollTop:false}}); }} catch (e) {{}}
}}
function selectedDashboardScope() {{
  const scopeEl = document.getElementById('scope');
  const customDateEl = document.getElementById('customDateFilter');
  const summaryStartEl = document.getElementById('summaryStartDate');
  const summaryEndEl = document.getElementById('summaryEndDate');
  const customDate = customDateEl ? (customDateEl.value || '') : '';
  const summaryStart = summaryStartEl ? (summaryStartEl.value || '') : '';
  const summaryEnd = summaryEndEl ? (summaryEndEl.value || '') : '';
  const rangeScope = (summaryStart || summaryEnd) ? `range:${{summaryStart}}..${{summaryEnd}}` : '';
  return rangeScope || customDate || (scopeEl ? scopeEl.value : '') || 'today';
}}
function buildCrossCompanyAccountExcelUrl(queryText) {{
  const parts = selectedChatParts();
  const flow = document.getElementById('flowFilter').value || parts.flow_type || 'all';
  const scope = (currentSnapshot && currentSnapshot.scope) || selectedDashboardScope();
  return '/api/export' + query({{bot_key:'__all__', flow_type:flow, scope, cross_account_search:queryText || ''}});
}}
function buildExcelUrl() {{
  const parts = selectedChatParts();
  const exportBot = (document.getElementById('exportCompanyFilter') || {{}}).value || '';
  const bot = exportBot || selectedBotKey() || parts.bot_key;
  const scope = document.getElementById('scope').value || 'today';
  const flow = document.getElementById('flowFilter').value || parts.flow_type || 'all';
  const start_date = document.getElementById('exportStartDate').value || '';
  const end_date = document.getElementById('exportEndDate').value || '';
  let chat = (parts.bot_key === bot && (flow === 'all' || parts.flow_type === flow)) ? parts.chat_id : '';
  if (currentSnapshot && String(currentSnapshot.selected_bot_key || '') === String(bot || '')) {{
    chat = currentSnapshot.selected_chat_id || chat;
  }}
  return '/api/export' + query({{chat_id:chat, bot_key:bot, flow_type:flow, scope, start_date, end_date}});
}}
function updateExcel() {{
  document.getElementById('excel').href = buildExcelUrl();
}}
function exportExcel() {{
  updateExcel();
  const status = document.getElementById('statusline');
  if (status) status.textContent = 'กำลังส่งออก Excel ตามบริษัท/ช่วงวันที่ที่เลือก...';
  return true;
}}
function clearSlipSearch() {{
  document.getElementById('slipSearch').value = '';
  load({{scrollTop:true}});
}}
async function closeOpenPeriod() {{
  const parts = selectedChatParts();
  const chat = parts.chat_id;
  const bot = selectedBotKey() || parts.bot_key;
  const note = document.getElementById('closeNote').value || 'dashboard close';
  const status = document.getElementById('statusline');
  if (!chat) {{ status.textContent = 'ยังไม่มีห้องให้ปิดรอบ'; return; }}
  if (!await dashboardConfirm('ยืนยันปิดรอบ/เคลียร์ยอดเปิดของห้องนี้? ประวัติจะไม่ถูกลบ', 'ยืนยันปิดรอบ', true)) return;
  status.textContent = 'กำลังปิดรอบ...';
  const res = await fetch('/api/close'+query(), {{
    method: 'POST',
    headers: postHeaders(),
    body: JSON.stringify({{chat_id: chat, bot_key: bot, company_name: (currentSnapshot && currentSnapshot.company_accounts && currentSnapshot.company_accounts[0] && currentSnapshot.company_accounts[0].company_name) || '', note}})
  }});
  const data = await res.json();
  if (!res.ok || !data.ok) {{ status.textContent = 'ปิดรอบไม่สำเร็จ: ' + (data.error || res.status); return; }}
  status.textContent = 'ปิดรอบแล้ว ' + data.closed_count + ' สลิป · ' + money(data.total_amount);
  await load({{scrollTop:true}});
}}
document.getElementById('botFilter').addEventListener('change', () => load({{scrollTop:true}}));
document.getElementById('exportCompanyFilter').addEventListener('change', () => {{
  const bot = document.getElementById('exportCompanyFilter').value || '__all__';
  const flow = document.getElementById('flowFilter').value || 'all';
  document.getElementById('botFilter').value = bot;
  document.getElementById('chat').value = bot + '||' + flow;
  load({{scrollTop:true}});
}});
document.getElementById('flowFilter').addEventListener('change', () => load({{scrollTop:true}}));
document.getElementById('chat').addEventListener('change', () => load({{scrollTop:true}}));
document.getElementById('scope').addEventListener('change', () => {{
  document.getElementById('customDateFilter').value = '';
  document.getElementById('summaryStartDate').value = '';
  document.getElementById('summaryEndDate').value = '';
  load({{scrollTop:true}});
}});
const customDateEl = document.getElementById('customDateFilter');
const summaryStartDateEl = document.getElementById('summaryStartDate');
const summaryEndDateEl = document.getElementById('summaryEndDate');
customDateEl.addEventListener('change', () => {{
  if (customDateEl.value) {{
    document.getElementById('scope').value = 'all';
    summaryStartDateEl.value = '';
    summaryEndDateEl.value = '';
  }}
  load({{scrollTop:true}});
}});
summaryStartDateEl.addEventListener('change', () => {{ if (summaryStartDateEl.value || summaryEndDateEl.value) {{ customDateEl.value = ''; document.getElementById('scope').value = 'all'; }} load({{scrollTop:true}}); }});
summaryEndDateEl.addEventListener('change', () => {{ if (summaryStartDateEl.value || summaryEndDateEl.value) {{ customDateEl.value = ''; document.getElementById('scope').value = 'all'; }} load({{scrollTop:true}}); }});
document.getElementById('slipFilter').addEventListener('change', () => load({{scrollTop:true}}));
document.getElementById('reconcileCompanyFilter').addEventListener('change', updateReconcileScopePreview);
document.getElementById('reconcileFlowFilter').addEventListener('change', updateReconcileScopePreview);
document.getElementById('reconcileDateScope').addEventListener('change', updateReconcileScopePreview);
document.getElementById('exportStartDate').addEventListener('change', updateExcel);
document.getElementById('exportEndDate').addEventListener('change', updateExcel);
document.getElementById('slipSearch').addEventListener('keydown', (event) => {{ if (event.key === 'Enter') {{ window.__accountSearchMode = 'scoped'; load({{scrollTop:true}}); }} }});
(function () {{
  const pendingFilterEl = document.getElementById('pendingStatusFilter');
  if (pendingFilterEl) pendingFilterEl.addEventListener('change', () => loadPendingActions({{scrollTop:false}}));
}})();
try {{ if (location.search && /[?&]token=/.test(location.search)) {{ const cleaned = location.search.replace(/([?&])token=[^&]*/g, '$1').replace(/[?&]$/, '').replace(/&&+/g, '&').replace(/^\\?&/, '?'); window.history.replaceState({{}}, '', location.pathname + (cleaned && cleaned !== '?' ? cleaned : '') + location.hash); }} }} catch (e) {{}}
load({{home:true, scrollTop:true, smooth:false}}); setInterval(() => load({{lite:true}}), 10000);
loadPendingActions({{scrollTop:false}}); refreshPendingBadge(); setInterval(refreshPendingBadge, 30000);
</script>
</body>
</html>"""


class DashboardHandler(BaseHTTPRequestHandler):
    server_version = "AuditslipDashboard/1.0"

    def log_message(self, fmt: str, *args: Any) -> None:
        # Access logging disabled to prevent token leakage in URL query strings.
        return

    def token_from_request(self) -> str:
        # Priority: (1) HttpOnly cookie, (2) Authorization: Bearer, (3) ?token= legacy fallback
        # accepted ONLY on '/' or '/index.html' to prevent token leakage via Referer header on
        # /api/* responses. After the first hit to '/', the cookie is set and the URL is scrubbed.
        cookie = self.headers.get("Cookie", "")
        for part in cookie.split(";"):
            if "=" in part:
                k, v = part.strip().split("=", 1)
                if k == COOKIE_NAME:
                    return v
        auth = self.headers.get("Authorization", "")
        if auth.lower().startswith("bearer "):
            return auth.split(" ", 1)[1].strip()
        parsed = urlparse(self.path)
        if parsed.path in {"/", "/index.html"}:
            q = parse_qs(parsed.query)
            if q.get("token"):
                return q["token"][0]
        return ""

    def authorized(self) -> bool:
        # Resolve role; any non-empty role means an authorized request.
        return bool(self.actor_role())

    def actor_role(self) -> str:
        """Resolve the role for the current request.

        - Look up sha256(token) in dashboard_tokens; if active, return stored role.
        - If the token IS registered (revoked or otherwise inactive): return "".
          The DB is the source of truth — legacy DASHBOARD_TOKEN cannot bypass a
          revoke once it has been registered.
        - If the token is NOT registered AT ALL but matches the legacy
          DASHBOARD_TOKEN env, return "admin" as a bootstrap fallback for fresh
          installs before bootstrap_dashboard_admin_token has run.
        - Otherwise return "".
        """
        tok = self.token_from_request()
        if not tok:
            return ""
        try:
            role = lookup_dashboard_token_role(DB_PATH, tok)
        except Exception:
            role = ""
        if role:
            return role
        # Token might be registered but revoked — DO NOT fall back to legacy.
        try:
            registered = dashboard_token_is_registered(DB_PATH, tok)
        except Exception:
            registered = False
        if registered:
            return ""
        # Pre-bootstrap legacy fallback only.
        if DASHBOARD_TOKEN:
            import hmac as _hmac
            if _hmac.compare_digest(tok, DASHBOARD_TOKEN):
                return "admin"
        owner_role = dashboard_owner_session_role(tok)
        if owner_role:
            return owner_role
        return ""

    def require_role(self, *allowed: str) -> bool:
        return self.actor_role() in set(allowed)

    def role_or_401(self, *allowed: str) -> bool:
        if self.require_role(*allowed):
            return True
        self.send_bytes(HTTPStatus.UNAUTHORIZED, b"Unauthorized", "text/plain; charset=utf-8")
        return False

    def actor_fingerprint(self) -> str:
        tok = self.token_from_request()
        if not tok:
            return ""
        return hashlib.sha256(tok.encode("utf-8", "replace")).hexdigest()[:12]

    def parse_approval_param(self) -> str:
        """Read ?approval=request|execute from URL query (default 'request')."""
        parsed = urlparse(self.path)
        q = parse_qs(parsed.query)
        val = (q.get("approval") or ["request"])[0].strip().lower()
        return val if val in {"request", "execute"} else "request"

    def cookie_attrs(self) -> str:
        attrs = ["HttpOnly", "SameSite=Strict", "Path=/", "Max-Age=86400"]
        if self.headers.get("X-Forwarded-Proto", "").lower() == "https":
            attrs.append("Secure")
        return "; ".join(attrs)

    def csrf_authorized(self) -> bool:
        return self.headers.get("X-Auditslip-Action", "") == "dashboard"

    def security_headers(self) -> Dict[str, str]:
        return {
            "Content-Security-Policy": "default-src 'self'; base-uri 'none'; frame-ancestors 'none'; form-action 'self'; img-src 'self' data: https: http:; style-src 'self' 'unsafe-inline'; script-src 'self' 'unsafe-inline'",
            "X-Frame-Options": "DENY",
            "Referrer-Policy": "no-referrer",
            "X-Content-Type-Options": "nosniff",
            "Cache-Control": "private, no-store",
        }

    def send_bytes(self, status: int, body: bytes, content_type: str, extra_headers: Dict[str, str] | None = None) -> None:
        self.send_response(status)
        headers = self.security_headers()
        headers.update(extra_headers or {})
        headers["Content-Type"] = content_type
        headers["Content-Length"] = str(len(body))
        for k, v in headers.items():
            self.send_header(k, v)
        if self.token_from_request() == DASHBOARD_TOKEN:
            self.send_header("Set-Cookie", f"{COOKIE_NAME}={DASHBOARD_TOKEN}; {self.cookie_attrs()}")
        self.end_headers()
        self.wfile.write(body)

    def send_json(self, obj: Any, status: int = 200) -> None:
        self.send_bytes(status, json.dumps(obj, ensure_ascii=False).encode("utf-8"), "application/json; charset=utf-8")

    def read_simple_payload(self) -> Dict[str, Any]:
        length = int(self.headers.get("Content-Length", "0") or 0)
        raw = self.rfile.read(length).decode("utf-8") if length else ""
        if not raw:
            return {}
        content_type = self.headers.get("Content-Type", "")
        if "application/json" in content_type:
            try:
                obj = json.loads(raw)
                return obj if isinstance(obj, dict) else {}
            except Exception:
                return {}
        return {k: v[0] for k, v in parse_qs(raw).items()}

    def set_owner_session_cookie(self) -> str:
        session_token = dashboard_owner_session_token()
        return f"{COOKIE_NAME}={session_token}; {self.cookie_attrs()}"

    def clear_owner_session_cookie(self) -> str:
        return f"{COOKIE_NAME}=; HttpOnly; SameSite=Strict; Path=/; Max-Age=0"

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        q = parse_qs(parsed.query)
        if parsed.path == "/api/health":
            if health_env_bool("AUDITSLIP_HEALTH_QUICK_DEFAULT", False) or q.get("quick") in (["1"], ["true"], ["yes"], ["on"]):
                self.send_json(dashboard_quick_health(DB_PATH))
            else:
                self.send_json(dashboard_operational_health(DB_PATH))
            return
        if q.get("token") and parsed.path in {"/", "/index.html"}:
            clean_q = {k: v for k, v in q.items() if k != "token"}
            new_query = urlencode([(k, vv) for k, vs in clean_q.items() for vv in vs])
            new_loc = parsed.path + (f"?{new_query}" if new_query else "")
            self.send_response(HTTPStatus.FOUND)
            for k, v in self.security_headers().items():
                self.send_header(k, v)
            self.send_header("Set-Cookie", f"{COOKIE_NAME}={DASHBOARD_TOKEN}; {self.cookie_attrs()}")
            self.send_header("Location", new_loc)
            self.send_header("Content-Length", "0")
            self.end_headers()
            return
        if parsed.path in {"/", "/index.html"}:
            self.send_bytes(200, render_dashboard_html().encode("utf-8"), "text/html; charset=utf-8")
            return
        if parsed.path == "/api/tokens":
            if not self.role_or_401("admin"):
                return
            try:
                self.send_json({"ok": True, "tokens": list_dashboard_tokens(DB_PATH)})
            except Exception as exc:
                logger.exception("api_tokens_list failed")
                self.send_json({"ok": False, "error": safe_error(exc)}, 400)
            return
        if parsed.path == "/api/summary":
            # Lazy expiry hook (cheap; runs at most once per request).
            try:
                expire_old_pending_actions(DB_PATH)
            except Exception:
                logger.exception("expire_old_pending_actions failed during /api/summary")
            q = parse_qs(parsed.query)
            chat_id = (q.get("chat_id") or [""])[0]
            bot_key = (q.get("bot_key") or [""])[0]
            scope = (q.get("scope") or ["today"])[0]
            flow_type = (q.get("flow_type") or ["all"])[0]
            slip_filter = (q.get("slip_filter") or ["all"])[0]
            slip_search = (q.get("slip_search") or [""])[0]
            account_search_mode = (q.get("account_search_mode") or [""])[0] or "scoped"
            detail_level = (q.get("detail") or q.get("detail_level") or ["full"])[0]
            self.send_json(dashboard_snapshot(DB_PATH, chat_id=chat_id, bot_key=bot_key, flow_type=flow_type, scope=scope, slip_filter=slip_filter, slip_search=slip_search, account_search_mode=account_search_mode, detail_level=detail_level))
            return
        if parsed.path == "/api/ledger":
            q = parse_qs(parsed.query)
            chat_id = (q.get("chat_id") or [""])[0]
            bot_key = (q.get("bot_key") or [""])[0]
            account_key = (q.get("account_key") or [""])[0]
            date_from = (q.get("date_from") or [""])[0]
            date_to = (q.get("date_to") or [""])[0]
            flow_type = (q.get("flow_type") or ["all"])[0]
            try:
                limit = max(1, min(2000, int((q.get("limit") or ["500"])[0])))
            except (TypeError, ValueError):
                limit = 500
            try:
                self.send_json(account_ledger_rows(DB_PATH, bot_key=bot_key, chat_id=chat_id, account_key=account_key, date_from=date_from, date_to=date_to, flow_type=flow_type, limit=limit))
            except Exception as exc:
                logger.exception("api_ledger failed")
                self.send_json({"ok": False, "error": safe_error(exc), "rows": []}, 400)
            return
        # --- audit employee endpoints ---
        if parsed.path == "/api/audit/reconcile":
            q = parse_qs(parsed.query)
            bot_key = (q.get("bot_key") or [""])[0]
            chat_id = (q.get("chat_id") or [""])[0]
            account_key = (q.get("account_key") or [""])[0]
            scope = (q.get("scope") or ["open"])[0]
            flow_type = (q.get("flow_type") or ["all"])[0]
            try:
                self.send_json(audit_employee_component.reconcile_slips_ledger(
                    DB_PATH, bot_key=bot_key, chat_id=chat_id, account_key=account_key, scope=scope, flow_type=flow_type
                ))
            except Exception as exc:
                logger.exception("api_audit_reconcile failed")
                self.send_json({"ok": False, "error": safe_error(exc)}, 400)
            return
        if parsed.path == "/api/audit/daily-variance":
            q = parse_qs(parsed.query)
            bot_key = (q.get("bot_key") or [""])[0]
            chat_id = (q.get("chat_id") or [""])[0]
            scope = (q.get("scope") or ["open"])[0]
            flow_type = (q.get("flow_type") or ["all"])[0]
            try:
                threshold = max(0.0, float((q.get("threshold") or ["100"])[0]))
            except (TypeError, ValueError):
                threshold = 100.0
            try:
                self.send_json(audit_employee_component.employee_daily_variance(
                    DB_PATH, bot_key=bot_key, chat_id=chat_id, scope=scope, flow_type=flow_type, threshold=threshold
                ))
            except Exception as exc:
                logger.exception("api_audit_daily_variance failed")
                self.send_json({"ok": False, "error": safe_error(exc)}, 400)
            return
        if parsed.path == "/api/audit/cross-dup":
            q = parse_qs(parsed.query)
            bot_key = (q.get("bot_key") or [""])[0]
            scope = (q.get("scope") or ["open"])[0]
            try:
                self.send_json(audit_employee_component.cross_bot_duplicates(
                    DB_PATH, bot_key=bot_key, scope=scope
                ))
            except Exception as exc:
                logger.exception("api_audit_cross_dup failed")
                self.send_json({"ok": False, "error": safe_error(exc)}, 400)
            return
        if parsed.path == "/api/pending":
            actor_role = self.actor_role()
            if actor_role not in {"admin", "operator", "auditor"}:
                # Public read-only dashboard polling must not reveal approval queue details.
                self.send_json({"ok": True, "redacted": True, "items": [], "count": 0, "current_role": actor_role or "public", "simple_approval_enabled": False})
                return
            try:
                expire_old_pending_actions(DB_PATH)
            except Exception:
                logger.exception("expire_old_pending_actions failed during /api/pending")
            q = parse_qs(parsed.query)
            status_filter = (q.get("status") or [""])[0]
            try:
                limit = int((q.get("limit") or ["200"])[0])
            except (TypeError, ValueError):
                limit = 200
            rows = list_pending_actions(DB_PATH, status=status_filter, limit=limit)
            actor_fp = self.actor_fingerprint()
            self.send_json({"ok": True, "items": rows, "count": len(rows), "current_actor": actor_fp, "current_role": actor_role, "simple_approval_enabled": simple_approval_enabled(DB_PATH, actor_fp, actor_role)})
            return
        if parsed.path == "/api/slip-image":
            q = parse_qs(parsed.query)
            slip_id = (q.get("id") or [""])[0]
            try:
                body, mime = fetch_slip_image(DB_PATH, slip_id)
                self.send_bytes(200, body, mime, {"Cache-Control": "private, max-age=86400"})
            except FileNotFoundError as exc:
                self.send_json({"ok": False, "error": str(exc)}, 404)
            except Exception as exc:
                logger.exception("api_slip_image failed slip_id=%s", slip_id)
                self.send_json({"ok": False, "error": safe_error(exc)}, 404)
            return
        if parsed.path in {"/api/export", "/api/export/preview"}:
            q = parse_qs(parsed.query)
            chat_id = (q.get("chat_id") or [""])[0]
            bot_key = (q.get("bot_key") or ["default"])[0] or "default"
            scope = (q.get("scope") or ["open"])[0]
            flow_type = (q.get("flow_type") or ["all"])[0]
            start_date = (q.get("start_date") or [""])[0]
            end_date = (q.get("end_date") or [""])[0]
            cross_account_search = (q.get("cross_account_search") or [""])[0]
            dry_run = parsed.path == "/api/export/preview" or clean_display((q.get("dry_run") or q.get("preview") or [""])[0]).lower() in {"1", "true", "yes", "y", "on"}
            try:
                if dry_run:
                    company_name = ""
                    if chat_id and not clean_display(cross_account_search):
                        selection = resolve_export_selection(DB_PATH, chat_id=chat_id, bot_key=bot_key, flow_type=flow_type)
                        if not selection.get("ok"):
                            self.send_json(selection, 404)
                            return
                        chat_id = str(selection["chat_id"])
                        bot_key = str(selection["bot_key"])
                        company_name = str(selection.get("company_name") or APP_NAME)
                    self.send_json(export_dashboard_preview(DB_PATH, bot_key=bot_key, chat_id=chat_id, flow_type=flow_type, scope=scope, start_date=start_date, end_date=end_date, cross_account_search=cross_account_search, company_name=company_name))
                    return
                if clean_display(cross_account_search):
                    path = export_cross_company_account_slips_excel(DB_PATH, flow_type=flow_type, scope=scope, search=cross_account_search)
                    body = path.read_bytes()
                    self.send_bytes(
                        200,
                        body,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        {"Content-Disposition": f'attachment; filename="{path.name}"'},
                    )
                    return
                requested_all = clean_display(bot_key) in {"__all__", "all"}
                if requested_all and not chat_id:
                    path = export_dashboard_zip_by_company(DB_PATH, bot_key=bot_key, flow_type=flow_type, scope=scope, start_date=start_date, end_date=end_date)
                    body = path.read_bytes()
                    self.send_bytes(
                        200,
                        body,
                        "application/zip",
                        {"Content-Disposition": f'attachment; filename="{path.name}"'},
                    )
                    return
                if chat_id:
                    selection = resolve_export_selection(DB_PATH, chat_id=chat_id, bot_key=bot_key, flow_type=flow_type)
                    if not selection.get("ok"):
                        self.send_json(selection, 404)
                        return
                    chat_id = str(selection["chat_id"])
                    bot_key = str(selection["bot_key"])
                    company_name = str(selection.get("company_name") or APP_NAME)
                else:
                    company_name = ""
                path = export_dashboard_excel(DB_PATH, bot_key=bot_key, chat_id=chat_id, flow_type=flow_type, scope=scope, start_date=start_date, end_date=end_date, company_name=company_name)
                body = path.read_bytes()
                self.send_bytes(
                    200,
                    body,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    {"Content-Disposition": f'attachment; filename="{path.name}"'},
                )
            except Exception as exc:
                logger.exception("api_export failed")
                self.send_json({"ok": False, "error": safe_error(exc)}, 400)
            return
        if parsed.path == "/api/audit-chain/verify":
            if not self.role_or_401("admin", "auditor"):
                return
            try:
                self.send_json(verify_mutation_chain(DB_PATH))
            except Exception as exc:
                logger.exception("api_audit_chain_verify failed")
                self.send_json({"ok": False, "error": safe_error(exc)}, 500)
            return
        if parsed.path == "/api/audit-chain/tail":
            if not self.role_or_401("admin", "auditor"):
                return
            q = parse_qs(parsed.query)
            limit_str = (q.get("limit") or ["50"])[0]
            try:
                self.send_json({"ok": True, "entries": mutation_chain_tail(DB_PATH, limit=int(limit_str or 50))})
            except Exception as exc:
                logger.exception("api_audit_chain_tail failed")
                self.send_json({"ok": False, "error": safe_error(exc)}, 500)
            return
        self.send_json({"ok": False, "error": "not found"}, 404)

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/api/login":
            if not self.csrf_authorized():
                self.send_json({"ok": False, "error": "missing action header"}, HTTPStatus.FORBIDDEN)
                return
            payload = self.read_simple_payload()
            username = clean_display(payload.get("username"))
            password = str(payload.get("password") or "")
            if not dashboard_owner_credentials_valid(username, password):
                self.send_json({"ok": False, "error": "invalid credentials"}, HTTPStatus.UNAUTHORIZED)
                return
            body = {"ok": True, "role": "admin", "username": DASHBOARD_OWNER_USER, "actor": hashlib.sha256(dashboard_owner_session_token().encode("utf-8")).hexdigest()[:12]}
            self.send_bytes(200, json.dumps(body, ensure_ascii=False).encode("utf-8"), "application/json; charset=utf-8", {"Set-Cookie": self.set_owner_session_cookie()})
            return
        if parsed.path == "/api/logout":
            if not self.csrf_authorized():
                self.send_json({"ok": False, "error": "missing action header"}, HTTPStatus.FORBIDDEN)
                return
            body = {"ok": True, "role": ""}
            self.send_bytes(200, json.dumps(body, ensure_ascii=False).encode("utf-8"), "application/json; charset=utf-8", {"Set-Cookie": self.clear_owner_session_cookie()})
            return
        if not self.authorized():
            self.send_bytes(HTTPStatus.UNAUTHORIZED, b"Unauthorized", "text/plain; charset=utf-8")
            return
        if not self.csrf_authorized():
            self.send_json({"ok": False, "error": "missing action header"}, HTTPStatus.FORBIDDEN)
            return
        content_type = self.headers.get("Content-Type", "")
        payload: Dict[str, Any] = {}
        uploaded_excel_path = ""
        uploaded_statement_path = ""
        if "multipart/form-data" in content_type:
            form = cgi.FieldStorage(fp=self.rfile, headers=self.headers, environ={"REQUEST_METHOD": "POST", "CONTENT_TYPE": content_type})
            for key in ["chat_id", "bot_key", "company_name", "flow_type", "scope", "excel_path", "statement_path", "note", "bank", "account_no", "account_name", "daily_limit"]:
                if key in form and not getattr(form[key], "filename", None):
                    payload[key] = form.getfirst(key, "")
            if "excel" in form and getattr(form["excel"], "filename", None):
                base = Path(os.environ.get("AUDITSLIP_BACKEND_IMPORT_DIR", "/root/projects/auditslip/imports/backend"))
                base.mkdir(parents=True, exist_ok=True)
                filename = re.sub(r"[^A-Za-z0-9_.-]+", "-", Path(form["excel"].filename).name) or "backend.xlsx"
                uploaded_excel_path = str(base / f"upload-{int(time.time())}-{filename}")
                with open(uploaded_excel_path, "wb") as out:
                    out.write(form["excel"].file.read())
            if "statement" in form and getattr(form["statement"], "filename", None):
                base = Path(os.environ.get("AUDITSLIP_BACKEND_IMPORT_DIR", "/root/projects/auditslip/imports/backend"))
                base.mkdir(parents=True, exist_ok=True)
                filename = re.sub(r"[^A-Za-z0-9_.-]+", "-", Path(form["statement"].filename).name) or "statement.xlsx"
                uploaded_statement_path = str(base / f"statement-{int(time.time())}-{filename}")
                with open(uploaded_statement_path, "wb") as out:
                    out.write(form["statement"].file.read())
        else:
            length = int(self.headers.get("Content-Length", "0") or 0)
            raw = self.rfile.read(length).decode("utf-8") if length else ""
            if raw:
                if "application/json" in content_type:
                    try:
                        payload = json.loads(raw)
                    except Exception:
                        payload = {}
                else:
                    payload = {k: v[0] for k, v in parse_qs(raw).items()}
        if parsed.path == "/api/close":
            if not self.role_or_401("admin"):
                return
            actor_fp = self.actor_fingerprint()
            approval_mode = self.parse_approval_param()
            pending_id_executed: int = 0
            if approval_mode == "request":
                request_id = uuid.uuid4().hex[:12]
                pending_id = create_pending_action(
                    DB_PATH,
                    action="period.close",
                    payload=payload,
                    requested_by=actor_fp,
                    request_id=request_id,
                )
                record_endpoint_mutation(DB_PATH, "close.request", actor=actor_fp, request_id=request_id, chat_id=str(payload.get("chat_id") or ""), bot_key=str(payload.get("bot_key") or "default"), payload=payload, result_status="pending", result_summary=f"pending_id={pending_id}")
                auto_result = simple_auto_execute_pending(DB_PATH, pending_id, actor_fp, self.actor_role())
                if auto_result:
                    status_code = int(auto_result.pop("status_code", 200)) if not auto_result.get("ok") else 200
                    self.send_json(auto_result, status_code)
                    return
                self.send_json({"ok": True, "status": "pending", "pending_id": pending_id, "request_id": request_id, "expires_in_hours": PENDING_ACTION_TTL_HOURS})
                return
            # approval_mode == "execute"
            try:
                pending_id_executed = int(payload.get("pending_id") or 0)
            except (TypeError, ValueError):
                pending_id_executed = 0
            if not pending_id_executed:
                self.send_json({"ok": False, "error": "pending_id required for execute"}, 400)
                return
            expire_old_pending_actions(DB_PATH)
            pending_row = load_pending_action(DB_PATH, pending_id_executed)
            if not pending_row or pending_row.get("action") != "period.close":
                self.send_json({"ok": False, "error": "pending action not found"}, 404)
                return
            if pending_row.get("status") != "approved":
                self.send_json({"ok": False, "error": f"not approved (status={pending_row.get('status')})"}, 409)
                return
            # Use the originally-requested payload, NOT the new request body, to prevent target swap.
            stored_payload = pending_action_payload(pending_row)
            chat_id = str(stored_payload.get("chat_id") or "")
            bot_key = str(stored_payload.get("bot_key") or "default")
            company_name = str(stored_payload.get("company_name") or "")
            note = str(stored_payload.get("note") or "dashboard close")
            try:
                result = dashboard_close_period(DB_PATH, chat_id, note, bot_key=bot_key, company_name=company_name)
            except Exception as exc:
                logger.exception("api_close failed")
                close_req_id = str(pending_row.get("request_id") or "")
                record_endpoint_mutation(DB_PATH, "close", actor=actor_fp, request_id=close_req_id, chat_id=chat_id, bot_key=bot_key, payload=stored_payload, result_status="error", result_summary=safe_error(exc))
                self.send_json({"ok": False, "error": safe_error(exc), "pending_id": pending_id_executed}, 400)
                return
            close_req_id = str(pending_row.get("request_id") or "")
            record_endpoint_mutation(DB_PATH, "close", actor=actor_fp, request_id=close_req_id, chat_id=chat_id, bot_key=bot_key, payload=stored_payload, result_status=("ok" if result.get("ok") else "error"), result_summary=str(result.get("settlement_id") or result.get("error") or ""))
            if result.get("ok"):
                mark_pending_executed(DB_PATH, pending_id_executed, executed_result=str(result.get("settlement_id") or ""))
            if isinstance(result, dict):
                result["pending_id"] = pending_id_executed
            self.send_json(result, 200 if result.get("ok") else 400)
            return
        if parsed.path == "/api/account-limit":
            if not self.role_or_401("admin"):
                return
            actor_fp = self.actor_fingerprint()
            approval_mode = self.parse_approval_param()
            if approval_mode != "execute":
                payload_err = account_limit_payload_error(payload)
                if payload_err:
                    self.send_json({"ok": False, "error": payload_err}, 400)
                    return
                request_id = uuid.uuid4().hex[:12]
                try:
                    result = save_account_limit(
                        DB_PATH,
                        str(payload.get("chat_id") or ""),
                        str(payload.get("limit_key") or ""),
                        str(payload.get("display_name") or ""),
                        str(payload.get("bank") or ""),
                        str(payload.get("account") or ""),
                        parse_number(payload.get("limit_amount")),
                    )
                    record_endpoint_mutation(DB_PATH, "account_limit", actor=actor_fp, request_id=request_id, chat_id=str(payload.get("chat_id") or ""), payload=payload, result_status=("ok" if result.get("ok") else "error"), result_summary=str(result.get("limit_key") or result.get("error") or ""))
                    if isinstance(result, dict):
                        result["request_id"] = request_id
                        if result.get("ok"):
                            result["status"] = "saved"
                    self.send_json(result, 200 if result.get("ok") else int(result.get("status_code") or 400))
                except Exception as exc:
                    logger.exception("api_account_limit failed")
                    record_endpoint_mutation(DB_PATH, "account_limit", actor=actor_fp, request_id=request_id, chat_id=str(payload.get("chat_id") or ""), payload=payload, result_status="error", result_summary=safe_error(exc))
                    self.send_json({"ok": False, "error": safe_error(exc), "request_id": request_id}, 400)
                return
            # Legacy support: execute an already-approved account.limit pending action.
            try:
                pending_id_executed = int(payload.get("pending_id") or 0)
            except (TypeError, ValueError):
                pending_id_executed = 0
            if not pending_id_executed:
                self.send_json({"ok": False, "error": "pending_id required for execute"}, 400)
                return
            expire_old_pending_actions(DB_PATH)
            pending_row = load_pending_action(DB_PATH, pending_id_executed)
            if not pending_row or pending_row.get("action") != "account.limit":
                self.send_json({"ok": False, "error": "pending action not found"}, 404)
                return
            if pending_row.get("status") != "approved":
                self.send_json({"ok": False, "error": f"not approved (status={pending_row.get('status')})"}, 409)
                return
            stored_payload = pending_action_payload(pending_row)
            request_id = str(pending_row.get("request_id") or uuid.uuid4().hex[:12])
            try:
                result = save_account_limit(
                    DB_PATH,
                    str(stored_payload.get("chat_id") or ""),
                    str(stored_payload.get("limit_key") or ""),
                    str(stored_payload.get("display_name") or ""),
                    str(stored_payload.get("bank") or ""),
                    str(stored_payload.get("account") or ""),
                    parse_number(stored_payload.get("limit_amount")),
                )
                record_endpoint_mutation(DB_PATH, "account_limit", actor=actor_fp, request_id=request_id, chat_id=str(stored_payload.get("chat_id") or ""), payload=stored_payload, result_status=("ok" if result.get("ok") else "error"), result_summary=str(result.get("limit_key") or result.get("error") or ""))
                if result.get("ok"):
                    mark_pending_executed(DB_PATH, pending_id_executed, executed_result=str(result.get("limit_key") or ""))
                if isinstance(result, dict):
                    result["request_id"] = request_id
                    result["pending_id"] = pending_id_executed
                self.send_json(result)
            except Exception as exc:
                logger.exception("api_account_limit failed")
                record_endpoint_mutation(DB_PATH, "account_limit", actor=actor_fp, request_id=request_id, chat_id=str(stored_payload.get("chat_id") or ""), payload=stored_payload, result_status="error", result_summary=safe_error(exc))
                self.send_json({"ok": False, "error": safe_error(exc), "request_id": request_id, "pending_id": pending_id_executed}, 400)
            return
        if parsed.path == "/api/company-account":
            if not self.role_or_401("admin"):
                return
            actor_fp = self.actor_fingerprint()
            approval_mode = self.parse_approval_param()
            if approval_mode == "request":
                request_id = uuid.uuid4().hex[:12]
                pending_id = create_pending_action(
                    DB_PATH,
                    action="company.account",
                    payload=payload,
                    requested_by=actor_fp,
                    request_id=request_id,
                )
                record_endpoint_mutation(DB_PATH, "company_account.request", actor=actor_fp, request_id=request_id, chat_id=str(payload.get("chat_id") or ""), bot_key=str(payload.get("bot_key") or "default"), payload=payload, result_status="pending", result_summary=f"pending_id={pending_id}")
                auto_result = simple_auto_execute_pending(DB_PATH, pending_id, actor_fp, self.actor_role())
                if auto_result:
                    status_code = int(auto_result.pop("status_code", 200)) if not auto_result.get("ok") else 200
                    self.send_json(auto_result, status_code)
                    return
                self.send_json({"ok": True, "status": "pending", "pending_id": pending_id, "request_id": request_id, "expires_in_hours": PENDING_ACTION_TTL_HOURS})
                return
            # approval_mode == "execute"
            try:
                pending_id_executed = int(payload.get("pending_id") or 0)
            except (TypeError, ValueError):
                pending_id_executed = 0
            if not pending_id_executed:
                self.send_json({"ok": False, "error": "pending_id required for execute"}, 400)
                return
            expire_old_pending_actions(DB_PATH)
            pending_row = load_pending_action(DB_PATH, pending_id_executed)
            if not pending_row or pending_row.get("action") != "company.account":
                self.send_json({"ok": False, "error": "pending action not found"}, 404)
                return
            if pending_row.get("status") != "approved":
                self.send_json({"ok": False, "error": f"not approved (status={pending_row.get('status')})"}, 409)
                return
            stored_payload = pending_action_payload(pending_row)
            request_id = str(pending_row.get("request_id") or uuid.uuid4().hex[:12])
            try:
                result = save_company_account(
                    DB_PATH,
                    bot_key=str(stored_payload.get("bot_key") or "default"),
                    chat_id=str(stored_payload.get("chat_id") or ""),
                    company_name=str(stored_payload.get("company_name") or ""),
                    bank=str(stored_payload.get("bank") or ""),
                    account_no=str(stored_payload.get("account_no") or ""),
                    account_name=str(stored_payload.get("account_name") or ""),
                    daily_limit=parse_number(stored_payload.get("daily_limit")),
                )
                record_endpoint_mutation(DB_PATH, "company_account", actor=actor_fp, request_id=request_id, chat_id=str(stored_payload.get("chat_id") or ""), bot_key=str(stored_payload.get("bot_key") or "default"), payload=stored_payload, result_status=("ok" if result.get("ok") else "error"), result_summary=str(result.get("account_key") or result.get("error") or ""))
                if result.get("ok"):
                    mark_pending_executed(DB_PATH, pending_id_executed, executed_result=str(result.get("account_key") or ""))
                if isinstance(result, dict):
                    result["request_id"] = request_id
                    result["pending_id"] = pending_id_executed
                self.send_json(result)
            except Exception as exc:
                logger.exception("api_company_account failed")
                record_endpoint_mutation(DB_PATH, "company_account", actor=actor_fp, request_id=request_id, chat_id=str(stored_payload.get("chat_id") or ""), bot_key=str(stored_payload.get("bot_key") or "default"), payload=stored_payload, result_status="error", result_summary=safe_error(exc))
                self.send_json({"ok": False, "error": safe_error(exc), "request_id": request_id, "pending_id": pending_id_executed}, 400)
            return
        if parsed.path == "/api/duplicate/unmark":
            if not self.role_or_401("admin", "operator"):
                return
            request_id = uuid.uuid4().hex[:12]
            slip_id_in = str(payload.get("id") or payload.get("slip_id") or "")
            bot_key_in = str(payload.get("bot_key") or "")
            try:
                result = unmark_duplicate_slip(DB_PATH, slip_id_in, bot_key_in)
            except Exception as exc:
                logger.exception("api_duplicate_unmark failed")
                record_endpoint_mutation(DB_PATH, "unmark_dup", actor=self.actor_fingerprint(), request_id=request_id, bot_key=bot_key_in, slip_id=slip_id_in, payload=payload, result_status="error", result_summary=safe_error(exc))
                self.send_json({"ok": False, "error": safe_error(exc), "request_id": request_id}, 400)
                return
            record_endpoint_mutation(DB_PATH, "unmark_dup", actor=self.actor_fingerprint(), request_id=request_id, bot_key=bot_key_in, slip_id=slip_id_in, payload=payload, result_status=("ok" if result.get("ok") else "error"), result_summary=str(result.get("previous_duplicate_of") or result.get("error") or ""))
            if isinstance(result, dict):
                result["request_id"] = request_id
            self.send_json(result, 200 if result.get("ok") else 400)
            return
        if parsed.path == "/api/slip/delete":
            if not self.role_or_401("admin"):
                return
            actor_fp = self.actor_fingerprint()
            approval_mode = self.parse_approval_param()
            if approval_mode == "request":
                request_id = uuid.uuid4().hex[:12]
                pending_result = request_pending_action_once(
                    DB_PATH,
                    action="slip.delete",
                    payload=payload,
                    requested_by=actor_fp,
                    request_id=request_id,
                )
                slip_id_in = str(payload.get("id") or payload.get("slip_id") or "")
                record_endpoint_mutation(
                    DB_PATH,
                    "delete.request",
                    actor=actor_fp,
                    request_id=str(pending_result.get("request_id") or request_id),
                    bot_key=str(payload.get("bot_key") or ""),
                    slip_id=slip_id_in,
                    payload=payload,
                    result_status="pending",
                    result_summary=f"pending_id={pending_result.get('pending_id')} already_pending={pending_result.get('already_pending')}",
                )
                auto_result = simple_auto_execute_pending(DB_PATH, int(pending_result.get("pending_id") or 0), actor_fp, self.actor_role()) if not pending_result.get("already_pending") else {}
                if auto_result:
                    status_code = int(auto_result.pop("status_code", 200)) if not auto_result.get("ok") else 200
                    self.send_json(auto_result, status_code)
                    return
                self.send_json(pending_result)
                return
            # approval_mode == "execute"
            try:
                pending_id_executed = int(payload.get("pending_id") or 0)
            except (TypeError, ValueError):
                pending_id_executed = 0
            if not pending_id_executed:
                self.send_json({"ok": False, "error": "pending_id required for execute"}, 400)
                return
            expire_old_pending_actions(DB_PATH)
            pending_row = load_pending_action(DB_PATH, pending_id_executed)
            if not pending_row or pending_row.get("action") != "slip.delete":
                self.send_json({"ok": False, "error": "pending action not found"}, 404)
                return
            if pending_row.get("status") != "approved":
                self.send_json({"ok": False, "error": f"not approved (status={pending_row.get('status')})"}, 409)
                return
            stored_payload = pending_action_payload(pending_row)
            request_id = str(pending_row.get("request_id") or uuid.uuid4().hex[:12])
            slip_id_in = str(stored_payload.get("id") or stored_payload.get("slip_id") or "")
            bot_key_in = str(stored_payload.get("bot_key") or "")
            try:
                result = delete_dashboard_slip(
                    DB_PATH,
                    slip_id_in,
                    bot_key_in,
                    str(stored_payload.get("reason") or "dashboard operator delete"),
                )
            except Exception as exc:
                logger.exception("api_slip_delete failed")
                record_endpoint_mutation(DB_PATH, "delete", actor=actor_fp, request_id=request_id, bot_key=bot_key_in, slip_id=slip_id_in, payload=stored_payload, result_status="error", result_summary=safe_error(exc))
                self.send_json({"ok": False, "error": safe_error(exc), "request_id": request_id, "pending_id": pending_id_executed}, 400)
                return
            record_endpoint_mutation(DB_PATH, "delete", actor=actor_fp, request_id=request_id, bot_key=bot_key_in, slip_id=slip_id_in, chat_id=str(result.get("chat_id") or ""), payload=stored_payload, result_status=("ok" if result.get("ok") else "error"), result_summary=str(result.get("previous_status") or result.get("error") or ""))
            if result.get("ok"):
                mark_pending_executed(DB_PATH, pending_id_executed, executed_result=str(result.get("previous_status") or ""))
            if isinstance(result, dict):
                result["request_id"] = request_id
                result["pending_id"] = pending_id_executed
            self.send_json(result, 200 if result.get("ok") else 400)
            return
        if parsed.path == "/api/slip/reprocess":
            if not self.role_or_401("admin", "operator"):
                return
            request_id = uuid.uuid4().hex[:12]
            slip_id_in = str(payload.get("id") or payload.get("slip_id") or "")
            bot_key_in = str(payload.get("bot_key") or "")
            try:
                result = reprocess_dashboard_slip(
                    DB_PATH,
                    slip_id_in,
                    bot_key_in,
                )
                record_endpoint_mutation(DB_PATH, "reprocess", actor=self.actor_fingerprint(), request_id=request_id, bot_key=bot_key_in, slip_id=slip_id_in, payload=payload, result_status=("ok" if result.get("ok") else "error"), result_summary=str(result.get("status") or result.get("error") or ""))
                if isinstance(result, dict):
                    result["request_id"] = request_id
                self.send_json(result, 200 if result.get("ok") else 400)
            except Exception as exc:
                logger.exception("api_slip_reprocess failed")
                record_endpoint_mutation(DB_PATH, "reprocess", actor=self.actor_fingerprint(), request_id=request_id, bot_key=bot_key_in, slip_id=slip_id_in, payload=payload, result_status="error", result_summary=safe_error(exc))
                self.send_json({"ok": False, "error": safe_error(exc), "request_id": request_id}, 400)
            return
        if parsed.path == "/api/slip/mark-reviewed":
            if not self.role_or_401("admin", "operator", "auditor"):
                return
            slip_id_in = str(payload.get("id") or payload.get("slip_id") or "")
            note = str(payload.get("note") or "")
            try:
                result = mark_slip_reviewed(DB_PATH, slip_id_in, reviewed_by=self.actor_fingerprint(), note=note)
                record_endpoint_mutation(DB_PATH, "mark_reviewed", actor=self.actor_fingerprint(), slip_id=slip_id_in, payload=payload, result_status=("ok" if result.get("ok") else "error"), result_summary=str(result.get("error") or ""))
                self.send_json(result, 200 if result.get("ok") else 400)
            except Exception as exc:
                logger.exception("api_slip_mark_reviewed failed")
                self.send_json({"ok": False, "error": safe_error(exc)}, 400)
            return
        if parsed.path == "/api/bank-review/openai":
            if not self.role_or_401("admin", "operator"):
                return
            slip_id_in = str(payload.get("id") or payload.get("slip_id") or "")
            try:
                result = openai_bank_double_check_slip(
                    DB_PATH,
                    slip_id_in,
                    apply=str(payload.get("apply", "true")).strip().lower() not in {"0", "false", "no"},
                )
                record_mutation(DB_PATH, "bank_review", actor=self.actor_fingerprint(), slip_id=slip_id_in, payload=payload, result_status=("ok" if result.get("ok") else "error"), result_summary=json.dumps(result.get("applied") or result.get("error") or "", ensure_ascii=False, default=str)[:480])
                self.send_json(result, 200 if result.get("ok") else 400)
            except Exception as exc:
                logger.exception("api_bank_review_openai failed")
                record_mutation(DB_PATH, "bank_review", actor=self.actor_fingerprint(), slip_id=slip_id_in, payload=payload, result_status="error", result_summary=safe_error(exc))
                self.send_json({"ok": False, "error": safe_error(exc)}, 400)
            return
        if parsed.path == "/api/bank-review/openai-all":
            if not self.role_or_401("admin"):
                return
            try:
                q = parse_qs(parsed.query)
                chat_id_b = str(payload.get("chat_id") or (q.get("chat_id") or [""])[0] or "")
                bot_key_b = str(payload.get("bot_key") or (q.get("bot_key") or [""])[0] or "")
                result = openai_bank_recheck_scope(
                    DB_PATH,
                    chat_id=chat_id_b,
                    bot_key=bot_key_b,
                    scope=str(payload.get("scope") or (q.get("scope") or ["open"])[0] or "open"),
                    flow_type=str(payload.get("flow_type") or (q.get("flow_type") or ["all"])[0] or "all"),
                    search=str(payload.get("slip_search") or (q.get("slip_search") or [""])[0] or ""),
                    apply=str(payload.get("apply", "true")).strip().lower() not in {"0", "false", "no"},
                )
                ok_count = result.get("ok_count") or 0
                fail_count = result.get("fail_count") or 0
                status = "ok" if (result.get("ok") and fail_count == 0) else ("partial" if ok_count > 0 else "error")
                record_mutation(DB_PATH, "bank_review_batch", actor=self.actor_fingerprint(), chat_id=chat_id_b, bot_key=bot_key_b, payload=payload, result_status=status, result_summary=f"ok={ok_count} fail={fail_count} total={result.get('total_count') or 0}")
                self.send_json(result, 200 if result.get("ok") else 207)
            except Exception as exc:
                logger.exception("api_bank_review_openai_all failed")
                record_mutation(DB_PATH, "bank_review_batch", actor=self.actor_fingerprint(), payload=payload, result_status="error", result_summary=safe_error(exc))
                self.send_json({"ok": False, "error": safe_error(exc)}, 400)
            return
        if parsed.path == "/api/ledger/import":
            if not self.role_or_401("admin"):
                return
            actor_fp = self.actor_fingerprint()
            approval_mode = self.parse_approval_param()
            if approval_mode == "request":
                req_payload = dict(payload) if isinstance(payload, dict) else {}
                if uploaded_statement_path and not req_payload.get("statement_path"):
                    req_payload["statement_path"] = str(uploaded_statement_path)
                statement_path = safe_statement_file_path(str(req_payload.get("statement_path") or ""))
                if not statement_path.exists():
                    self.send_json({"ok": False, "error": f"statement not found: {statement_path}"}, 404)
                    return
                request_id = uuid.uuid4().hex[:12]
                pending_id = create_pending_action(
                    DB_PATH,
                    action="ledger.import",
                    payload=req_payload,
                    requested_by=actor_fp,
                    request_id=request_id,
                )
                record_endpoint_mutation(
                    DB_PATH,
                    "ledger_import.request",
                    actor=actor_fp,
                    request_id=request_id,
                    bot_key=str(req_payload.get("bot_key") or ""),
                    payload=req_payload,
                    result_status="pending",
                    result_summary=f"pending_id={pending_id}",
                )
                auto_result = simple_auto_execute_pending(DB_PATH, pending_id, actor_fp, self.actor_role())
                if auto_result:
                    status_code = int(auto_result.pop("status_code", 200)) if not auto_result.get("ok") else 200
                    self.send_json(auto_result, status_code)
                    return
                self.send_json({"ok": True, "status": "pending", "pending_id": pending_id, "request_id": request_id, "expires_in_hours": PENDING_ACTION_TTL_HOURS, "action": "ledger.import"})
                return
            try:
                pending_id_executed = int(payload.get("pending_id") or 0)
            except (TypeError, ValueError):
                pending_id_executed = 0
            if not pending_id_executed:
                self.send_json({"ok": False, "error": "pending_id required for execute"}, 400)
                return
            row = load_pending_action(DB_PATH, pending_id_executed)
            if not row or row.get("action") != "ledger.import":
                self.send_json({"ok": False, "error": "pending action not found"}, 404)
                return
            result = execute_pending_action(DB_PATH, pending_id_executed, actor_fp)
            status_code = int(result.pop("status_code", 200)) if not result.get("ok") else 200
            self.send_json(result, status_code)
            return
        if parsed.path == "/api/ledger/preview":
            if not self.role_or_401("admin"):
                return
            try:
                req_payload = dict(payload) if isinstance(payload, dict) else {}
                if uploaded_statement_path and not req_payload.get("statement_path"):
                    req_payload["statement_path"] = str(uploaded_statement_path)
                statement_path = safe_statement_file_path(str(req_payload.get("statement_path") or ""))
                if not statement_path.exists():
                    self.send_json({"ok": False, "error": f"statement not found: {statement_path}", "dry_run": True}, 404)
                    return
                result = preview_bank_ledger_import(
                    DB_PATH,
                    statement_path,
                    bot_key=str(req_payload.get("bot_key") or ""),
                    company_name=str(req_payload.get("company_name") or ""),
                    bank=str(req_payload.get("bank") or ""),
                    account_no=str(req_payload.get("account_no") or ""),
                    account_name=str(req_payload.get("account_name") or ""),
                    flow_type=str(req_payload.get("flow_type") or "all"),
                    scope=str(req_payload.get("scope") or "all"),
                )
                result["approval_required"] = False
                self.send_json(result, 200 if result.get("ok") else 400)
            except Exception as exc:
                logger.exception("api_ledger_preview failed")
                self.send_json({"ok": False, "error": safe_error(exc), "dry_run": True}, 400)
            return
        if parsed.path == "/api/reconcile/statement":
            if not self.role_or_401("admin"):
                return
            try:
                req_payload = dict(payload) if isinstance(payload, dict) else {}
                if uploaded_excel_path and not req_payload.get("excel_path"):
                    req_payload["excel_path"] = str(uploaded_excel_path)
                if uploaded_statement_path and not req_payload.get("statement_path"):
                    req_payload["statement_path"] = str(uploaded_statement_path)
                chat_id = str(req_payload.get("chat_id") or "")
                bot_key = str(req_payload.get("bot_key") or "")
                flow_type = str(req_payload.get("flow_type") or "all")
                scope = str(req_payload.get("scope") or "all")
                excel_path = safe_backend_excel_path(str(req_payload.get("excel_path") or ""))
                statement_path = safe_statement_file_path(str(req_payload.get("statement_path") or ""))
                if not excel_path.exists():
                    self.send_json({"ok": False, "error": f"excel not found: {excel_path}"}, 404)
                    return
                if not statement_path.exists():
                    self.send_json({"ok": False, "error": f"statement not found: {statement_path}"}, 404)
                    return
                result = reconcile_backend_slips_statement(DB_PATH, excel_path, statement_path, chat_id=chat_id, scope=scope, bot_key=bot_key, flow_type=flow_type)
                result["approval_required"] = False
                self.send_json(result, 200 if result.get("ok") else 400)
            except Exception as exc:
                logger.exception("api_reconcile_statement failed")
                self.send_json({"ok": False, "error": safe_error(exc)}, 400)
            return
        if parsed.path in {"/api/reconcile", "/api/reconcile/preview"}:
            if not self.role_or_401("admin"):
                return
            actor_fp = self.actor_fingerprint()
            dry_run = parsed.path == "/api/reconcile/preview" or clean_display((parse_qs(parsed.query).get("dry_run") or parse_qs(parsed.query).get("preview") or [payload.get("dry_run") or payload.get("preview") or ""])[0]).lower() in {"1", "true", "yes", "y", "on"}
            if dry_run:
                try:
                    req_payload = dict(payload) if isinstance(payload, dict) else {}
                    if uploaded_excel_path and not req_payload.get("excel_path"):
                        req_payload["excel_path"] = str(uploaded_excel_path)
                    chat_id = str(req_payload.get("chat_id") or "")
                    bot_key = str(req_payload.get("bot_key") or "")
                    flow_type = str(req_payload.get("flow_type") or "all")
                    scope = str(req_payload.get("scope") or "all")
                    excel_path = safe_backend_excel_path(str(req_payload.get("excel_path") or ""))
                    if not excel_path.exists():
                        self.send_json({"ok": False, "error": f"excel not found: {excel_path}", "dry_run": True}, 404)
                        return
                    result = reconcile_backend_excel(DB_PATH, excel_path, chat_id=chat_id, scope=scope, bot_key=bot_key, flow_type=flow_type)
                    result["dry_run"] = True
                    result["approval_required"] = False
                    self.send_json(result, 200 if result.get("ok") else 400)
                except Exception as exc:
                    logger.exception("api_reconcile_preview failed")
                    self.send_json({"ok": False, "error": safe_error(exc), "dry_run": True}, 400)
                return
            approval_mode = self.parse_approval_param()
            if approval_mode == "request":
                request_id = uuid.uuid4().hex[:12]
                # Capture excel_path resolution into the stored payload so that the eventual
                # execute branch reconciles against the same file. uploaded_excel_path may be a
                # temp file written by this request; we persist its string form.
                req_payload = dict(payload) if isinstance(payload, dict) else {}
                if uploaded_excel_path and not req_payload.get("excel_path"):
                    req_payload["excel_path"] = str(uploaded_excel_path)
                pending_id = create_pending_action(
                    DB_PATH,
                    action="reconcile.run",
                    payload=req_payload,
                    requested_by=actor_fp,
                    request_id=request_id,
                )
                record_endpoint_mutation(DB_PATH, "reconcile.request", actor=actor_fp, request_id=request_id, chat_id=str(req_payload.get("chat_id") or ""), bot_key=str(req_payload.get("bot_key") or ""), payload=req_payload, result_status="pending", result_summary=f"pending_id={pending_id}")
                auto_result = simple_auto_execute_pending(DB_PATH, pending_id, actor_fp, self.actor_role())
                if auto_result:
                    status_code = int(auto_result.pop("status_code", 200)) if not auto_result.get("ok") else 200
                    self.send_json(auto_result, status_code)
                    return
                self.send_json({"ok": True, "status": "pending", "pending_id": pending_id, "request_id": request_id, "expires_in_hours": PENDING_ACTION_TTL_HOURS})
                return
            # approval_mode == "execute"
            try:
                pending_id_executed = int(payload.get("pending_id") or 0)
            except (TypeError, ValueError):
                pending_id_executed = 0
            if not pending_id_executed:
                self.send_json({"ok": False, "error": "pending_id required for execute"}, 400)
                return
            expire_old_pending_actions(DB_PATH)
            pending_row = load_pending_action(DB_PATH, pending_id_executed)
            if not pending_row or pending_row.get("action") != "reconcile.run":
                self.send_json({"ok": False, "error": "pending action not found"}, 404)
                return
            if pending_row.get("status") != "approved":
                self.send_json({"ok": False, "error": f"not approved (status={pending_row.get('status')})"}, 409)
                return
            # Use the originally-requested payload so an attacker re-posting the execute
            # request cannot swap arguments after approval.
            payload = pending_action_payload(pending_row)
            request_id = str(pending_row.get("request_id") or uuid.uuid4().hex[:12])
            try:
                chat_id = str(payload.get("chat_id") or "")
                bot_key = str(payload.get("bot_key") or "")
                flow_type = str(payload.get("flow_type") or "all")
                scope = str(payload.get("scope") or "all")
                excel_path = safe_backend_excel_path(str(payload.get("excel_path") or ""))
                if not excel_path.exists():
                    record_endpoint_mutation(DB_PATH, "reconcile", actor=actor_fp, request_id=request_id, chat_id=chat_id, bot_key=bot_key, payload=payload, result_status="error", result_summary="excel not found")
                    self.send_json({"ok": False, "error": f"excel not found: {excel_path}", "request_id": request_id, "pending_id": pending_id_executed}, 404)
                    return
                result = reconcile_backend_excel(DB_PATH, excel_path, chat_id=chat_id, scope=scope, bot_key=bot_key, flow_type=flow_type)
                summary = f"diff={result.get('diff_amount')} matched={result.get('matched', {}).get('count')} missing={result.get('missing', {}).get('count')} extra={result.get('extra', {}).get('count')}"
                record_endpoint_mutation(DB_PATH, "reconcile", actor=actor_fp, request_id=request_id, chat_id=chat_id, bot_key=bot_key, payload=payload, result_status=("ok" if result.get("ok") else "error"), result_summary=summary)
                if result.get("ok"):
                    mark_pending_executed(DB_PATH, pending_id_executed, executed_result=summary)
                if isinstance(result, dict):
                    result["request_id"] = request_id
                    result["pending_id"] = pending_id_executed
                self.send_json(result)
            except Exception as exc:
                logger.exception("api_reconcile failed")
                record_endpoint_mutation(DB_PATH, "reconcile", actor=actor_fp, request_id=request_id, payload=payload, result_status="error", result_summary=safe_error(exc))
                self.send_json({"ok": False, "error": safe_error(exc), "request_id": request_id, "pending_id": pending_id_executed}, 400)
            return
        if parsed.path == "/api/tokens/create":
            if not self.role_or_401("admin"):
                return
            request_id = uuid.uuid4().hex[:12]
            # Redact role/label only — never echo or log the raw token.
            mutation_payload = {"role": str(payload.get("role") or ""), "label": str(payload.get("label") or "")}
            try:
                result = create_dashboard_token(
                    DB_PATH,
                    str(payload.get("role") or ""),
                    str(payload.get("label") or ""),
                )
                token_hash_prefix = ""
                if isinstance(result, dict) and result.get("ok"):
                    token_hash_prefix = str(result.get("token_hash_prefix") or "")
                record_endpoint_mutation(DB_PATH, "token.create", actor=self.actor_fingerprint(), request_id=request_id, payload=mutation_payload, result_status=("ok" if (isinstance(result, dict) and result.get("ok")) else "error"), result_summary=token_hash_prefix or (isinstance(result, dict) and str(result.get("error") or "")) or "")
                if isinstance(result, dict):
                    result["request_id"] = request_id
                self.send_json(result, 200 if (isinstance(result, dict) and result.get("ok")) else 400)
            except Exception as exc:
                logger.exception("api_tokens_create failed")
                record_endpoint_mutation(DB_PATH, "token.create", actor=self.actor_fingerprint(), request_id=request_id, payload=mutation_payload, result_status="error", result_summary=safe_error(exc))
                self.send_json({"ok": False, "error": safe_error(exc), "request_id": request_id}, 400)
            return
        if parsed.path == "/api/tokens/revoke":
            if not self.role_or_401("admin"):
                return
            request_id = uuid.uuid4().hex[:12]
            token_prefix = str(payload.get("token_hash_prefix") or "")
            mutation_payload = {"token_hash_prefix": token_prefix}
            try:
                result = revoke_dashboard_token(DB_PATH, token_prefix)
                record_endpoint_mutation(DB_PATH, "token.revoke", actor=self.actor_fingerprint(), request_id=request_id, payload=mutation_payload, result_status=("ok" if (isinstance(result, dict) and result.get("ok")) else "error"), result_summary=token_prefix + " " + (isinstance(result, dict) and str(result.get("error") or "")))
                if isinstance(result, dict):
                    result["request_id"] = request_id
                self.send_json(result, 200 if (isinstance(result, dict) and result.get("ok")) else 400)
            except Exception as exc:
                logger.exception("api_tokens_revoke failed")
                record_endpoint_mutation(DB_PATH, "token.revoke", actor=self.actor_fingerprint(), request_id=request_id, payload=mutation_payload, result_status="error", result_summary=safe_error(exc))
                self.send_json({"ok": False, "error": safe_error(exc), "request_id": request_id}, 400)
            return
        if parsed.path == "/api/pending/approve":
            if not self.role_or_401("admin", "auditor"):
                return
            actor_fp = self.actor_fingerprint()
            try:
                pending_id = int(payload.get("pending_id") or 0)
            except (TypeError, ValueError):
                pending_id = 0
            if not pending_id:
                self.send_json({"ok": False, "error": "pending_id required"}, 400)
                return
            approval_mode = self.parse_approval_param()
            if approval_mode == "execute":
                row = load_pending_action(DB_PATH, pending_id)
                if row and row.get("status") == "pending":
                    simple_ok = simple_approval_enabled(DB_PATH, actor_fp, self.actor_role())
                    approval = approve_pending_action(DB_PATH, pending_id, actor_fp, allow_self_approval=simple_ok)
                    status_code = int(approval.pop("status", 200)) if not approval.get("ok") else 200
                    record_endpoint_mutation(DB_PATH, "pending.approve", actor=actor_fp, request_id=str(pending_id), payload=payload, result_status=("ok" if approval.get("ok") else "error"), result_summary=str(approval.get("error") or approval.get("approved_at") or ""))
                    if not approval.get("ok"):
                        self.send_json(approval, status_code)
                        return
                result = execute_pending_action(DB_PATH, pending_id, actor_fp)
                status_code = int(result.pop("status_code", 200)) if not result.get("ok") else 200
                self.send_json(result, status_code)
                return
            result = approve_pending_action(DB_PATH, pending_id, actor_fp, allow_self_approval=simple_approval_enabled(DB_PATH, actor_fp, self.actor_role()))
            status_code = int(result.pop("status", 200)) if not result.get("ok") else 200
            record_endpoint_mutation(DB_PATH, "pending.approve", actor=actor_fp, request_id=str(pending_id), payload=payload, result_status=("ok" if result.get("ok") else "error"), result_summary=str(result.get("error") or result.get("approved_at") or ""))
            self.send_json(result, status_code)
            return
        if parsed.path == "/api/pending/reject":
            if not self.role_or_401("admin", "auditor"):
                return
            actor_fp = self.actor_fingerprint()
            try:
                pending_id = int(payload.get("pending_id") or 0)
            except (TypeError, ValueError):
                pending_id = 0
            if not pending_id:
                self.send_json({"ok": False, "error": "pending_id required"}, 400)
                return
            reason = str(payload.get("reason") or "")
            result = reject_pending_action(DB_PATH, pending_id, actor_fp, reason)
            status_code = int(result.pop("status", 200)) if not result.get("ok") else 200
            record_endpoint_mutation(DB_PATH, "pending.reject", actor=actor_fp, request_id=str(pending_id), payload=payload, result_status=("ok" if result.get("ok") else "error"), result_summary=str(result.get("error") or result.get("reason") or ""))
            self.send_json(result, status_code)
            return
        if parsed.path == "/api/pending/cancel":
            actor_fp = self.actor_fingerprint()
            try:
                pending_id = int(payload.get("pending_id") or 0)
            except (TypeError, ValueError):
                pending_id = 0
            if not pending_id:
                self.send_json({"ok": False, "error": "pending_id required"}, 400)
                return
            result = cancel_pending_action(DB_PATH, pending_id, actor_fp)
            status_code = int(result.pop("status", 200)) if not result.get("ok") else 200
            record_endpoint_mutation(DB_PATH, "pending.cancel", actor=actor_fp, request_id=str(pending_id), payload=payload, result_status=("ok" if result.get("ok") else "error"), result_summary=str(result.get("error") or ""))
            self.send_json(result, status_code)
            return
        self.send_json({"ok": False, "error": "not found"}, 404)




def configure_bank_ledger_component() -> None:
    bank_ledger_component.configure(
        DB_PATH=DB_PATH,
        clean_display=clean_display,
        clean_company_name=clean_company_name,
        display_bank=display_bank,
        account_key_for=account_key_for,
        normalize_match_date=normalize_match_date,
        normalize_match_text=normalize_match_text,
        normalize_flow_type=normalize_flow_type,
        flow_label=flow_label,
        sqlite_table_exists=sqlite_table_exists,
        rows_to_dicts=rows_to_dicts,
        scope_date_range=scope_date_range,
        scope_to_date=scope_to_date,
        connect=connect,
        parse_statement_file=parse_statement_file,
        filter_statement_reconcile_rows=filter_statement_reconcile_rows,
        slip_reconcile_rows=slip_reconcile_rows,
        amount_time_date_match=amount_time_date_match,
        reconcile_daily_summary=reconcile_daily_summary,
    )


configure_bank_ledger_component()

def main() -> None:
    if not DASHBOARD_TOKEN and not dashboard_owner_login_enabled():
        raise SystemExit("AUDITSLIP_DASHBOARD_OWNER_PASSWORD is required when no dashboard token is configured")
    try:
        with connect(DB_PATH) as conn:
            ensure_dashboard_performance_indexes(conn)
            conn.commit()
        ensure_dashboard_tokens_table(DB_PATH)
        if DASHBOARD_TOKEN:
            bootstrap_dashboard_admin_token(DB_PATH, DASHBOARD_TOKEN)
    except Exception:
        logger.exception("dashboard bootstrap failed")
    httpd = ThreadingHTTPServer((HOST, PORT), DashboardHandler)
    print(f"Auditslip dashboard listening on http://{HOST}:{PORT}", flush=True)
    httpd.serve_forever()


if __name__ == "__main__":
    main()
